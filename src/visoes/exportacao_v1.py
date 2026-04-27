"""
exportacao_v1.py — Exportação Excel V1-específica · A-V1 Fase 7 · D-175 · D-176.

Consome F-APRESENT integralmente. Estrutura de 6 abas (Mockup-V1 + S-V1 §2.11):
  1. Resumo Executivo     · bespoke · 9 seções
  2. Resumo por Agrupador · ListObject · CONDICIONAL (omitida se vazia)
  3. Mapa de Conciliação  · ListObject · 1 linha por registro
  4. Análise Analítica    · ListObject · expansão por campo (4 colunas × N campos)
  5. Ponte de Conciliação · bespoke vertical · 1 sub-Ponte por campo elegível
  6. Diagnóstico          · bespoke · 6 seções · ÚLTIMA aba (D-017)

Capability 11 D-205 (formato_adaptativo_por_unidade) usada via formatos.
Capability 1 (paletas) via apresentacao.paletas. Vocabulário user-facing
respeitando Bloco 1.1 V1 + Bloco 3 estendido (P-V1 §2.2 · §2.5 · §2.6).
"""
from __future__ import annotations

import os
import sys
import time
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

sys.path.insert(0, str(Path(__file__).parent.parent))

from apresentacao import (  # noqa: E402
    escrever_titulo_aba,
    formatar_moeda_br,
    formatar_percentual_br,
    gerar_nome_arquivo,
    obter_paleta,
)
from apresentacao.formatos import (  # noqa: E402
    FORMATO_CONTAGEM,
    FORMATO_MONETARIO_BR,
    FORMATO_PERCENTUAL,
    formatar_diferenca_por_unidade,
    formatar_valor_por_unidade,
    number_format_diferenca,
    number_format_valor,
)
from apresentacao.templates.familia_a._shared import (  # noqa: E402
    bordas_finas as _bordas_finas,
    mesclar_card as _mesclar_card,
    renderizar_cabecalho_secao as _renderizar_cabecalho_secao,
    renderizar_secao_como_tabela as _renderizar_secao_como_tabela,
)
from contratos import ConfigExportacao, ExportacaoResult  # noqa: E402
from visoes.visao_v1 import (  # noqa: E402
    CasoLogicoV1,
    ClassificacaoRegistroV1,
    ConciliacaoV1Result,
    StatusPonteV1,
    UnidadeCanonica,
)


# ===========================================================================
# Vocabulário user-facing V1 (P-V1 §2.2 · Bloco 3 estendido)
# ===========================================================================


def _label_classificacao_v1(
    classificacao: ClassificacaoRegistroV1,
    origem_ux: str,
    comparado_ux: str,
    rotulo_amig: bool,
) -> str:
    """Substituição dinâmica · P-V1 §2.2 · F-APRESENT capability 2 V1."""
    if classificacao == ClassificacaoRegistroV1.CONCILIADO:
        return "Conciliado"
    if classificacao == ClassificacaoRegistroV1.DIVERGENTE_VALOR:
        return "Divergente por valor"
    if classificacao == ClassificacaoRegistroV1.SO_ORIGEM:
        return f"Saiu do {origem_ux}" if rotulo_amig else "Só na Origem"
    if classificacao == ClassificacaoRegistroV1.SO_COMPARADO:
        return f"Apareceu no {comparado_ux}" if rotulo_amig else "Só no Comparado"
    if classificacao == ClassificacaoRegistroV1.DIVERGENCIA_DUPLICIDADE:
        return "Divergência por duplicidade"
    if classificacao == ClassificacaoRegistroV1.DIVERGENCIA_AMBIGUIDADE:
        return "Divergência por ambiguidade de match"
    return classificacao.value


def _label_caso_logico(caso: CasoLogicoV1) -> str:
    if caso == CasoLogicoV1.MESMA_ABA_EM_COLUNAS:
        return "Mesma aba · Origem e Comparado em colunas distintas"
    return "Abas distintas · match executado"


def _label_modo_match(modo: str) -> str:
    return {
        "EXATO": "Exato (igualdade total)",
        "CONTEM": "Contém",
        "INICIA_COM": "Inicia com",
        "TERMINA_COM": "Termina com",
    }.get(modo, modo)


def _label_status_ponte(status: StatusPonteV1) -> str:
    if status == StatusPonteV1.FECHA:
        return "✅ Ponte fecha em todos os campos comparados"
    return "⚠️ Ponte com resíduo · ver Aba 5 · Ponte de Conciliação"


# ===========================================================================
# Helpers visuais locais
# ===========================================================================


def _ajustar_larguras(ws: Worksheet, min_: int = 12, max_: int = 50) -> None:
    """Auto-fit de larguras (D-202 · ignora masters de merges multi-coluna)."""
    masters_merge: set = set()
    for mr in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = mr.bounds
        if max_col > min_col:
            masters_merge.add((min_row, min_col))

    larguras: Dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            if (cell.row, cell.column) in masters_merge:
                continue
            valor_str = str(cell.value)
            tamanho = max(
                (len(parte) for parte in valor_str.split("\n")),
                default=len(valor_str),
            )
            tamanho = min(max_, tamanho + 2)
            tamanho = max(min_, tamanho)
            col = cell.column
            if tamanho > larguras.get(col, 0):
                larguras[col] = tamanho
    for col, largura in larguras.items():
        ws.column_dimensions[get_column_letter(col)].width = largura


def _aplicar_format_valor(cell, unidade: UnidadeCanonica) -> None:
    """Aplica number_format adaptativo por unidade (capability 11)."""
    cell.number_format = number_format_valor(unidade.value)


def _aplicar_format_diferenca(cell, unidade: UnidadeCanonica) -> None:
    """Aplica number_format de diferença adaptativo por unidade."""
    cell.number_format = number_format_diferenca(unidade.value)


def _to_float(d: Optional[Decimal]) -> Optional[float]:
    if d is None:
        return None
    return float(d)


# ===========================================================================
# Aba 1 · Resumo Executivo · 9 seções (Mockup-V1 §3 · S-V1 §2.10)
# ===========================================================================


def _renderizar_aba_resumo_executivo_v1(
    wb: Workbook,
    v1: ConciliacaoV1Result,
    paleta,
    arquivo_nome: Optional[str],
    aba_consumida: Optional[str],
) -> None:
    """Aba 1 bespoke · 9 seções narrativas + leitura qualitativa."""
    ws = wb.create_sheet("Resumo Executivo")
    ws.sheet_view.showGridLines = False
    cr = v1.conciliacao_realizada
    origem_ux = cr.origem_ux or "Origem"
    comparado_ux = cr.comparado_ux or "Comparado"
    rotulo_amig = cr.rotulo_amigavel_declarado
    largura_util = 8

    # Título
    titulo = f"Conciliação de Bases · {origem_ux} × {comparado_ux}"
    escrever_titulo_aba(ws, 1, 1, largura_util, titulo, paleta)

    # Subtítulo (Mockup §3.2.1)
    timestamp = datetime.now().strftime("%d/%m/%Y às %H:%M")
    n_arq = cr.n_arquivos
    caso = cr.caso_logico_inferido
    if n_arq == 2 and caso == CasoLogicoV1.ABAS_DISTINTAS:
        subtitulo = (
            f"Gerado em {timestamp} · Origem: {cr.arquivo_origem} · {cr.aba_origem} · "
            f"Comparado: {cr.arquivo_comparado} · {cr.aba_comparado}"
        )
    elif n_arq == 1 and caso == CasoLogicoV1.ABAS_DISTINTAS:
        subtitulo = (
            f"Gerado em {timestamp} · Arquivo: {cr.arquivo_origem} · "
            f"Origem: {cr.aba_origem} · Comparado: {cr.aba_comparado}"
        )
    else:  # MESMA_ABA_EM_COLUNAS
        subtitulo = (
            f"Gerado em {timestamp} · Arquivo: {cr.arquivo_origem} · "
            f"Aba: {cr.aba_origem} · Origem e Comparado em colunas distintas"
        )
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=largura_util)
    cell_sub = ws.cell(row=2, column=1, value=subtitulo)
    cell_sub.font = Font(
        name=paleta.fonte_familia,
        size=paleta.fonte_tamanho_corpo,
        italic=True,
        color=paleta.cor_neutra_escura,
    )
    cell_sub.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    if v1.modelo_aplicado:
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=largura_util)
        ws.cell(
            row=3, column=1,
            value=f"Modelo de configuração: {v1.modelo_aplicado.nome_modelo}"
        ).font = Font(
            name=paleta.fonte_familia,
            size=paleta.fonte_tamanho_corpo,
            italic=True,
            color=paleta.cor_neutra_escura,
        )
        linha = 5
    else:
        linha = 4

    # § 2 · Taxa de Conciliação · KPI primário
    contagem = v1.contagem_por_classificacao
    n_total = sum(contagem.values()) or 1
    n_concil = contagem.get(ClassificacaoRegistroV1.CONCILIADO, 0)
    taxa = n_concil / n_total

    linha = _renderizar_cabecalho_secao(
        ws, linha, 1, largura_util, "Taxa de Conciliação", paleta,
    )
    # Card centralizado · 3 linhas merge
    _mesclar_card(
        ws,
        linha,
        1,
        linha + 2,
        largura_util,
        "Taxa de Conciliação Geral",
        formatar_percentual_br(taxa, conversao_fracao=True),
        paleta,
    )
    # Sub-texto manual abaixo do card
    ws.merge_cells(
        start_row=linha + 3, start_column=1, end_row=linha + 3, end_column=largura_util
    )
    sub = ws.cell(
        row=linha + 3, column=1,
        value=f"{n_concil:,} de {n_total:,} registros conciliados".replace(",", "."),
    )
    sub.font = Font(
        name=paleta.fonte_familia,
        size=paleta.fonte_tamanho_corpo - 1,
        italic=True,
        color=paleta.cor_neutra_escura,
    )
    sub.alignment = Alignment(horizontal="center", vertical="center")
    linha += 5
    # Tabela de decomposição (6 classes + total)
    rows_dec = []
    for cls in [
        ClassificacaoRegistroV1.CONCILIADO,
        ClassificacaoRegistroV1.DIVERGENTE_VALOR,
        ClassificacaoRegistroV1.SO_ORIGEM,
        ClassificacaoRegistroV1.SO_COMPARADO,
        ClassificacaoRegistroV1.DIVERGENCIA_DUPLICIDADE,
        ClassificacaoRegistroV1.DIVERGENCIA_AMBIGUIDADE,
    ]:
        n = contagem.get(cls, 0)
        pct = n / n_total
        rows_dec.append(
            (
                _label_classificacao_v1(cls, origem_ux, comparado_ux, rotulo_amig),
                f"{n:,}".replace(",", "."),
                formatar_percentual_br(pct, conversao_fracao=True),
            )
        )
    rows_dec.append(("Total processado", f"{n_total:,}".replace(",", "."), "100,00%"))
    linha = _escrever_minitabela(
        ws, linha, 1, largura_util,
        cabecalho=("Classificação", "N registros", "% do total"),
        linhas=rows_dec,
        paleta=paleta,
    )
    if v1.sintese_diagnostico.n_tolerancia_absorvida > 0:
        ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=largura_util)
        ws.cell(
            row=linha, column=1,
            value=(
                f"Dos {n_concil:,} conciliados · "
                f"{v1.sintese_diagnostico.n_tolerancia_absorvida:,} tiveram diferença "
                f"absorvida pela tolerância "
                f"(soma {formatar_moeda_br(float(v1.sintese_diagnostico.valor_tolerancia_absorvida))})"
            ).replace(",", "."),
        ).font = Font(
            name=paleta.fonte_familia,
            size=paleta.fonte_tamanho_corpo - 1,
            italic=True,
            color=paleta.cor_neutra_escura,
        )
        linha += 2

    # § 3 · Volumetria
    linha = _renderizar_cabecalho_secao(ws, linha, 1, largura_util, "Volumetria", paleta)
    rows_vol = [
        (f"Registros · {origem_ux}", f"{cr.n_registros_origem:,}".replace(",", ".")),
        (f"Registros · {comparado_ux}", f"{cr.n_registros_comparado:,}".replace(",", ".")),
        ("Processados após match", f"{cr.n_processados:,}".replace(",", ".")),
    ]
    linha = _renderizar_secao_como_tabela(
        ws, linha, "(detalhes)", rows_vol, 1, largura_util, paleta,
    )

    # § 4 · Status da Ponte (NOVO V1)
    linha = _renderizar_cabecalho_secao(ws, linha, 1, largura_util, "Status da Ponte", paleta)
    label_ponte = _label_status_ponte(v1.status_ponte_geral)
    ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=largura_util)
    cell_p = ws.cell(row=linha, column=1, value=label_ponte)
    cell_p.font = Font(
        name=paleta.fonte_familia,
        size=paleta.fonte_tamanho_secao,
        bold=True,
        color=(
            paleta.cor_destaque
            if v1.status_ponte_geral == StatusPonteV1.FECHA
            else 'C00000'
        ),
    )
    cell_p.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[linha].height = 28
    linha += 2

    # § 5 · Valor financeiro por campo comparado
    linha = _renderizar_cabecalho_secao(
        ws, linha, 1, largura_util, "Valor por campo comparado", paleta,
    )
    rows_val = []
    for vpc in v1.valor_por_campo:
        unid = vpc.unidade
        rows_val.append(
            (
                vpc.nome_analitico,
                formatar_valor_por_unidade(_to_float(vpc.soma_origem), unid.value),
                formatar_valor_por_unidade(_to_float(vpc.soma_comparado), unid.value),
                formatar_diferenca_por_unidade(
                    _to_float(vpc.diferenca_liquida), unid.value
                ),
                formatar_valor_por_unidade(_to_float(vpc.sigma_diferenca), unid.value),
                (
                    "—"
                    if vpc.n_tolerancia_absorvida == 0
                    else f"{vpc.n_tolerancia_absorvida} reg · "
                    + formatar_valor_por_unidade(
                        _to_float(vpc.valor_tolerancia_absorvida), unid.value
                    )
                ),
            )
        )
    linha = _escrever_minitabela(
        ws, linha, 1, largura_util,
        cabecalho=(
            "Campo",
            f"Soma · {origem_ux}",
            f"Soma · {comparado_ux}",
            "Diferença líquida",
            "Σ |Diferença|",
            "Tolerância absorvida",
        ),
        linhas=rows_val,
        paleta=paleta,
    )

    # § 6 · Cobertura por base
    linha = _renderizar_cabecalho_secao(
        ws, linha, 1, largura_util, "Cobertura por base", paleta,
    )
    if v1.cobertura is None:
        ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=largura_util)
        ws.cell(
            row=linha, column=1,
            value="Cobertura 100% por construção (Caso: Mesma aba em colunas)",
        ).font = Font(
            name=paleta.fonte_familia, size=paleta.fonte_tamanho_corpo,
            italic=True, color=paleta.cor_neutra_escura,
        )
        linha += 2
    else:
        cob = v1.cobertura
        rows_cob = [
            (
                f"Cobertura · {origem_ux}",
                f"{cob.n_origem_com_par:,} de {cr.n_registros_origem:,} "
                f"({formatar_percentual_br(float(cob.cobertura_origem_pct), conversao_fracao=True)})"
                .replace(",", "."),
            ),
            (
                f"Cobertura · {comparado_ux}",
                f"{cob.n_comparado_com_par:,} de {cr.n_registros_comparado:,} "
                f"({formatar_percentual_br(float(cob.cobertura_comparado_pct), conversao_fracao=True)})"
                .replace(",", "."),
            ),
        ]
        linha = _renderizar_secao_como_tabela(
            ws, linha, "(detalhes)", rows_cob, 1, largura_util, paleta,
        )

    # § 7 · Resumo por agrupador executivo (CONDICIONAL)
    if v1.resumo_por_agrupador_executivo:
        agrups_exec = cr.agrupadores_resumo_executivo
        titulo_sec = (
            f"Resumo por {' × '.join(agrups_exec)}"
            if agrups_exec
            else "Resumo por agrupador"
        )
        linha = _renderizar_cabecalho_secao(ws, linha, 1, largura_util, titulo_sec, paleta)
        rows_res = []
        for lr in v1.resumo_por_agrupador_executivo[:30]:  # cap 30 linhas no Resumo
            valores = " · ".join(
                f"{k}={v}" for k, v in (lr.valores_agrupador or {}).items()
            )
            rows_res.append(
                (
                    valores or "(sem valor)",
                    str(lr.n_conciliados),
                    str(lr.n_divergentes_valor),
                    formatar_diferenca_por_unidade(
                        float(lr.diferenca_liquida_total), "MONETARIO_BRL"
                    ),
                )
            )
        linha = _escrever_minitabela(
            ws, linha, 1, largura_util,
            cabecalho=(
                " × ".join(agrups_exec) if agrups_exec else "Agrupador",
                "N Conciliados", "N Divergentes", "Diferença líquida",
            ),
            linhas=rows_res,
            paleta=paleta,
        )

    # § 8 · Síntese do Diagnóstico
    linha = _renderizar_cabecalho_secao(
        ws, linha, 1, largura_util, "Síntese do Diagnóstico", paleta,
    )
    sd = v1.sintese_diagnostico
    rows_sint = [
        (
            "Tolerância absorvida",
            f"{sd.n_tolerancia_absorvida} registros · "
            + formatar_moeda_br(float(sd.valor_tolerancia_absorvida)),
        ),
        (
            "Duplicidades detectadas",
            f"{sd.n_chaves_duplicadas} chaves afetando "
            f"{sd.n_registros_afetados_duplicidade} registros",
        ),
        (
            "Ambiguidades de match",
            f"{sd.n_chaves_ambiguas} chaves afetando "
            f"{sd.n_registros_afetados_ambiguidade} registros",
        ),
        ("Warnings ativos", str(sd.n_warnings_ativos)),
    ]
    linha = _renderizar_secao_como_tabela(
        ws, linha, "(detalhes)", rows_sint, 1, largura_util, paleta,
    )

    # § 9 · Configuração aplicada
    linha = _renderizar_cabecalho_secao(
        ws, linha, 1, largura_util, "Configuração aplicada", paleta,
    )
    ca = v1.config_aplicada
    rows_cfg = [
        (
            "Agrupadores de match",
            " · ".join(
                f"{a.rotulo_analitico} ({_label_modo_match(a.modo_match.value)})"
                for a in ca.agrupadores_match
            ) or "—",
        ),
    ]
    if ca.agrupadores_resumo_executivo:
        rows_cfg.append(
            ("Agrupadores do Resumo", " · ".join(ca.agrupadores_resumo_executivo))
        )
    for c in ca.campos_comparados:
        rows_cfg.append(
            (
                f"Campo: {c.nome_analitico}",
                f"tipo {c.tipo_logico.value} · unidade {c.unidade.value} · "
                f"tolerância {c.tolerancia}",
            )
        )
    rows_cfg.append(("Caso lógico", _label_caso_logico(ca.caso_logico_inferido)))
    rows_cfg.append(("Paleta executiva", ca.paleta_aplicada))
    linha = _renderizar_secao_como_tabela(
        ws, linha, "(detalhes)", rows_cfg, 1, largura_util, paleta,
    )

    # Bloco final · Leitura Qualitativa
    linha = _renderizar_cabecalho_secao(
        ws, linha, 1, largura_util, "Leitura Qualitativa", paleta,
    )
    if v1.leitura_qualitativa and v1.leitura_qualitativa.texto:
        ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=largura_util)
        cell_lq = ws.cell(row=linha, column=1, value=v1.leitura_qualitativa.texto)
        cell_lq.font = Font(
            name=paleta.fonte_familia,
            size=paleta.fonte_tamanho_corpo,
            color=paleta.cor_neutra_escura,
        )
        cell_lq.alignment = Alignment(
            horizontal="left", vertical="top", wrap_text=True, indent=1,
        )
        ws.row_dimensions[linha].height = 80

    _ajustar_larguras(ws, min_=14, max_=42)


def _escrever_minitabela(
    ws: Worksheet,
    linha: int,
    col_ini: int,
    col_fim: int,
    cabecalho: Tuple[str, ...],
    linhas: List[Tuple],
    paleta,
) -> int:
    """Mini-tabela manual (não-ListObject) · cabeçalho + linhas + zebra."""
    n_cols = len(cabecalho)
    bordas = _bordas_finas(paleta)
    # Cabeçalho
    for i, c in enumerate(cabecalho):
        cell = ws.cell(row=linha, column=col_ini + i, value=c)
        cell.fill = PatternFill("solid", fgColor=paleta.cor_primaria)
        cell.font = Font(
            name=paleta.fonte_familia,
            size=paleta.fonte_tamanho_corpo,
            bold=True,
            color="FFFFFF",
        )
        cell.alignment = Alignment(
            horizontal="center" if i > 0 else "left", vertical="center", indent=1,
        )
        cell.border = bordas
    # Preenche colunas vazias até col_fim
    for i in range(n_cols, col_fim - col_ini + 1):
        cell = ws.cell(row=linha, column=col_ini + i, value="")
        cell.fill = PatternFill("solid", fgColor=paleta.cor_primaria)
        cell.border = bordas
    linha += 1
    for idx_linha, tupla in enumerate(linhas):
        cor_fundo = paleta.cor_secundaria if idx_linha % 2 == 0 else "FFFFFF"
        for i, val in enumerate(tupla[:n_cols]):
            cell = ws.cell(row=linha, column=col_ini + i, value=val)
            cell.fill = PatternFill("solid", fgColor=cor_fundo)
            cell.font = Font(
                name=paleta.fonte_familia,
                size=paleta.fonte_tamanho_corpo,
                color=paleta.cor_neutra_escura,
            )
            cell.alignment = Alignment(
                horizontal="right" if i > 0 else "left",
                vertical="center", indent=1, wrap_text=True,
            )
            cell.border = bordas
        for i in range(n_cols, col_fim - col_ini + 1):
            cell = ws.cell(row=linha, column=col_ini + i, value="")
            cell.fill = PatternFill("solid", fgColor=cor_fundo)
            cell.border = bordas
        linha += 1
    ws.row_dimensions[linha].height = 8
    return linha + 1


# ===========================================================================
# Aba 2 · Resumo por Agrupador (CONDICIONAL · ListObject)
# ===========================================================================


def _renderizar_aba_resumo_agrupador_v1(
    wb: Workbook,
    v1: ConciliacaoV1Result,
    paleta,
) -> None:
    """Aba 2 · CONDICIONAL · só renderiza quando há linhas."""
    if not v1.resumo_por_agrupador_executivo:
        return
    ws = wb.create_sheet("Resumo por Agrupador")
    ws.sheet_view.showGridLines = False
    cr = v1.conciliacao_realizada
    origem_ux = cr.origem_ux or "Origem"
    comparado_ux = cr.comparado_ux or "Comparado"
    agrups_exec = cr.agrupadores_resumo_executivo
    titulo = (
        f"Resumo por {' × '.join(agrups_exec)}"
        if agrups_exec
        else "Resumo por agrupador"
    )
    escrever_titulo_aba(ws, 1, 1, 8, titulo, paleta)
    ws.row_dimensions[2].height = 8

    # Cabeçalho
    headers = list(agrups_exec) if agrups_exec else ["Agrupador"]
    headers += [
        "N Conciliados", "N Divergentes",
        f"N Saiu do {origem_ux}" if cr.rotulo_amigavel_declarado else "N Só Origem",
        f"N Apareceu no {comparado_ux}" if cr.rotulo_amigavel_declarado else "N Só Comparado",
    ]
    for c in cr.campos_comparados:
        nome = c.nome_analitico
        headers += [
            f"Soma {origem_ux} · {nome}",
            f"Soma {comparado_ux} · {nome}",
            f"Diferença líquida · {nome}",
            f"Σ |Diferença| · {nome}",
        ]

    linha = 3
    bordas = _bordas_finas(paleta)
    for j, h in enumerate(headers):
        cell = ws.cell(row=linha, column=j + 1, value=h)
        cell.fill = PatternFill("solid", fgColor=paleta.cor_primaria)
        cell.font = Font(name=paleta.fonte_familia, bold=True, color="FFFFFF",
                         size=paleta.fonte_tamanho_corpo)
        cell.alignment = Alignment(horizontal="center", vertical="center", indent=1)
        cell.border = bordas
    linha += 1

    n_id = len(agrups_exec) if agrups_exec else 1
    for idx_linha, lr in enumerate(v1.resumo_por_agrupador_executivo):
        cor_fundo = paleta.cor_secundaria if idx_linha % 2 == 0 else "FFFFFF"
        # Identificadores
        if agrups_exec:
            for j, ag_nome in enumerate(agrups_exec):
                val = lr.valores_agrupador.get(ag_nome, "")
                cell = ws.cell(row=linha, column=j + 1, value=val)
                cell.fill = PatternFill("solid", fgColor=cor_fundo)
                cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
                cell.border = bordas
        else:
            cell = ws.cell(row=linha, column=1, value="(todos)")
            cell.fill = PatternFill("solid", fgColor=cor_fundo)
            cell.border = bordas

        # Métricas estruturais
        col = n_id + 1
        for val in [lr.n_conciliados, lr.n_divergentes_valor, lr.n_so_origem, lr.n_so_comparado]:
            cell = ws.cell(row=linha, column=col, value=val)
            cell.fill = PatternFill("solid", fgColor=cor_fundo)
            cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)
            cell.border = bordas
            cell.number_format = FORMATO_CONTAGEM
            col += 1
        # Métricas por campo
        for m in lr.metricas_por_campo:
            for val_dec in [m.soma_origem, m.soma_comparado, m.diferenca_liquida, m.sigma_diferenca]:
                cell = ws.cell(row=linha, column=col, value=_to_float(val_dec))
                cell.fill = PatternFill("solid", fgColor=cor_fundo)
                cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)
                cell.border = bordas
                _aplicar_format_valor(cell, m.unidade)
                col += 1
        linha += 1

    _ajustar_larguras(ws, min_=12, max_=32)


# ===========================================================================
# Aba 3 · Mapa de Conciliação (ListObject · 1 linha por registro)
# ===========================================================================


def _renderizar_aba_mapa_conciliacao_v1(
    wb: Workbook,
    v1: ConciliacaoV1Result,
    paleta,
) -> None:
    """Aba 3 · todos os registros processados · 1 linha cada."""
    ws = wb.create_sheet("Mapa de Conciliação")
    ws.sheet_view.showGridLines = False
    cr = v1.conciliacao_realizada
    origem_ux = cr.origem_ux or "Origem"
    comparado_ux = cr.comparado_ux or "Comparado"
    rotulo_amig = cr.rotulo_amigavel_declarado
    escrever_titulo_aba(ws, 1, 1, 6, "Mapa de Conciliação", paleta)
    ws.row_dimensions[2].height = 8

    # Cabeçalho · agrupadores + Classificação + 3 métricas + Observações
    nomes_agrups = [a.rotulo_analitico for a in cr.agrupadores_match]
    headers = list(nomes_agrups) + [
        "Classificação",
        "Diferença total",
        "Σ |Diferença|",
        "Variação total %",
        "Observações",
    ]
    linha = 3
    bordas = _bordas_finas(paleta)
    for j, h in enumerate(headers):
        cell = ws.cell(row=linha, column=j + 1, value=h)
        cell.fill = PatternFill("solid", fgColor=paleta.cor_primaria)
        cell.font = Font(name=paleta.fonte_familia, bold=True, color="FFFFFF",
                         size=paleta.fonte_tamanho_corpo)
        cell.alignment = Alignment(horizontal="center", vertical="center", indent=1)
        cell.border = bordas
    linha += 1
    for idx_linha, reg in enumerate(v1.classificacao_por_registro):
        cor_fundo = paleta.cor_secundaria if idx_linha % 2 == 0 else "FFFFFF"
        col = 1
        for ag_nome in nomes_agrups:
            val = reg.valores_agrupadores.get(ag_nome, "")
            cell = ws.cell(row=linha, column=col, value=val)
            cell.fill = PatternFill("solid", fgColor=cor_fundo)
            cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            cell.border = bordas
            col += 1
        cell_class = ws.cell(
            row=linha, column=col,
            value=_label_classificacao_v1(
                reg.classificacao_estrutural, origem_ux, comparado_ux, rotulo_amig
            ),
        )
        cell_class.fill = PatternFill("solid", fgColor=cor_fundo)
        cell_class.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        cell_class.border = bordas
        col += 1
        for val in [reg.diferenca_total_registro, reg.sigma_diferenca_total_registro]:
            cell = ws.cell(row=linha, column=col, value=_to_float(val))
            cell.fill = PatternFill("solid", fgColor=cor_fundo)
            cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)
            cell.border = bordas
            # Default monetário (multi-campo agregado)
            cell.number_format = FORMATO_MONETARIO_BR
            col += 1
        cell = ws.cell(
            row=linha, column=col,
            value=_to_float(reg.variacao_total_registro_pct),
        )
        cell.fill = PatternFill("solid", fgColor=cor_fundo)
        cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        cell.border = bordas
        cell.number_format = FORMATO_PERCENTUAL
        col += 1
        cell = ws.cell(row=linha, column=col, value=reg.observacoes or "—")
        cell.fill = PatternFill("solid", fgColor=cor_fundo)
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        cell.border = bordas
        linha += 1

    _ajustar_larguras(ws, min_=12, max_=42)


# ===========================================================================
# Aba 4 · Análise Analítica (ListObject · expansão por campo)
# ===========================================================================


def _renderizar_aba_analise_analitica_v1(
    wb: Workbook,
    v1: ConciliacaoV1Result,
    paleta,
) -> None:
    """Aba 4 · 1 linha por registro · 4 colunas por campo (Valor O · Valor C · Diferença · Status)."""
    ws = wb.create_sheet("Análise Analítica")
    ws.sheet_view.showGridLines = False
    cr = v1.conciliacao_realizada
    origem_ux = cr.origem_ux or "Origem"
    comparado_ux = cr.comparado_ux or "Comparado"
    rotulo_amig = cr.rotulo_amigavel_declarado
    escrever_titulo_aba(ws, 1, 1, 8, "Análise Analítica", paleta)
    ws.row_dimensions[2].height = 8

    nomes_agrups = [a.rotulo_analitico for a in cr.agrupadores_match]
    campos = cr.campos_comparados
    headers = list(nomes_agrups) + ["Classificação"]
    for c in campos:
        headers += [
            f"Valor {origem_ux} · {c.nome_analitico}",
            f"Valor {comparado_ux} · {c.nome_analitico}",
            f"Diferença · {c.nome_analitico}",
            f"Status · {c.nome_analitico}",
        ]

    linha = 3
    bordas = _bordas_finas(paleta)
    for j, h in enumerate(headers):
        cell = ws.cell(row=linha, column=j + 1, value=h)
        cell.fill = PatternFill("solid", fgColor=paleta.cor_primaria)
        cell.font = Font(name=paleta.fonte_familia, bold=True, color="FFFFFF",
                         size=paleta.fonte_tamanho_corpo)
        cell.alignment = Alignment(horizontal="center", vertical="center", indent=1)
        cell.border = bordas
    linha += 1
    for idx_linha, reg in enumerate(v1.classificacao_por_registro):
        cor_fundo = paleta.cor_secundaria if idx_linha % 2 == 0 else "FFFFFF"
        col = 1
        for ag_nome in nomes_agrups:
            cell = ws.cell(row=linha, column=col, value=reg.valores_agrupadores.get(ag_nome, ""))
            cell.fill = PatternFill("solid", fgColor=cor_fundo)
            cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            cell.border = bordas
            col += 1
        cell_class = ws.cell(
            row=linha, column=col,
            value=_label_classificacao_v1(
                reg.classificacao_estrutural, origem_ux, comparado_ux, rotulo_amig
            ),
        )
        cell_class.fill = PatternFill("solid", fgColor=cor_fundo)
        cell_class.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        cell_class.border = bordas
        col += 1
        for i_c, c in enumerate(campos):
            cel_data = (
                reg.valores_por_campo[i_c]
                if i_c < len(reg.valores_por_campo)
                else None
            )
            valores = (
                [
                    _to_float(cel_data.valor_origem),
                    _to_float(cel_data.valor_comparado),
                    _to_float(cel_data.diferenca),
                    cel_data.status_campo.value if cel_data else "—",
                ]
                if cel_data
                else [None, None, None, "—"]
            )
            for k, v in enumerate(valores):
                cell = ws.cell(row=linha, column=col, value=v)
                cell.fill = PatternFill("solid", fgColor=cor_fundo)
                cell.border = bordas
                if k < 2:
                    cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)
                    _aplicar_format_valor(cell, c.unidade)
                elif k == 2:
                    cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)
                    _aplicar_format_diferenca(cell, c.unidade)
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
                col += 1
        linha += 1
    _ajustar_larguras(ws, min_=12, max_=36)


# ===========================================================================
# Aba 5 · Ponte de Conciliação (bespoke · 1 sub-Ponte por campo elegível)
# ===========================================================================


def _renderizar_aba_ponte_conciliacao_v1(
    wb: Workbook,
    v1: ConciliacaoV1Result,
    paleta,
) -> None:
    """Aba 5 · decomposição matemática vertical."""
    ws = wb.create_sheet("Ponte de Conciliação")
    ws.sheet_view.showGridLines = False
    cr = v1.conciliacao_realizada
    origem_ux = cr.origem_ux or "Origem"
    comparado_ux = cr.comparado_ux or "Comparado"
    escrever_titulo_aba(ws, 1, 1, 6, "Ponte de Conciliação", paleta)
    ws.row_dimensions[2].height = 8

    linha = 3

    # Nota global · campos PERCENTUAL/ADIMENSIONAL/RAZAO omitidos (Q1.B · D-210)
    campos_omitidos = [
        c.nome_analitico
        for c in cr.campos_comparados
        if c.unidade.value in ("PERCENTUAL", "ADIMENSIONAL", "RAZAO")
    ]
    if campos_omitidos:
        ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=6)
        ws.cell(
            row=linha, column=1,
            value=(
                "Nota: campos com unidade PERCENTUAL/ADIMENSIONAL/RAZAO são omitidos "
                f"da Ponte (Q1.B · D-210). Omitidos: {', '.join(campos_omitidos)}"
            ),
        ).font = Font(
            name=paleta.fonte_familia,
            size=paleta.fonte_tamanho_corpo - 1,
            italic=True,
            color=paleta.cor_neutra_escura,
        )
        linha += 2

    if not v1.pontes:
        ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=6)
        ws.cell(
            row=linha, column=1,
            value="Nenhuma ponte aplicável neste resultado (todos os campos omitidos).",
        ).font = Font(
            name=paleta.fonte_familia, italic=True,
            color=paleta.cor_neutra_escura,
        )
        return

    for ponte in v1.pontes:
        unid = ponte.unidade
        titulo_ponte = f"{ponte.nome_analitico}  ·  {unid.value}"
        linha = _renderizar_cabecalho_secao(ws, linha, 1, 6, titulo_ponte, paleta)

        # 7 linhas de decomposição
        rows = [
            (f"Soma · {origem_ux}", _to_float(ponte.saldo_origem)),
            (f"(−) Registros Só em {origem_ux}", _to_float(ponte.ajuste_so_origem)),
            ("(−) Diferença líquida em Divergentes por valor",
             _to_float(ponte.ajuste_divergentes_valor)),
            ("(−) Tolerância absorvida", _to_float(ponte.ajuste_tolerancia_absorvida)),
            (f"(+) Registros Só em {comparado_ux}", _to_float(ponte.ajuste_so_comparado)),
            (f"= Soma · {comparado_ux} (calculado)",
             _to_float(ponte.saldo_comparado_esperado)),
            (f"Soma · {comparado_ux} (real)", _to_float(ponte.saldo_comparado_real)),
        ]
        bordas = _bordas_finas(paleta)
        for idx, (rotulo, val) in enumerate(rows):
            cor_fundo = paleta.cor_secundaria if idx % 2 == 0 else "FFFFFF"
            ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=4)
            c_r = ws.cell(row=linha, column=1, value=rotulo)
            c_r.fill = PatternFill("solid", fgColor=cor_fundo)
            c_r.font = Font(
                name=paleta.fonte_familia,
                size=paleta.fonte_tamanho_corpo,
                bold=(idx >= 5),
                color=paleta.cor_neutra_escura,
            )
            c_r.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            c_v = ws.cell(row=linha, column=5, value=val)
            c_v.fill = PatternFill("solid", fgColor=cor_fundo)
            c_v.font = Font(
                name=paleta.fonte_familia,
                size=paleta.fonte_tamanho_corpo,
                bold=(idx >= 5),
                color=paleta.cor_neutra_escura,
            )
            c_v.alignment = Alignment(horizontal="right", vertical="center", indent=1)
            _aplicar_format_valor(c_v, unid)
            for col in range(1, 7):
                ws.cell(row=linha, column=col).border = bordas
            linha += 1

        # Verificação · resíduo + status
        cor_status = (
            paleta.cor_destaque if ponte.fecha else 'C00000'
        )
        ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=4)
        c_v_res = ws.cell(
            row=linha, column=1,
            value="✅ Fecha (resíduo absorvido pela tolerância)" if ponte.fecha
            else f"⚠️ Resíduo não atribuído",
        )
        c_v_res.font = Font(
            name=paleta.fonte_familia,
            size=paleta.fonte_tamanho_corpo,
            bold=True,
            color=cor_status,
        )
        c_v_res.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        cell_res_val = ws.cell(row=linha, column=5, value=_to_float(ponte.residuo))
        cell_res_val.font = Font(
            name=paleta.fonte_familia, bold=True, color=cor_status,
        )
        cell_res_val.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        _aplicar_format_valor(cell_res_val, unid)
        linha += 2

    _ajustar_larguras(ws, min_=14, max_=44)


# ===========================================================================
# Aba 6 · Diagnóstico (bespoke · 6 seções · ÚLTIMA)
# ===========================================================================


def _renderizar_aba_diagnostico_v1(
    wb: Workbook,
    v1: ConciliacaoV1Result,
    paleta,
    arquivo_nome: Optional[str],
    aba_consumida: Optional[str],
) -> None:
    """Aba 6 · 6 seções temáticas (DCV-V1 §6.7)."""
    ws = wb.create_sheet("Diagnóstico")
    ws.sheet_view.showGridLines = False
    cr = v1.conciliacao_realizada
    escrever_titulo_aba(ws, 1, 1, 6, "Diagnóstico", paleta)
    ws.row_dimensions[2].height = 8
    linha = 3

    # § 1 · Estrutura detectada
    linha = _renderizar_cabecalho_secao(ws, linha, 1, 6, "Estrutura detectada", paleta)
    rows = [
        ("Caso lógico", _label_caso_logico(cr.caso_logico_inferido)),
        ("Arquivos", f"{cr.n_arquivos} arquivo(s)"),
        ("Origem", f"{cr.arquivo_origem} · {cr.aba_origem}"),
        ("Comparado", f"{cr.arquivo_comparado} · {cr.aba_comparado}"),
        ("Registros · Origem", f"{cr.n_registros_origem:,}".replace(",", ".")),
        ("Registros · Comparado", f"{cr.n_registros_comparado:,}".replace(",", ".")),
        ("Processados", f"{cr.n_processados:,}".replace(",", ".")),
        ("Data/hora geração", datetime.now().strftime("%d/%m/%Y %H:%M")),
    ]
    linha = _renderizar_secao_como_tabela(ws, linha, "(detalhes)", rows, 1, 6, paleta)

    # § 2 · Tolerâncias absorvidas (resumo)
    linha = _renderizar_cabecalho_secao(ws, linha, 1, 6, "Tolerâncias absorvidas", paleta)
    sd = v1.sintese_diagnostico
    rows2 = [
        ("Total de células absorvidas", str(sd.n_tolerancia_absorvida)),
        ("Soma absoluta absorvida", formatar_moeda_br(float(sd.valor_tolerancia_absorvida))),
    ]
    linha = _renderizar_secao_como_tabela(ws, linha, "(detalhes)", rows2, 1, 6, paleta)

    # § 3 · Configuração aplicada
    linha = _renderizar_cabecalho_secao(ws, linha, 1, 6, "Configuração aplicada", paleta)
    ca = v1.config_aplicada
    rows3 = []
    rows3.append(
        ("Agrupadores de match",
         " · ".join(
             f"{a.rotulo_analitico} ({_label_modo_match(a.modo_match.value)})"
             for a in ca.agrupadores_match
         ) or "—"),
    )
    if ca.agrupadores_resumo_executivo:
        rows3.append(
            ("Agrupadores do Resumo Executivo",
             " · ".join(ca.agrupadores_resumo_executivo)),
        )
    for c in ca.campos_comparados:
        rows3.append(
            (f"Campo · {c.nome_analitico}",
             f"tipo {c.tipo_logico.value} · unidade {c.unidade.value} · "
             f"tolerância {c.tolerancia}"),
        )
    rows3.append(("Paleta aplicada", ca.paleta_aplicada))
    if ca.epsilon_por_unidade:
        eps_str = " · ".join(
            f"{u.value if hasattr(u, 'value') else u}: {v}"
            for u, v in ca.epsilon_por_unidade.items()
        )
        rows3.append(("Épsilon por unidade", eps_str))
    linha = _renderizar_secao_como_tabela(ws, linha, "(detalhes)", rows3, 1, 6, paleta)

    # § 4 · Modos de match aplicados
    linha = _renderizar_cabecalho_secao(ws, linha, 1, 6, "Modos de match aplicados", paleta)
    rows4 = []
    for a in ca.agrupadores_match:
        rows4.append(
            (a.rotulo_analitico,
             _label_modo_match(a.modo_match.value)),
        )
    if not rows4:
        rows4 = [("—", "—")]
    linha = _renderizar_secao_como_tabela(ws, linha, "(detalhes)", rows4, 1, 6, paleta)

    # § 5 · Warnings emitidos
    linha = _renderizar_cabecalho_secao(ws, linha, 1, 6, "Warnings emitidos", paleta)
    if not v1.warnings_emitidos:
        rows5 = [("(nenhum warning)", "")]
    else:
        rows5 = []
        for w in v1.warnings_emitidos:
            rows5.append(
                (f"{w.codigo} · {w.severidade}",
                 f"{w.n_ocorrencias} ocorrência(s)"
                 if w.n_ocorrencias > 0
                 else "0 ocorrências · nenhuma a reportar"),
            )
    linha = _renderizar_secao_como_tabela(ws, linha, "(detalhes)", rows5, 1, 6, paleta)

    # § 6 · Status da Ponte · nota de resíduo
    linha = _renderizar_cabecalho_secao(
        ws, linha, 1, 6, "Status da Ponte · nota de resíduo", paleta,
    )
    if v1.status_ponte_geral == StatusPonteV1.FECHA:
        rows6 = [
            ("Status geral", "✅ Todas as Pontes fecham dentro do épsilon aplicado"),
            ("Resíduos", "Nenhum a reportar"),
        ]
    else:
        rows6 = [("Status geral", "⚠️ Resíduo presente em pelo menos 1 campo")]
        for p in v1.pontes:
            if not p.fecha:
                rows6.append(
                    (f"Resíduo · {p.nome_analitico}",
                     formatar_valor_por_unidade(_to_float(p.residuo), p.unidade.value)),
                )
    linha = _renderizar_secao_como_tabela(ws, linha, "(detalhes)", rows6, 1, 6, paleta)

    _ajustar_larguras(ws, min_=14, max_=42)


# ===========================================================================
# Construção do Workbook · 6 abas (5 quando agrupadores executivos vazios)
# ===========================================================================


def _construir_workbook_v1(
    v1: ConciliacaoV1Result,
    paleta_nome: str,
    arquivo_nome: Optional[str] = None,
    aba_consumida: Optional[str] = None,
) -> Tuple[Workbook, List[str]]:
    """Monta workbook · ordem fixa de abas (Mockup-V1 + S-V1 §2.11)."""
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    paleta = obter_paleta(paleta_nome)
    capabilities: List[str] = []

    _renderizar_aba_resumo_executivo_v1(wb, v1, paleta, arquivo_nome, aba_consumida)
    capabilities.extend(["CAP-RESUMO-EXECUTIVO-V1", "CAP-FORMATO-ADAPTATIVO"])

    if v1.resumo_por_agrupador_executivo:
        _renderizar_aba_resumo_agrupador_v1(wb, v1, paleta)
        capabilities.append("CAP-RESUMO-AGRUPADOR-V1")

    _renderizar_aba_mapa_conciliacao_v1(wb, v1, paleta)
    capabilities.append("CAP-MAPA-CONCILIACAO-V1")

    _renderizar_aba_analise_analitica_v1(wb, v1, paleta)
    capabilities.append("CAP-ANALISE-ANALITICA-V1")

    _renderizar_aba_ponte_conciliacao_v1(wb, v1, paleta)
    capabilities.append("CAP-PONTE-CONCILIACAO-V1")

    _renderizar_aba_diagnostico_v1(wb, v1, paleta, arquivo_nome, aba_consumida)
    capabilities.append("CAP-DIAGNOSTICO-V1")

    return wb, capabilities


# ===========================================================================
# Função pública (entry point)
# ===========================================================================


def exportar_resultado_v1(
    v1_result: ConciliacaoV1Result,
    caminho_saida: str,
    paleta_nome: str = "azul",
    configuracao: Optional[ConfigExportacao] = None,
    origem_rotulo: Optional[str] = None,
    comparado_rotulo: Optional[str] = None,
    arquivo_nome_origem: Optional[str] = None,
    aba_consumida: Optional[str] = None,
    usar_nome_executivo: bool = False,
) -> ExportacaoResult:
    """Exportação V1 · 6 abas (5 sem agrupador executivo) · F-APRESENT integral.

    Mesma assinatura que `exportar_resultado_v2` (D-175 · D-176).
    """
    t0 = time.monotonic()
    nome_efetivo = (paleta_nome or "azul").strip().lower()

    wb, capabilities_acionadas = _construir_workbook_v1(
        v1_result, nome_efetivo,
        arquivo_nome=arquivo_nome_origem,
        aba_consumida=aba_consumida,
    )

    if usar_nome_executivo and origem_rotulo and comparado_rotulo:
        contexto = f"{origem_rotulo} vs {comparado_rotulo}"
        caminho_final = gerar_nome_arquivo(
            nome_visao_user_facing="Conciliacao de Bases",
            contexto=contexto,
            data=date.today(),
            diretorio_saida=Path(caminho_saida).parent,
        )
    else:
        caminho_final = Path(caminho_saida)

    os.makedirs(caminho_final.parent, exist_ok=True)
    wb.save(str(caminho_final))

    t1 = time.monotonic()
    tempo = round(t1 - t0, 4)
    tamanho = os.path.getsize(str(caminho_final))
    numero_abas = len(wb.sheetnames)

    return ExportacaoResult(
        caminho_arquivo=str(caminho_final),
        tamanho_bytes=tamanho,
        numero_abas=numero_abas,
        tempo_geracao_segundos=tempo,
        warnings_gerados=[],
        capabilities_acionadas=capabilities_acionadas,
    )
