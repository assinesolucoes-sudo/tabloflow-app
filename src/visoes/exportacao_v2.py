"""
exportacao_v2.py — Exportação Excel V2-específica · D-173 · D-175 · D-176.

Consome F-APRESENT integralmente. Substitui a chamada de app_v2 ao
`exportacao.exportar_resultado` legado (este permanece intocado para V1/V3–V11
enquanto cada visão não declarar seu próprio renderizador).

Estrutura de 4 abas (D-165 · absorve Parâmetros em Diagnóstico):
  1ª  Resumo Executivo     · bespoke · 5 blocos executivos + 2 gráficos
  2ª  Matriz de Confronto  · ListObject · totalsRowShown · badges · formato condicional
  3ª  Base Analítica       · ListObject · totalsRowShown · coluna Observações
  4ª  Diagnóstico          · bespoke 6 seções · respeita contrato de unidade D-166

Decisões-chave desta sessão (4-ter):
  - Resumo Executivo NÃO usa `renderizar_resumo_executivo` (capability 7)
    porque a estrutura §3.1 do prompt é diferente e porque a capability
    serializa `chave_agrupadores` como `str(dict)` em fallback ·
    escape implementado via renderer próprio com primitivas de tipografia.
  - Diagnóstico NÃO usa `renderizar_diagnostico` (capability 10)
    porque seção 5 formata TODOS thresholds como percentual (bug conhecido) ·
    bespoke respeita D-166: "Limite de valores na coluna de comparação"
    é contagem absoluta · os demais são percentuais.
  - Nome de arquivo executivo (D-176 · P-12): "Analise Comparativa - {origem} vs {comparado} - DD-MM-AAAA.xlsx"
"""
from __future__ import annotations

import math
import os
import re
import sys
import time
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

sys.path.insert(0, str(Path(__file__).parent.parent))

from apresentacao import (
    MAPEAMENTO_V2,
    aplicar_badge,
    aplicar_campo,
    aplicar_formato_contagem,
    aplicar_formato_monetario,
    aplicar_formato_percentual,
    aplicar_paleta,
    aplicar_secao,
    aplicar_titulo_aba,
    aplicar_valor,
    carregar_vocabulario_bilingue,
    criar_grafico_top_variacoes,
    criar_tabela_executiva,
    escrever_secao,
    escrever_titulo_aba,
    formatar_moeda_br,
    formatar_percentual_br,
    gerar_nome_arquivo,
    obter_paleta,
    traduzir,
)
from apresentacao.templates.familia_a import (
    construir_leitura_qualitativa,
    renderizar_concentracao,
    renderizar_grafico_variacoes,
    renderizar_onde_se_concentra,
    renderizar_saude_comparacao,
    renderizar_variacoes_destaque,
)
from apresentacao.templates.familia_a._shared import (
    LABEL_SEMANTICA_SAUDE as _LABEL_SEMANTICA_SAUDE,
    bordas_finas as _bordas_finas,
    calcular_altura_leitura_qualitativa as _calcular_altura_leitura_qualitativa,
    categorias_saude_para_exibir as _categorias_saude_para_exibir,
    contrair_de as _contrair_de,
    mesclar_card as _mesclar_card,
    renderizar_cabecalho_secao as _renderizar_cabecalho_secao,
    renderizar_secao_como_tabela as _renderizar_secao_como_tabela,
    rotular_agrupador as _rotular_agrupador,
)
from apresentacao.formatos import (
    FORMATO_CONTAGEM,
    FORMATO_MONETARIO_BR,
    FORMATO_PERCENTUAL,
    THRESHOLDS_CONTRATO_FUNDACAO,
    formatar_diferenca_por_unidade,
    formatar_threshold_por_contrato,
    formatar_valor_por_unidade,
    label_total_card,
    number_format_diferenca,
    number_format_valor,
    number_format_variacao,
    rotulo_diferenca,
    rotulo_variacao,
    valor_diferenca_para_celula,
    valor_total_card,
)
from contratos import ConfigExportacao, ExportacaoResult
from visoes.visao_v2 import V2Result


# ===========================================================================
# Tradutores / fallbacks user-facing
# ===========================================================================

# Fallback para classificações ESTRUTURAIS (bloco 3 do vocabulário cobre,
# mas as semânticas V2 ficam fora). Frases alinhadas com `_SEMA_LABEL_TO_CLASS`.
_FALLBACK_CLASSIF_ESTRUTURAL: Dict[str, str] = {
    "PRESENTE_AMBOS": "Presente nos dois lados",
    "AUSENTE_ORIGEM": "Ausente na origem",
    "AUSENTE_COMPARADO": "Ausente no comparado",
    "NULO_ORIGEM": "Sem valor na origem",
    "NULO_COMPARADO": "Sem valor no comparado",
    "NULO_AMBOS": "Sem valor nos dois lados",
}
_FALLBACK_CLASSIF_SEMANTICA: Dict[str, str] = {
    # MAIOR_MELHOR / MENOR_MELHOR · qualitativo
    "POSITIVO": "Melhorou",
    "NEGATIVO": "Piorou",
    "NEUTRO": "Estável",   # legado · ainda emitido em fallback genérico
    "NAO_APLICAVEL": "Não aplicável",
    # NEUTRA · 4 estruturais distintos preservados (D-187)
    "AUMENTOU": "Aumentou",
    "REDUZIU": "Reduziu",
    "ESTAVEL": "Estável",
}

_LEITURA_QUALITATIVA_USER_FACING: Dict[str, str] = {
    "Melhoria Geral": "A comparação indica melhoria geral na maior parte dos casos.",
    "Deterioração Geral": "A comparação indica deterioração na maior parte dos casos.",
    "Resultado Misto": "O resultado é misto · há ganhos e perdas distribuídos entre os casos.",
    "Resultado Estável": "A maior parte dos casos ficou estável entre os dois lados.",
    "Alta Taxa de Mudança de Estado": "A maior parte das combinações mudou de estado.",
    "Mudanças Parciais": "Algumas combinações mudaram de estado · a maior parte permaneceu.",
    "Estados Estáveis": "Os estados permaneceram estáveis na maior parte das combinações.",
}

# Thresholds do vocabulário com o contrato de unidade D-166 explícito.
# D-202 · constante promovida para `apresentacao.formatos.THRESHOLDS_CONTRATO_FUNDACAO`.
# Mantido alias local para compatibilidade com importadores externos (testes).
_THRESHOLDS_CONTRATO: Dict[str, Tuple[str, str]] = THRESHOLDS_CONTRATO_FUNDACAO


def _strip_aspas(texto: str) -> str:
    """Remove aspas duplas literais de início/fim (vocabulário v2 as inclui)."""
    if not texto:
        return texto
    t = texto.strip()
    if len(t) >= 2 and t[0] == '"' and t[-1] == '"':
        return t[1:-1]
    return t


def _traduzir_classif_estrutural(codigo: str, vocabulario) -> str:
    if not codigo:
        return "—"
    bloco = vocabulario.get("classificacoes", {}) if vocabulario else {}
    if codigo in bloco:
        return _strip_aspas(bloco[codigo])
    return _FALLBACK_CLASSIF_ESTRUTURAL.get(codigo, codigo.replace("_", " ").capitalize())


def _traduzir_classif_semantica(codigo: str, vocabulario) -> str:
    if not codigo:
        return "—"
    return _FALLBACK_CLASSIF_SEMANTICA.get(codigo, codigo.replace("_", " ").capitalize())


def _formatar_threshold(chave: str, valor: Any) -> Tuple[str, str]:
    """D-202 · delega para `apresentacao.formatos.formatar_threshold_por_contrato`."""
    return formatar_threshold_por_contrato(chave, valor)


# ===========================================================================
# Utilitários de worksheet
# ===========================================================================

def _ajustar_larguras(ws: Worksheet, min_: int = 12, max_: int = 50) -> None:
    """Auto-fit de larguras baseado no conteúdo textual das células.

    Sessão 8.2 · C-3 · Pula masters de merges multi-coluna · o texto
    de prosa (Leitura qualitativa, banners, narrativas mescladas em A:H)
    se estende visualmente pela largura combinada das 8 colunas; usar o
    `len(texto)` para auto-fit da COLUNA A artificialmente esticava A
    para o limite (50) e distorcia cards/tabelas. O wrap_text + altura
    explícita da linha cuidam da quebra dentro do range mesclado.
    """
    masters_merge_multi_col: set = set()
    for mr in ws.merged_cells.ranges:
        # max_col / min_col existem em MergedCellRange via bounds (top/bottom/left/right)
        # Usar bounds para clareza: (min_col, min_row, max_col, max_row)
        min_col, min_row, max_col, max_row = mr.bounds
        if max_col > min_col:
            masters_merge_multi_col.add((min_row, min_col))

    larguras: Dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            if (cell.row, cell.column) in masters_merge_multi_col:
                continue
            # Ignora células mescladas "slave" (valor None na apparent cell)
            valor_str = str(cell.value)
            col = cell.column
            tamanho = max(
                (len(parte) for parte in valor_str.split("\n")),
                default=len(valor_str),
            )
            tamanho = min(max_, tamanho + 2)
            tamanho = max(min_, tamanho)
            if tamanho > larguras.get(col, 0):
                larguras[col] = tamanho
    for col, largura in larguras.items():
        ws.column_dimensions[get_column_letter(col)].width = largura


# Helpers visuais (`_bordas_finas` · `_renderizar_cabecalho_secao` ·
# `_renderizar_secao_como_tabela` · `_calcular_altura_leitura_qualitativa` ·
# `_mesclar_card` · `_contrair_de`) movidos em D-202 etapa 5 para
# `apresentacao.templates.familia_a._shared` · re-exportados como aliases
# locais no bloco de imports do topo deste módulo.


def _construir_leitura_qualitativa_v2(
    v2: V2Result,
    origem_ux: str,
    comparado_ux: str,
) -> str:
    """D-202 · delegação para template Família A parametrizado."""
    comp = v2.comparacao_realizada
    return construir_leitura_qualitativa(
        unidade=comp.unidade,
        tipo_campo=comp.tipo_campo,
        semantica=comp.semantica_campo,
        ancora=v2.numeros_ancora,
        dist_sem=v2.distribuicao_classificacoes_semanticas,
        delta_sem=v2.delta_por_classificacao_semantica,
        base_analitica=v2.base_analitica,
        concentracao=v2.concentracao,
        onde_se_concentra=v2.onde_se_concentra,
        origem_ux=origem_ux,
        comparado_ux=comparado_ux,
    )


# ===========================================================================
# Seções dedicadas do Resumo Executivo · E2 · E3a · E3b (Sessão 8)
# Templates promovidos em D-202 etapa 5 para
# `apresentacao.templates.familia_a` · funções abaixo são wrappers de
# delegação preservando a interface V2 (recebe V2Result).
# ===========================================================================


def _renderizar_secao_saude_comparacao(
    ws: Worksheet,
    linha: int,
    largura_util: int,
    v2: V2Result,
    paleta,
) -> int:
    """D-202 · delegação para template Família A."""
    comp = v2.comparacao_realizada
    return renderizar_saude_comparacao(
        ws=ws,
        linha=linha,
        largura_util=largura_util,
        unidade=comp.unidade,
        semantica=comp.semantica_campo,
        dist_sem=v2.distribuicao_classificacoes_semanticas,
        delta_sem=v2.delta_por_classificacao_semantica,
        dist_estru=v2.distribuicao_classificacoes_estruturais,
        paleta=paleta,
    )


def _renderizar_secao_distribuicao_estrutural(
    ws: Worksheet,
    linha: int,
    largura_util: int,
    v2: V2Result,
    vocabulario,
    paleta,
) -> int:
    """Distribuição estrutural · seção 3 legada · usada apenas para
    tipo_campo=ESTADO_SITUACAO em V2 (E2 · D-192). Mantida intacta para
    consumo futuro de V1 e V11."""
    linha = _renderizar_cabecalho_secao(
        ws, linha, 1, largura_util, "Como os casos se distribuem", paleta,
    )

    dist = v2.distribuicao_classificacoes_estruturais or {}
    total_casos = sum(int(v or 0) for v in dist.values()) or 1
    bordas = _bordas_finas(paleta)
    cor_zebra = paleta.cor_secundaria

    col_cat_ini = 1
    col_cat_fim = largura_util - 4
    col_casos = largura_util - 3
    col_part_ini = largura_util - 2
    col_part_fim = largura_util

    def _header(linha_h: int) -> None:
        if col_cat_fim > col_cat_ini:
            ws.merge_cells(start_row=linha_h, start_column=col_cat_ini,
                           end_row=linha_h, end_column=col_cat_fim)
        ws.cell(row=linha_h, column=col_cat_ini, value="Categoria")
        ws.cell(row=linha_h, column=col_casos, value="Casos")
        if col_part_fim > col_part_ini:
            ws.merge_cells(start_row=linha_h, start_column=col_part_ini,
                           end_row=linha_h, end_column=col_part_fim)
        ws.cell(row=linha_h, column=col_part_ini, value="Participação")
        for col in range(1, largura_util + 1):
            c = ws.cell(row=linha_h, column=col)
            c.fill = PatternFill("solid", fgColor=paleta.cor_secundaria)
            c.font = Font(
                name=paleta.fonte_familia, size=paleta.fonte_tamanho_corpo,
                bold=True, color=paleta.cor_destaque,
            )
            c.border = bordas
            if col == col_cat_ini:
                align_h = "left"
            elif col == col_casos:
                align_h = "center"
            else:
                align_h = "right"
            c.alignment = Alignment(horizontal=align_h, vertical="center", indent=1)

    _header(linha)
    linha += 1

    idx_linha_dist = 0
    for categoria, qtd in dist.items():
        if not categoria:
            continue
        cor_fundo = cor_zebra if (idx_linha_dist % 2 == 0) else "FFFFFF"
        rot = _traduzir_classif_estrutural(categoria, vocabulario)
        n = int(qtd or 0)
        part = n / total_casos if total_casos else 0

        if col_cat_fim > col_cat_ini:
            ws.merge_cells(start_row=linha, start_column=col_cat_ini,
                           end_row=linha, end_column=col_cat_fim)
        c_cat = ws.cell(row=linha, column=col_cat_ini, value=rot)
        aplicar_valor(c_cat, paleta)
        c_cat.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        c_n = ws.cell(row=linha, column=col_casos, value=n)
        c_n.number_format = FORMATO_CONTAGEM
        aplicar_valor(c_n, paleta)
        c_n.alignment = Alignment(horizontal="center", vertical="center")

        if col_part_fim > col_part_ini:
            ws.merge_cells(start_row=linha, start_column=col_part_ini,
                           end_row=linha, end_column=col_part_fim)
        c_part = ws.cell(row=linha, column=col_part_ini, value=part)
        c_part.number_format = FORMATO_PERCENTUAL
        aplicar_valor(c_part, paleta)
        c_part.alignment = Alignment(horizontal="right", vertical="center", indent=1)

        for col in range(1, largura_util + 1):
            c = ws.cell(row=linha, column=col)
            c.fill = PatternFill("solid", fgColor=cor_fundo)
            c.border = bordas
        linha += 1
        idx_linha_dist += 1

    ws.row_dimensions[linha].height = 8
    return linha + 1


def _renderizar_secao_concentracao(
    ws: Worksheet,
    linha: int,
    largura_util: int,
    v2: V2Result,
    paleta,
) -> int:
    """D-202 · delegação para template Família A."""
    return renderizar_concentracao(
        ws=ws,
        linha=linha,
        largura_util=largura_util,
        concentracao=v2.concentracao,
        paleta=paleta,
    )


def _renderizar_secao_onde_se_concentra(
    ws: Worksheet,
    linha: int,
    largura_util: int,
    v2: V2Result,
    paleta,
) -> int:
    """D-202 · delegação para template Família A."""
    return renderizar_onde_se_concentra(
        ws=ws,
        linha=linha,
        largura_util=largura_util,
        unidade=v2.comparacao_realizada.unidade,
        onde_se_concentra=v2.onde_se_concentra,
        paleta=paleta,
    )


# ===========================================================================
# ABA 1 · Resumo Executivo (bespoke · D-175)
# ===========================================================================

def _renderizar_resumo_executivo_v2(
    ws: Worksheet, v2: V2Result, paleta, vocabulario,
) -> None:
    """
    Renderiza Resumo Executivo conforme §3.1 do prompt de Sessão 4-ter:
      1. Cabeçalho
      2. 4 números-âncora em cards visuais
      3. Distribuição + pizza
      4. Top variações + barras
      5. Leitura qualitativa
      6. Qualidade estrutural (panorâmica, sem lista detalhada)

    NÃO usa `renderizar_resumo_executivo` (capability 7) — escopo distinto
    declarado no prompt + contorno a bug conhecido de serialização de dict.
    """
    LARGURA_UTIL = 8  # A..H · cards ocupam 2 colunas cada
    ws.sheet_view.showGridLines = False

    linha = 1
    # --------------------------------------------------------------------
    # 1 · Cabeçalho (título + data BR)
    # --------------------------------------------------------------------
    comp = v2.comparacao_realizada
    origem = comp.origem_rotulo_ux or comp.origem_rotulo_tecnico or "origem"
    comparado = comp.comparado_rotulo_ux or comp.comparado_rotulo_tecnico or "comparado"
    titulo = f"Análise comparativa · {origem} × {comparado}"
    linha = escrever_titulo_aba(ws, linha, 1, LARGURA_UTIL, titulo, paleta)
    agora = datetime.now()
    subtitulo = f"Gerado em {agora.day:02d}/{agora.month:02d}/{agora.year:04d} às {agora.hour:02d}:{agora.minute:02d}"
    c_sub = ws.cell(row=linha, column=1, value=subtitulo)
    c_sub.font = Font(
        name=paleta.fonte_familia, size=paleta.fonte_tamanho_auxiliar,
        bold=False, italic=True, color=paleta.cor_neutra_escura,
    )
    ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=LARGURA_UTIL)
    linha += 2

    # --------------------------------------------------------------------
    # 2 · Números-âncora · banner de seção + 4 cards (C-4)
    # --------------------------------------------------------------------
    ancora = v2.numeros_ancora
    comp = v2.comparacao_realizada
    unidade = comp.unidade
    linha = _renderizar_cabecalho_secao(ws, linha, 1, LARGURA_UTIL, "Números principais", paleta)

    if ancora.total_origem is not None or ancora.total_comparado is not None:
        # E1 · cards adaptados pela unidade · 'Total' vira 'Média' para PERCENTUAL.
        rot_card_total = label_total_card(unidade)
        # Conta linhas com diferenca não-nula (proxy de PRESENTE_AMBOS para média)
        n_pa = 0
        if v2.base_analitica is not None and "diferenca" in v2.base_analitica.columns:
            n_pa = int(v2.base_analitica["diferenca"].dropna().shape[0])
        valor_orig_card = valor_total_card(ancora.total_origem, n_pa, unidade)
        valor_comp_card = valor_total_card(ancora.total_comparado, n_pa, unidade)
        if unidade == "PERCENTUAL" and valor_orig_card is not None and valor_comp_card is not None:
            dif_card = valor_comp_card - valor_orig_card
            var_card = (dif_card / valor_orig_card) if valor_orig_card not in (None, 0) else None
        else:
            dif_card = ancora.diferenca_total
            var_card = ancora.variacao_total_pct
        pares_cards: List[Tuple[str, str]] = [
            (f"{rot_card_total} · {origem}", formatar_valor_por_unidade(valor_orig_card, unidade)),
            (f"{rot_card_total} · {comparado}", formatar_valor_por_unidade(valor_comp_card, unidade)),
            (rotulo_diferenca(unidade), formatar_diferenca_por_unidade(dif_card, unidade)),
            (
                rotulo_variacao(unidade),
                formatar_percentual_br(var_card, conversao_fracao=True),
            ),
        ]
    else:
        pares_cards = [
            ("Combinações analisadas", f"{ancora.total_combinacoes_analisadas or 0:,}".replace(",", ".")),
            ("Com mudança", f"{ancora.combinacoes_com_mudanca or 0:,}".replace(",", ".")),
            ("Estáveis", f"{ancora.combinacoes_estaveis or 0:,}".replace(",", ".")),
            (
                "% mudança",
                formatar_percentual_br(ancora.pct_mudanca, conversao_fracao=True),
            ),
        ]

    # 4 cards · delimitados com borda fina para ficarem como "caixa" da seção
    linha_cards_topo = linha
    bordas_card = _bordas_finas(paleta)
    for i, (rot, val) in enumerate(pares_cards):
        col_ini = 1 + i * 2
        col_fim = col_ini + 1
        _mesclar_card(ws, linha_cards_topo, col_ini, linha_cards_topo + 1, col_fim, rot, val, paleta)
        for col in range(col_ini, col_fim + 1):
            for l in (linha_cards_topo, linha_cards_topo + 1):
                ws.cell(row=l, column=col).border = bordas_card
    ws.row_dimensions[linha_cards_topo].height = 22
    ws.row_dimensions[linha_cards_topo + 1].height = 32
    ws.row_dimensions[linha_cards_topo + 2].height = 8  # respiro
    linha = linha_cards_topo + 3

    # --------------------------------------------------------------------
    # 3 · Saúde da comparação (numérico) ou Distribuição estrutural (Estado)
    # E2 · D-192 · router por tipo_campo
    # --------------------------------------------------------------------
    bordas = _bordas_finas(paleta)
    cor_zebra = paleta.cor_secundaria

    if comp.tipo_campo == "ESTADO_SITUACAO":
        linha = _renderizar_secao_distribuicao_estrutural(
            ws, linha, LARGURA_UTIL, v2, vocabulario, paleta,
        )
    else:
        linha = _renderizar_secao_saude_comparacao(
            ws, linha, LARGURA_UTIL, v2, paleta,
        )
        # E3a · Concentração (oculta automaticamente quando v2.concentracao=None)
        linha = _renderizar_secao_concentracao(
            ws, linha, LARGURA_UTIL, v2, paleta,
        )
        # E3b · Onde se concentra (oculta automaticamente quando agrupador inválido)
        linha = _renderizar_secao_onde_se_concentra(
            ws, linha, LARGURA_UTIL, v2, paleta,
        )

    # --------------------------------------------------------------------
    # 4 · Top variações em destaque · banner + tabela + barras (C-3 + C-4)
    # D-202 · delegado para template Família A `renderizar_variacoes_destaque`.
    # --------------------------------------------------------------------
    top_list: List[Any] = list(getattr(v2, "top_variacoes", []) or [])
    agrupadores = list(v2.agrupadores_aplicados or [])
    linha, top_list = renderizar_variacoes_destaque(
        ws=ws, linha=linha, largura_util=LARGURA_UTIL,
        top_list=top_list, agrupadores=agrupadores, unidade=unidade,
        origem_ux=origem, comparado_ux=comparado, paleta=paleta,
    )

    # --------------------------------------------------------------------
    # 5 · Leitura qualitativa (prosa) · seção como tabela (C-4)
    # E3c · template parametrizado consumindo dados das ENTREGAS 2 + 3a + 3b
    # --------------------------------------------------------------------
    frase = _construir_leitura_qualitativa_v2(
        v2=v2,
        origem_ux=origem,
        comparado_ux=comparado,
    )
    linha_inicial_leitura = linha
    linha = _renderizar_secao_como_tabela(
        ws, linha, "Leitura qualitativa",
        [(frase,)],
        col_inicial=1, col_final=LARGURA_UTIL,
        paleta=paleta, aplicar_zebra=False,
    )
    # Sessão 8.3 · C-1 (P-30) · Altura robusta + alignment top na célula de
    # conteúdo. Apenas a linha do conteúdo recebe altura customizada (cabeçalho
    # já tem 22pt da função, respiro já tem 8pt da função, vizinhas preservam
    # default Excel). Sobrescrevemos vertical=center→top porque texto longo
    # com wrap fica visualmente melhor alinhado pelo topo.
    linha_conteudo_leitura = linha_inicial_leitura + 1
    ws.row_dimensions[linha_conteudo_leitura].height = (
        _calcular_altura_leitura_qualitativa(frase)
    )
    cel_leitura = ws.cell(row=linha_conteudo_leitura, column=1)
    cel_leitura.alignment = Alignment(
        horizontal="left", vertical="top",
        wrap_text=True, indent=1,
    )

    # --------------------------------------------------------------------
    # 6 · Qualidade estrutural (panorâmica) · seção como tabela (C-4)
    # --------------------------------------------------------------------
    qualidade = v2.resumo_executivo.bloco_6_qualidade_estrutural
    total = int(qualidade.total_warnings or 0)
    ajustes = int(qualidade.ajustes_aplicados or 0)
    if total == 0 and ajustes == 0:
        panorama = "Nenhum aviso estrutural gerado · nenhum ajuste automático aplicado."
    else:
        parte_avisos = (
            f"A análise gerou {total:,}".replace(",", ".") +
            (" aviso estrutural." if total == 1 else " avisos estruturais.")
        )
        if ajustes:
            parte_ajustes = (
                f" Foram aplicados {ajustes:,}".replace(",", ".") +
                (" ajuste automático." if ajustes == 1 else " ajustes automáticos.")
            )
        else:
            parte_ajustes = ""
        panorama = parte_avisos + parte_ajustes + " Nenhum bloqueio foi escapado."
    linha = _renderizar_secao_como_tabela(
        ws, linha, "Qualidade estrutural",
        [(panorama,)],
        col_inicial=1, col_final=LARGURA_UTIL,
        paleta=paleta, aplicar_zebra=False,
    )

    # --------------------------------------------------------------------
    # 7 · Variações em destaque · gráfico (Sub-sessão 8.4 · P-36)
    # D-202 · delegado para template Família A `renderizar_grafico_variacoes`.
    # --------------------------------------------------------------------
    if top_list:
        linha = renderizar_grafico_variacoes(
            ws=ws, linha=linha, largura_util=LARGURA_UTIL,
            top_list=top_list, agrupadores=agrupadores, unidade=unidade,
            paleta=paleta,
        )

    # Larguras finais
    _ajustar_larguras(ws)
    ws.freeze_panes = "A4"


# ===========================================================================
# ABA 2 · Matriz de Confronto (ListObject · badges · formato condicional)
# ===========================================================================

def _montar_colunas_base(
    v2: V2Result,
) -> List[Tuple[str, str, str]]:
    """
    Retorna [(identificador_no_df, rótulo_user_facing, unidade), ...] para as
    abas tabulares (Matriz + Base). Inclui agrupadores dinâmicos primeiro.
    """
    comp = v2.comparacao_realizada
    origem = comp.origem_rotulo_ux or comp.origem_rotulo_tecnico or "origem"
    comparado = comp.comparado_rotulo_ux or comp.comparado_rotulo_tecnico or "comparado"

    cols: List[Tuple[str, str, str]] = []
    df_cols = set(v2.base_analitica.columns) if v2.base_analitica is not None else set()

    for a in (v2.agrupadores_aplicados or []):
        if a in df_cols:
            cols.append((a, _rotular_agrupador(a), "texto"))

    tipo_campo = comp.tipo_campo
    if tipo_campo == "ESTADO_SITUACAO":
        if "estado_origem" in df_cols:
            cols.append(("estado_origem", f"Estado · {origem}", "texto"))
        if "estado_comparado" in df_cols:
            cols.append(("estado_comparado", f"Estado · {comparado}", "texto"))

    # E1 · tags adaptativas · "valor"/"diferenca"/"variacao" resolvem formato e
    # rótulo via `comp.unidade` no momento de aplicar formato (D-190 · C.D8).
    if "valor_origem" in df_cols:
        cols.append(("valor_origem", f"Valor · {origem}", "valor"))
    if "valor_comparado" in df_cols:
        cols.append(("valor_comparado", f"Valor · {comparado}", "valor"))
    if "diferenca" in df_cols:
        cols.append(("diferenca", rotulo_diferenca(comp.unidade), "diferenca"))
    if "variacao_percentual" in df_cols and tipo_campo != "ESTADO_SITUACAO":
        cols.append(("variacao_percentual", rotulo_variacao(comp.unidade), "variacao"))
    if "classificacao_estrutural" in df_cols:
        cols.append(("classificacao_estrutural", "Situação estrutural", "classificacao"))
    if "classificacao_semantica" in df_cols and tipo_campo != "ESTADO_SITUACAO":
        cols.append(("classificacao_semantica", "Leitura qualitativa", "classificacao"))
    return cols


def _escrever_tabela_dados(
    ws: Worksheet,
    linha_header: int,
    colunas: List[Tuple[str, str, str]],
    df: pd.DataFrame,
    paleta,
    vocabulario,
    extra_col_obs: bool = False,
) -> Tuple[int, int]:
    """
    Escreve cabeçalho + dados de uma tabela tabular a partir do DataFrame.
    Retorna (linha_fim_dados, n_cols).

    extra_col_obs=True adiciona coluna "Observações" no fim (sintetizada
    a partir de classificacao_estrutural quando a row não é 'Presente nos dois lados').
    """
    # Headers
    for idx, (_tec, rot, _unidade) in enumerate(colunas, start=1):
        c = ws.cell(row=linha_header, column=idx, value=rot)
        aplicar_campo(c, paleta)
    n_cols = len(colunas)
    if extra_col_obs:
        n_cols += 1
        c = ws.cell(row=linha_header, column=n_cols, value="Observações")
        aplicar_campo(c, paleta)

    linha = linha_header + 1
    for _, row in df.iterrows():
        for idx, (tec, _rot, unidade) in enumerate(colunas, start=1):
            valor = row.get(tec)
            celula = ws.cell(row=linha, column=idx)
            if valor is None or (isinstance(valor, float) and pd.isna(valor)):
                celula.value = "—"
                continue
            if unidade == "classificacao":
                codigo = str(valor)
                if tec == "classificacao_estrutural":
                    celula.value = _traduzir_classif_estrutural(codigo, vocabulario)
                else:
                    celula.value = _traduzir_classif_semantica(codigo, vocabulario)
                aplicar_badge(celula, codigo, paleta, MAPEAMENTO_V2)
            else:
                celula.value = valor
        # Observações coluna extra
        if extra_col_obs:
            classif = row.get("classificacao_estrutural")
            if classif and classif != "PRESENTE_AMBOS":
                texto = _traduzir_classif_estrutural(str(classif), vocabulario)
            else:
                texto = ""
            cel = ws.cell(row=linha, column=n_cols)
            cel.value = texto or "—"
            aplicar_valor(cel, paleta)
        linha += 1
    linha_fim = linha - 1
    return linha_fim, n_cols


def _resolver_number_format(tag_coluna: str, unidade_contrato: str) -> Optional[str]:
    """E1 · resolve number_format Excel a partir da tag semântica + unidade do contrato.

    Tags semânticas adaptativas:
      - "valor"     → formato do valor segundo unidade declarada
      - "diferenca" → formato da diferença segundo unidade declarada
      - "variacao"  → formato da variação relativa (sempre percentual nos casos
        previstos atualmente)

    Tags fixas (preservam contrato pré-Sessão 8):
      - "monetario", "percentual", "contagem"

    Tags sem formato numérico ("texto", "classificacao") retornam None.
    """
    if tag_coluna == "valor":
        return number_format_valor(unidade_contrato)
    if tag_coluna == "diferenca":
        return number_format_diferenca(unidade_contrato)
    if tag_coluna == "variacao":
        return number_format_variacao(unidade_contrato)
    if tag_coluna == "monetario":
        return FORMATO_MONETARIO_BR
    if tag_coluna == "percentual":
        return FORMATO_PERCENTUAL
    if tag_coluna == "contagem":
        return FORMATO_CONTAGEM
    return None


def _aplicar_formatos_tabela(
    ws: Worksheet,
    linha_header: int,
    linha_fim_dados: int,
    colunas: List[Tuple[str, str, str]],
    unidade_contrato: str = "MONETARIO_BRL",
) -> None:
    for idx, (_tec, _rot, tag) in enumerate(colunas, start=1):
        letra = get_column_letter(idx)
        rng = ws[f"{letra}{linha_header + 1}":f"{letra}{linha_fim_dados}"]
        fmt = _resolver_number_format(tag, unidade_contrato)
        if fmt is None:
            continue
        # PERCENTUAL + tag "diferenca": formato 'p.p' literal · valor cru
        # (fração) precisa ser pré-multiplicado por 100 para renderizar.
        rescale_diferenca = (tag == "diferenca" and unidade_contrato == "PERCENTUAL")
        for row in rng:
            for cel in row:
                cel.number_format = fmt
                if rescale_diferenca and isinstance(cel.value, (int, float)):
                    cel.value = float(cel.value) * 100.0


def _aplicar_formatos_totals_row(
    ws: Worksheet,
    linha_totais: int,
    colunas: List[Tuple[str, str, str]],
    unidade_contrato: str = "MONETARIO_BRL",
) -> None:
    """
    Aplica number_format nas células da linha de totais para herdar o contrato
    de unidade D-166 da coluna de dados correspondente.

    `criar_tabela_executiva` (capability 3) adiciona a totalsRow com fórmulas
    SUBTOTAL mas não propaga number_format das células de dados; sem isso,
    totais aparecem crus (1037289 ao invés de R$ 1.037.289,00). Tags
    adaptativas resolvem formato pelo contrato `unidade` da ComparacaoV2 ·
    default MONETARIO_BRL preserva comportamento anterior.
    """
    for idx, (_tec, _rot, tag) in enumerate(colunas, start=1):
        letra = get_column_letter(idx)
        celula = ws[f"{letra}{linha_totais}"]
        fmt = _resolver_number_format(tag, unidade_contrato)
        if fmt is not None:
            celula.number_format = fmt


def _renderizar_matriz_confronto_v2(
    ws: Worksheet, v2: V2Result, paleta, vocabulario,
) -> None:
    ws.sheet_view.showGridLines = False
    colunas = _montar_colunas_base(v2)
    if not colunas or v2.base_analitica is None or v2.base_analitica.empty:
        # Apenas título e mensagem
        escrever_titulo_aba(ws, 1, 1, 4, "Matriz de Confronto", paleta)
        c = ws.cell(row=3, column=1, value="Nenhum dado analítico para exibir.")
        aplicar_valor(c, paleta)
        return

    linha = escrever_titulo_aba(ws, 1, 1, len(colunas), "Matriz de Confronto", paleta)
    linha += 1  # respiro

    unidade_contrato = v2.comparacao_realizada.unidade

    linha_header = linha
    linha_fim, n_cols = _escrever_tabela_dados(
        ws, linha_header, colunas, v2.base_analitica, paleta, vocabulario,
        extra_col_obs=False,
    )

    # Formatos monetário/percentual
    _aplicar_formatos_tabela(ws, linha_header, linha_fim, colunas, unidade_contrato)

    # Tabela nativa (ListObject)
    if linha_fim >= linha_header + 1:
        ref = f"A{linha_header}:{get_column_letter(n_cols)}{linha_fim}"
        totais: Dict[str, str] = {}
        for (_tec, rot, tag) in colunas:
            if tag in ("valor", "diferenca", "monetario"):
                # Para PERCENTUAL · soma de percentuais não é semanticamente
                # útil · mas preservamos "sum" no Excel (totalsRow exibe valor
                # com number_format adaptativo) · análise vai ao card "Média".
                totais[rot] = "sum"
            elif tag in ("variacao", "percentual"):
                totais[rot] = "average"
            else:
                totais[rot] = "none"
        try:
            criar_tabela_executiva(
                ws, range_ref=ref, nome="TMatrizConfrontoV2",
                totais_por_coluna=totais, paleta_nome=paleta.nome,
            )
        except ValueError:
            # nome duplicado ou range inválido · mantém dados sem tabela nativa
            pass

        # Aplica number_format na linha de totais. `criar_tabela_executiva`
        # adiciona a totalsRow em `linha_fim + 1` com fórmulas SUBTOTAL, mas
        # não propaga number_format das células de dados.
        # TODO-FAPRESENT-CLEANUP: promover para capability 3 (criar_tabela_executiva
        # deveria aplicar number_format à totalsRow herdando das células de dados
        # da mesma coluna, respeitando `unidade` de cada TableColumn).
        _aplicar_formatos_totals_row(ws, linha_fim + 1, colunas, unidade_contrato)

    # Formatação condicional em Diferença (gradiente)
    idx_dif = None
    idx_var = None
    for i, (tec, _r, _u) in enumerate(colunas, start=1):
        if tec == "diferenca":
            idx_dif = i
        if tec == "variacao_percentual":
            idx_var = i
    if idx_dif and linha_fim >= linha_header + 1:
        letra = get_column_letter(idx_dif)
        rule = ColorScaleRule(
            start_type="min", start_color="F4CCCC",
            mid_type="percentile", mid_value=50, mid_color="EDF3F9",
            end_type="max", end_color="CFE2F3",
        )
        ws.conditional_formatting.add(
            f"{letra}{linha_header + 1}:{letra}{linha_fim}", rule,
        )
    if idx_var and linha_fim >= linha_header + 1:
        letra = get_column_letter(idx_var)
        rule = ColorScaleRule(
            start_type="min", start_color="F4CCCC",
            mid_type="percentile", mid_value=50, mid_color="EDF3F9",
            end_type="max", end_color="CFE2F3",
        )
        ws.conditional_formatting.add(
            f"{letra}{linha_header + 1}:{letra}{linha_fim}", rule,
        )

    ws.freeze_panes = ws.cell(row=linha_header + 1, column=1).coordinate
    _ajustar_larguras(ws)


# ===========================================================================
# ABA 3 · Base Analítica (ListObject · coluna Observações)
# ===========================================================================

def _renderizar_base_analitica_v2(
    ws: Worksheet, v2: V2Result, paleta, vocabulario,
) -> None:
    ws.sheet_view.showGridLines = False
    colunas = _montar_colunas_base(v2)
    if not colunas or v2.base_analitica is None or v2.base_analitica.empty:
        escrever_titulo_aba(ws, 1, 1, 4, "Base Analítica", paleta)
        c = ws.cell(row=3, column=1, value="Nenhum dado analítico para exibir.")
        aplicar_valor(c, paleta)
        return

    # Título ocupa header + 1 coluna para "Observações"
    largura_util = len(colunas) + 1
    linha = escrever_titulo_aba(ws, 1, 1, largura_util, "Base Analítica", paleta)
    linha += 1

    unidade_contrato = v2.comparacao_realizada.unidade

    linha_header = linha
    linha_fim, n_cols = _escrever_tabela_dados(
        ws, linha_header, colunas, v2.base_analitica, paleta, vocabulario,
        extra_col_obs=True,
    )

    _aplicar_formatos_tabela(ws, linha_header, linha_fim, colunas, unidade_contrato)

    # Tabela nativa
    if linha_fim >= linha_header + 1:
        ref = f"A{linha_header}:{get_column_letter(n_cols)}{linha_fim}"
        totais: Dict[str, str] = {}
        for (_tec, rot, tag) in colunas:
            if tag in ("valor", "diferenca", "monetario"):
                totais[rot] = "sum"
            elif tag in ("variacao", "percentual"):
                totais[rot] = "average"
            else:
                totais[rot] = "none"
        totais["Observações"] = "none"
        try:
            criar_tabela_executiva(
                ws, range_ref=ref, nome="TBaseAnaliticaV2",
                totais_por_coluna=totais, paleta_nome=paleta.nome,
            )
        except ValueError:
            pass

        # number_format na totalsRow.
        # TODO-FAPRESENT-CLEANUP: promover para capability 3 (criar_tabela_executiva
        # deveria aplicar number_format à totalsRow herdando das células de dados
        # da mesma coluna, respeitando `unidade` de cada TableColumn).
        _aplicar_formatos_totals_row(ws, linha_fim + 1, colunas, unidade_contrato)

    ws.freeze_panes = ws.cell(row=linha_header + 1, column=1).coordinate
    _ajustar_larguras(ws)


# ===========================================================================
# ABA 4 · Diagnóstico (bespoke · 6 seções · respeita D-166)
# ===========================================================================

_NOMES_MODO_BASE_V2 = {
    "POR_COLUNAS": "Dois lados em colunas distintas",
    "POR_LINHAS":  "Dois lados empilhados em coluna identificadora",
    "SIMPLES":     "Envio simples (um arquivo)",
    "DUAL":        "Envio dual (dois arquivos)",
    True:          "Base pré-agregada",
    False:         "Base transacional",
}

_NOMES_TIPO_CAMPO_V2 = {
    "NUMERICO_ADITIVO":     "Valor somável",
    "NUMERICO_RELATIVO":    "Valor percentual ou taxa",
    "NUMERICO_NAO_ADITIVO": "Indicador não somável",
    "ESTADO_SITUACAO":      "Categoria ou rótulo",
}


def _traduzir_modo_base(valor: Any) -> str:
    if valor is None:
        return "—"
    if valor in _NOMES_MODO_BASE_V2:
        return _NOMES_MODO_BASE_V2[valor]
    if isinstance(valor, str) and valor.upper() in _NOMES_MODO_BASE_V2:
        return _NOMES_MODO_BASE_V2[valor.upper()]
    return str(valor).replace("_", " ").capitalize()


def _traduzir_tipo_campo(valor: Any) -> str:
    if valor is None:
        return "—"
    if isinstance(valor, str) and valor in _NOMES_TIPO_CAMPO_V2:
        return _NOMES_TIPO_CAMPO_V2[valor]
    return str(valor).replace("_", " ").capitalize()


def _escrever_pares_diag(
    ws: Worksheet, linha: int, pares: List[Tuple[str, str]], paleta,
) -> int:
    for rotulo, valor in pares:
        c_r = ws.cell(row=linha, column=1, value=rotulo)
        c_v = ws.cell(row=linha, column=2, value=valor)
        aplicar_campo(c_r, paleta)
        aplicar_valor(c_v, paleta)
        linha += 1
    return linha


def _escrever_frase_diag(
    ws: Worksheet, linha: int, texto: str, paleta, col_fim: int = 4,
) -> int:
    ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=col_fim)
    c = ws.cell(row=linha, column=1, value=texto)
    aplicar_valor(c, paleta)
    return linha + 1


def _renderizar_diagnostico_v2(
    ws: Worksheet, v2: V2Result, paleta, vocabulario,
    paleta_nome_selecionada: str,
    arquivo_nome: Optional[str],
    aba_consumida: Optional[str],
) -> None:
    """
    Diagnóstico V2-específico · 6 seções user-facing que respeitam o contrato
    de unidade D-166 (seção 5) e não dependem da capability 10 (workaround
    do bug de formatação indiscriminada de thresholds).

    Cada seção é renderizada como bloco de tabela estilizada (cabeçalho
    colorido + zebra + bordas), via helper `_renderizar_secao_como_tabela`.
    Rótulo na col 1, valor mesclado entre col 2 e col COL_FIM · preserva
    contrato dos testes existentes (ws.cell(row=label_row, column=2).value).
    """
    ws.sheet_view.showGridLines = False
    COL_INI = 1
    COL_FIM = 4

    linha = 1
    linha = escrever_titulo_aba(ws, linha, 1, COL_FIM, "Diagnóstico", paleta)
    linha += 1

    comp = v2.comparacao_realizada
    config_usada = v2.config_usada or {}

    # --------------------------------------------------------------------
    # Seção 1 · Como a análise foi feita
    # --------------------------------------------------------------------
    pares_s1: List[Tuple[str, str]] = [
        ("Arquivo", arquivo_nome or config_usada.get("arquivo") or "—"),
        ("Aba consumida", aba_consumida or config_usada.get("aba_consumida") or "—"),
        ("Modo da base", _traduzir_modo_base(config_usada.get("estrutura_entrada"))),
        (
            "Agrupadores aplicados",
            " · ".join(_rotular_agrupador(a) for a in (v2.agrupadores_aplicados or [])) or "—",
        ),
        ("Campo analisado", comp.campo_analisado or "—"),
        ("Tipo de medida", _traduzir_tipo_campo(comp.tipo_campo)),
        ("Coluna origem", comp.origem_rotulo_ux or comp.origem_rotulo_tecnico or "—"),
        ("Coluna comparado", comp.comparado_rotulo_ux or comp.comparado_rotulo_tecnico or "—"),
    ]
    linha = _renderizar_secao_como_tabela(
        ws, linha, "Como a análise foi feita", pares_s1, COL_INI, COL_FIM, paleta,
    )

    # --------------------------------------------------------------------
    # Seção 2 · Ajustes automáticos do motor
    # --------------------------------------------------------------------
    warnings_list = list(v2.warnings or [])
    ajustes_leves = [w for w in warnings_list if _cat(w) in ("ajuste_leve",)]
    diagnostico_obj = getattr(v2, "diagnostico", None)
    ajustes_motor = []
    if diagnostico_obj is not None:
        ajustes_motor = list(getattr(diagnostico_obj, "ajustes_aplicados", []) or [])

    pares_s2: List[Tuple] = []
    for aj in ajustes_motor:
        tipo = getattr(aj, "tipo_ajuste", "") or ""
        descricao = getattr(aj, "descricao", "") or ""
        rot = _humanizar_tipo_ajuste(tipo)
        valor = descricao or f"{getattr(aj, 'linhas_afetadas', 0):,} linhas".replace(",", ".")
        pares_s2.append((rot, valor))
    for w in ajustes_leves:
        pares_s2.append(("Ajuste automático", _microcopy(w)))
    if not pares_s2:
        pares_s2 = [("Nenhum ajuste estrutural foi necessário.",)]
    linha = _renderizar_secao_como_tabela(
        ws, linha, "Ajustes automáticos do motor", pares_s2, COL_INI, COL_FIM, paleta,
    )

    # --------------------------------------------------------------------
    # Seção 3 · Pontos de atenção
    # --------------------------------------------------------------------
    atencao = [
        w for w in warnings_list
        if _cat(w) in ("alerta_estrutural_leve", "alerta_estrutural", "escape_acionado")
    ]
    pares_s3: List[Tuple] = []
    for w in atencao:
        cat_rot = _traduzir_categoria_warning(_cat(w))
        pares_s3.append((cat_rot, _microcopy(w)))
    if not pares_s3:
        pares_s3 = [("Nenhum ponto de atenção identificado nesta análise.",)]
    linha = _renderizar_secao_como_tabela(
        ws, linha, "Pontos de atenção", pares_s3, COL_INI, COL_FIM, paleta,
    )

    # --------------------------------------------------------------------
    # Seção 4 · Decisões do usuário
    # --------------------------------------------------------------------
    pares_s4: List[Tuple] = []
    estados_nao = list(comp.estados_nao_escolhidos or [])
    if estados_nao:
        pares_s4.append((
            "Estados excluídos da comparação",
            " · ".join(str(e) for e in estados_nao),
        ))
    decisoes = [w for w in warnings_list if _cat(w) == "decisao_usuario"]
    for w in decisoes:
        pares_s4.append(("Decisão registrada", _microcopy(w)))
    if not pares_s4:
        pares_s4 = [("Análise executada com configurações padrão.",)]
    linha = _renderizar_secao_como_tabela(
        ws, linha, "Decisões do usuário", pares_s4, COL_INI, COL_FIM, paleta,
    )

    # --------------------------------------------------------------------
    # Seção 5 · Configurações avançadas aplicadas (D-166 unit contract)
    # --------------------------------------------------------------------
    pares_s5: List[Tuple[str, str]] = []
    paleta_rot = {
        "azul": "Azul executivo",
        "verde": "Verde executivo",
        "cinza": "Cinza executivo",
        "vinho": "Vinho executivo",
    }.get((paleta_nome_selecionada or "").strip().lower(), paleta.rotulo_user_facing)
    pares_s5.append(("Paleta selecionada", paleta_rot))
    thresholds = config_usada.get("thresholds", {}) or {}
    for chave in (
        "limiar_estabilidade_pct",
        "limiar_nulo_massivo_pct",
        "limite_valores_discriminador_alerta",
        "limite_variacao_extrema",
        "limite_variacao_extrema_pct",
    ):
        if chave in thresholds:
            rot, val = _formatar_threshold(chave, thresholds[chave])
            if any(r == rot for r, _ in pares_s5):
                continue
            pares_s5.append((rot, val))
    linha = _renderizar_secao_como_tabela(
        ws, linha, "Configurações avançadas aplicadas", pares_s5, COL_INI, COL_FIM, paleta,
    )

    # --------------------------------------------------------------------
    # Seção 6 · Qualidade estrutural da análise
    # --------------------------------------------------------------------
    qualidade = v2.resumo_executivo.bloco_6_qualidade_estrutural
    total = int(qualidade.total_warnings or 0)
    por_cat = dict(qualidade.warnings_por_categoria or {})
    pares_s6: List[Tuple[str, str]] = [
        ("Total de avisos", f"{total:,}".replace(",", ".")),
    ]
    for cat_tec, qtd in por_cat.items():
        rot = _traduzir_categoria_warning(cat_tec)
        pares_s6.append((f"Avisos · {rot}", f"{int(qtd):,}".replace(",", ".")))
    pares_s6.append((
        "Bloqueios escapados",
        "Não" if not qualidade.tem_bloqueios_escapados else "Sim",
    ))
    if diagnostico_obj is not None:
        tempo_total = getattr(diagnostico_obj, "tempo_total_seg", None)
        if tempo_total is not None:
            try:
                pares_s6.append((
                    "Tempo total de execução",
                    f"{float(tempo_total):.2f} s",
                ))
            except (TypeError, ValueError):
                pass
    linha = _renderizar_secao_como_tabela(
        ws, linha, "Qualidade estrutural da análise", pares_s6, COL_INI, COL_FIM, paleta,
    )

    # Larguras
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 22


# ---------------------------------------------------------------------------
# Helpers seção 2/3 — warnings
# ---------------------------------------------------------------------------

def _cat(w: Any) -> str:
    cat = getattr(w, "categoria", None)
    if cat is None and isinstance(w, dict):
        cat = w.get("categoria")
    if cat is None:
        return "informativo"
    if hasattr(cat, "value"):
        return str(cat.value)
    return str(cat)


def _microcopy(w: Any) -> str:
    micro = getattr(w, "microcopy", None)
    if micro is None and isinstance(w, dict):
        micro = w.get("microcopy")
    if not micro:
        return "—"
    return str(micro)


def _humanizar_tipo_ajuste(tipo: str) -> str:
    tabela = {
        "INTERVALO_AJUSTADO_INICIO": "Intervalo ajustado (início)",
        "INTERVALO_AJUSTADO_FIM":    "Intervalo ajustado (fim)",
        "LINHA_EXCLUIDA_NULO":       "Linhas excluídas por valor ausente",
        "COLUNA_RENOMEADA":          "Coluna renomeada",
        "TIPO_COERCIDO":             "Tipo convertido",
        "DUPLICATA_REMOVIDA":        "Duplicatas removidas",
        "CONSOLIDACAO_APLICADA":     "Consolidação aplicada",
    }
    if tipo in tabela:
        return tabela[tipo]
    if not tipo:
        return "Ajuste aplicado"
    return tipo.replace("_", " ").capitalize()


def _traduzir_categoria_warning(cat: str) -> str:
    tabela = {
        "informativo":               "Informativo",
        "ajuste_leve":               "Ajuste automático",
        "alerta_estrutural_leve":    "Alerta estrutural leve",
        "alerta_estrutural":         "Alerta estrutural",
        "decisao_usuario":           "Decisão do usuário",
        "escape_acionado":           "Escape acionado",
    }
    return tabela.get(cat, cat.replace("_", " ").capitalize())


# ===========================================================================
# API pública
# ===========================================================================

def _construir_workbook_v2(
    v2_result: V2Result,
    paleta_nome: str,
    arquivo_nome: Optional[str] = None,
    aba_consumida: Optional[str] = None,
) -> Tuple[Workbook, List[str]]:
    """Constrói Workbook completo em memória · preserva marcador de paleta."""
    capabilities_acionadas: List[str] = []

    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    nome_normalizado = (paleta_nome or "azul").strip().lower()
    paleta = obter_paleta(nome_normalizado)
    aplicar_paleta(wb, paleta)
    capabilities_acionadas.append("CAP-APLICAR-PALETA")

    vocabulario = carregar_vocabulario_bilingue()

    ws_resumo = wb.create_sheet("Resumo Executivo")
    _renderizar_resumo_executivo_v2(ws_resumo, v2_result, paleta, vocabulario)
    capabilities_acionadas.append("CAP-RESUMO-EXECUTIVO-V2")

    ws_cv = wb.create_sheet("Matriz de Confronto")
    _renderizar_matriz_confronto_v2(ws_cv, v2_result, paleta, vocabulario)
    capabilities_acionadas.append("CAP-MATRIZ-CONFRONTO-V2")

    ws_ba = wb.create_sheet("Base Analítica")
    _renderizar_base_analitica_v2(ws_ba, v2_result, paleta, vocabulario)
    capabilities_acionadas.append("CAP-BASE-ANALITICA-V2")

    ws_diag = wb.create_sheet("Diagnóstico")
    _renderizar_diagnostico_v2(
        ws_diag, v2_result, paleta, vocabulario,
        paleta_nome_selecionada=nome_normalizado,
        arquivo_nome=arquivo_nome,
        aba_consumida=aba_consumida,
    )
    capabilities_acionadas.append("CAP-DIAGNOSTICO-V2")

    return wb, capabilities_acionadas


def exportar_resultado_v2(
    v2_result: V2Result,
    caminho_saida: str,
    paleta_nome: str = "azul",
    configuracao: Optional[ConfigExportacao] = None,
    origem_rotulo: Optional[str] = None,
    comparado_rotulo: Optional[str] = None,
    arquivo_nome_origem: Optional[str] = None,
    aba_consumida: Optional[str] = None,
    usar_nome_executivo: bool = False,
) -> ExportacaoResult:
    """
    Exportação V2-específica · consome F-APRESENT integralmente (D-173 · D-175 · D-176).

    Parâmetros:
      v2_result · V2Result obrigatório
      caminho_saida · path de destino. Se `usar_nome_executivo=True`, o nome
        de arquivo final é derivado de (origem_rotulo, comparado_rotulo) via
        gerar_nome_arquivo; caso contrário, caminho_saida é usado literal.
      paleta_nome · "azul"/"verde"/"cinza"/"vinho" · default "azul"
      origem_rotulo · rótulo UX do lado origem (para nome executivo)
      comparado_rotulo · rótulo UX do lado comparado (para nome executivo)
      arquivo_nome_origem · nome do arquivo de origem (seção 1 do Diagnóstico)
      aba_consumida · nome da aba consumida (seção 1 do Diagnóstico)
      usar_nome_executivo · quando True, reescreve o nome do arquivo via D-176.

    Retorna ExportacaoResult.
    """
    t0 = time.monotonic()
    _config = configuracao or ConfigExportacao(paleta=paleta_nome)
    nome_efetivo = (paleta_nome or _config.paleta or "azul").strip().lower()

    wb, capabilities_acionadas = _construir_workbook_v2(
        v2_result, nome_efetivo,
        arquivo_nome=arquivo_nome_origem,
        aba_consumida=aba_consumida,
    )

    # Nome executivo D-176 (substitui caminho_saida)
    if usar_nome_executivo and origem_rotulo and comparado_rotulo:
        contexto = f"{origem_rotulo} vs {comparado_rotulo}"
        caminho_final = gerar_nome_arquivo(
            nome_visao_user_facing="Analise Comparativa",
            contexto=contexto,
            data=date.today(),
            diretorio_saida=Path(caminho_saida).parent,
        )
    else:
        caminho_final = Path(caminho_saida)

    os.makedirs(caminho_final.parent, exist_ok=True)
    wb.save(str(caminho_final))

    t1 = time.monotonic()
    tempo = round(t1 - t0, 4)
    tamanho = os.path.getsize(str(caminho_final))
    numero_abas = len(wb.sheetnames)

    return ExportacaoResult(
        caminho_arquivo=str(caminho_final),
        tamanho_bytes=tamanho,
        numero_abas=numero_abas,
        tempo_geracao_segundos=tempo,
        warnings_gerados=[],
        capabilities_acionadas=capabilities_acionadas,
    )
