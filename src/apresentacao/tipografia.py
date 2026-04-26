"""
F-APRESENT · capability 9 · Hierarquia tipográfica (D-159 · D-165).

4 níveis tipográficos canônicos consumidos por Resumo Executivo (capability 7)
e Diagnóstico narrativo (capability 10):

  titulo_aba   · linha 1 da aba · nome user-facing · merge horizontal
  secao        · subdivisão da aba · linha em branco antes · 12pt bold
  campo        · rótulo chave-valor · 10pt semibold
  valor        · valor ou corpo de texto · 10pt regular

Cada helper recebe `cell` + `Paleta` e aplica Font + Alignment coerentes com
a paleta selecionada. A tipografia respeita `Paleta.fonte_familia` (Calibri ·
universal cross-plataforma) e os tamanhos declarados em `Paleta.fonte_tamanho_*`
do catálogo D-164.

Design:
  - Helpers individuais (não marcadores stateful) · chamada direta da
    capability 7/10 na renderização · simples · testável · sem side effects
    em estruturas compartilhadas
  - `aplicar_hierarquia_tipografica(ws, paleta)` mantido como wrapper de
    conveniência (interface canônica D-159) · reservado para evolução futura
    quando capabilities declararem marcadores na célula · no P1 é no-op seguro
"""
from __future__ import annotations

from typing import Literal

from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .paletas import Paleta


NivelTipografico = Literal["titulo_aba", "secao", "campo", "valor"]

NIVEIS_CANONICOS = ("titulo_aba", "secao", "campo", "valor")


def _validar(cell: Cell, paleta: Paleta, funcao: str) -> None:
    if cell is None:
        raise TypeError(f"cell não pode ser None em {funcao}")
    if paleta is None:
        raise TypeError(f"paleta não pode ser None em {funcao}")


def aplicar_titulo_aba(cell: Cell, paleta: Paleta) -> None:
    """Título da aba · linha 1 · cor primária · bold · tamanho titulo da paleta.

    O merge horizontal do título sobre a largura útil fica a cargo do
    chamador · capability 9 trata apenas de fonte + alinhamento da célula
    de ancoragem.
    """
    _validar(cell, paleta, "aplicar_titulo_aba")
    cell.font = Font(
        name=paleta.fonte_familia,
        size=paleta.fonte_tamanho_titulo,
        bold=True,
        color=paleta.cor_primaria,
    )
    cell.alignment = Alignment(horizontal="left", vertical="center")


def aplicar_secao(cell: Cell, paleta: Paleta) -> None:
    """Título de seção · tamanho seção da paleta · bold · cor destaque da paleta.

    Convenção de composição: o chamador deixa uma linha em branco antes
    de uma seção nova · mantém respiro visual entre blocos.
    """
    _validar(cell, paleta, "aplicar_secao")
    cell.font = Font(
        name=paleta.fonte_familia,
        size=paleta.fonte_tamanho_secao,
        bold=True,
        color=paleta.cor_destaque,
    )
    cell.alignment = Alignment(horizontal="left", vertical="center")


def aplicar_campo(cell: Cell, paleta: Paleta) -> None:
    """Rótulo de chave-valor · tamanho corpo · bold · cor neutra escura.

    "Semibold" não tem suporte direto em openpyxl · usamos `bold=True` com
    tamanho corpo padrão · produz peso visual intermediário coerente.
    """
    _validar(cell, paleta, "aplicar_campo")
    cell.font = Font(
        name=paleta.fonte_familia,
        size=paleta.fonte_tamanho_corpo,
        bold=True,
        color=paleta.cor_neutra_escura,
    )
    cell.alignment = Alignment(
        horizontal="left", vertical="center", wrap_text=True,
    )


def aplicar_valor(cell: Cell, paleta: Paleta) -> None:
    """Valor de chave-valor ou corpo de texto · tamanho corpo · regular."""
    _validar(cell, paleta, "aplicar_valor")
    cell.font = Font(
        name=paleta.fonte_familia,
        size=paleta.fonte_tamanho_corpo,
        bold=False,
        color=paleta.cor_neutra_escura,
    )
    cell.alignment = Alignment(
        horizontal="left", vertical="center", wrap_text=True,
    )


def aplicar_hierarquia_tipografica(ws: Worksheet, paleta: Paleta) -> None:
    """
    Wrapper de conveniência declarado em D-159 · interface mínima canônica.

    No P1 é no-op seguro · capabilities renderizadoras (7 e 10) chamam os
    helpers individuais (aplicar_titulo_aba · aplicar_secao · aplicar_campo
    · aplicar_valor) durante a montagem de cada aba.

    Fica reservado para evolução futura: quando uma capability gravar marcadores
    de nível na célula (ex: `cell.style = "tabloflow_secao"`), este wrapper
    varre a worksheet aplicando o estilo correspondente. Hoje os helpers
    individuais cobrem 100% dos casos de P1.
    """
    if ws is None:
        raise TypeError("ws não pode ser None em aplicar_hierarquia_tipografica")
    if paleta is None:
        raise TypeError("paleta não pode ser None em aplicar_hierarquia_tipografica")
    # no-op intencional · ver docstring


APLICADORES = {
    "titulo_aba": aplicar_titulo_aba,
    "secao": aplicar_secao,
    "campo": aplicar_campo,
    "valor": aplicar_valor,
}


def aplicar_nivel(nivel: NivelTipografico, cell: Cell, paleta: Paleta) -> None:
    """
    Dispatcher para o aplicador correspondente ao nível.

    Útil para chamadores que recebem o nível como parâmetro (ex: renderização
    data-driven). Levanta ValueError se nível fora do catálogo canônico.
    """
    if nivel not in APLICADORES:
        raise ValueError(
            f"nível '{nivel}' fora do catálogo canônico · opções: {list(NIVEIS_CANONICOS)}"
        )
    APLICADORES[nivel](cell, paleta)


def escrever_titulo_aba(
    ws: Worksheet,
    linha: int,
    col_ini: int,
    col_fim: int,
    texto: str,
    paleta: Paleta,
) -> int:
    """
    Escreve título da aba em merge horizontal de col_ini a col_fim · aplica
    estilo titulo_aba · retorna a próxima linha livre (linha + 1).

    Conveniência estrutural: abas executivas começam com 1 título merged
    sobre a largura útil · este helper encapsula merge + estilo + avanço.
    """
    if ws is None:
        raise TypeError("ws não pode ser None em escrever_titulo_aba")
    if col_fim < col_ini:
        raise ValueError(
            f"col_fim ({col_fim}) precisa ser >= col_ini ({col_ini})"
        )
    if col_fim > col_ini:
        ws.merge_cells(
            start_row=linha, start_column=col_ini,
            end_row=linha, end_column=col_fim,
        )
    celula = ws.cell(row=linha, column=col_ini, value=texto)
    aplicar_titulo_aba(celula, paleta)
    ws.row_dimensions[linha].height = 28
    return linha + 1


def escrever_secao(
    ws: Worksheet,
    linha: int,
    col_ini: int,
    texto: str,
    paleta: Paleta,
) -> int:
    """Escreve título de seção numa célula simples · aplica estilo secao."""
    if ws is None:
        raise TypeError("ws não pode ser None em escrever_secao")
    celula = ws.cell(row=linha, column=col_ini, value=texto)
    aplicar_secao(celula, paleta)
    return linha + 1


def escrever_campo_valor(
    ws: Worksheet,
    linha: int,
    col_campo: int,
    col_valor: int,
    rotulo: str,
    valor: str,
    paleta: Paleta,
) -> int:
    """
    Escreve par rótulo-valor em duas colunas com estilos campo e valor.
    Retorna próxima linha livre.
    """
    if ws is None:
        raise TypeError("ws não pode ser None em escrever_campo_valor")
    c_campo = ws.cell(row=linha, column=col_campo, value=rotulo)
    c_valor = ws.cell(row=linha, column=col_valor, value=valor)
    aplicar_campo(c_campo, paleta)
    aplicar_valor(c_valor, paleta)
    return linha + 1
