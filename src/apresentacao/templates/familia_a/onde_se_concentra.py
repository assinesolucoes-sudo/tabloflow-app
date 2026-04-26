"""Sub-template Família A · Onde se concentra · D-202 etapa 5.

Movido e parametrizado de
`visoes.exportacao_v2._renderizar_secao_onde_se_concentra`.
"""
from __future__ import annotations

from typing import Any, Dict, Optional

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from apresentacao.formatos import (
    formatar_diferenca_por_unidade,
    formatar_valor_por_unidade,
    number_format_diferenca,
    number_format_valor,
    number_format_variacao,
    rotulo_diferenca,
    rotulo_variacao,
    valor_diferenca_para_celula,
)
from apresentacao.tipografia import aplicar_valor

from ._shared import (
    bordas_finas,
    renderizar_cabecalho_secao,
    rotular_agrupador,
)


def renderizar_onde_se_concentra(
    *,
    ws: Worksheet,
    linha: int,
    largura_util: int,
    unidade: str,
    onde_se_concentra: Optional[Dict[str, Any]],
    paleta,
) -> int:
    """E3b · 'Onde se concentra · Top 3 por agrupador escolhido'.

    Tabela de 4 colunas: Categoria · Δ · Variação % · Direção (↑/↓/→).

    PERCENTUAL exibe 'Δ médio' (média das diferenças individuais · faz sentido
    em p.p) ao invés de 'Δ' (soma · viola C.D3 quando agregada em p.p).
    """
    osc = onde_se_concentra
    if not osc or not osc.get("top"):
        return linha

    agrupador = osc.get("agrupador") or ""
    titulo = f"Onde se concentra · por {rotular_agrupador(agrupador)}"
    usa_media = unidade == "PERCENTUAL"
    rotulo_col_delta = "Δ médio" if usa_media else rotulo_diferenca(unidade)

    linha = renderizar_cabecalho_secao(
        ws, linha, 1, largura_util, titulo, paleta,
    )
    bordas = bordas_finas(paleta)
    cor_zebra = paleta.cor_secundaria

    # Layout · 4 segmentos
    col_cat_ini  = 1
    col_cat_fim  = max(2, largura_util - 5)
    col_delta_ini = col_cat_fim + 1
    col_delta_fim = col_cat_fim + 2
    col_var_ini  = col_cat_fim + 3
    col_var_fim  = col_cat_fim + 4
    col_dir_ini  = col_cat_fim + 5
    col_dir_fim  = largura_util

    def _header(linha_h: int) -> None:
        if col_cat_fim > col_cat_ini:
            ws.merge_cells(start_row=linha_h, start_column=col_cat_ini,
                           end_row=linha_h, end_column=col_cat_fim)
        ws.cell(row=linha_h, column=col_cat_ini, value=rotular_agrupador(agrupador) or "Categoria")
        if col_delta_fim > col_delta_ini:
            ws.merge_cells(start_row=linha_h, start_column=col_delta_ini,
                           end_row=linha_h, end_column=col_delta_fim)
        ws.cell(row=linha_h, column=col_delta_ini, value=rotulo_col_delta)
        if col_var_fim > col_var_ini:
            ws.merge_cells(start_row=linha_h, start_column=col_var_ini,
                           end_row=linha_h, end_column=col_var_fim)
        ws.cell(row=linha_h, column=col_var_ini, value=rotulo_variacao(unidade))
        if col_dir_fim > col_dir_ini:
            ws.merge_cells(start_row=linha_h, start_column=col_dir_ini,
                           end_row=linha_h, end_column=col_dir_fim)
        ws.cell(row=linha_h, column=col_dir_ini, value="Direção")
        for col in range(1, largura_util + 1):
            c = ws.cell(row=linha_h, column=col)
            c.fill = PatternFill("solid", fgColor=paleta.cor_secundaria)
            c.font = Font(
                name=paleta.fonte_familia, size=paleta.fonte_tamanho_corpo,
                bold=True, color=paleta.cor_destaque,
            )
            c.border = bordas
            align_h = "left" if col <= col_cat_fim else ("right" if col < col_dir_ini else "center")
            c.alignment = Alignment(horizontal=align_h, vertical="center", indent=1)

    _header(linha)
    linha += 1

    top_list = osc.get("top", []) or []
    for idx_t, item in enumerate(top_list):
        cor_fundo = cor_zebra if (idx_t % 2 == 0) else "FFFFFF"
        cat = item.get("categoria") or "—"
        if usa_media:
            delta_raw = float(item.get("delta_medio") or 0)
        else:
            delta_raw = float(item.get("delta_soma") or item.get("delta") or 0)
        var_rel = item.get("variacao_relativa")

        # Categoria
        if col_cat_fim > col_cat_ini:
            ws.merge_cells(start_row=linha, start_column=col_cat_ini,
                           end_row=linha, end_column=col_cat_fim)
        c_cat = ws.cell(row=linha, column=col_cat_ini, value=str(cat))
        aplicar_valor(c_cat, paleta)
        c_cat.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        if col_delta_fim > col_delta_ini:
            ws.merge_cells(start_row=linha, start_column=col_delta_ini,
                           end_row=linha, end_column=col_delta_fim)
        if usa_media:
            valor_celula = valor_diferenca_para_celula(delta_raw, unidade)
            fmt_celula = number_format_diferenca(unidade)
        else:
            valor_celula = delta_raw
            fmt_celula = number_format_valor(unidade)
        c_d = ws.cell(row=linha, column=col_delta_ini, value=valor_celula)
        c_d.number_format = fmt_celula
        aplicar_valor(c_d, paleta)
        c_d.alignment = Alignment(horizontal="right", vertical="center", indent=1)

        # Variação %
        if col_var_fim > col_var_ini:
            ws.merge_cells(start_row=linha, start_column=col_var_ini,
                           end_row=linha, end_column=col_var_fim)
        c_v = ws.cell(row=linha, column=col_var_ini)
        if var_rel is None:
            c_v.value = "—"
        else:
            c_v.value = float(var_rel)
            c_v.number_format = number_format_variacao(unidade)
        aplicar_valor(c_v, paleta)
        c_v.alignment = Alignment(horizontal="right", vertical="center", indent=1)

        # Direção
        if delta_raw > 0:
            direcao = "↑ Puxa para cima"
        elif delta_raw < 0:
            direcao = "↓ Puxa para baixo"
        else:
            direcao = "→ Estável"
        if col_dir_fim > col_dir_ini:
            ws.merge_cells(start_row=linha, start_column=col_dir_ini,
                           end_row=linha, end_column=col_dir_fim)
        c_dir = ws.cell(row=linha, column=col_dir_ini, value=direcao)
        aplicar_valor(c_dir, paleta)
        c_dir.alignment = Alignment(horizontal="center", vertical="center", indent=0)

        for col in range(1, largura_util + 1):
            c = ws.cell(row=linha, column=col)
            c.fill = PatternFill("solid", fgColor=cor_fundo)
            c.border = bordas
        linha += 1

    # Rodapé "(outras N ...)"
    outras_count = int(osc.get("outras_count") or 0)
    outras_soma = float(osc.get("outras_delta_soma") or 0.0)
    outras_medio = float(osc.get("outras_delta_medio") or 0.0)
    outras_dom = bool(osc.get("outras_dominante"))
    if outras_count > 0:
        if usa_media:
            soma_str = formatar_diferenca_por_unidade(outras_medio, unidade)
            dom_txt = "" if not outras_dom else " · INFLUÊNCIA DOMINANTE"
            rodape = (
                f"(outras {outras_count} categoria(s) têm Δ médio {soma_str}"
                f"{dom_txt} · sem influência dominante)"
                if not outras_dom else
                f"(outras {outras_count} categoria(s) têm Δ médio {soma_str}{dom_txt})"
            )
        else:
            soma_str = formatar_valor_por_unidade(outras_soma, unidade)
            dom_txt = "" if not outras_dom else " · INFLUÊNCIA DOMINANTE"
            rodape = (
                f"(outras {outras_count} categoria(s) somam Δ {soma_str}"
                f"{dom_txt} · sem influência dominante)"
                if not outras_dom else
                f"(outras {outras_count} categoria(s) somam Δ {soma_str}{dom_txt})"
            )
        if largura_util > 1:
            ws.merge_cells(start_row=linha, start_column=1,
                           end_row=linha, end_column=largura_util)
        c = ws.cell(row=linha, column=1, value=rodape)
        c.font = Font(
            name=paleta.fonte_familia, size=paleta.fonte_tamanho_auxiliar,
            italic=True, color=paleta.cor_neutra_escura,
        )
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c.fill = PatternFill("solid", fgColor="FFFFFF")
        for col in range(1, largura_util + 1):
            ws.cell(row=linha, column=col).border = bordas
        linha += 1

    ws.row_dimensions[linha].height = 8
    return linha + 1
