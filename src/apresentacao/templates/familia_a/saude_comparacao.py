"""Sub-template Família A · Saúde da comparação · D-202 etapa 5.

Movido e parametrizado de
`visoes.exportacao_v2._renderizar_secao_saude_comparacao`.
"""
from __future__ import annotations

from typing import Dict, List, Optional

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from apresentacao.formatos import (
    FORMATO_CONTAGEM,
    FORMATO_PERCENTUAL,
    number_format_valor,
)
from apresentacao.tipografia import aplicar_valor

from ._shared import (
    LABEL_SEMANTICA_SAUDE,
    bordas_finas,
    categorias_saude_para_exibir,
    renderizar_cabecalho_secao,
)


def renderizar_saude_comparacao(
    *,
    ws: Worksheet,
    linha: int,
    largura_util: int,
    unidade: str,
    semantica: str,
    dist_sem: Optional[Dict[str, int]],
    delta_sem: Optional[Dict[str, float]],
    dist_estru: Optional[Dict[str, int]],
    paleta,
) -> int:
    """E2 · seção 'Saúde da comparação' · numérico.

    PERCENTUAL oculta coluna 'Δ total' (somar p.p de várias linhas viola C.D3).
    Demais unidades mantêm coluna usando `number_format_valor(unidade)`.

    Parâmetros explícitos (D-202):
      - dist_sem: {classif → contagem}
      - delta_sem: {classif → soma de Δ}
      - dist_estru: {classif_estrutural → contagem} (rodapé)
    """
    dist_sem = dist_sem or {}
    delta_sem = delta_sem or {}
    dist_estru = dist_estru or {}
    exibe_delta = unidade != "PERCENTUAL"

    linha = renderizar_cabecalho_secao(
        ws, linha, 1, largura_util, "Saúde da comparação", paleta,
    )

    bordas = bordas_finas(paleta)
    cor_zebra = paleta.cor_secundaria

    # Layout adaptativo
    if exibe_delta:
        col_cat_ini  = 1
        col_cat_fim  = largura_util - 5
        col_casos    = largura_util - 4
        col_part_ini = largura_util - 3
        col_part_fim = largura_util - 2
        col_delta_ini = largura_util - 1
        col_delta_fim = largura_util
    else:
        col_cat_ini  = 1
        col_cat_fim  = largura_util - 3
        col_casos    = largura_util - 2
        col_part_ini = largura_util - 1
        col_part_fim = largura_util
        col_delta_ini = 0
        col_delta_fim = 0

    def _header(linha_h: int) -> None:
        if col_cat_fim > col_cat_ini:
            ws.merge_cells(start_row=linha_h, start_column=col_cat_ini,
                           end_row=linha_h, end_column=col_cat_fim)
        ws.cell(row=linha_h, column=col_cat_ini,  value="Categoria")
        ws.cell(row=linha_h, column=col_casos,    value="Casos")
        if col_part_fim > col_part_ini:
            ws.merge_cells(start_row=linha_h, start_column=col_part_ini,
                           end_row=linha_h, end_column=col_part_fim)
        ws.cell(row=linha_h, column=col_part_ini, value="Participação")
        if exibe_delta:
            if col_delta_fim > col_delta_ini:
                ws.merge_cells(start_row=linha_h, start_column=col_delta_ini,
                               end_row=linha_h, end_column=col_delta_fim)
            ws.cell(row=linha_h, column=col_delta_ini, value="Δ total")
        for col in range(1, largura_util + 1):
            c = ws.cell(row=linha_h, column=col)
            c.fill = PatternFill("solid", fgColor=paleta.cor_secundaria)
            c.font = Font(
                name=paleta.fonte_familia, size=paleta.fonte_tamanho_corpo,
                bold=True, color=paleta.cor_destaque,
            )
            c.border = bordas
            if col == col_cat_ini:
                align_h = "left"
            elif col == col_casos:
                align_h = "center"
            else:
                align_h = "right"
            c.alignment = Alignment(horizontal=align_h, vertical="center", indent=1)

    _header(linha)
    linha += 1

    chaves = categorias_saude_para_exibir(semantica, dist_sem)
    total_pa = sum(int(dist_sem.get(k, 0)) for k in chaves) or 1

    idx_linha = 0
    for chave in chaves:
        n = int(dist_sem.get(chave, 0))
        delta = float(delta_sem.get(chave, 0.0))
        part = n / total_pa if total_pa else 0
        rot = LABEL_SEMANTICA_SAUDE.get(chave, chave)

        cor_fundo = cor_zebra if (idx_linha % 2 == 0) else "FFFFFF"

        if col_cat_fim > col_cat_ini:
            ws.merge_cells(start_row=linha, start_column=col_cat_ini,
                           end_row=linha, end_column=col_cat_fim)
        c_cat = ws.cell(row=linha, column=col_cat_ini, value=rot)
        aplicar_valor(c_cat, paleta)
        c_cat.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        c_n = ws.cell(row=linha, column=col_casos, value=n)
        c_n.number_format = FORMATO_CONTAGEM
        aplicar_valor(c_n, paleta)
        c_n.alignment = Alignment(horizontal="center", vertical="center")

        if col_part_fim > col_part_ini:
            ws.merge_cells(start_row=linha, start_column=col_part_ini,
                           end_row=linha, end_column=col_part_fim)
        c_part = ws.cell(row=linha, column=col_part_ini, value=part)
        c_part.number_format = FORMATO_PERCENTUAL
        aplicar_valor(c_part, paleta)
        c_part.alignment = Alignment(horizontal="right", vertical="center", indent=1)

        if exibe_delta:
            if col_delta_fim > col_delta_ini:
                ws.merge_cells(start_row=linha, start_column=col_delta_ini,
                               end_row=linha, end_column=col_delta_fim)
            c_delta = ws.cell(row=linha, column=col_delta_ini, value=delta)
            c_delta.number_format = number_format_valor(unidade)
            aplicar_valor(c_delta, paleta)
            c_delta.alignment = Alignment(horizontal="right", vertical="center", indent=1)

        for col in range(1, largura_util + 1):
            c = ws.cell(row=linha, column=col)
            c.fill = PatternFill("solid", fgColor=cor_fundo)
            c.border = bordas
        linha += 1
        idx_linha += 1

    # Rodapé · "Total comparado: X · Não comparáveis: Y"
    n_pa = int(dist_estru.get("PRESENTE_AMBOS", 0))
    n_aus_o = int(dist_estru.get("AUSENTE_ORIGEM", 0))
    n_aus_c = int(dist_estru.get("AUSENTE_COMPARADO", 0))
    n_nul_o = int(dist_estru.get("NULO_ORIGEM", 0))
    n_nul_c = int(dist_estru.get("NULO_COMPARADO", 0))
    n_nul_a = int(dist_estru.get("NULO_AMBOS", 0))
    n_nao_comparaveis = n_aus_o + n_aus_c + n_nul_o + n_nul_c + n_nul_a

    rodapes: List[str] = []
    rodapes.append(
        f"Total comparado: {n_pa:,} casos · Não comparáveis: {n_nao_comparaveis:,} casos"
        .replace(",", ".")
    )
    if n_nao_comparaveis > 0:
        partes_nao_comp: List[str] = []
        if n_aus_o:
            partes_nao_comp.append(f"{n_aus_o} ausente(s) na origem")
        if n_aus_c:
            partes_nao_comp.append(f"{n_aus_c} ausente(s) no comparado")
        if n_nul_o:
            partes_nao_comp.append(f"{n_nul_o} sem valor na origem")
        if n_nul_c:
            partes_nao_comp.append(f"{n_nul_c} sem valor no comparado")
        if n_nul_a:
            partes_nao_comp.append(f"{n_nul_a} sem valor em ambos")
        if partes_nao_comp:
            rodapes.append("Decomposição: " + " · ".join(partes_nao_comp) + ".")
        rodapes.append("ver detalhes na aba Diagnóstico")

    for rodape in rodapes:
        if largura_util > 1:
            ws.merge_cells(start_row=linha, start_column=1,
                           end_row=linha, end_column=largura_util)
        c = ws.cell(row=linha, column=1, value=rodape)
        c.font = Font(
            name=paleta.fonte_familia, size=paleta.fonte_tamanho_auxiliar,
            italic=True, color=paleta.cor_neutra_escura,
        )
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c.fill = PatternFill("solid", fgColor="FFFFFF")
        for col in range(1, largura_util + 1):
            ws.cell(row=linha, column=col).border = bordas
        linha += 1

    ws.row_dimensions[linha].height = 8
    return linha + 1
