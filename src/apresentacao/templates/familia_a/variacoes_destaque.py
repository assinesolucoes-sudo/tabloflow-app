"""Sub-template Família A · Variações em destaque · D-202 etapa 5.

Movido e parametrizado da seção de top variações + BarChart de
`visoes.exportacao_v2._renderizar_resumo_executivo_v2`.

Renderiza:
  - Tabela top variações (1 linha por item · agrupadores + valor_origem +
    valor_comparado + diferenca + variacao_percentual).
  - Bloco gráfico no final · BarChart 22cm × 16cm com tabela auxiliar
    Rótulo/Δ ao lado.
"""
from __future__ import annotations

from typing import Any, List, Tuple

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from apresentacao.formatos import (
    number_format_diferenca,
    number_format_valor,
    number_format_variacao,
    rotulo_diferenca,
    rotulo_variacao,
    valor_diferenca_para_celula,
)
from apresentacao.graficos import criar_grafico_top_variacoes
from apresentacao.tipografia import aplicar_valor

from ._shared import (
    bordas_finas,
    renderizar_cabecalho_secao,
    rotular_agrupador,
)


def renderizar_variacoes_destaque(
    *,
    ws: Worksheet,
    linha: int,
    largura_util: int,
    top_list: List[Any],
    agrupadores: List[str],
    unidade: str,
    origem_ux: str,
    comparado_ux: str,
    paleta,
) -> Tuple[int, List[Any]]:
    """Renderiza tabela top variações com cabeçalho colorido.

    Retorna (próxima_linha, top_list_efetiva). Quando top_list está vazio,
    renderiza linha única "Nenhuma variação significativa para destacar."
    e retorna lista vazia.
    """
    bordas = bordas_finas(paleta)
    cor_zebra = paleta.cor_secundaria

    linha = renderizar_cabecalho_secao(
        ws, linha, 1, largura_util, "Variações em destaque", paleta,
    )

    # Header de tabela top variações
    headers_top: List[Tuple[str, str]] = []
    for a in agrupadores:
        headers_top.append((a, rotular_agrupador(a)))
    headers_top.extend([
        ("valor_origem", f"Valor · {origem_ux}"),
        ("valor_comparado", f"Valor · {comparado_ux}"),
        ("diferenca", rotulo_diferenca(unidade)),
        ("variacao_percentual", rotulo_variacao(unidade)),
    ])
    n_cols_top = len(headers_top)
    for idx, (_tec, rot) in enumerate(headers_top, start=1):
        c = ws.cell(row=linha, column=idx, value=rot)
        c.fill = PatternFill("solid", fgColor=paleta.cor_secundaria)
        c.font = Font(
            name=paleta.fonte_familia, size=paleta.fonte_tamanho_corpo,
            bold=True, color=paleta.cor_destaque,
        )
        c.border = bordas
        c.alignment = Alignment(
            horizontal="left" if idx <= len(agrupadores) else "right",
            vertical="center", indent=1,
        )
    linha += 1

    if not top_list:
        if n_cols_top > 1:
            ws.merge_cells(start_row=linha, start_column=1,
                           end_row=linha, end_column=n_cols_top)
        c = ws.cell(row=linha, column=1, value="Nenhuma variação significativa para destacar.")
        aplicar_valor(c, paleta)
        c.fill = PatternFill("solid", fgColor="FFFFFF")
        for col in range(1, n_cols_top + 1):
            ws.cell(row=linha, column=col).border = bordas
        linha += 1
    else:
        for idx_top, t in enumerate(top_list):
            cor_fundo = cor_zebra if (idx_top % 2 == 0) else "FFFFFF"
            chave = getattr(t, "chave_agrupadores", {}) or {}
            if not isinstance(chave, dict):
                chave = {}
            col_idx = 1
            for a in agrupadores:
                val = chave.get(a)
                c = ws.cell(row=linha, column=col_idx)
                c.value = str(val) if val is not None else "—"
                aplicar_valor(c, paleta)
                c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
                col_idx += 1
            vo = getattr(t, "valor_origem", None)
            vc = getattr(t, "valor_comparado", None)
            dif = getattr(t, "diferenca", None)
            var = getattr(t, "variacao_percentual", None)

            def _escrever_num(c_pos, valor, fmt):
                c = ws.cell(row=linha, column=c_pos)
                if valor is not None:
                    c.value = valor
                    c.number_format = fmt
                else:
                    c.value = "—"
                aplicar_valor(c, paleta)
                c.alignment = Alignment(horizontal="right", vertical="center", indent=1)

            _escrever_num(col_idx, vo, number_format_valor(unidade))
            col_idx += 1
            _escrever_num(col_idx, vc, number_format_valor(unidade))
            col_idx += 1
            _escrever_num(
                col_idx,
                valor_diferenca_para_celula(dif, unidade),
                number_format_diferenca(unidade),
            )
            col_idx += 1
            _escrever_num(col_idx, var, number_format_variacao(unidade))
            col_idx += 1

            for col in range(1, n_cols_top + 1):
                c = ws.cell(row=linha, column=col)
                c.fill = PatternFill("solid", fgColor=cor_fundo)
                c.border = bordas
            linha += 1
    ws.row_dimensions[linha].height = 8
    linha += 1

    return linha, list(top_list)


def renderizar_grafico_variacoes(
    *,
    ws: Worksheet,
    linha: int,
    largura_util: int,
    top_list: List[Any],
    agrupadores: List[str],
    unidade: str,
    paleta,
) -> int:
    """Sub-sessão 8.4 · P-36 · gráfico isolado no final do Resumo Executivo.

    Tabela auxiliar Rótulo/Δ + BarChart 22cm × 16cm. Reserva linhas vazias
    suficientes para gráfico não vazar.
    """
    if not top_list:
        return linha

    bordas = bordas_finas(paleta)
    cor_zebra = paleta.cor_secundaria

    linha = renderizar_cabecalho_secao(
        ws, linha, 1, largura_util, "Variações em destaque · gráfico", paleta,
    )

    # Tabela auxiliar
    linha_aux_header = linha
    for col_idx, (rot, align) in enumerate(
        [("Rótulo", "left"), (rotulo_diferenca(unidade), "right")], start=1
    ):
        c = ws.cell(row=linha_aux_header, column=col_idx, value=rot)
        c.fill = PatternFill("solid", fgColor=paleta.cor_secundaria)
        c.font = Font(
            name=paleta.fonte_familia, size=paleta.fonte_tamanho_corpo,
            bold=True, color=paleta.cor_destaque,
        )
        c.border = bordas
        c.alignment = Alignment(horizontal=align, vertical="center", indent=1)
    linha_aux = linha_aux_header + 1
    for idx_aux, t in enumerate(top_list):
        chave = getattr(t, "chave_agrupadores", {}) or {}
        if not isinstance(chave, dict):
            chave = {}
        rotulo_compacto = " · ".join(str(chave.get(a, "—")) for a in agrupadores) or "—"
        dif = getattr(t, "diferenca", None)
        cor_fundo = cor_zebra if (idx_aux % 2 == 0) else "FFFFFF"
        c_r = ws.cell(row=linha_aux, column=1, value=rotulo_compacto)
        c_r.fill = PatternFill("solid", fgColor=cor_fundo)
        c_r.font = Font(
            name=paleta.fonte_familia, size=paleta.fonte_tamanho_corpo,
            bold=False, color=paleta.cor_neutra_escura,
        )
        c_r.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c_r.border = bordas
        dif_celula = valor_diferenca_para_celula(dif, unidade) if dif is not None else 0
        c_v = ws.cell(row=linha_aux, column=2, value=dif_celula)
        c_v.fill = PatternFill("solid", fgColor=cor_fundo)
        c_v.number_format = number_format_diferenca(unidade)
        c_v.font = Font(
            name=paleta.fonte_familia, size=paleta.fonte_tamanho_corpo,
            bold=False, color=paleta.cor_neutra_escura,
        )
        c_v.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        c_v.border = bordas
        linha_aux += 1
    linha_aux_dados_fim = linha_aux - 1

    # BarChart 22cm × 16cm ancorado em D{linha_aux_header}
    try:
        chart = criar_grafico_top_variacoes(
            ws, anchor_cell=f"D{linha_aux_header}",
            dados_range=f"A{linha_aux_header}:B{linha_aux_dados_fim}",
            paleta_nome=paleta.nome,
            titulo="Variações em destaque",
        )
        if chart is not None:
            fmt_eixo = number_format_diferenca(unidade)
            if unidade == "MONETARIO_BRL":
                fmt_eixo = 'R$ #,##0'
            elif unidade == "PERCENTUAL":
                fmt_eixo = '+0" p.p";[Red]-0" p.p"'
            elif unidade == "QUANTIDADE":
                fmt_eixo = '#,##0'
            elif unidade == "TEMPO_DIAS":
                fmt_eixo = '#,##0 "dias"'
            elif unidade == "TEMPO_HORAS":
                fmt_eixo = '#,##0 "h"'
            elif unidade == "MULTIPLICADOR":
                fmt_eixo = '0.0"x"'
            elif unidade == "RAZAO":
                fmt_eixo = '0.000'
            chart.x_axis.number_format = fmt_eixo
            chart.y_axis.delete = False
            chart.y_axis.majorTickMark = "out"
            chart.height = 16
            chart.width = 22
    except Exception:  # noqa: BLE001
        pass

    return max(linha_aux_dados_fim, linha_aux_header + 12) + 3
