"""Helpers visuais compartilhados entre os 5 sub-templates de Família A · D-202.

Movidos de `visoes.exportacao_v2` em D-202 etapa 5. Funções puras de
estilização (bordas · cabeçalhos · cards · seção como tabela · altura
de prosa). Constantes e helpers semânticos (LABEL_SEMANTICA_SAUDE ·
categorias_saude_para_exibir · contrair_de · rotular_agrupador).
"""
from __future__ import annotations

from typing import Dict, List, Tuple

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet


# ---------------------------------------------------------------------------
# Bordas e cabeçalhos
# ---------------------------------------------------------------------------

def bordas_finas(paleta) -> Border:
    """Borda fina nos 4 lados · usa cor_neutra_clara da paleta."""
    lado = Side(border_style="thin", color=paleta.cor_neutra_clara)
    return Border(left=lado, right=lado, top=lado, bottom=lado)


def renderizar_cabecalho_secao(
    ws: Worksheet, linha: int, col_ini: int, col_fim: int,
    titulo: str, paleta,
) -> int:
    """Cabeçalho de seção como banner (fill cor_primária · texto branco)."""
    ws.merge_cells(
        start_row=linha, start_column=col_ini,
        end_row=linha, end_column=col_fim,
    )
    c = ws.cell(row=linha, column=col_ini, value=titulo)
    c.fill = PatternFill("solid", fgColor=paleta.cor_primaria)
    c.font = Font(
        name=paleta.fonte_familia,
        size=paleta.fonte_tamanho_secao,
        bold=True,
        color="FFFFFF",
    )
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[linha].height = 22
    bordas = bordas_finas(paleta)
    for col in range(col_ini, col_fim + 1):
        ws.cell(row=linha, column=col).border = bordas
    return linha + 1


def renderizar_secao_como_tabela(
    ws: Worksheet,
    linha_inicial: int,
    titulo_secao: str,
    linhas_conteudo: List[Tuple],
    col_inicial: int,
    col_final: int,
    paleta,
    aplicar_zebra: bool = True,
) -> int:
    """Seção como bloco de tabela estilizada · cabeçalho colorido + corpo
    com zebra e bordas. Tuplas de 1 elemento → prosa mesclada · 2+ → par
    rótulo/valor. Retorna próxima linha livre."""
    cor_zebra = paleta.cor_secundaria
    bordas = bordas_finas(paleta)

    linha = renderizar_cabecalho_secao(
        ws, linha_inicial, col_inicial, col_final, titulo_secao, paleta,
    )

    for idx_linha, tupla in enumerate(linhas_conteudo):
        cor_fundo = cor_zebra if (idx_linha % 2 == 0 and aplicar_zebra) else "FFFFFF"
        if len(tupla) <= 1:
            texto = tupla[0] if tupla else ""
            if col_final > col_inicial:
                ws.merge_cells(
                    start_row=linha, start_column=col_inicial,
                    end_row=linha, end_column=col_final,
                )
            cel = ws.cell(row=linha, column=col_inicial, value=texto)
            cel.fill = PatternFill("solid", fgColor=cor_fundo)
            cel.font = Font(
                name=paleta.fonte_familia,
                size=paleta.fonte_tamanho_corpo,
                bold=False,
                color=paleta.cor_neutra_escura,
            )
            cel.alignment = Alignment(
                horizontal="left", vertical="center",
                wrap_text=True, indent=1,
            )
            for col in range(col_inicial, col_final + 1):
                c = ws.cell(row=linha, column=col)
                c.fill = PatternFill("solid", fgColor=cor_fundo)
                c.border = bordas
        else:
            rotulo = tupla[0]
            valor = tupla[1]
            c_r = ws.cell(row=linha, column=col_inicial, value=rotulo)
            c_r.fill = PatternFill("solid", fgColor=cor_fundo)
            c_r.font = Font(
                name=paleta.fonte_familia,
                size=paleta.fonte_tamanho_corpo,
                bold=True,
                color=paleta.cor_neutra_escura,
            )
            c_r.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            c_r.border = bordas
            if col_final > col_inicial:
                ws.merge_cells(
                    start_row=linha, start_column=col_inicial + 1,
                    end_row=linha, end_column=col_final,
                )
            c_v = ws.cell(row=linha, column=col_inicial + 1, value=valor)
            c_v.fill = PatternFill("solid", fgColor=cor_fundo)
            c_v.font = Font(
                name=paleta.fonte_familia,
                size=paleta.fonte_tamanho_corpo,
                bold=False,
                color=paleta.cor_neutra_escura,
            )
            c_v.alignment = Alignment(
                horizontal="left", vertical="center",
                wrap_text=True, indent=1,
            )
            for col in range(col_inicial + 1, col_final + 1):
                c = ws.cell(row=linha, column=col)
                c.fill = PatternFill("solid", fgColor=cor_fundo)
                c.border = bordas
        linha += 1

    ws.row_dimensions[linha].height = 8
    return linha + 1


def calcular_altura_leitura_qualitativa(texto: str) -> float:
    """Sessão 8.3 · C-1 · altura robusta para célula mesclada A:H com wrap_text.
    Excel não auto-fita altura de células mescladas · estimativa pessimista
    (90 chars/linha visual) com margem 1.4x."""
    if not texto:
        return 30.0
    chars_por_linha_visual = 90
    num_linhas = max(2, (len(texto) // chars_por_linha_visual) + 1)
    num_linhas_seguro = int(num_linhas * 1.4) + 1
    altura_por_linha_pt = 16
    altura_total = num_linhas_seguro * altura_por_linha_pt + 12
    return min(220, max(55, altura_total))


def mesclar_card(
    ws: Worksheet, linha_ini: int, col_ini: int, linha_fim: int, col_fim: int,
    rotulo: str, valor: str, paleta,
) -> None:
    """Card visual: linha superior = rótulo (cinza pequeno) ·
    linha inferior = valor (cor primária grande)."""
    if col_fim > col_ini:
        ws.merge_cells(
            start_row=linha_ini, start_column=col_ini,
            end_row=linha_ini, end_column=col_fim,
        )
    c_rot = ws.cell(row=linha_ini, column=col_ini, value=rotulo)
    c_rot.font = Font(
        name=paleta.fonte_familia, size=paleta.fonte_tamanho_auxiliar,
        bold=False, color=paleta.cor_neutra_escura,
    )
    c_rot.alignment = Alignment(horizontal="center", vertical="center")
    c_rot.fill = PatternFill("solid", fgColor=paleta.cor_secundaria)

    if linha_fim > linha_ini + 1 or col_fim > col_ini:
        ws.merge_cells(
            start_row=linha_ini + 1, start_column=col_ini,
            end_row=linha_fim, end_column=col_fim,
        )
    c_val = ws.cell(row=linha_ini + 1, column=col_ini, value=valor)
    c_val.font = Font(
        name=paleta.fonte_familia, size=paleta.fonte_tamanho_titulo + 2,
        bold=True, color=paleta.cor_primaria,
    )
    c_val.alignment = Alignment(horizontal="center", vertical="center")
    c_val.fill = PatternFill("solid", fgColor="FFFFFF")


# ---------------------------------------------------------------------------
# Helpers semânticos · Família A
# ---------------------------------------------------------------------------

# Rótulos user-facing para classificacao_semantica · 7 valores D-187.
LABEL_SEMANTICA_SAUDE: Dict[str, str] = {
    "POSITIVO":      "Melhorou",
    "NEGATIVO":      "Piorou",
    "NEUTRO":        "Estável",
    "NAO_APLICAVEL": "Não aplicável",
    "AUMENTOU":      "Aumentou",
    "REDUZIU":       "Reduziu",
    "ESTAVEL":       "Estável",
}

ORDEM_SAUDE_QUALITATIVA = ["POSITIVO", "NEGATIVO", "NEUTRO"]
ORDEM_SAUDE_NEUTRA      = ["AUMENTOU", "REDUZIU", "ESTAVEL"]


def categorias_saude_para_exibir(
    semantica_campo: str,
    dist_sem: Dict[str, int],
) -> List[str]:
    """Retorna a ordem das chaves a exibir conforme semantica_campo."""
    if semantica_campo == "NEUTRO":
        return [k for k in ORDEM_SAUDE_NEUTRA if dist_sem.get(k, 0) > 0]
    chaves = [k for k in ORDEM_SAUDE_QUALITATIVA if dist_sem.get(k, 0) > 0]
    if not chaves:
        return ORDEM_SAUDE_QUALITATIVA[:]
    return chaves


def contrair_de(rotulo: str) -> str:
    """Sessão 8.1 · resolve contração 'de o/a' → 'do/da' por 1ª letra.
    Heurística: vogal 'a' → 'da'; demais vogais → 'de'; consoante → 'do'."""
    if not rotulo:
        return f"de {rotulo}"
    primeira = rotulo[0].lower()
    if primeira in "aáàâã":
        return f"da {rotulo}"
    if primeira in "eéêiíoóôõuú":
        return f"de {rotulo}"
    return f"do {rotulo}"


def rotular_agrupador(nome_tec: str) -> str:
    """Humaniza nome de coluna técnica (Centro_Custo → Centro de Custo)."""
    if not nome_tec:
        return ""
    especiais = {
        "Centro_Custo": "Centro de Custo",
        "centro_custo": "Centro de Custo",
        "Mes": "Mês",
        "mes": "Mês",
    }
    if nome_tec in especiais:
        return especiais[nome_tec]
    texto = nome_tec.replace("_", " ").strip()
    if not texto:
        return nome_tec
    return texto[0].upper() + texto[1:]
