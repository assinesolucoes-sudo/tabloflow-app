"""
F-APRESENT · capability 8 · Badges semânticos (D-159).

Aplica fill + fonte coerentes com a paleta selecionada a uma célula contendo
classificação semântica (ex: PRESENTE_AMBOS · AUSENTE_ORIGEM · POSITIVO).

Design:
  - BadgeStyle · dataclass frozen · par (fill_hex, font_hex) · hasheável
  - CATALOGO_BADGES · dict {paleta_nome: {semantica: BadgeStyle}} · 4 paletas × 4
    semânticas (positivo · negativo · neutro · atencao) · cores declaradas
    sem perder identidade cromática de cada paleta
  - MAPEAMENTO_V2 · dict {classificacao_tecnica: semantica} · específico de V2
    · capability 8 não conhece visão · mapping é injetado · V1/V11 declaram os
    seus em arquivos próprios
  - aplicar_badge recebe mapping opcional · default MAPEAMENTO_V2 para compat
    · classificação desconhecida cai em fallback "neutro" + warning de log
    (C.5 · nunca quebra renderização)

Interface canônica (D-159):
    aplicar_badge(celula, classificacao, paleta)
"""
from __future__ import annotations

import logging
from dataclasses import dataclass
from typing import Dict, Literal, Mapping, Optional

from openpyxl.cell.cell import Cell
from openpyxl.styles import Font, PatternFill

from .paletas import Paleta


_logger = logging.getLogger(__name__)


Semantica = Literal["positivo", "negativo", "neutro", "atencao"]

SEMANTICAS_CANONICAS = ("positivo", "negativo", "neutro", "atencao")


@dataclass(frozen=True)
class BadgeStyle:
    """Par (fill, font) de um badge · hex RGB sem '#' · hasheável."""
    fill_hex: str
    font_hex: str

    def __post_init__(self) -> None:
        for campo, valor in (("fill_hex", self.fill_hex), ("font_hex", self.font_hex)):
            if not isinstance(valor, str) or len(valor) != 6:
                raise ValueError(
                    f"{campo}='{valor}' inválido · exigido hex RGB de 6 caracteres sem '#'"
                )
            try:
                int(valor, 16)
            except ValueError as exc:
                raise ValueError(f"{campo}='{valor}' não é hex válido") from exc


# ---------------------------------------------------------------------------
# Catálogo de badges por paleta × semântica
#
# Cada paleta tem 4 tons semânticos declarados. Cores escolhidas para:
#   - manter identidade cromática da paleta (azul mantém azul · vinho mantém
#     vinho · cinza monocromático · verde em tons de verde)
#   - produzir contraste WCAG AA com a fonte em cada badge
#   - atenção sempre em âmbar (F6C85F) · cross-paleta · universalmente lido
#     como "cuidado" em relatórios executivos
# ---------------------------------------------------------------------------

CATALOGO_BADGES: Dict[str, Dict[str, BadgeStyle]] = {
    "azul": {
        # verde escuro sobre secundária azul clara · sinal positivo legível
        "positivo": BadgeStyle(fill_hex="D8E8F5", font_hex="1E6E3B"),
        # vermelho escuro sobre fundo azul muito claro · negativo visível
        "negativo": BadgeStyle(fill_hex="F4E3E1", font_hex="9B1E14"),
        # neutro · cinza-azulado suave · fundo secundário azul claro
        "neutro":   BadgeStyle(fill_hex="EDF3F9", font_hex="2F2F2F"),
        # atenção · âmbar universal
        "atencao":  BadgeStyle(fill_hex="FFF4D1", font_hex="7A5A00"),
    },
    "verde": {
        "positivo": BadgeStyle(fill_hex="D4EAD8", font_hex="1B4D30"),
        "negativo": BadgeStyle(fill_hex="F4D9D5", font_hex="8F1C14"),
        # neutro · verde muito claro · preserva identidade cromática
        "neutro":   BadgeStyle(fill_hex="EDF3EE", font_hex="2F2F2F"),
        "atencao":  BadgeStyle(fill_hex="FFF4D1", font_hex="7A5A00"),
    },
    "cinza": {
        # monocromático · positivo verde suave dentro do cinza institucional
        "positivo": BadgeStyle(fill_hex="DDE6DD", font_hex="1B4D30"),
        "negativo": BadgeStyle(fill_hex="E6D4D4", font_hex="8F1C14"),
        "neutro":   BadgeStyle(fill_hex="EDEDED", font_hex="2F2F2F"),
        "atencao":  BadgeStyle(fill_hex="F7E9C4", font_hex="7A5A00"),
    },
    "vinho": {
        # tons quentes coerentes com vinho · positivo terra-verde
        "positivo": BadgeStyle(fill_hex="D9E4D4", font_hex="1B4D30"),
        # negativo mais profundo · vinho escuro próprio
        "negativo": BadgeStyle(fill_hex="E8CAD0", font_hex="6D1F2D"),
        "neutro":   BadgeStyle(fill_hex="F2E1E4", font_hex="2F2F2F"),
        "atencao":  BadgeStyle(fill_hex="F5E0C3", font_hex="7A5A00"),
    },
}


# ---------------------------------------------------------------------------
# Mapeamento classificação técnica → semântica (por visão)
#
# capability 8 é genérica · cada visão declara o seu mapping. V1/V11
# declararão os seus em arquivos próprios quando forem implementadas.
# ---------------------------------------------------------------------------

MAPEAMENTO_V2: Dict[str, Semantica] = {
    # Classificação estrutural · D-022/D-023 · V2
    "PRESENTE_AMBOS":    "neutro",
    "AUSENTE_ORIGEM":    "atencao",
    "AUSENTE_COMPARADO": "atencao",
    "NULO_ORIGEM":       "atencao",
    "NULO_COMPARADO":    "atencao",
    "NULO_AMBOS":        "atencao",

    # Classificação semântica · T-SEMA · V2
    "POSITIVO":       "positivo",
    "NEGATIVO":       "negativo",
    "NEUTRO":         "neutro",
    "NAO_APLICAVEL":  "neutro",
    # NEUTRA × estrutural (D-187 · semantica sem viés · badge neutra)
    "AUMENTOU":       "neutro",
    "REDUZIU":        "neutro",

    # Sinônimos de superfície · apoiam renderização livre
    "AUMENTO":       "positivo",
    "CRESCIMENTO":   "positivo",
    "ACIMA":         "positivo",
    "REDUCAO":       "negativo",
    "QUEDA":         "negativo",
    "ABAIXO":        "negativo",
    "ESTAVEL":       "neutro",
    "NORMAL":        "neutro",
    "NA_FAIXA":      "neutro",
    "AUSENTE":       "atencao",
    "INCONSISTENTE": "atencao",
    "INDETERMINADO": "atencao",
}


def _resolver_semantica(
    classificacao: str,
    mapeamento: Mapping[str, str],
) -> Semantica:
    """Resolve classificação para semântica · fallback 'neutro' com warning."""
    if classificacao is None:
        _logger.warning(
            "classificacao None em aplicar_badge · fallback 'neutro' (C.5)"
        )
        return "neutro"
    chave = str(classificacao).strip().upper()
    if chave in mapeamento:
        sem = mapeamento[chave]
        if sem not in SEMANTICAS_CANONICAS:
            _logger.warning(
                "mapping entregou semântica '%s' fora do catálogo · fallback 'neutro'",
                sem,
            )
            return "neutro"
        return sem  # type: ignore[return-value]
    _logger.warning(
        "classificação '%s' sem mapeamento · fallback 'neutro' (C.5)",
        classificacao,
    )
    return "neutro"


def _obter_estilo(paleta: Paleta, semantica: Semantica) -> BadgeStyle:
    if paleta.nome not in CATALOGO_BADGES:
        _logger.warning(
            "paleta '%s' sem catálogo de badges · usando azul como fallback",
            paleta.nome,
        )
        paleta_cat = CATALOGO_BADGES["azul"]
    else:
        paleta_cat = CATALOGO_BADGES[paleta.nome]
    if semantica not in paleta_cat:
        _logger.warning(
            "semântica '%s' ausente na paleta '%s' · usando neutro",
            semantica, paleta.nome,
        )
        return paleta_cat["neutro"]
    return paleta_cat[semantica]


def aplicar_badge(
    celula: Cell,
    classificacao: str,
    paleta: Paleta,
    mapeamento: Optional[Mapping[str, str]] = None,
) -> None:
    """
    Aplica fill + fonte coerentes com a paleta selecionada a uma célula
    contendo classificação semântica.

    Interface canônica D-159:
        aplicar_badge(celula, classificacao, paleta)

    Argumento opcional `mapeamento` permite que cada visão declare seu
    próprio mapa classificação → semântica. Default: MAPEAMENTO_V2 (V2 é
    a primeira visão que consome F-APRESENT; V1/V11 declararão os seus).

    Comportamento:
      - classificação mapeada · aplica BadgeStyle correspondente
      - classificação desconhecida · fallback 'neutro' + warning em log (C.5)
      - celula=None · TypeError explícito (C.2)
      - paleta=None · TypeError explícito (C.2)

    Compatibilidade com Table (capability 3): a aplicação de PatternFill +
    Font numa célula DENTRO de um openpyxl.Table não quebra autoFilter nem
    totalsRow · sobrescreve o estilo da linha mas mantém a célula como
    parte do range da Table.
    """
    if celula is None:
        raise TypeError("celula não pode ser None em aplicar_badge")
    if paleta is None:
        raise TypeError("paleta não pode ser None em aplicar_badge")

    mapa = mapeamento if mapeamento is not None else MAPEAMENTO_V2
    semantica = _resolver_semantica(classificacao, mapa)
    estilo = _obter_estilo(paleta, semantica)

    celula.fill = PatternFill(
        fill_type="solid",
        start_color=estilo.fill_hex,
        end_color=estilo.fill_hex,
    )
    celula.font = Font(
        name=paleta.fonte_familia,
        size=paleta.fonte_tamanho_corpo,
        bold=True,
        color=estilo.font_hex,
    )
