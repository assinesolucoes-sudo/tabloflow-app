"""
F-APRESENT · capability 7 · Resumo Executivo narrativo (D-125 · D-163 · OBS-10).

Renderiza ResumoExecutivoPadrao como prosa executiva em worksheet Excel ·
substitui o JSON dump atual (OBS-10 de VV-V2).

Estrutura: título + 6 blocos fixos (D-125) com traduções user-facing ·
estilo aplicado conforme Paleta · ZERO vazamento técnico.

A suite de invariantes varre cada amostra célula-a-célula procurando
padrões do bloco 7 do vocabulário · esta função deve produzir saída
100% user-facing.
"""
from __future__ import annotations

import logging
from typing import Any, Dict, List, Optional, Tuple

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from contratos import ResumoExecutivoPadrao

from .formatos import (
    FORMATO_CONTAGEM,
    FORMATO_MONETARIO_BR,
    FORMATO_PERCENTUAL,
    formatar_data_br,
    formatar_moeda_br,
    formatar_percentual_br,
)
from .paletas import Paleta
from .vocabulario import eh_termo_proibido, traduzir


_logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Estilos derivados da Paleta
# ---------------------------------------------------------------------------

def _font_titulo(paleta: Paleta) -> Font:
    return Font(
        name=paleta.fonte_familia,
        size=paleta.fonte_tamanho_titulo,
        bold=True,
        color=paleta.cor_neutra_escura,
    )


def _font_subtitulo(paleta: Paleta) -> Font:
    return Font(
        name=paleta.fonte_familia,
        size=paleta.fonte_tamanho_secao,
        bold=True,
        color=paleta.cor_primaria,
    )


def _font_corpo(paleta: Paleta) -> Font:
    return Font(
        name=paleta.fonte_familia,
        size=paleta.fonte_tamanho_corpo,
        color=paleta.cor_neutra_escura,
    )


def _font_corpo_bold(paleta: Paleta) -> Font:
    return Font(
        name=paleta.fonte_familia,
        size=paleta.fonte_tamanho_corpo,
        bold=True,
        color=paleta.cor_neutra_escura,
    )


def _font_tabela_header(paleta: Paleta) -> Font:
    return Font(
        name=paleta.fonte_familia,
        size=paleta.fonte_tamanho_corpo,
        bold=True,
        color=paleta.cor_texto_sobre_primaria,
    )


def _fill_primaria(paleta: Paleta) -> PatternFill:
    return PatternFill(
        fill_type="solid",
        start_color=paleta.cor_primaria,
        end_color=paleta.cor_primaria,
    )


def _fill_secundaria(paleta: Paleta) -> PatternFill:
    return PatternFill(
        fill_type="solid",
        start_color=paleta.cor_secundaria,
        end_color=paleta.cor_secundaria,
    )


# ---------------------------------------------------------------------------
# Tradução resiliente
# ---------------------------------------------------------------------------

def _traduzir_com_contextos(
    termo: str,
    contextos: List[str],
    vocabulario: Dict[str, Dict[str, str]],
) -> str:
    """Tenta traduzir procurando em lista ordenada de contextos, depois global."""
    for ctx in contextos:
        bloco = vocabulario.get(ctx, {})
        if termo in bloco:
            return bloco[termo]
    # Global (sem contexto)
    return traduzir(termo, contexto=None, vocabulario=vocabulario)


def _rotular_visao(visao_id: str, vocabulario: Dict[str, Dict[str, str]]) -> str:
    """Traduz identificador técnico 'V2' → 'Análise Comparativa entre Referências' etc."""
    rotulos = {
        "V1": "Análise de Conciliação",
        "V2": "Análise Comparativa entre Referências",
        "V11": "Análise de Aderência",
    }
    return rotulos.get(visao_id, f"Análise {visao_id}")


def _rotular_modo_upload(modo: str, vocabulario: Dict[str, Dict[str, str]]) -> str:
    """SIMPLES/DUAL não estão em vocabulário v1 · fallback explícito user-facing."""
    if modo == "SIMPLES":
        return "Envio simples (um arquivo)"
    if modo == "DUAL":
        return "Envio em duas referências (dois arquivos)"
    _logger.warning("modo_upload desconhecido: %s", modo)
    return f"[{modo}]"


def _rotular_agrupadores(agrupadores: List[str]) -> str:
    """Lista de agrupadores formatada como 'Mes · Loja · Produto'."""
    if not agrupadores:
        return "(sem agrupadores)"
    return " · ".join(str(a) for a in agrupadores)


def _rotular_threshold(chave: str, vocabulario: Dict[str, Dict[str, str]]) -> str:
    """Thresholds têm bloco próprio · fallback marcador."""
    bloco = vocabulario.get("thresholds", {})
    if chave in bloco:
        return bloco[chave]
    _logger.warning("threshold sem tradução: %s", chave)
    return f"[{chave}]"


def _rotular_classificacao(
    codigo: str, vocabulario: Dict[str, Dict[str, str]]
) -> str:
    bloco = vocabulario.get("classificacoes", {})
    if codigo in bloco:
        return bloco[codigo]
    # Não warning · visão pode trazer classificações custom
    return codigo.replace("_", " ").title()


def _rotular_categoria_warning(
    codigo: str, vocabulario: Dict[str, Dict[str, str]]
) -> str:
    """Categorias de warning · tradução leve."""
    traducoes = {
        "informativo": "Informativos",
        "ajuste_leve": "Ajustes leves aplicados",
        "decisao_usuario": "Decisões do usuário",
        "ausencia": "Ausências detectadas",
        "nulo": "Valores ausentes",
    }
    return traducoes.get(codigo, codigo.replace("_", " ").capitalize())


def _rotular_numero_ancora(chave: str) -> str:
    """Chaves do bloco 2 (dict Dict[str, Union[float,int,str]]) · rótulos user-facing.

    Usa acentuação portuguesa (cedilha · til) para evitar colisão com
    identificadores Python ascii-only do bloco 7.1.
    """
    traducoes = {
        "total_origem": "Total na origem",
        "total_comparado": "Total no comparado",
        "diferenca_total": "Diferença total",
        "diferenca": "Diferença",
        "variacao_total": "Variação total",
        "variacao_percentual": "Variação",
        "total_casos": "Total de casos analisados",
        "total_linhas_base_analitica": "Total de linhas da base analítica",
        "top_variacoes": "Variações em destaque",
        "valor_origem": "Valor na origem",
        "valor_comparado": "Valor no comparado",
        "chave_agrupadores": "Agrupadores",
    }
    if chave in traducoes:
        return traducoes[chave]
    # Fallback: humanizar snake_case → "Capitalizar Palavras"
    return " ".join(p.capitalize() for p in chave.split("_"))


# ---------------------------------------------------------------------------
# Heurística de formatação de valor
# ---------------------------------------------------------------------------

def _eh_monetario_por_chave(chave: str) -> bool:
    marcadores = ("total", "valor", "receita", "diferenca", "ancora", "soma")
    chave_low = chave.lower()
    return any(m in chave_low for m in marcadores) and "variacao" not in chave_low

def _eh_percentual_por_chave(chave: str) -> bool:
    marcadores = ("variacao", "pct", "percentual", "taxa", "proporcao")
    return any(m in chave.lower() for m in marcadores)


def _formatar_valor_com_heuristica(chave: str, valor: Any) -> str:
    if isinstance(valor, str):
        # Sanitizar: se string tem padrão proibido, deixamos como está e
        # a suite de invariantes vai flagar (o motivo de estar flagando).
        return valor
    if valor is None:
        return "-"
    # D-202 · dict (ex: chave_agrupadores) deve render como pares legíveis ·
    # nunca cair em str(dict) que produz `{'col': 'val'}` na célula.
    if isinstance(valor, dict):
        if not valor:
            return "-"
        partes = [
            f"{_rotular_numero_ancora(str(sk))}: {_formatar_valor_com_heuristica(str(sk), sv)}"
            for sk, sv in valor.items()
        ]
        return " · ".join(partes)
    if isinstance(valor, (list, tuple)):
        if not valor:
            return "-"
        return " · ".join(_formatar_valor_com_heuristica(chave, v) for v in valor)
    if _eh_percentual_por_chave(chave):
        return formatar_percentual_br(valor, conversao_fracao=True)
    if _eh_monetario_por_chave(chave):
        return formatar_moeda_br(valor)
    if isinstance(valor, (int,)) or (isinstance(valor, float) and float(valor).is_integer()):
        n = int(valor)
        return f"{n:,}".replace(",", ".")
    try:
        return f"{float(valor):,.2f}".replace(",", "@").replace(".", ",").replace("@", ".")
    except (TypeError, ValueError):
        return str(valor)


# ---------------------------------------------------------------------------
# Utilitários de escrita na worksheet
# ---------------------------------------------------------------------------

def _merge_e_escrever(
    ws: Worksheet, linha: int, col_ini: int, col_fim: int, valor: str, font: Font,
    alignment: Optional[Alignment] = None, fill: Optional[PatternFill] = None,
) -> int:
    """Mescla colunas da linha · escreve valor · aplica estilo. Retorna próxima linha."""
    if col_fim > col_ini:
        ws.merge_cells(
            start_row=linha, start_column=col_ini,
            end_row=linha, end_column=col_fim,
        )
    c = ws.cell(row=linha, column=col_ini, value=valor)
    c.font = font
    if alignment is not None:
        c.alignment = alignment
    if fill is not None:
        c.fill = fill
    # Aplica fill também às demais células mescladas (openpyxl requer)
    for col in range(col_ini + 1, col_fim + 1):
        cc = ws.cell(row=linha, column=col)
        if fill is not None:
            cc.fill = fill
    return linha + 1


def _escrever_tabela_2col(
    ws: Worksheet, linha_inicial: int, col_ini: int,
    linhas: List[Tuple[str, str]], paleta: Paleta,
) -> int:
    """
    Escreve tabela label/valor de 2 colunas com header da paleta.
    Linhas alternadas recebem fill secundário para legibilidade.
    Larguras de coluna NÃO são tocadas aqui · setadas uma única vez ao
    final de renderizar_resumo_executivo com valores fixos documentados.
    Retorna próxima linha livre.
    """
    linha = linha_inicial
    for i, (label, valor) in enumerate(linhas):
        c_label = ws.cell(row=linha, column=col_ini, value=label)
        c_valor = ws.cell(row=linha, column=col_ini + 1, value=valor)
        c_label.font = _font_corpo_bold(paleta)
        c_valor.font = _font_corpo(paleta)
        c_valor.alignment = Alignment(horizontal="right", vertical="center")
        c_label.alignment = Alignment(wrap_text=True, vertical="center")
        if i % 2 == 0:
            fill = _fill_secundaria(paleta)
            c_label.fill = fill
            c_valor.fill = fill
        linha += 1
    return linha


def _escrever_tabela_3col(
    ws: Worksheet, linha_inicial: int, col_ini: int,
    linhas: List[Tuple[str, str, str]], paleta: Paleta,
    headers: Optional[Tuple[str, str, str]] = None,
) -> int:
    linha = linha_inicial
    if headers is not None:
        for j, h in enumerate(headers):
            c = ws.cell(row=linha, column=col_ini + j, value=h)
            c.font = _font_tabela_header(paleta)
            c.fill = _fill_primaria(paleta)
            c.alignment = Alignment(horizontal="center", vertical="center")
        linha += 1
    for i, tripla in enumerate(linhas):
        for j, v in enumerate(tripla):
            c = ws.cell(row=linha, column=col_ini + j, value=v)
            c.font = _font_corpo_bold(paleta) if j == 0 else _font_corpo(paleta)
            if j == 0:
                c.alignment = Alignment(wrap_text=True, vertical="center")
            else:
                c.alignment = Alignment(horizontal="right", vertical="center")
            if i % 2 == 0:
                c.fill = _fill_secundaria(paleta)
        linha += 1

    return linha


# ---------------------------------------------------------------------------
# Renderizador principal
# ---------------------------------------------------------------------------

def renderizar_resumo_executivo(
    ws: Worksheet,
    resumo: ResumoExecutivoPadrao,
    paleta: Paleta,
    vocabulario: Dict[str, Dict[str, str]],
) -> None:
    """
    Renderiza ResumoExecutivoPadrao como prosa executiva na worksheet.

    Estrutura (topo → base):
      Linha 1: Título "Resumo Executivo · {Visão} · {Data BR}"
      Bloco 1: Como foi analisado (parágrafos)
      Bloco 2: Números principais (tabela 2 col)
      Bloco 3: Como os casos se distribuem (tabela 3 col)
      Bloco 4: O que chama atenção (lista adaptativa)
      Bloco 5: Leitura qualitativa (parágrafo + thresholds)
      Bloco 6: Qualidade do dado (tabela 2 col)

    Regras invioláveis (bloco 7 de vocabulario_bilingue.md):
      - Nenhum valor literal em CAPS_SNAKE, datetime(...), D-XXX, {'...', ...
      - Percentuais sempre via formatar_percentual_br (ou number_format nativo)
      - Enums sempre traduzidos via vocabulário
    """
    if ws is None:
        raise TypeError("ws não pode ser None em renderizar_resumo_executivo")
    if paleta is None:
        raise TypeError("paleta não pode ser None em renderizar_resumo_executivo")
    if vocabulario is None:
        raise TypeError("vocabulario não pode ser None em renderizar_resumo_executivo")

    cab = resumo.bloco_1_cabecalho
    visao_rotulo = _rotular_visao(cab.visao, vocabulario)
    data_br = formatar_data_br(cab.data_execucao)

    COL_INI = 1
    COL_FIM = 4  # 4 colunas de largura útil

    linha = 1

    # Título
    titulo = f"Resumo Executivo · {visao_rotulo} · {data_br}"
    linha = _merge_e_escrever(
        ws, linha, COL_INI, COL_FIM, titulo,
        font=_font_titulo(paleta),
        alignment=Alignment(horizontal="left", vertical="center"),
    )
    ws.row_dimensions[linha - 1].height = 28
    linha += 1  # respiro

    # BLOCO 1 · Como foi analisado
    linha = _merge_e_escrever(
        ws, linha, COL_INI, COL_FIM, "Como foi analisado",
        font=_font_subtitulo(paleta),
    )
    frase_1 = (
        f"Análise executada em {data_br}, no modo "
        f"{_rotular_modo_upload(cab.modo_upload, vocabulario)}."
    )
    linha = _merge_e_escrever(
        ws, linha, COL_INI, COL_FIM, frase_1, font=_font_corpo(paleta),
    )

    if cab.agrupadores:
        frase_2 = (
            f"Agrupadores aplicados: {_rotular_agrupadores(cab.agrupadores)}."
        )
    else:
        frase_2 = "Nenhum agrupador foi aplicado."
    linha = _merge_e_escrever(
        ws, linha, COL_INI, COL_FIM, frase_2, font=_font_corpo(paleta),
    )

    if cab.medida_principal:
        frase_3 = f"Medida principal: {cab.medida_principal}."
        linha = _merge_e_escrever(
            ws, linha, COL_INI, COL_FIM, frase_3, font=_font_corpo(paleta),
        )

    linha += 1

    # BLOCO 2 · Números principais
    linha = _merge_e_escrever(
        ws, linha, COL_INI, COL_FIM, "Números principais",
        font=_font_subtitulo(paleta),
    )
    numeros_linhas: List[Tuple[str, str]] = []
    for chave, valor in resumo.bloco_2_numeros_ancora.items():
        label = _rotular_numero_ancora(chave)
        valor_fmt = _formatar_valor_com_heuristica(chave, valor)
        numeros_linhas.append((label, valor_fmt))
    if numeros_linhas:
        linha = _escrever_tabela_2col(
            ws, linha, COL_INI, numeros_linhas, paleta,
        )
    else:
        linha = _merge_e_escrever(
            ws, linha, COL_INI, COL_FIM, "(sem números-âncora publicados pela visão)",
            font=_font_corpo(paleta),
        )
    linha += 1

    # BLOCO 3 · Como os casos se distribuem
    linha = _merge_e_escrever(
        ws, linha, COL_INI, COL_FIM, "Como os casos se distribuem",
        font=_font_subtitulo(paleta),
    )
    dist = resumo.bloco_3_distribuicao
    total_casos = sum(
        v for v in dist.values() if isinstance(v, (int, float))
    ) or 0
    linhas_dist: List[Tuple[str, str, str]] = []
    for codigo, qtd in dist.items():
        rotulo = _rotular_classificacao(codigo, vocabulario)
        if isinstance(qtd, (int, float)):
            qtd_fmt = f"{int(qtd):,}".replace(",", ".")
            pct = (qtd / total_casos) if total_casos else 0
            pct_fmt = formatar_percentual_br(pct, conversao_fracao=True)
        else:
            qtd_fmt = str(qtd)
            pct_fmt = "-"
        linhas_dist.append((rotulo, qtd_fmt, pct_fmt))
    if linhas_dist:
        linha = _escrever_tabela_3col(
            ws, linha, COL_INI, linhas_dist, paleta,
            headers=("Categoria", "Casos", "Participação"),
        )
    else:
        linha = _merge_e_escrever(
            ws, linha, COL_INI, COL_FIM,
            "(sem distribuição publicada pela visão)",
            font=_font_corpo(paleta),
        )
    linha += 1

    # BLOCO 4 · O que chama atenção
    linha = _merge_e_escrever(
        ws, linha, COL_INI, COL_FIM, "O que chama atenção",
        font=_font_subtitulo(paleta),
    )
    elementos = resumo.bloco_4_elementos_destacados
    if not elementos:
        linha = _merge_e_escrever(
            ws, linha, COL_INI, COL_FIM, "Nenhum elemento destacado.",
            font=_font_corpo(paleta),
        )
    else:
        for grupo_nome, grupo_conteudo in elementos.items():
            rotulo_grupo = _rotular_numero_ancora(grupo_nome)
            linha = _merge_e_escrever(
                ws, linha, COL_INI, COL_FIM, rotulo_grupo,
                font=_font_corpo_bold(paleta),
            )
            if isinstance(grupo_conteudo, list):
                linhas_tabela: List[Tuple[str, str, str]] = []
                for item in grupo_conteudo[:10]:
                    if isinstance(item, dict):
                        rotulo = item.get("rotulo") or item.get("label") or "-"
                        valores = [
                            f"{k}: {_formatar_valor_com_heuristica(k, v)}"
                            for k, v in item.items()
                            if k not in ("rotulo", "label")
                        ]
                        # Primeira chave de valor vira coluna do meio
                        valor_a = valores[0] if valores else "-"
                        valor_b = valores[1] if len(valores) > 1 else ""
                        # Formatar "k: v" mantendo user-facing: traduz k
                        # se achar em número-âncora heurístico
                        def _render(texto: str) -> str:
                            if not texto or ":" not in texto:
                                return texto
                            k, v = texto.split(":", 1)
                            return f"{_rotular_numero_ancora(k.strip())}: {v.strip()}"
                        valor_a_fmt = _render(valor_a)
                        valor_b_fmt = _render(valor_b)
                        linhas_tabela.append((str(rotulo), valor_a_fmt, valor_b_fmt))
                    else:
                        linhas_tabela.append((str(item), "", ""))
                if linhas_tabela:
                    linha = _escrever_tabela_3col(
                        ws, linha, COL_INI, linhas_tabela, paleta,
                    )
            elif isinstance(grupo_conteudo, dict):
                pares = [
                    (_rotular_numero_ancora(k), _formatar_valor_com_heuristica(k, v))
                    for k, v in grupo_conteudo.items()
                ]
                linha = _escrever_tabela_2col(
                    ws, linha, COL_INI, pares, paleta,
                )
            else:
                linha = _merge_e_escrever(
                    ws, linha, COL_INI, COL_FIM, str(grupo_conteudo),
                    font=_font_corpo(paleta),
                )
            linha += 1
    linha += 1

    # BLOCO 5 · Leitura qualitativa
    linha = _merge_e_escrever(
        ws, linha, COL_INI, COL_FIM, "Leitura qualitativa",
        font=_font_subtitulo(paleta),
    )
    leitura = resumo.bloco_5_leitura_qualitativa
    classif_fmt = _rotular_classificacao(leitura.classificacao_ativa, vocabulario)
    frase_leitura = f"Classificação qualitativa ativa: {classif_fmt}."
    linha = _merge_e_escrever(
        ws, linha, COL_INI, COL_FIM, frase_leitura,
        font=_font_corpo(paleta),
    )
    if leitura.alguma_leitura_alterada_por_edicao:
        linha = _merge_e_escrever(
            ws, linha, COL_INI, COL_FIM,
            "Atenção: limites editados pelo usuário afetaram esta leitura.",
            font=_font_corpo_bold(paleta),
        )
    if leitura.thresholds_usados:
        linha = _merge_e_escrever(
            ws, linha, COL_INI, COL_FIM, "Limites aplicados:",
            font=_font_corpo_bold(paleta),
        )
        pares_thresh: List[Tuple[str, str]] = []
        for chave, valor in leitura.thresholds_usados.items():
            pares_thresh.append((
                _rotular_threshold(chave, vocabulario),
                formatar_percentual_br(valor, conversao_fracao=True),
            ))
        linha = _escrever_tabela_2col(ws, linha, COL_INI, pares_thresh, paleta)
    linha += 1

    # BLOCO 6 · Qualidade do dado
    linha = _merge_e_escrever(
        ws, linha, COL_INI, COL_FIM, "Qualidade do dado",
        font=_font_subtitulo(paleta),
    )
    qual = resumo.bloco_6_qualidade_estrutural
    qualidade_linhas: List[Tuple[str, str]] = [
        ("Total de avisos emitidos", f"{int(qual.total_warnings):,}".replace(",", ".")),
        ("Ajustes automáticos aplicados", f"{int(qual.ajustes_aplicados):,}".replace(",", ".")),
        ("Há bloqueios escapados pelo usuário?", "Sim" if qual.tem_bloqueios_escapados else "Não"),
    ]
    for cat, qtd in qual.warnings_por_categoria.items():
        rotulo_cat = _rotular_categoria_warning(cat, vocabulario)
        qualidade_linhas.append((
            f"Avisos · {rotulo_cat}",
            f"{int(qtd):,}".replace(",", "."),
        ))
    linha = _escrever_tabela_2col(ws, linha, COL_INI, qualidade_linhas, paleta)

    # -----------------------------------------------------------------------
    # Larguras de coluna fixas · declaradas uma vez aqui.
    #
    # Valores estáveis escolhidos para que o Resumo Executivo abra LEGÍVEL
    # sem necessidade de ajuste manual (defeito identificado em P0):
    #   A = 42 · labels de tabela e primeira coluna de todo bloco · precisa
    #           comportar rótulos longos como "Limite de variação extrema
    #           (variações maiores são destacadas)" (~60 chars · com
    #           wrap_text cabe em 2 linhas de 42)
    #   B = 55 · valores (monetários "R$ 1.234.567,89" · prosa em blocos
    #           quando mesclada com C/D · conteúdo principal)
    #   C = 24 · terceira coluna de tabelas 3-col (contagens · participação)
    #   D = 24 · quarta coluna de tabelas 3-col ou limite da mescla de prosa
    # -----------------------------------------------------------------------
    LARGURAS_COLUNAS = {"A": 42, "B": 55, "C": 24, "D": 24}
    for letra, largura in LARGURAS_COLUNAS.items():
        ws.column_dimensions[letra].width = largura

    # Oculta gridlines para visual limpo
    ws.sheet_view.showGridLines = False
