"""
F-APRESENT · Gerador das 4 amostras de demonstração (D-164).

Uso:
    python -m src.apresentacao.demo_amostras

Produz 4 arquivos em demos/:
    amostra_paleta_azul.xlsx
    amostra_paleta_verde.xlsx
    amostra_paleta_cinza.xlsx
    amostra_paleta_vinho.xlsx

Cada arquivo tem 4 abas:
    1. Resumo Executivo (capability 7)
    2. Base Analítica (capabilities 3 · 4 · 5 · 6)
    3. Coração Visual (placeholder · P1)
    4. Diagnóstico (placeholder · P1)

As 4 amostras compartilham o mesmo V2Result-like simulado · só a paleta muda ·
permite inspeção lado a lado pela Usuária.
"""
from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from contratos import (
    CabecalhoExecucao,
    LeituraQualitativa,
    QualidadeEstrutural,
    ResumoExecutivoPadrao,
)

from .colunas_adaptativas import (
    ColunaAdaptativa,
    montar_colunas_adaptativas,
    se_config_igual,
    sempre,
)
from .formatos import (
    FORMATO_CONTAGEM,
    FORMATO_MONETARIO_BR,
    FORMATO_PERCENTUAL,
    aplicar_formato_contagem,
    aplicar_formato_monetario,
    aplicar_formato_percentual,
)
from .paletas import CATALOGO_PALETAS, Paleta, aplicar_paleta
from .resumo_executivo import renderizar_resumo_executivo
from .tabelas import criar_tabela_executiva
from .vocabulario import carregar_vocabulario_bilingue, traduzir


_logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# V2Result simulado · mesmo em todas as amostras (só paleta muda)
# ---------------------------------------------------------------------------

def _construir_base_analitica_demo() -> pd.DataFrame:
    """Base Analítica sintética com estrutura representativa de V2."""
    linhas: List[Dict[str, Any]] = []
    meses = ["2025-01", "2025-02", "2025-03"]
    lojas = ["Loja Centro", "Loja Norte", "Loja Sul", "Loja Leste"]
    import random
    random.seed(42)

    for mes in meses:
        for loja in lojas:
            origem = round(random.uniform(500.0, 5000.0), 2)
            comparado = round(origem * random.uniform(0.85, 1.20), 2)
            diferenca = round(comparado - origem, 2)
            variacao = round(diferenca / origem if origem else 0.0, 4)

            # Mistura algumas classificações estruturais simuladas
            r = random.random()
            if r < 0.75:
                classif_estrutural = "PRESENTE_AMBOS"
            elif r < 0.85:
                classif_estrutural = "AUSENTE_ORIGEM"
                origem = 0.0
                diferenca = comparado
            elif r < 0.92:
                classif_estrutural = "AUSENTE_COMPARADO"
                comparado = 0.0
                diferenca = -origem
            else:
                classif_estrutural = "NULO_ORIGEM"

            linhas.append({
                "chave_agrupadores": f"{mes}|{loja}",
                "valor_origem": origem,
                "valor_comparado": comparado,
                "diferenca": diferenca,
                "variacao_percentual": variacao,
                "classificacao_estrutural": classif_estrutural,
            })

    return pd.DataFrame(linhas)


def _construir_resumo_demo(base_analitica: pd.DataFrame) -> ResumoExecutivoPadrao:
    total_origem = float(base_analitica["valor_origem"].sum())
    total_comparado = float(base_analitica["valor_comparado"].sum())
    diferenca_total = total_comparado - total_origem
    variacao_total = (diferenca_total / total_origem) if total_origem else 0.0

    distribuicao: Dict[str, int] = {}
    for codigo in base_analitica["classificacao_estrutural"]:
        distribuicao[codigo] = distribuicao.get(codigo, 0) + 1
    # Garante chaves canônicas presentes
    for codigo in ("PRESENTE_AMBOS", "AUSENTE_ORIGEM", "AUSENTE_COMPARADO",
                   "NULO_ORIGEM", "NULO_COMPARADO", "NULO_AMBOS"):
        distribuicao.setdefault(codigo, 0)

    top_variacoes_df = (
        base_analitica.assign(_abs=lambda d: d["variacao_percentual"].abs())
        .sort_values("_abs", ascending=False)
        .head(5)
    )
    top_variacoes = [
        {
            "rotulo": row["chave_agrupadores"],
            "variacao_percentual": row["variacao_percentual"],
            "diferenca": row["diferenca"],
        }
        for _, row in top_variacoes_df.iterrows()
    ]

    return ResumoExecutivoPadrao(
        bloco_1_cabecalho=CabecalhoExecucao(
            visao="V2",
            data_execucao=datetime(2026, 4, 23, 14, 32),
            modo_upload="SIMPLES",
            agrupadores=["Mes", "Loja"],
            medida_principal="Vendas",
        ),
        bloco_2_numeros_ancora={
            "total_origem": total_origem,
            "total_comparado": total_comparado,
            "diferenca_total": diferenca_total,
            "variacao_total": variacao_total,
            "total_casos": int(len(base_analitica)),
        },
        bloco_3_distribuicao=distribuicao,
        bloco_4_elementos_destacados={
            "top_variacoes": top_variacoes,
        },
        bloco_5_leitura_qualitativa=LeituraQualitativa(
            classificacao_ativa="ESTAVEL" if abs(variacao_total) <= 0.05 else "SIGNIFICATIVA",
            thresholds_usados={
                "limiar_estabilidade_pct": 0.05,
                "limite_variacao_extrema": 0.20,
            },
            alguma_leitura_alterada_por_edicao=False,
        ),
        bloco_6_qualidade_estrutural=QualidadeEstrutural(
            total_warnings=3,
            warnings_por_categoria={"informativo": 1, "ajuste_leve": 2},
            ajustes_aplicados=1,
            tem_bloqueios_escapados=False,
        ),
    )


def _construir_resultado_demo() -> Dict[str, Any]:
    """Retorna dict V2Result-like reutilizável pelas 4 amostras."""
    base = _construir_base_analitica_demo()
    resumo = _construir_resumo_demo(base)
    config_usada = {
        "estrutura_entrada": "POR_COLUNAS",
        "origem_rotulo_ux": "Orçado",
        "comparado_rotulo_ux": "Realizado",
        "agrupadores": ["Mes", "Loja"],
        "medida_principal": "Vendas",
    }
    return {"base_analitica": base, "resumo": resumo, "config_usada": config_usada}


# ---------------------------------------------------------------------------
# Esquema de colunas adaptativas para a aba Base Analítica
# ---------------------------------------------------------------------------

def _esquema_v2_colunas() -> List[ColunaAdaptativa]:
    return [
        ColunaAdaptativa(
            identificador="chave_agrupadores",
            cabecalho_user_facing="Agrupadores",
            unidade="texto",
            condicao=sempre(),
            funcao_total="none",
            ordem_sugerida=1,
        ),
        ColunaAdaptativa(
            identificador="estado_origem",
            cabecalho_user_facing="Estado Origem",
            unidade="texto",
            condicao=se_config_igual("estrutura_entrada", "POR_LINHAS"),
            funcao_total="none",
            ordem_sugerida=2,
        ),
        ColunaAdaptativa(
            identificador="estado_comparado",
            cabecalho_user_facing="Estado Comparado",
            unidade="texto",
            condicao=se_config_igual("estrutura_entrada", "POR_LINHAS"),
            funcao_total="none",
            ordem_sugerida=3,
        ),
        ColunaAdaptativa(
            identificador="valor_origem",
            cabecalho_user_facing="Orçado (R$)",
            unidade="monetario",
            condicao=sempre(),
            funcao_total="sum",
            ordem_sugerida=4,
        ),
        ColunaAdaptativa(
            identificador="valor_comparado",
            cabecalho_user_facing="Realizado (R$)",
            unidade="monetario",
            condicao=sempre(),
            funcao_total="sum",
            ordem_sugerida=5,
        ),
        ColunaAdaptativa(
            identificador="diferenca",
            cabecalho_user_facing="Diferença (R$)",
            unidade="monetario",
            condicao=sempre(),
            funcao_total="sum",
            ordem_sugerida=6,
        ),
        ColunaAdaptativa(
            identificador="variacao_percentual",
            cabecalho_user_facing="Variação",
            unidade="percentual",
            condicao=sempre(),
            funcao_total="average",
            ordem_sugerida=7,
        ),
        ColunaAdaptativa(
            identificador="classificacao_estrutural",
            cabecalho_user_facing="Classificação estrutural",
            unidade="classificacao",
            condicao=sempre(),
            funcao_total="none",
            ordem_sugerida=8,
        ),
    ]


# ---------------------------------------------------------------------------
# Renderização da aba Base Analítica
# ---------------------------------------------------------------------------

def _renderizar_base_analitica(
    ws,
    base_analitica: pd.DataFrame,
    config_usada: Dict[str, Any],
    paleta: Paleta,
    vocabulario: Dict[str, Dict[str, str]],
) -> None:
    esquema_completo = _esquema_v2_colunas()
    colunas = montar_colunas_adaptativas(config_usada, esquema_completo)

    # Cabeçalho
    for j, coluna in enumerate(colunas, start=1):
        ws.cell(row=1, column=j, value=coluna.cabecalho_user_facing)

    # Linhas de dados
    for i, (_, row) in enumerate(base_analitica.iterrows(), start=2):
        for j, coluna in enumerate(colunas, start=1):
            valor = row.get(coluna.identificador)
            # Traduções user-facing de campos classificação
            if coluna.unidade == "classificacao" and isinstance(valor, str):
                valor = traduzir(valor, contexto="classificacoes", vocabulario=vocabulario)
            ws.cell(row=i, column=j, value=valor)

    # Range atual (sem linha de totais)
    n_linhas = len(base_analitica)
    ultima_col = get_column_letter(len(colunas))
    range_ref = f"A1:{ultima_col}{1 + n_linhas}"

    totais_por_coluna: Dict[str, str] = {}
    for coluna in colunas:
        totais_por_coluna[coluna.cabecalho_user_facing] = coluna.funcao_total

    # Nome Excel-válido único
    nome_tabela = f"t_base_analitica_{paleta.nome}"
    criar_tabela_executiva(
        ws, range_ref, nome_tabela, totais_por_coluna, paleta_nome=paleta.nome,
    )

    # Formatação por unidade · aplicar em todas as células da coluna (data + linha total)
    for j, coluna in enumerate(colunas, start=1):
        col_letter = get_column_letter(j)
        # linha 1 é header · linha 2..n+1 são dados · linha n+2 é totals
        celulas = [
            ws.cell(row=r, column=j)
            for r in range(2, n_linhas + 3)
        ]
        if coluna.unidade == "monetario":
            aplicar_formato_monetario(celulas)
        elif coluna.unidade == "percentual":
            aplicar_formato_percentual(celulas, conversao_fracao=True)
        elif coluna.unidade == "contagem":
            aplicar_formato_contagem(celulas)

        # Largura
        larguras_padrao = {
            "monetario": 18, "percentual": 14, "contagem": 14,
            "texto": 22, "classificacao": 24, "data": 14, "booleano": 10,
        }
        ws.column_dimensions[col_letter].width = larguras_padrao.get(
            coluna.unidade, 18
        )

    # IMPORTANTE: NÃO aplicar Font manual na linha do cabeçalho. O
    # TableStyleMedium* já define fill cor_primaria + fonte cor_texto_sobre_primaria
    # (branca) + negrito. Sobrescrever Font aqui (mesmo "só com bold=True")
    # anula a cor branca do estilo e força a fonte a preto → texto preto
    # sobre fill escuro = barra preta visual (defeito P0 corrigido).

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Placeholders P1
# ---------------------------------------------------------------------------

def _renderizar_placeholder(ws, titulo: str, mensagem: str, paleta: Paleta) -> None:
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 90

    titulo_font = Font(
        name=paleta.fonte_familia,
        size=paleta.fonte_tamanho_titulo,
        bold=True,
        color=paleta.cor_neutra_escura,
    )
    corpo_font = Font(
        name=paleta.fonte_familia,
        size=paleta.fonte_tamanho_corpo,
        color=paleta.cor_neutra_escura,
    )
    fill_destaque = PatternFill(
        fill_type="solid",
        start_color=paleta.cor_secundaria,
        end_color=paleta.cor_secundaria,
    )

    c = ws.cell(row=2, column=2, value=titulo)
    c.font = titulo_font
    ws.row_dimensions[2].height = 28

    ws.cell(row=3, column=2, value="").fill = fill_destaque
    c = ws.cell(row=5, column=2, value=mensagem)
    c.font = corpo_font
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[5].height = 80


def _renderizar_coracao_visual_placeholder(ws, paleta: Paleta) -> None:
    _renderizar_placeholder(
        ws,
        titulo="Coração Visual",
        mensagem=(
            "O Coração Visual apresenta o gráfico central desta análise · "
            "o ponto focal que o tomador de decisão vê primeiro. Esta aba "
            "será detalhada na próxima versão do produto."
        ),
        paleta=paleta,
    )


def _renderizar_diagnostico_placeholder(ws, paleta: Paleta) -> None:
    _renderizar_placeholder(
        ws,
        titulo="Diagnóstico",
        mensagem=(
            "O Diagnóstico apresenta a leitura narrativa detalhada da análise · "
            "sinais estruturais, ajustes aplicados e recomendações. Esta aba "
            "será detalhada na próxima versão do produto."
        ),
        paleta=paleta,
    )


# ---------------------------------------------------------------------------
# Gerador por paleta
# ---------------------------------------------------------------------------

def gerar_amostra(
    paleta: Paleta,
    diretorio_saida: Path,
    resultado_simulado: Dict[str, Any],
    vocabulario: Dict[str, Dict[str, str]],
) -> Path:
    wb = Workbook()
    aplicar_paleta(wb, paleta)

    # 1. Resumo Executivo
    ws_resumo = wb.active
    ws_resumo.title = "Resumo Executivo"
    renderizar_resumo_executivo(
        ws_resumo, resultado_simulado["resumo"], paleta, vocabulario,
    )

    # 2. Base Analítica
    ws_base = wb.create_sheet("Base Analítica")
    _renderizar_base_analitica(
        ws_base,
        resultado_simulado["base_analitica"],
        resultado_simulado["config_usada"],
        paleta,
        vocabulario,
    )

    # 3. Coração Visual (placeholder P1)
    ws_cv = wb.create_sheet("Coração Visual")
    _renderizar_coracao_visual_placeholder(ws_cv, paleta)

    # 4. Diagnóstico (placeholder P1)
    ws_diag = wb.create_sheet("Diagnóstico")
    _renderizar_diagnostico_placeholder(ws_diag, paleta)

    caminho = diretorio_saida / f"amostra_paleta_{paleta.nome}.xlsx"
    wb.save(caminho)
    return caminho


# ---------------------------------------------------------------------------
# Entrada pública
# ---------------------------------------------------------------------------

def gerar_todas_as_amostras(
    diretorio_saida: str | Path = "demos",
) -> List[Path]:
    """
    Gera as 4 amostras em `diretorio_saida`. Idempotente · sobrescreve existentes.
    Retorna lista de caminhos gerados na ordem azul · verde · cinza · vinho.
    """
    diretorio = Path(diretorio_saida)
    diretorio.mkdir(parents=True, exist_ok=True)

    resultado = _construir_resultado_demo()
    vocabulario = carregar_vocabulario_bilingue()

    caminhos: List[Path] = []
    for nome_paleta in ("azul", "verde", "cinza", "vinho"):
        paleta = CATALOGO_PALETAS[nome_paleta]
        caminho = gerar_amostra(paleta, diretorio, resultado, vocabulario)
        caminhos.append(caminho)
    return caminhos


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format="%(levelname)s · %(message)s")
    caminhos = gerar_todas_as_amostras()
    print(f"Geradas {len(caminhos)} amostras:")
    for c in caminhos:
        tamanho_kb = c.stat().st_size / 1024
        print(f"  · {c} ({tamanho_kb:.1f} KB)")
