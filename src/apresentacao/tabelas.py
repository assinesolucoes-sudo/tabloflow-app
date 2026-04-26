"""
F-APRESENT · capability 3 · Tabela Excel nativa + totais dinâmicos (D-166).

Toda aba tabular do Excel executivo é openpyxl.Table (ListObject) com
totalsRowShown=True · evita AutoFilter simples que perde totais dinâmicos
ao filtrar (OBS-11 de VV-V2).

Mapeamento paleta → estilo de tabela (estável · declarado):
    azul  → TableStyleMedium2
    verde → TableStyleMedium4
    cinza → TableStyleMedium1
    vinho → TableStyleMedium3
"""
from __future__ import annotations

import logging
import re
from typing import Dict, Literal, Optional

from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl.worksheet.filters import AutoFilter
from openpyxl.worksheet.table import Table, TableColumn, TableFormula, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

from .paletas import CATALOGO_PALETAS


_logger = logging.getLogger(__name__)

FuncaoTotal = Literal[
    "sum", "average", "count", "countNums",
    "min", "max", "stdDev", "var", "custom", "none",
]

# Conjunto canônico de funções válidas na linha de totais Excel
_FUNCOES_VALIDAS = {
    "sum", "average", "count", "countNums",
    "min", "max", "stdDev", "var", "custom", "none",
}

# Mapa função Excel → código SUBTOTAL (série 1XX ignora linhas ocultas/filtradas).
# Setar apenas tableColumn.totalsRowFunction no XML não preenche a célula da
# linha de totais · Excel e openpyxl exigem também a fórmula SUBTOTAL
# explícita na célula para que o valor apareça ao abrir e recalcule sob filtro.
_SUBTOTAL_POR_FUNCAO: dict[str, int] = {
    "average": 101,
    "countNums": 102,
    "count": 103,
    "max": 104,
    "min": 105,
    "stdDev": 107,
    "sum": 109,
    "var": 110,
}

_REGEX_NOME_TABELA_VALIDO = re.compile(r"^[A-Za-z_][A-Za-z0-9_]*$")


def _parsear_range(range_ref: str) -> tuple[int, int, int, int]:
    """Retorna (linha_inicial, coluna_inicial, linha_final, coluna_final) 1-based."""
    if ":" not in range_ref:
        raise ValueError(f"range_ref inválido: {range_ref!r} · esperado 'A1:E20'")
    ini, fim = range_ref.split(":", 1)
    col_ini, lin_ini = coordinate_from_string(ini)
    col_fim, lin_fim = coordinate_from_string(fim)
    return (
        int(lin_ini),
        column_index_from_string(col_ini),
        int(lin_fim),
        column_index_from_string(col_fim),
    )


def _cabecalhos_do_range(ws: Worksheet, range_ref: str) -> list[str]:
    lin_ini, col_ini, _, col_fim = _parsear_range(range_ref)
    cabecalhos = []
    for c in range(col_ini, col_fim + 1):
        val = ws.cell(row=lin_ini, column=c).value
        cabecalhos.append(str(val) if val is not None else "")
    return cabecalhos


def _ref_com_totals(range_ref: str) -> str:
    lin_ini, col_ini, lin_fim, col_fim = _parsear_range(range_ref)
    col_ini_letra = get_column_letter(col_ini)
    col_fim_letra = get_column_letter(col_fim)
    return f"{col_ini_letra}{lin_ini}:{col_fim_letra}{lin_fim + 1}"


def estilo_para_paleta(paleta_nome: Optional[str]) -> str:
    """Retorna o nome do TableStyleInfo correspondente à paleta."""
    if paleta_nome is None:
        return "TableStyleMedium2"
    if paleta_nome not in CATALOGO_PALETAS:
        _logger.warning(
            "paleta '%s' desconhecida em estilo_para_paleta · usando default azul",
            paleta_nome,
        )
        return "TableStyleMedium2"
    return CATALOGO_PALETAS[paleta_nome].estilo_tabela


def criar_tabela_executiva(
    ws: Worksheet,
    range_ref: str,
    nome: str,
    totais_por_coluna: Dict[str, FuncaoTotal],
    paleta_nome: Optional[str] = None,
    custom_totals: Optional[Dict[str, str]] = None,
) -> Table:
    """
    Cria tabela Excel nativa na worksheet no range dado (sem linha de totais).

    Comportamento:
      - Constrói Table com displayName=nome · ref inclui linha de totais (range_ref + 1)
      - Para cada coluna declarada em totais_por_coluna, define totalsRowFunction
      - totalsRowShown=True
      - TableStyleInfo = estilo_para_paleta(paleta_nome)
      - Retorna o Table construído

    Validações estritas (C.2):
      - range_ref precisa ter >=2 linhas (cabeçalho + 1 dado) · senão ValueError
      - nome deve ser identificador Excel válido · senão ValueError
      - nome deve ser único na worksheet · senão ValueError
      - chaves de totais_por_coluna devem existir no cabeçalho · senão ValueError
      - funcao "custom" exige entrada em custom_totals · senão ValueError
      - colunas do range sem entrada em totais_por_coluna recebem "none" com warning
    """
    if ws is None:
        raise TypeError("ws não pode ser None em criar_tabela_executiva")
    if not isinstance(nome, str) or not nome:
        raise ValueError(f"nome inválido: {nome!r}")
    if not _REGEX_NOME_TABELA_VALIDO.match(nome):
        raise ValueError(
            f"nome '{nome}' não é identificador Excel válido "
            f"(letras/dígitos/underscore · começa por letra)"
        )

    lin_ini, col_ini, lin_fim, col_fim = _parsear_range(range_ref)
    if lin_fim - lin_ini < 1:
        raise ValueError(
            f"range_ref '{range_ref}' tem apenas {lin_fim - lin_ini + 1} linha(s) · "
            f"tabela executiva exige cabeçalho + ao menos 1 linha de dados"
        )

    cabecalhos = _cabecalhos_do_range(ws, range_ref)
    cabecalhos_set = set(cabecalhos)

    for chave in totais_por_coluna.keys():
        if chave not in cabecalhos_set:
            raise ValueError(
                f"coluna '{chave}' em totais_por_coluna não existe no cabeçalho "
                f"{cabecalhos}"
            )

    for chave, func in totais_por_coluna.items():
        if func not in _FUNCOES_VALIDAS:
            raise ValueError(
                f"função total '{func}' em coluna '{chave}' inválida · "
                f"válidas: {sorted(_FUNCOES_VALIDAS)}"
            )
        if func == "custom":
            if not custom_totals or chave not in custom_totals:
                raise ValueError(
                    f"coluna '{chave}' marcada 'custom' exige entrada em custom_totals"
                )

    # Checa nome único na worksheet
    tabelas_existentes = list(ws.tables)
    if nome in tabelas_existentes:
        raise ValueError(f"nome de tabela '{nome}' já existe na worksheet")

    # Colunas sem entrada em totais_por_coluna recebem "none" com warning
    for cab in cabecalhos:
        if cab and cab not in totais_por_coluna:
            _logger.warning(
                "coluna '%s' sem função de total declarada · aplicando 'none' (C.2)",
                cab,
            )

    ref_final = _ref_com_totals(range_ref)

    # Construir TableColumn explícitos · openpyxl exige quando totalsRow é usado
    colunas_tabela: list[TableColumn] = []
    for i, cab in enumerate(cabecalhos, start=1):
        nome_col = cab if cab else f"Coluna{i}"
        tc = TableColumn(id=i, name=nome_col)
        if cab in totais_por_coluna:
            func = totais_por_coluna[cab]
            if func == "none":
                tc.totalsRowFunction = None
            elif func == "custom":
                tc.totalsRowFunction = None
                formula = (custom_totals or {}).get(cab, "")
                tc.totalsRowFormula = TableFormula(attr_text=formula)
            else:
                tc.totalsRowFunction = func
        else:
            tc.totalsRowFunction = None
        colunas_tabela.append(tc)

    tabela = Table(
        displayName=nome,
        ref=ref_final,
        tableColumns=colunas_tabela,
    )
    tabela.totalsRowShown = True
    tabela.totalsRowCount = 1

    # Ativa o botão "Filtrar" (setinha ▼) em cada cabeçalho por default.
    # Sem Table.autoFilter explícito o Excel abre com filtros desligados ·
    # usuária precisaria marcar Design da Tabela > Botão Filtrar manualmente.
    # O ref do autoFilter cobre cabeçalho + dados · EXCLUI a linha de totais
    # (senão o botão aparece também na totals row · visual incorreto).
    tabela.autoFilter = AutoFilter(ref=range_ref)

    tabela.tableStyleInfo = TableStyleInfo(
        name=estilo_para_paleta(paleta_nome),
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )

    ws.add_table(tabela)

    # Preenche a célula da linha de totais para cada coluna com função declarada.
    # Sem essa gravação a linha de totais fica visualmente VAZIA ao abrir o
    # arquivo · metadata XML sozinha não basta (defeito identificado em P0).
    # SUBTOTAL 1XX respeita filtros · totalize recalcula nativamente.
    linha_totais = lin_fim + 1
    for j, cab in enumerate(cabecalhos):
        if not cab or cab not in totais_por_coluna:
            continue
        func = totais_por_coluna[cab]
        col_letra = get_column_letter(col_ini + j)
        range_dados = f"{col_letra}{lin_ini + 1}:{col_letra}{lin_fim}"
        celula_total = ws.cell(row=linha_totais, column=col_ini + j)
        if func == "custom":
            formula = (custom_totals or {}).get(cab, "")
            if formula:
                celula_total.value = formula if formula.startswith("=") else f"={formula}"
        elif func in _SUBTOTAL_POR_FUNCAO:
            codigo = _SUBTOTAL_POR_FUNCAO[func]
            celula_total.value = f"=SUBTOTAL({codigo},{range_dados})"

    return tabela
