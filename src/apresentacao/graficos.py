"""
F-APRESENT · capability 11 · Gráficos executivos nativos + nomenclatura (D-176).

Duas funções de gráfico nativo openpyxl ancoradas em células específicas,
mais um gerador de nome de arquivo Excel executivo (padrão P-12 / D-176).

API:
  criar_grafico_distribuicao(ws, anchor_cell, dados_range, paleta_nome, titulo)
  criar_grafico_top_variacoes(ws, anchor_cell, dados_range, paleta_nome, titulo)
  gerar_nome_arquivo(nome_visao_user_facing, contexto, data, diretorio_saida)

Princípios:
  - Gráficos usam as cores da paleta do catálogo F-APRESENT (D-164) · coerência
    visual com o restante do workbook (tabelas, headers, badges).
  - Nomes de arquivo nunca contêm timestamps ISO nem códigos técnicos.
    Caracteres inválidos de filesystem Windows (\\ / : * ? " < > |) são
    substituídos por hífen.
"""
from __future__ import annotations

import re
from datetime import date
from pathlib import Path
from typing import Optional, Union

from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.marker import DataPoint
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.colors import ColorChoice
from openpyxl.drawing.fill import ColorChoice as FillColorChoice  # noqa: F401
from openpyxl.drawing.fill import PatternFillProperties, SolidColorFillProperties
from openpyxl.worksheet.worksheet import Worksheet

from .paletas import CATALOGO_PALETAS, obter_paleta


# ---------------------------------------------------------------------------
# Paleta de cores semânticas por categoria (pizza de distribuição)
# ---------------------------------------------------------------------------

_CORES_SEMANTICAS_CATEGORIAS = {
    "Presente nos dois lados": "4F81BD",
    "Ausente na origem": "E7A977",
    "Ausente no comparado": "E7A977",
    "Sem valor na origem": "D9823B",
    "Sem valor no comparado": "D9823B",
    "Sem valor nos dois lados": "B0B0B0",
    "Melhorou": "5B9E5B",
    "Piorou": "C06060",
    "Estável": "B0B0B0",
    "Não aplicável": "D9D9D9",
}


def _cor_por_rotulo(rotulo: str, fallback: str) -> str:
    for chave, cor in _CORES_SEMANTICAS_CATEGORIAS.items():
        if rotulo.strip().lower() == chave.lower():
            return cor
    return fallback


# ---------------------------------------------------------------------------
# Capability 11.1 · Pizza de distribuição
# ---------------------------------------------------------------------------

def criar_grafico_distribuicao(
    ws: Worksheet,
    anchor_cell: str,
    dados_range: str,
    paleta_nome: str,
    titulo: str = "Como os casos se distribuem",
) -> PieChart:
    """
    Insere um PieChart nativo na worksheet ancorado em anchor_cell.

    dados_range: range A1-notation que inclui header + linhas de dados, com
    2 colunas: [rótulo da categoria, contagem]. Exemplo: "B5:C10".

    A primeira linha é tratada como header. As cores das fatias seguem o
    mapa semântico F-APRESENT (positivo=verde, negativo=vermelho, neutro=cinza)
    quando o rótulo bate; caso contrário, cai na cor primária da paleta.
    """
    if ws is None:
        raise TypeError("ws não pode ser None em criar_grafico_distribuicao")
    if not anchor_cell:
        raise ValueError("anchor_cell é obrigatório")
    if not dados_range or ":" not in dados_range:
        raise ValueError(f"dados_range inválido: {dados_range!r}")

    paleta = obter_paleta((paleta_nome or "azul").strip().lower())

    ini, fim = dados_range.split(":")
    col_ini_letra = _extrair_letras(ini)
    col_fim_letra = _extrair_letras(fim)
    lin_ini = _extrair_numero(ini)
    lin_fim = _extrair_numero(fim)
    col_ini_idx = _letra_para_indice(col_ini_letra)
    col_fim_idx = _letra_para_indice(col_fim_letra)

    chart = PieChart()
    chart.title = titulo
    chart.height = 7.5
    chart.width = 11

    rotulos_ref = Reference(
        ws, min_col=col_ini_idx, max_col=col_ini_idx,
        min_row=lin_ini + 1, max_row=lin_fim,
    )
    valores_ref = Reference(
        ws, min_col=col_fim_idx, max_col=col_fim_idx,
        min_row=lin_ini, max_row=lin_fim,
    )
    chart.add_data(valores_ref, titles_from_data=True)
    chart.set_categories(rotulos_ref)

    serie = chart.series[0] if chart.series else None
    if serie is not None:
        rotulos: list[str] = []
        for r in range(lin_ini + 1, lin_fim + 1):
            cel = ws.cell(row=r, column=col_ini_idx)
            rotulos.append(str(cel.value) if cel.value is not None else "")
        for idx, rotulo in enumerate(rotulos):
            cor_hex = _cor_por_rotulo(rotulo, paleta.cor_primaria)
            dp = DataPoint(idx=idx)
            dp.graphicalProperties = GraphicalProperties(solidFill=cor_hex)
            serie.dPt.append(dp)

    chart.dataLabels = DataLabelList(showPercent=True)

    ws.add_chart(chart, anchor_cell)
    return chart


# ---------------------------------------------------------------------------
# Capability 11.2 · Barras horizontais de top variações
# ---------------------------------------------------------------------------

def criar_grafico_top_variacoes(
    ws: Worksheet,
    anchor_cell: str,
    dados_range: str,
    paleta_nome: str,
    titulo: str = "Variações em destaque",
) -> BarChart:
    """
    Insere BarChart horizontal nativo · cores por sinal (positivo/negativo/neutro).

    dados_range: range 2 colunas [rótulo, valor_numerico_da_diferenca].
    Primeira linha é header; barras positivas vão em verde, negativas em
    vermelho, zero/None em cinza.
    """
    if ws is None:
        raise TypeError("ws não pode ser None em criar_grafico_top_variacoes")
    if not anchor_cell:
        raise ValueError("anchor_cell é obrigatório")
    if not dados_range or ":" not in dados_range:
        raise ValueError(f"dados_range inválido: {dados_range!r}")

    paleta = obter_paleta((paleta_nome or "azul").strip().lower())

    ini, fim = dados_range.split(":")
    col_ini_letra = _extrair_letras(ini)
    col_fim_letra = _extrair_letras(fim)
    lin_ini = _extrair_numero(ini)
    lin_fim = _extrair_numero(fim)
    col_ini_idx = _letra_para_indice(col_ini_letra)
    col_fim_idx = _letra_para_indice(col_fim_letra)

    chart = BarChart()
    chart.type = "bar"
    chart.style = 10
    chart.title = titulo
    chart.y_axis.title = None
    chart.x_axis.title = None
    chart.legend = None
    chart.height = 10
    chart.width = 16

    rotulos_ref = Reference(
        ws, min_col=col_ini_idx, max_col=col_ini_idx,
        min_row=lin_ini + 1, max_row=lin_fim,
    )
    valores_ref = Reference(
        ws, min_col=col_fim_idx, max_col=col_fim_idx,
        min_row=lin_ini, max_row=lin_fim,
    )
    chart.add_data(valores_ref, titles_from_data=True)
    chart.set_categories(rotulos_ref)

    serie = chart.series[0] if chart.series else None
    if serie is not None:
        valores: list[Optional[float]] = []
        for r in range(lin_ini + 1, lin_fim + 1):
            cel = ws.cell(row=r, column=col_fim_idx)
            try:
                valores.append(float(cel.value) if cel.value is not None else None)
            except (TypeError, ValueError):
                valores.append(None)
        for idx, v in enumerate(valores):
            if v is None or v == 0:
                cor = "B0B0B0"
            elif v > 0:
                cor = "5B9E5B"
            else:
                cor = "C06060"
            dp = DataPoint(idx=idx)
            dp.graphicalProperties = GraphicalProperties(solidFill=cor)
            serie.dPt.append(dp)

    ws.add_chart(chart, anchor_cell)
    return chart


# ---------------------------------------------------------------------------
# Capability 11.3 · Nome executivo de arquivo (D-176 · P-12 opção 2)
# ---------------------------------------------------------------------------

_REGEX_INVALIDOS_FS = re.compile(r'[\\/:*?"<>|]')


def gerar_nome_arquivo(
    nome_visao_user_facing: str,
    contexto: Optional[str],
    data: date,
    diretorio_saida: Union[str, Path],
) -> Path:
    """
    Gera nome executivo do arquivo Excel (D-176 · P-12 opção 2).

    Padrão com contexto:
        "<NomeVisao> - <Contexto> - DD-MM-AAAA.xlsx"
    Padrão sem contexto:
        "<NomeVisao> - DD-MM-AAAA.xlsx"

    Caracteres inválidos de filesystem (\\ / : * ? " < > |) são substituídos
    por hífen. Retorna Path absoluto (diretorio_saida / nome_limpo).
    """
    if not nome_visao_user_facing or not nome_visao_user_facing.strip():
        raise ValueError("nome_visao_user_facing é obrigatório")
    if data is None:
        raise ValueError("data é obrigatória")
    diretorio = Path(diretorio_saida)

    data_br = data.strftime("%d-%m-%Y")
    nome_base = nome_visao_user_facing.strip()

    if contexto and contexto.strip():
        nome = f"{nome_base} - {contexto.strip()} - {data_br}.xlsx"
    else:
        nome = f"{nome_base} - {data_br}.xlsx"

    nome_limpo = _REGEX_INVALIDOS_FS.sub("-", nome)
    return diretorio / nome_limpo


# ---------------------------------------------------------------------------
# Utilitários internos
# ---------------------------------------------------------------------------

_REGEX_LETRAS = re.compile(r"^([A-Z]+)")
_REGEX_NUMERO = re.compile(r"(\d+)$")


def _extrair_letras(ref: str) -> str:
    m = _REGEX_LETRAS.match(ref.upper())
    if not m:
        raise ValueError(f"ref inválida (letras): {ref!r}")
    return m.group(1)


def _extrair_numero(ref: str) -> int:
    m = _REGEX_NUMERO.search(ref)
    if not m:
        raise ValueError(f"ref inválida (numero): {ref!r}")
    return int(m.group(1))


def _letra_para_indice(letras: str) -> int:
    idx = 0
    for c in letras:
        idx = idx * 26 + (ord(c) - ord("A") + 1)
    return idx
