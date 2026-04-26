"""
F-APRESENT · capability 10 · Diagnóstico narrativo em 6 seções (D-159 · D-165).

Renderiza a aba Diagnóstico com 6 seções fixas user-facing (D-165 · cópia
literal · não alterar):

  1. Como foi analisado
  2. Ajustes do motor
  3. Pontos de atenção
  4. Decisões do usuário
  5. Configurações avançadas aplicadas
  6. Qualidade estrutural

Interface canônica (D-159):
    renderizar_diagnostico(ws, config_usada, resolucao_estrutural,
                           modelo_aplicado, t_diag, warnings, paleta,
                           vocabulario)

Regras invioláveis (vocabulario_bilingue.md v2 · blocos 7 e 8):
  - Nenhum enum técnico literal (POR_COLUNAS · AUSENTE_ORIGEM · etc)
  - Nenhum código interno (D-XXX · B-V2-* · W-V2-* · T-DIAG · F-APRESENT)
  - Nenhum `None`/`NaN` literal · sempre `—` com microcopy contextual
  - Nenhum dict Python serializado · renderização sempre estruturada
  - Nenhum datetime cru · via formatar_data_br

A capability é genérica e NÃO conhece a visão. Cada A-VN desempacota os
campos do seu V{N}Result e chama `renderizar_diagnostico(...)` passando
os 7 argumentos declarados na interface canônica.
"""
from __future__ import annotations

import logging
from typing import Any, Dict, List, Optional, Tuple

from openpyxl.worksheet.worksheet import Worksheet

from .formatos import (
    formatar_data_br,
    formatar_moeda_br,
    formatar_percentual_br,
    formatar_threshold_por_contrato,
)
from .paletas import Paleta
from .tipografia import (
    aplicar_campo,
    aplicar_titulo_aba,
    aplicar_valor,
    escrever_secao,
    escrever_titulo_aba,
)
from .vocabulario import traduzir


_logger = logging.getLogger(__name__)


# 6 seções fixas · D-165 · cópia literal
NOMES_SECOES = (
    "Como foi analisado",
    "Ajustes do motor",
    "Pontos de atenção",
    "Decisões do usuário",
    "Configurações avançadas aplicadas",
    "Qualidade estrutural",
)

# Gravidade canônica das categorias de warning (contratos · CategoriaWarning)
# - leve/info vai para Seção 2 (Ajustes do motor)
# - média/alta vai para Seção 3 (Pontos de atenção)
_CATEGORIAS_AJUSTE = {"informativo", "ajuste_leve"}
_CATEGORIAS_ATENCAO = {
    "alerta_estrutural_leve",
    "alerta_estrutural",
    "decisao_usuario",
    "escape_acionado",
}


# ---------------------------------------------------------------------------
# Helpers de formatação user-facing
# ---------------------------------------------------------------------------

def formatar_valor_ou_traco(
    valor: Any, origem: str = "generico",
) -> str:
    """
    Converte `None`/`NaN`/string vazia em `—` com microcopy contextual
    conforme bloco 8 de vocabulario_bilingue.md v2.

    Origens reconhecidas:
      - "generico"         → "—"
      - "ausencia"         → "— (não consta)"
      - "nulo"             → "— (sem valor)"
      - "impossivel"       → "— (não calculável)"
    """
    if valor is None:
        return _traco_por_origem(origem)
    if isinstance(valor, float):
        # NaN check sem importar math · NaN != NaN
        if valor != valor:
            return _traco_por_origem(origem)
    if isinstance(valor, str):
        texto = valor.strip()
        if texto == "" or texto.lower() in ("none", "null", "nan", "(null)"):
            return _traco_por_origem(origem)
        return texto
    return str(valor)


def _traco_por_origem(origem: str) -> str:
    tabela = {
        "ausencia":   "— (não consta)",
        "nulo":       "— (sem valor)",
        "impossivel": "— (não calculável)",
    }
    return tabela.get(origem, "—")


def _traduzir_modo_base(
    modo: Optional[str], vocabulario: Dict[str, Dict[str, str]],
) -> str:
    """
    Traduz o código técnico do modo da base via bloco 'modos_base'.

    Aceita sinônimos do SIMPLES/POR_COLUNAS/POR_LINHAS/LIVRE_DEDUPLICADA.
    """
    if not modo:
        return formatar_valor_ou_traco(None, origem="ausencia")
    bloco = vocabulario.get("modos_base", {})
    if modo in bloco:
        return bloco[modo]
    # Sinônimos user-facing explícitos para modos não contemplados no v2
    extras = {
        "SIMPLES":            "Envio simples (um arquivo)",
        "POR_COLUNAS":        "Duas colunas lado a lado",
        "POR_LINHAS":         "Duas referências em uma coluna",
        "LIVRE_DEDUPLICADA":  "Registros individuais sem agrupamento",
    }
    if modo in extras:
        return extras[modo]
    # Fallback via traduzir (log + marcador)
    return traduzir(modo, contexto="modos_base", vocabulario=vocabulario)


def _traduzir_tipo_campo(
    tipo: Optional[str], vocabulario: Dict[str, Dict[str, str]],
) -> str:
    if not tipo:
        return formatar_valor_ou_traco(None, origem="ausencia")
    bloco = vocabulario.get("tipos_campo", {})
    if tipo in bloco:
        return bloco[tipo]
    return traduzir(tipo, contexto="tipos_campo", vocabulario=vocabulario)


def _traduzir_categoria_warning(codigo: str) -> str:
    """Traduz código canônico de CategoriaWarning para rótulo user-facing."""
    tabela = {
        "informativo":               "Informativo",
        "ajuste_leve":               "Ajuste automático aplicado",
        "alerta_estrutural_leve":    "Alerta estrutural leve",
        "alerta_estrutural":         "Alerta estrutural",
        "decisao_usuario":           "Decisão do usuário",
        "escape_acionado":           "Escape acionado",
    }
    return tabela.get(codigo, codigo.replace("_", " ").capitalize())


def _traduzir_threshold(
    chave: str, vocabulario: Dict[str, Dict[str, str]],
) -> str:
    bloco = vocabulario.get("thresholds", {})
    if chave in bloco:
        return bloco[chave]
    return traduzir(chave, contexto="thresholds", vocabulario=vocabulario)


def _formatar_lista(valores: Optional[List[Any]], conector: str = " · ") -> str:
    if not valores:
        return formatar_valor_ou_traco(None, origem="ausencia")
    return conector.join(str(v) for v in valores)


def _formatar_colunas_mapeadas(
    mapa: Optional[Dict[str, str]],
) -> List[Tuple[str, str]]:
    """Transforma dict técnico→user-facing em pares (user_facing, user_facing).

    A chave do dict é nome técnico interno (coluna do contrato) · o valor é o
    rótulo user-facing da coluna na base da Usuária. Só o user-facing é
    exibido · o nome técnico fica invisível (bloco 8.1).
    """
    if not mapa:
        return []
    # Exibimos só os user-facing · técnico fica invisível (bloco 8.1)
    return [("Coluna mapeada", valor) for valor in mapa.values() if valor]


# ---------------------------------------------------------------------------
# Escrita genérica de blocos campo-valor na worksheet
# ---------------------------------------------------------------------------

def _escrever_pares(
    ws: Worksheet, linha: int, col_campo: int, col_valor: int,
    pares: List[Tuple[str, str]], paleta: Paleta,
) -> int:
    for rotulo, valor in pares:
        c_campo = ws.cell(row=linha, column=col_campo, value=rotulo)
        c_valor = ws.cell(row=linha, column=col_valor, value=valor)
        aplicar_campo(c_campo, paleta)
        aplicar_valor(c_valor, paleta)
        linha += 1
    return linha


def _escrever_frase(
    ws: Worksheet, linha: int, col_ini: int, col_fim: int,
    texto: str, paleta: Paleta,
) -> int:
    """Escreve uma frase em merge horizontal com estilo valor."""
    if col_fim > col_ini:
        ws.merge_cells(
            start_row=linha, start_column=col_ini,
            end_row=linha, end_column=col_fim,
        )
    celula = ws.cell(row=linha, column=col_ini, value=texto)
    aplicar_valor(celula, paleta)
    return linha + 1


# ---------------------------------------------------------------------------
# Extração resiliente de campos canônicos em config_usada
# ---------------------------------------------------------------------------

def _get_cfg(config: Any, chave: str, default: Any = None) -> Any:
    """
    Lê campo canônico de config_usada · aceita dict OU objeto com atributo.

    capability 10 é genérica · V{N}Result.config_usada pode chegar como
    dict (caso típico) · como subobjeto Pydantic (caso possível) · ou como
    estrutura mista. Default `None` é sinal de ausência · não é preenchido
    com dummy.
    """
    if config is None:
        return default
    if isinstance(config, dict):
        return config.get(chave, default)
    return getattr(config, chave, default)


# ---------------------------------------------------------------------------
# Seção 1 · Como foi analisado
# ---------------------------------------------------------------------------

def _renderizar_secao_1(
    ws: Worksheet, linha: int, col_campo: int, col_valor: int,
    col_fim: int, config_usada: Any, paleta: Paleta,
    vocabulario: Dict[str, Dict[str, str]],
) -> int:
    linha = escrever_secao(ws, linha, col_campo, NOMES_SECOES[0], paleta)

    pares: List[Tuple[str, str]] = []

    arquivo = _get_cfg(config_usada, "arquivo")
    pares.append((
        "Arquivo",
        formatar_valor_ou_traco(arquivo, origem="ausencia"),
    ))

    aba = _get_cfg(config_usada, "aba_consumida")
    pares.append((
        "Aba consumida",
        formatar_valor_ou_traco(aba, origem="ausencia"),
    ))

    modo = _get_cfg(config_usada, "modo_base")
    pares.append((
        "Modo da base",
        _traduzir_modo_base(modo, vocabulario),
    ))

    agrupadores = _get_cfg(config_usada, "agrupadores")
    pares.append((
        "Agrupadores aplicados",
        _formatar_lista(agrupadores) if agrupadores else "Nenhum agrupador",
    ))

    campo_analisado = _get_cfg(config_usada, "campo_analisado")
    pares.append((
        "Campo analisado",
        formatar_valor_ou_traco(campo_analisado, origem="ausencia"),
    ))

    tipo_medida = _get_cfg(config_usada, "tipo_medida")
    pares.append((
        "Tipo de medida",
        _traduzir_tipo_campo(tipo_medida, vocabulario),
    ))

    linha = _escrever_pares(ws, linha, col_campo, col_valor, pares, paleta)

    # Colunas mapeadas · renderizadas como sub-lista quando existirem
    mapeadas = _formatar_colunas_mapeadas(_get_cfg(config_usada, "colunas_mapeadas"))
    if mapeadas:
        linha = _escrever_pares(ws, linha, col_campo, col_valor, mapeadas, paleta)

    return linha + 1  # respiro após a seção


# ---------------------------------------------------------------------------
# Seção 2 · Ajustes do motor
# ---------------------------------------------------------------------------

def _renderizar_secao_2(
    ws: Worksheet, linha: int, col_campo: int, col_valor: int,
    col_fim: int, t_diag: Any, warnings_leves: List[Any], paleta: Paleta,
    vocabulario: Dict[str, Dict[str, str]],
) -> int:
    linha = escrever_secao(ws, linha, col_campo, NOMES_SECOES[1], paleta)

    ajustes = _extrair_ajustes(t_diag)

    if not ajustes and not warnings_leves:
        linha = _escrever_frase(
            ws, linha, col_campo, col_fim,
            "Nenhum ajuste estrutural foi necessário.",
            paleta,
        )
        return linha + 1

    pares: List[Tuple[str, str]] = []
    for aj in ajustes:
        tipo = _ler_campo(aj, "tipo_ajuste", "—")
        descricao = _ler_campo(aj, "descricao", "")
        linhas_aj = _ler_campo(aj, "linhas_afetadas", 0)
        rotulo = _humanizar_tipo_ajuste(tipo)
        valor = descricao if descricao else f"{int(linhas_aj):,} linhas".replace(",", ".")
        pares.append((rotulo, valor))

    for w in warnings_leves:
        codigo = _ler_campo(w, "codigo", "")
        microcopy = _ler_campo(w, "microcopy", "")
        categoria = _ler_campo(w, "categoria", "informativo")
        categoria_rotulo = _traduzir_categoria_warning(
            _normalizar_categoria(categoria)
        )
        valor = microcopy if microcopy else formatar_valor_ou_traco(None, origem="ausencia")
        # Código técnico do warning (W-V*-*) NÃO é exibido · bloco 7
        _ = codigo
        pares.append((categoria_rotulo, valor))

    linha = _escrever_pares(ws, linha, col_campo, col_valor, pares, paleta)
    return linha + 1


def _extrair_ajustes(t_diag: Any) -> List[Any]:
    if t_diag is None:
        return []
    if hasattr(t_diag, "ajustes_aplicados"):
        return list(getattr(t_diag, "ajustes_aplicados") or [])
    if isinstance(t_diag, dict):
        return list(t_diag.get("ajustes_aplicados") or [])
    return []


def _humanizar_tipo_ajuste(tipo: str) -> str:
    """Traduz tipo_ajuste técnico em rótulo user-facing."""
    if not tipo or tipo == "—":
        return "Ajuste aplicado"
    tabela = {
        "INTERVALO_AJUSTADO_INICIO": "Intervalo ajustado (início)",
        "INTERVALO_AJUSTADO_FIM":    "Intervalo ajustado (fim)",
        "LINHA_EXCLUIDA_NULO":       "Linhas excluídas por valor ausente",
        "COLUNA_RENOMEADA":          "Coluna renomeada",
        "TIPO_COERCIDO":             "Tipo convertido",
        "DUPLICATA_REMOVIDA":        "Duplicatas removidas",
        "CONSOLIDACAO_APLICADA":     "Consolidação aplicada",
    }
    if tipo in tabela:
        return tabela[tipo]
    # Sanear · não deixar caps-snake escapar para tela (bloco 8.2)
    return tipo.replace("_", " ").capitalize()


def _normalizar_categoria(categoria: Any) -> str:
    """Aceita Enum · str · retorna string canônica minúscula."""
    if categoria is None:
        return "informativo"
    if hasattr(categoria, "value"):
        return str(categoria.value)
    return str(categoria)


# ---------------------------------------------------------------------------
# Seção 3 · Pontos de atenção
# ---------------------------------------------------------------------------

def _renderizar_secao_3(
    ws: Worksheet, linha: int, col_campo: int, col_valor: int,
    col_fim: int, warnings_atencao: List[Any], paleta: Paleta,
    vocabulario: Dict[str, Dict[str, str]],
) -> int:
    linha = escrever_secao(ws, linha, col_campo, NOMES_SECOES[2], paleta)

    if not warnings_atencao:
        linha = _escrever_frase(
            ws, linha, col_campo, col_fim,
            "Nenhum ponto de atenção identificado nesta análise.",
            paleta,
        )
        return linha + 1

    # Agregar por categoria · contagem + 1 microcopy exemplo
    por_categoria: Dict[str, List[Any]] = {}
    for w in warnings_atencao:
        cat = _normalizar_categoria(_ler_campo(w, "categoria", "informativo"))
        por_categoria.setdefault(cat, []).append(w)

    pares: List[Tuple[str, str]] = []
    for cat, lista in por_categoria.items():
        rotulo = _traduzir_categoria_warning(cat)
        exemplo = _ler_campo(lista[0], "microcopy", "")
        if len(lista) == 1:
            valor = exemplo if exemplo else formatar_valor_ou_traco(None, origem="nulo")
        else:
            exemplo_txt = exemplo if exemplo else "—"
            valor = f"{len(lista)} ocorrências · exemplo: {exemplo_txt}"
        pares.append((rotulo, valor))

    linha = _escrever_pares(ws, linha, col_campo, col_valor, pares, paleta)
    return linha + 1


# ---------------------------------------------------------------------------
# Seção 4 · Decisões do usuário
# ---------------------------------------------------------------------------

def _renderizar_secao_4(
    ws: Worksheet, linha: int, col_campo: int, col_valor: int,
    col_fim: int, config_usada: Any, resolucao_estrutural: Any,
    modelo_aplicado: Any, t_diag: Any, paleta: Paleta,
    vocabulario: Dict[str, Dict[str, str]],
) -> int:
    linha = escrever_secao(ws, linha, col_campo, NOMES_SECOES[3], paleta)

    pares: List[Tuple[str, str]] = []

    # Resolução estrutural (caso detectado · decisão user-facing)
    if resolucao_estrutural is not None:
        tipo_caso = _ler_campo(resolucao_estrutural, "tipo_caso", "")
        escolha = _ler_campo(resolucao_estrutural, "escolha_usuario", "")
        rotulo_caso = _humanizar_tipo_caso(tipo_caso)
        pares.append((
            f"Resolução estrutural · {rotulo_caso}",
            formatar_valor_ou_traco(escolha, origem="ausencia"),
        ))

    # Modelo T-MODELO aplicado
    if modelo_aplicado is not None:
        nome_modelo = _ler_campo(modelo_aplicado, "nome_modelo", "")
        tipo_aplicacao = _ler_campo(modelo_aplicado, "tipo_aplicacao", "")
        pares.append((
            "Modelo aplicado",
            formatar_valor_ou_traco(nome_modelo, origem="ausencia"),
        ))
        pares.append((
            "Compatibilidade do modelo",
            _humanizar_tipo_aplicacao(tipo_aplicacao),
        ))
        campos_nao_casados = _ler_campo(
            modelo_aplicado, "campos_nao_casados", None,
        )
        if campos_nao_casados:
            pares.append((
                "Campos sem correspondência",
                _formatar_lista(campos_nao_casados),
            ))

    # Estados não escolhidos (Modo 4) · parte do Diagnóstico quando presente
    estados_nao_escolhidos = _get_cfg(config_usada, "estados_nao_escolhidos")
    if estados_nao_escolhidos:
        pares.append((
            "Estados excluídos da comparação",
            _formatar_lista(estados_nao_escolhidos),
        ))

    # Decisões de usuário catalogadas no diagnóstico
    decisoes = _extrair_decisoes(t_diag)
    for d in decisoes:
        contexto = _ler_campo(d, "contexto", "")
        escolhido = _ler_campo(d, "valor_escolhido", None)
        rotulo = _humanizar_contexto_decisao(contexto)
        valor = formatar_valor_ou_traco(escolhido, origem="ausencia")
        pares.append((rotulo, str(valor)))

    if not pares:
        linha = _escrever_frase(
            ws, linha, col_campo, col_fim,
            "Análise executada com configurações padrão.",
            paleta,
        )
        return linha + 1

    linha = _escrever_pares(ws, linha, col_campo, col_valor, pares, paleta)
    return linha + 1


def _humanizar_tipo_caso(tipo: str) -> str:
    tabela = {
        "NIVEL_AGRUPAMENTO_DIFERENTE": "nível de agrupamento diferente",
        "COLUNA_PRESENTE_EM_UM_LADO":  "coluna presente em um lado apenas",
        "TIPO_CAMPO_INCOMPATIVEL":     "tipo de campo incompatível",
        "VALOR_UNICO_DIVERGENTE":      "valor único divergente",
        "ORDEM_MAGNITUDE_DIVERGENTE":  "ordem de magnitude divergente",
    }
    if tipo in tabela:
        return tabela[tipo]
    if not tipo:
        return "caso estrutural"
    return tipo.replace("_", " ").lower()


def _humanizar_tipo_aplicacao(tipo: str) -> str:
    tabela = {
        "COMPLETA":     "Totalmente compatível",
        "PARCIAL":      "Parcialmente compatível",
        "INCOMPATIVEL": "Incompatível",
    }
    if tipo in tabela:
        return tabela[tipo]
    if not tipo:
        return formatar_valor_ou_traco(None, origem="ausencia")
    return tipo.replace("_", " ").capitalize()


def _humanizar_contexto_decisao(contexto: str) -> str:
    tabela = {
        "threshold_dominante_editado": "Limite editado manualmente",
        "escape_cardinalidade_acionado": "Escape acionado (cardinalidade)",
        "limiar_estabilidade_editado": "Limite de estabilidade editado",
        "limite_variacao_extrema_editado": "Limite de variação extrema editado",
    }
    if contexto in tabela:
        return tabela[contexto]
    if not contexto:
        return "Decisão"
    return contexto.replace("_", " ").capitalize()


def _extrair_decisoes(t_diag: Any) -> List[Any]:
    if t_diag is None:
        return []
    if hasattr(t_diag, "decisoes_usuario"):
        return list(getattr(t_diag, "decisoes_usuario") or [])
    if isinstance(t_diag, dict):
        return list(t_diag.get("decisoes_usuario") or [])
    return []


# ---------------------------------------------------------------------------
# Seção 5 · Configurações avançadas aplicadas
# ---------------------------------------------------------------------------

def _renderizar_secao_5(
    ws: Worksheet, linha: int, col_campo: int, col_valor: int,
    col_fim: int, config_usada: Any, paleta: Paleta,
    vocabulario: Dict[str, Dict[str, str]],
) -> int:
    linha = escrever_secao(ws, linha, col_campo, NOMES_SECOES[4], paleta)

    pares: List[Tuple[str, str]] = []

    # Paleta aplicada · sempre presente (passada como argumento no _head)
    nome_paleta_cfg = _get_cfg(config_usada, "paleta_aplicada")
    nome_paleta_exibido = nome_paleta_cfg or paleta.nome
    pares.append((
        "Paleta selecionada",
        _rotular_paleta(str(nome_paleta_exibido)),
    ))

    # Thresholds efetivamente usados · D-202 · respeita contrato D-166
    # (percentual vs contagem) via `formatar_threshold_por_contrato`.
    # Bug histórico: capability formatava TODOS thresholds como percentual.
    thresholds = _get_cfg(config_usada, "thresholds_usados")
    if thresholds:
        for chave, valor in thresholds.items():
            rotulo_contrato, valor_fmt = formatar_threshold_por_contrato(chave, valor)
            # Prioriza tradução do vocabulário · fallback para rótulo do contrato.
            rotulo_voc = _traduzir_threshold(chave, vocabulario)
            rotulo = rotulo_voc if not rotulo_voc.startswith("[") else rotulo_contrato
            pares.append((rotulo, valor_fmt))

    # Defaults sobrescritos · registrados explicitamente quando relevantes
    defaults_sobrescritos = _get_cfg(config_usada, "defaults_sobrescritos")
    if defaults_sobrescritos:
        for chave, info in defaults_sobrescritos.items():
            rotulo_voc = _traduzir_threshold(chave, vocabulario)
            rotulo_contrato, _ = formatar_threshold_por_contrato(chave, None)
            rotulo = rotulo_voc if not rotulo_voc.startswith("[") else rotulo_contrato
            if isinstance(info, dict):
                default = info.get("default")
                atual = info.get("atual")
                _, default_fmt = formatar_threshold_por_contrato(chave, default)
                _, atual_fmt = formatar_threshold_por_contrato(chave, atual)
                valor = f"Padrão: {default_fmt} · Aplicado: {atual_fmt}"
            else:
                valor = formatar_valor_ou_traco(info, origem="ausencia")
            pares.append((f"{rotulo} (sobrescrito)", valor))

    if len(pares) == 1:  # só a paleta · sem thresholds
        # frase explicativa complementar · C.2 nada silencioso
        pares.append((
            "Limites aplicados",
            "Padrões da Fundação · sem edições pelo usuário.",
        ))

    linha = _escrever_pares(ws, linha, col_campo, col_valor, pares, paleta)
    return linha + 1


def _rotular_paleta(nome: str) -> str:
    tabela = {
        "azul":  "Azul executivo",
        "verde": "Verde executivo",
        "cinza": "Cinza executivo",
        "vinho": "Vinho executivo",
    }
    return tabela.get(nome.lower(), nome.capitalize())


# ---------------------------------------------------------------------------
# Seção 6 · Qualidade estrutural
# ---------------------------------------------------------------------------

def _renderizar_secao_6(
    ws: Worksheet, linha: int, col_campo: int, col_valor: int,
    col_fim: int, warnings: List[Any], config_usada: Any,
    t_diag: Any, paleta: Paleta,
    vocabulario: Dict[str, Dict[str, str]],
) -> int:
    linha = escrever_secao(ws, linha, col_campo, NOMES_SECOES[5], paleta)

    total = len(warnings) if warnings else 0
    pares: List[Tuple[str, str]] = []

    # Contagem total · user-facing
    pares.append((
        "Total de avisos emitidos",
        f"{total:,}".replace(",", "."),
    ))

    # Agregado por categoria (contagem · não lista de códigos)
    por_cat: Dict[str, int] = {}
    for w in warnings or []:
        cat = _normalizar_categoria(_ler_campo(w, "categoria", "informativo"))
        por_cat[cat] = por_cat.get(cat, 0) + 1
    for cat, qtd in por_cat.items():
        pares.append((
            f"Avisos · {_traduzir_categoria_warning(cat)}",
            f"{qtd:,}".replace(",", "."),
        ))

    # Integridade estrutural (quando presente no t_diag)
    integridade = _extrair_integridade(t_diag)
    if integridade is not None:
        total_orig = _ler_campo(integridade, "total_linhas_base_original", None)
        total_proc = _ler_campo(integridade, "total_linhas_processadas", None)
        total_excl = _ler_campo(integridade, "total_linhas_excluidas", None)
        pares.append((
            "Linhas na base original",
            formatar_valor_ou_traco(
                f"{int(total_orig):,}".replace(",", ".")
                if total_orig is not None else None,
                origem="nulo",
            ),
        ))
        pares.append((
            "Linhas após processamento",
            formatar_valor_ou_traco(
                f"{int(total_proc):,}".replace(",", ".")
                if total_proc is not None else None,
                origem="nulo",
            ),
        ))
        pares.append((
            "Linhas excluídas no processamento",
            formatar_valor_ou_traco(
                f"{int(total_excl):,}".replace(",", ".")
                if total_excl is not None else None,
                origem="nulo",
            ),
        ))
        # Motivos de exclusão por categoria
        motivos = _ler_campo(integridade, "motivo_exclusao_por_categoria", None)
        if motivos:
            for cat, qtd in motivos.items():
                pares.append((
                    f"Excluídas · {_humanizar_motivo_exclusao(str(cat))}",
                    f"{int(qtd):,}".replace(",", "."),
                ))

    # Nulos por classificação (quando V{N}Result expõe em config_usada)
    nulos_por_cat = _get_cfg(config_usada, "nulos_por_classificacao")
    if nulos_por_cat:
        for cat, qtd in nulos_por_cat.items():
            rotulo = _rotular_classificacao_user_facing(str(cat), vocabulario)
            pares.append((
                f"Valores ausentes · {rotulo}",
                f"{int(qtd):,}".replace(",", "."),
            ))

    linha = _escrever_pares(ws, linha, col_campo, col_valor, pares, paleta)
    return linha + 1


def _extrair_integridade(t_diag: Any) -> Any:
    if t_diag is None:
        return None
    if hasattr(t_diag, "integridade"):
        return getattr(t_diag, "integridade")
    if isinstance(t_diag, dict):
        return t_diag.get("integridade")
    return None


def _humanizar_motivo_exclusao(motivo: str) -> str:
    tabela = {
        "NULO_CHAVE":     "Chave ausente",
        "NULO_VALOR":     "Valor ausente",
        "DUPLICATA":      "Duplicata",
        "TIPO_INVALIDO":  "Tipo inválido",
        "FORA_ESCOPO":    "Fora de escopo",
    }
    if motivo in tabela:
        return tabela[motivo]
    return motivo.replace("_", " ").capitalize()


def _rotular_classificacao_user_facing(
    codigo: str, vocabulario: Dict[str, Dict[str, str]],
) -> str:
    bloco = vocabulario.get("classificacoes", {})
    if codigo in bloco:
        return bloco[codigo]
    return codigo.replace("_", " ").capitalize()


# ---------------------------------------------------------------------------
# Utilitário genérico de leitura de campo (dict ou objeto)
# ---------------------------------------------------------------------------

def _ler_campo(origem: Any, chave: str, default: Any = None) -> Any:
    if origem is None:
        return default
    if isinstance(origem, dict):
        return origem.get(chave, default)
    return getattr(origem, chave, default)


# ---------------------------------------------------------------------------
# Classificador de warnings por gravidade (seção 2 vs seção 3)
# ---------------------------------------------------------------------------

def _split_warnings_por_gravidade(
    warnings: Optional[List[Any]],
) -> Tuple[List[Any], List[Any]]:
    """Retorna (leves_para_secao_2, atencao_para_secao_3)."""
    if not warnings:
        return [], []
    leves: List[Any] = []
    atencao: List[Any] = []
    for w in warnings:
        cat = _normalizar_categoria(_ler_campo(w, "categoria", "informativo"))
        if cat in _CATEGORIAS_AJUSTE:
            leves.append(w)
        elif cat in _CATEGORIAS_ATENCAO:
            atencao.append(w)
        else:
            # categorias fora do catálogo conhecido · classifica como atenção
            # (C.2 · nunca silenciar) · warning em log
            _logger.warning(
                "categoria de warning '%s' fora do catálogo · classificando como atenção",
                cat,
            )
            atencao.append(w)
    return leves, atencao


# ---------------------------------------------------------------------------
# Renderizador principal
# ---------------------------------------------------------------------------

def renderizar_diagnostico(
    ws: Worksheet,
    config_usada: Any,
    resolucao_estrutural: Any,
    modelo_aplicado: Any,
    t_diag: Any,
    warnings: Optional[List[Any]],
    paleta: Paleta,
    vocabulario: Dict[str, Dict[str, str]],
) -> None:
    """
    Renderiza a aba Diagnóstico com 6 seções fixas user-facing (D-165).

    Interface canônica D-159 · argumentos já desempacotados do V{N}Result
    pela A-VN que consome esta capability:

      - config_usada: dict (ou objeto) com campos canônicos esperados
        (arquivo · aba_consumida · modo_base · agrupadores · campo_analisado ·
        tipo_medida · colunas_mapeadas · paleta_aplicada · thresholds_usados ·
        defaults_sobrescritos · nulos_por_classificacao · estados_nao_escolhidos)
      - resolucao_estrutural: obj com tipo_caso · escolha_usuario · ou None
      - modelo_aplicado: obj com nome_modelo · tipo_aplicacao · ou None
      - t_diag: DiagnosticoVN (ou estrutura-like) · expõe ajustes_aplicados ·
        decisoes_usuario · integridade
      - warnings: lista de WarningEstrutural (ou dicts) · split interno por
        gravidade · leves → Seção 2 · atenção → Seção 3
      - paleta: Paleta do catálogo D-164
      - vocabulario: dict bilingue carregado via capability 2

    Regras invioláveis (blocos 7 e 8 de vocabulario_bilingue.md v2):
      - Nenhum código técnico literal em célula renderizada
      - Nenhum `None`/`NaN` literal · sempre `—` com microcopy contextual
      - Nenhum dict Python serializado · sempre estruturado

    Campos canônicos ausentes em `config_usada` são tratados como `—` com
    microcopy "(não consta)" · capability nunca inventa conteúdo.
    """
    if ws is None:
        raise TypeError("ws não pode ser None em renderizar_diagnostico")
    if paleta is None:
        raise TypeError("paleta não pode ser None em renderizar_diagnostico")
    if vocabulario is None:
        raise TypeError("vocabulario não pode ser None em renderizar_diagnostico")

    # Largura útil · 2 colunas de conteúdo (campo, valor) · wrapper até col 4
    # para frases merged (headers de seção) e título da aba.
    COL_CAMPO = 1
    COL_VALOR = 2
    COL_FIM = 4

    linha = 1

    # Título da aba
    linha = escrever_titulo_aba(
        ws, linha, COL_CAMPO, COL_FIM, "Diagnóstico", paleta,
    )
    linha += 1  # respiro

    # Split warnings por gravidade · feito UMA vez · reutilizado nas seções 2 e 3
    leves, atencao = _split_warnings_por_gravidade(warnings)

    # Seção 1
    linha = _renderizar_secao_1(
        ws, linha, COL_CAMPO, COL_VALOR, COL_FIM,
        config_usada, paleta, vocabulario,
    )
    # Seção 2
    linha = _renderizar_secao_2(
        ws, linha, COL_CAMPO, COL_VALOR, COL_FIM,
        t_diag, leves, paleta, vocabulario,
    )
    # Seção 3
    linha = _renderizar_secao_3(
        ws, linha, COL_CAMPO, COL_VALOR, COL_FIM,
        atencao, paleta, vocabulario,
    )
    # Seção 4
    linha = _renderizar_secao_4(
        ws, linha, COL_CAMPO, COL_VALOR, COL_FIM,
        config_usada, resolucao_estrutural, modelo_aplicado,
        t_diag, paleta, vocabulario,
    )
    # Seção 5
    linha = _renderizar_secao_5(
        ws, linha, COL_CAMPO, COL_VALOR, COL_FIM,
        config_usada, paleta, vocabulario,
    )
    # Seção 6
    linha = _renderizar_secao_6(
        ws, linha, COL_CAMPO, COL_VALOR, COL_FIM,
        warnings or [], config_usada, t_diag, paleta, vocabulario,
    )

    # Larguras de coluna fixas · consistência visual com Resumo Executivo
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 24

    ws.sheet_view.showGridLines = False
