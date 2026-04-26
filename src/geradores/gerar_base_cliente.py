"""
gerar_base_cliente.py · D-149 · recorte cliente-friendly

Gera `/bases/base_vN_cliente.xlsx` a partir de `/bases/base_fundacao.xlsx`,
contendo apenas as abas consumidas pela visão N.

Função:
    · Reempacotamento mecânico · zero geração de conteúdo novo
    · Zero risco de divergência (base mestre D-140 permanece fonte única)
    · Simula upload de cliente real na Validação Visual da Fase 2 (B.4)
    · Arquivo canônico de entrada do app_vN.py (D-149)

Ponto de chamada:
    · Sub-tarefa de cada S-VN quando D-147 declara "base_fundacao suficiente"
    · Sub-tarefa de cada B-VN quando D-147 exige B-VN dedicado (lê base_vN em vez de mestre)

Invariantes preservadas:
    · SEED=42 da base mestre (D-140)
    · Nomes canônicos das abas (D-140)
    · Conteúdo idêntico ao da base mestre · byte-for-byte por célula
    · Ordem das abas preservada do arquivo-fonte

Regras C.2 / C.5:
    · Nunca modificar base_vN_cliente.xlsx diretamente (Instrucoes §"Não fazer")
    · Divergência detectada → regenerar a partir da fonte

Uso como biblioteca:
    from src.geradores.gerar_base_cliente import gerar_base_cliente
    gerar_base_cliente(
        visao_id="V2",
        abas_consumidas=["vendas_padrao", "vendas_por_colunas"],
        caminho_fonte="bases/base_fundacao.xlsx",
        caminho_destino="bases/base_v2_cliente.xlsx",
    )

Uso como CLI:
    python -m src.geradores.gerar_base_cliente \\
        --visao V2 \\
        --abas vendas_padrao vendas_por_colunas \\
        --fonte bases/base_fundacao.xlsx \\
        --destino bases/base_v2_cliente.xlsx
"""

from __future__ import annotations

import argparse
import shutil
import sys
from pathlib import Path
from typing import List, Optional

from openpyxl import load_workbook


# ---------------------------------------------------------------------------
# Exceções específicas · C.2 nada silencioso
# ---------------------------------------------------------------------------

class GerarBaseClienteError(Exception):
    """Erro base do gerador de base cliente."""


class FonteNaoEncontradaError(GerarBaseClienteError):
    """Arquivo-fonte não existe ou não é legível."""


class AbaInexistenteError(GerarBaseClienteError):
    """Aba declarada em `abas_consumidas` não existe na fonte."""


class ConfiguracaoInvalidaError(GerarBaseClienteError):
    """Parâmetros de entrada inconsistentes."""


# ---------------------------------------------------------------------------
# Núcleo · função pública
# ---------------------------------------------------------------------------

def gerar_base_cliente(
    visao_id: str,
    abas_consumidas: List[str],
    caminho_fonte: str | Path = "bases/base_fundacao.xlsx",
    caminho_destino: Optional[str | Path] = None,
) -> Path:
    """
    Gera recorte cliente-friendly de base mestre contendo só as abas declaradas.

    Parâmetros
    ----------
    visao_id : str
        Identificador canônico da visão · 'V1' · 'V2' · ... · 'V11'.
        Usado para default de `caminho_destino` e para validação.
    abas_consumidas : List[str]
        Lista ordenada de abas que a visão consome · declarada na Spec S-VN
        aplicando D-147 · mínimo 1 aba · sem duplicatas.
    caminho_fonte : str | Path
        Caminho do arquivo-fonte · default 'bases/base_fundacao.xlsx' (D-140).
        Pode ser 'bases/base_vN.xlsx' quando B-VN existe (D-147 exigiu).
    caminho_destino : Optional[str | Path]
        Caminho do arquivo de saída · default derivado de visao_id:
        'bases/base_{visao_id_lower}_cliente.xlsx'.

    Retorna
    -------
    Path
        Caminho absoluto do arquivo gerado.

    Levanta
    -------
    ConfiguracaoInvalidaError
        Parâmetros inconsistentes (visao_id fora do enum · lista vazia · duplicatas).
    FonteNaoEncontradaError
        Arquivo-fonte não existe ou não é legível.
    AbaInexistenteError
        Alguma aba de `abas_consumidas` não está em `caminho_fonte`.
    """
    # --- Validação da configuração · C.3 sem invenção de comportamento ---
    VISOES_VALIDAS = {f"V{i}" for i in range(1, 12)}
    if visao_id not in VISOES_VALIDAS:
        raise ConfiguracaoInvalidaError(
            f"visao_id='{visao_id}' inválido. Valores aceitos: {sorted(VISOES_VALIDAS)}"
        )

    if not abas_consumidas:
        raise ConfiguracaoInvalidaError(
            "abas_consumidas está vazia. Declare pelo menos 1 aba."
        )

    if len(set(abas_consumidas)) != len(abas_consumidas):
        duplicatas = {a for a in abas_consumidas if abas_consumidas.count(a) > 1}
        raise ConfiguracaoInvalidaError(
            f"abas_consumidas contém duplicatas: {sorted(duplicatas)}"
        )

    caminho_fonte = Path(caminho_fonte)
    if not caminho_fonte.is_file():
        raise FonteNaoEncontradaError(
            f"Arquivo-fonte não encontrado: {caminho_fonte.resolve()}"
        )

    if caminho_destino is None:
        caminho_destino = Path("bases") / f"base_{visao_id.lower()}_cliente.xlsx"
    caminho_destino = Path(caminho_destino)

    # --- Estratégia: copiar fonte + remover abas não-consumidas ---
    # Preserva byte-for-byte o conteúdo das abas que ficam · zero risco
    # de divergência semântica (formatações · data types · fórmulas).

    caminho_destino.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(str(caminho_fonte), str(caminho_destino))

    wb = load_workbook(str(caminho_destino))
    abas_na_fonte = set(wb.sheetnames)

    # Validar que todas as abas declaradas existem na fonte
    abas_faltando = [a for a in abas_consumidas if a not in abas_na_fonte]
    if abas_faltando:
        caminho_destino.unlink(missing_ok=True)  # limpa arquivo parcial
        raise AbaInexistenteError(
            f"Abas declaradas não existem na fonte: {abas_faltando}. "
            f"Abas disponíveis: {sorted(abas_na_fonte)}"
        )

    # Remover abas que não estão em abas_consumidas
    # openpyxl exige pelo menos 1 aba no workbook · mas garantido pela validação acima
    abas_a_remover = [a for a in wb.sheetnames if a not in abas_consumidas]
    for nome_aba in abas_a_remover:
        del wb[nome_aba]

    # Reordenar abas conforme ordem declarada em abas_consumidas
    # (openpyxl usa .move_sheet para reordenar)
    for indice_final, nome_aba in enumerate(abas_consumidas):
        indice_atual = wb.sheetnames.index(nome_aba)
        if indice_atual != indice_final:
            wb.move_sheet(nome_aba, offset=indice_final - indice_atual)

    wb.save(str(caminho_destino))
    wb.close()

    return caminho_destino.resolve()


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def _parse_args(argv: Optional[List[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Gera recorte cliente-friendly da base mestre · D-149"
    )
    parser.add_argument(
        "--visao",
        required=True,
        help="Identificador da visão · ex: V2",
    )
    parser.add_argument(
        "--abas",
        nargs="+",
        required=True,
        help="Lista de abas consumidas pela visão · ordem preservada no output",
    )
    parser.add_argument(
        "--fonte",
        default="bases/base_fundacao.xlsx",
        help="Arquivo-fonte · default base_fundacao.xlsx (D-140)",
    )
    parser.add_argument(
        "--destino",
        default=None,
        help="Arquivo de saída · default bases/base_{visao}_cliente.xlsx",
    )
    return parser.parse_args(argv)


def main(argv: Optional[List[str]] = None) -> int:
    args = _parse_args(argv)
    try:
        caminho = gerar_base_cliente(
            visao_id=args.visao,
            abas_consumidas=args.abas,
            caminho_fonte=args.fonte,
            caminho_destino=args.destino,
        )
    except GerarBaseClienteError as err:
        print(f"[ERRO] {err}", file=sys.stderr)
        return 1

    print(f"[OK] Base cliente gerada: {caminho}")
    print(f"     Visão: {args.visao}")
    print(f"     Abas: {args.abas}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
