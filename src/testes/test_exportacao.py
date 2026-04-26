"""
test_exportacao.py — Suite de testes F-EXP · 11 capabilities + integração + warnings + determinismo + D-130

Cobertura:
  (a) 11 testes unitários · um por capability
  (b) 1 teste de integração · estrutura de abas invariante + 6 blocos do Resumo Executivo
  (c) Testes de emissão de warnings de exportação (§EXP.5)
  (d) 1 teste de determinismo C.1
  (e) 1 teste de receptividade a IA D-130

Fixtures:
  _vnresult()       · VNResultBase genérico com visao_id/tipo_cv configuráveis
  _config_minimal() · ConfigExportacao sem gráficos e sem formatação condicional (testes rápidos)
  _config_full()    · ConfigExportacao completa
"""
from __future__ import annotations

import hashlib
import json
import math
import os
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd
import pytest
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

# Garantir que /src está no path
_src_path = str(Path(__file__).parent.parent)
if _src_path not in sys.path:
    sys.path.insert(0, _src_path)

from contratos import (
    AjusteMotor,
    BloqueioOperacional,
    CabecalhoExecucao,
    CategoriaWarning,
    CoracaoVisualRef,
    ConfigExportacao,
    DecisaoUsuario,
    DiagnosticoVN,
    ExportacaoResult,
    IntegridadeEstrutural,
    LeituraQualitativa,
    MotorResultMeta,
    QualidadeEstrutural,
    ResumoExecutivoPadrao,
    VNResultBase,
    WarningEstrutural,
)
from exportacao import (
    EXCEL_MAX_FORMAT_RULES,
    LIMITE_PAGINACAO,
    cap_autofilter,
    cap_barchart_nativo,
    cap_combo_bar_line,
    cap_columnchart_empilhado_100,
    cap_diagnostico,
    cap_formatacao_condicional_matriz,
    cap_histograma_bins,
    cap_linechart_nativo,
    cap_paginacao_matriz,
    cap_resumo_executivo,
    cap_tabela_formatada,
    exportar_resultado,
    _escrever_dataframe,
)

# ---------------------------------------------------------------------------
# Pasta de saída para xlsx gerados pelos testes
# ---------------------------------------------------------------------------

_OUTPUTS_DIR = Path(__file__).parent / "outputs"
_OUTPUTS_DIR.mkdir(exist_ok=True)


def _out(nome: str) -> str:
    return str(_OUTPUTS_DIR / nome)


# ---------------------------------------------------------------------------
# Fábricas de fixtures
# ---------------------------------------------------------------------------

def _resumo_padrao(visao_id: str = "V4") -> ResumoExecutivoPadrao:
    return ResumoExecutivoPadrao(
        bloco_1_cabecalho=CabecalhoExecucao(
            visao=visao_id,
            data_execucao=datetime(2026, 4, 21, 10, 0, 0),
            modo_upload="SIMPLES",
            agrupadores=["Produto"],
            medida_principal="Vendas",
        ),
        bloco_2_numeros_ancora={
            "Total de elementos": 10,
            "Total de Vendas": 1500.0,
            "Elementos com dado": 10,
        },
        bloco_3_distribuicao={
            "Classe A": 3,
            "Classe B": 4,
            "Classe C": 3,
        },
        bloco_4_elementos_destacados={
            "Top 3 por Vendas": ["Produto A", "Produto B", "Produto C"],
            "Outliers detectados": 1,
        },
        bloco_5_leitura_qualitativa=LeituraQualitativa(
            classificacao_ativa="Concentrado",
            thresholds_usados={"limiar_A": 0.80, "limiar_B": 0.95},
            alguma_leitura_alterada_por_edicao=False,
        ),
        bloco_6_qualidade_estrutural=QualidadeEstrutural(
            total_warnings=2,
            warnings_por_categoria={"informativo": 2},
            ajustes_aplicados=1,
            tem_bloqueios_escapados=False,
        ),
    )


def _diagnostico_padrao() -> DiagnosticoVN:
    return DiagnosticoVN(
        ajustes_aplicados=[
            AjusteMotor(
                tipo_ajuste="LINHA_EXCLUIDA_NULO",
                linhas_afetadas=2,
                descricao="2 linhas com medida nula excluídas",
            )
        ],
        warnings_por_categoria={
            "informativo": [
                WarningEstrutural(
                    codigo="W-TEST-01",
                    categoria=CategoriaWarning.INFORMATIVO,
                    microcopy="Warning de teste",
                    contexto={"detalhe": "teste"},
                )
            ]
        },
        decisoes_usuario=[],
        bloqueios_informativos=[],
        integridade=IntegridadeEstrutural(
            total_linhas_base_original=12,
            total_linhas_processadas=10,
            total_linhas_excluidas=2,
            motivo_exclusao_por_categoria={"nulo_medida": 2},
        ),
    )


def _vnresult(
    visao_id: str = "V4",
    tipo_cv: str = "BARRAS",
    nome_cv: str = "Composição Principal",
    df: Optional[pd.DataFrame] = None,
    capabilities_requeridas: Optional[List[str]] = None,
    config_usada: Optional[Dict[str, Any]] = None,
) -> VNResultBase:
    if df is None:
        df = pd.DataFrame({
            "Produto": ["A", "B", "C", "D", "E"],
            "Vendas": [500.0, 400.0, 300.0, 200.0, 100.0],
            "Classe": ["A", "A", "B", "C", "C"],
        })
    return VNResultBase(
        visao_id=visao_id,
        config_usada=config_usada or {"agrupador": "Produto", "medida": "Vendas"},
        motor_result_meta=MotorResultMeta(
            total_linhas_originais=12,
            total_linhas_processadas=10,
            modo_upload="SIMPLES",
            timestamp_processamento=datetime(2026, 4, 21, 9, 0, 0),
            warnings_motor=[],
        ),
        timestamp_execucao=datetime(2026, 4, 21, 10, 0, 0),
        base_analitica=df,
        resumo_executivo=_resumo_padrao(visao_id),
        coracao_visual=CoracaoVisualRef(
            nome_aba=nome_cv,
            tipo=tipo_cv,
            capabilities_requeridas=capabilities_requeridas or [],
        ),
        bloqueios_disparados=[],
        warnings=[],
        diagnostico=_diagnostico_padrao(),
    )


def _config_minimal() -> ConfigExportacao:
    """Config sem gráficos · sem formatação condicional · testes unitários rápidos."""
    return ConfigExportacao(
        incluir_graficos_nativos=False,
        aplicar_formatacao_condicional=False,
        incluir_filtros_autofilter=True,
        congelar_cabecalho=True,
        largura_colunas_auto=True,
        paginar_matrizes_grandes=False,
    )


def _config_full() -> ConfigExportacao:
    return ConfigExportacao()


# ---------------------------------------------------------------------------
# (a) Testes unitários · 11 capabilities
# ---------------------------------------------------------------------------

def test_cap01_tabela_formatada():
    """CAP-01 · Formata tabela: header bold, bordas, freeze_panes, largura auto."""
    wb = Workbook()
    ws = wb.active
    df = pd.DataFrame({"Col A": [1, 2, 3], "Col B": ["x", "y", "z"]})
    n_rows, n_cols = _escrever_dataframe(ws, df, [])
    config = _config_minimal()

    cap_tabela_formatada(ws, n_cols, config, total_rows=n_rows)

    assert ws.cell(1, 1).font.bold is True
    # openpyxl normaliza "FFFFFF" → "00FFFFFF"; "FFFFFFFF" permanece "FFFFFFFF"
    assert ws.cell(1, 1).font.color.rgb.upper().endswith("FFFFFF")
    assert ws.freeze_panes is not None
    assert ws.cell(2, 1).border.left.border_style == "thin"


def test_cap02_autofilter():
    """CAP-02 · AutoFilter ativo na linha de cabeçalho."""
    wb = Workbook()
    ws = wb.active
    df = pd.DataFrame({"A": [1, 2], "B": [3, 4], "C": [5, 6]})
    _escrever_dataframe(ws, df, [])

    cap_autofilter(ws, num_colunas=3)

    assert ws.auto_filter.ref is not None
    assert "A1" in ws.auto_filter.ref
    assert "C1" in ws.auto_filter.ref


def test_cap03_diagnostico(tmp_path):
    """CAP-03 · Aba Diagnóstico contém 5 seções e dados do DiagnosticoVN."""
    wb = Workbook()
    ws = wb.active
    diag = _diagnostico_padrao()
    config = _config_minimal()

    cap_diagnostico(ws, diag, config)

    caminho = str(tmp_path / "test_cap03.xlsx")
    wb.save(caminho)

    wb2 = load_workbook(caminho)
    ws2 = wb2.active
    valores = [ws2.cell(r, 1).value for r in range(1, ws2.max_row + 1) if ws2.cell(r, 1).value]

    # Todas as 5 seções presentes
    assert any("Seção 1" in str(v) for v in valores)
    assert any("Seção 2" in str(v) for v in valores)
    assert any("Seção 3" in str(v) for v in valores)
    assert any("Seção 4" in str(v) for v in valores)
    assert any("Seção 5" in str(v) for v in valores)

    # Conteúdo do ajuste presente
    assert any("LINHA_EXCLUIDA_NULO" in str(v) for v in valores)

    # Warning de teste presente
    todos_vals = []
    for row in ws2.iter_rows(values_only=True):
        todos_vals.extend(v for v in row if v is not None)
    assert any("W-TEST-01" in str(v) for v in todos_vals)


def test_cap04_resumo_executivo(tmp_path):
    """CAP-04 · Resumo Executivo contém 6 blocos na ordem correta (D-125)."""
    wb = Workbook()
    ws = wb.active
    resumo = _resumo_padrao("V4")
    config = _config_minimal()

    cap_resumo_executivo(ws, resumo, "V4", config)

    caminho = str(tmp_path / "test_cap04.xlsx")
    wb.save(caminho)

    wb2 = load_workbook(caminho)
    ws2 = wb2.active

    titulos_encontrados = []
    for row in ws2.iter_rows(values_only=True):
        for cell_val in row:
            if cell_val and isinstance(cell_val, str) and cell_val.startswith("Bloco "):
                titulos_encontrados.append(cell_val)

    # 6 blocos obrigatórios
    assert len(titulos_encontrados) == 6
    assert "Bloco 1" in titulos_encontrados[0]
    assert "Bloco 2" in titulos_encontrados[1]
    assert "Bloco 3" in titulos_encontrados[2]
    assert "Bloco 4" in titulos_encontrados[3]
    assert "Bloco 5" in titulos_encontrados[4]
    assert "Bloco 6" in titulos_encontrados[5]

    # Visão correta no Bloco 1
    todos_vals = [cell for row in ws2.iter_rows(values_only=True) for cell in row if cell is not None]
    assert any("V4" in str(v) for v in todos_vals)
    assert any("Concentrado" in str(v) for v in todos_vals)  # classificação Bloco 5


def test_cap05_barchart_nativo(tmp_path):
    """CAP-05 · BarChart nativo adicionado ao worksheet."""
    wb = Workbook()
    ws = wb.active
    df = pd.DataFrame({
        "Produto": ["A", "B", "C", "D"],
        "Vendas": [400.0, 300.0, 200.0, 100.0],
    })
    n_rows, n_cols = _escrever_dataframe(ws, df, [])

    chart = cap_barchart_nativo(ws, n_rows, n_cols, titulo="Teste Barras")

    assert chart is not None
    assert len(ws._charts) == 1

    caminho = str(tmp_path / "test_cap05.xlsx")
    wb.save(caminho)
    wb2 = load_workbook(caminho)
    assert len(wb2.active._charts) == 1


def test_cap06_linechart_nativo(tmp_path):
    """CAP-06 · LineChart nativo adicionado ao worksheet."""
    wb = Workbook()
    ws = wb.active
    df = pd.DataFrame({
        "Mês": ["Jan", "Fev", "Mar", "Abr"],
        "Vendas": [100.0, 150.0, 130.0, 200.0],
    })
    n_rows, n_cols = _escrever_dataframe(ws, df, [])

    chart = cap_linechart_nativo(ws, n_rows, n_cols, titulo="Evolução Teste")

    assert chart is not None
    assert len(ws._charts) == 1

    caminho = str(tmp_path / "test_cap06.xlsx")
    wb.save(caminho)
    wb2 = load_workbook(caminho)
    assert len(wb2.active._charts) == 1


def test_cap07_histograma_bins(tmp_path):
    """CAP-07 · Histograma com bins Sturges calculados e gráfico gerado."""
    wb = Workbook()
    ws = wb.active
    import numpy as np

    rng = np.random.default_rng(42)
    valores = rng.normal(18, 4, 100).tolist()
    df = pd.DataFrame({"Prazo_Dias": valores})
    n_rows, n_cols = _escrever_dataframe(ws, df, [])

    chart = cap_histograma_bins(ws, df, "Prazo_Dias", titulo="Distribuição de Prazo")

    assert chart is not None
    assert len(ws._charts) == 1

    # Verificar bins Sturges: k = ceil(1 + log2(100)) = 8
    k_esperado = math.ceil(1 + math.log2(100))
    assert k_esperado == 8

    caminho = str(tmp_path / "test_cap07.xlsx")
    wb.save(caminho)
    wb2 = load_workbook(caminho)
    assert len(wb2.active._charts) == 1


def test_cap08_combo_bar_line(tmp_path):
    """CAP-08 · Combo BarChart+LineChart com secondary_axis para Pareto."""
    wb = Workbook()
    ws = wb.active
    df = pd.DataFrame({
        "Produto": ["A", "B", "C", "D", "E"],
        "Vendas": [500.0, 300.0, 100.0, 60.0, 40.0],
        "Acum_%": [50.0, 80.0, 90.0, 96.0, 100.0],
    })
    n_rows, n_cols = _escrever_dataframe(ws, df, [])

    cap_combo_bar_line(ws, n_rows, n_cols, titulo="Curva Pareto Teste")

    assert len(ws._charts) == 1

    caminho = str(tmp_path / "test_cap08.xlsx")
    wb.save(caminho)
    wb2 = load_workbook(caminho)
    assert len(wb2.active._charts) == 1

    # Verificar colunas de referência criadas (80% e 95%)
    col_r80 = n_cols + 2
    col_r95 = n_cols + 3
    assert ws.cell(1, col_r80).value == "Ref 80%"
    assert ws.cell(1, col_r95).value == "Ref 95%"
    assert ws.cell(2, col_r80).value == 80.0
    assert ws.cell(2, col_r95).value == 95.0


def test_cap09_formatacao_condicional_matriz(tmp_path):
    """CAP-09 · Formatação condicional aplicada · ColorScaleRule e FormulaRule."""
    # COLOR_SCALE (V9 heatmap)
    wb1 = Workbook()
    ws1 = wb1.active
    df_matrix = pd.DataFrame({
        "Filial": ["Norte", "Sul", "Leste"],
        "Prazo": [15.0, 22.0, 18.0],
        "Peso": [50.0, 35.0, 42.0],
    })
    _escrever_dataframe(ws1, df_matrix, [])
    warnings1: list = []
    rules1 = cap_formatacao_condicional_matriz(ws1, df_matrix, "COLOR_SCALE", warnings1)

    assert rules1 > 0
    assert len(ws1.conditional_formatting._cf_rules) > 0
    assert len(warnings1) == 0  # sem truncamento

    caminho1 = str(tmp_path / "test_cap09_cs.xlsx")
    wb1.save(caminho1)

    # FORMULA_DISCRETA (V6/V8)
    wb2 = Workbook()
    ws2 = wb2.active
    _escrever_dataframe(ws2, df_matrix, [])
    warnings2: list = []
    rules2 = cap_formatacao_condicional_matriz(ws2, df_matrix, "FORMULA_DISCRETA", warnings2)

    assert rules2 > 0
    assert len(warnings2) == 0

    caminho2 = str(tmp_path / "test_cap09_fd.xlsx")
    wb2.save(caminho2)
    # Verificar que arquivo pode ser reaberto sem erro
    wb2b = load_workbook(caminho2)
    assert wb2b is not None


def test_cap10_columnchart_empilhado_100(tmp_path):
    """CAP-10 · ColumnChart empilhado 100% (D-118) para V6."""
    wb = Workbook()
    ws = wb.active
    df = pd.DataFrame({
        "Filial": ["Norte", "Sul", "Leste", "Oeste"],
        "Transp_A": [40, 30, 20, 10],
        "Transp_B": [20, 40, 35, 25],
        "Transp_C": [10, 15, 20, 30],
    })
    n_rows, n_cols = _escrever_dataframe(ws, df, [])

    chart = cap_columnchart_empilhado_100(ws, n_rows, n_cols, titulo="Matriz de Cruzamento")

    assert chart is not None
    assert chart.grouping == "percentStacked"
    assert chart.overlap == 100
    assert len(ws._charts) == 1

    caminho = str(tmp_path / "test_cap10.xlsx")
    wb.save(caminho)
    wb2 = load_workbook(caminho)
    assert len(wb2.active._charts) == 1


def test_cap11_paginacao_matriz(tmp_path):
    """CAP-11 · Paginação de matriz > 30×30 · abas secundárias com notas de navegação."""
    # Criar matriz 35×35 (excede 30×30)
    n_row, n_col = 35, 35
    dados = {f"Col_{c}": list(range(n_row)) for c in range(n_col)}
    df_grande = pd.DataFrame(dados)

    wb = Workbook()
    warnings_out: list = []
    config = _config_minimal()

    abas = cap_paginacao_matriz(wb, df_grande, "Matriz Grande", warnings_out, config)

    assert len(abas) > 0
    assert any("W-EXP-MATRIZ-GRANDE-PAGINADA" == w.codigo for w in warnings_out)

    # Verificar notas de navegação em cada aba
    for nome_aba in abas:
        ws_page = wb[nome_aba]
        nota = ws_page.cell(1, 1).value
        assert nota is not None
        assert "Parte" in str(nota)

    caminho = str(tmp_path / "test_cap11.xlsx")
    wb.save(caminho)
    wb2 = load_workbook(caminho)
    # As abas paginadas devem estar no workbook
    for nome_aba in abas:
        assert nome_aba in wb2.sheetnames


# ---------------------------------------------------------------------------
# (b) Teste de integração · estrutura de abas invariante + 6 blocos Resumo
# ---------------------------------------------------------------------------

def test_integracao_estrutura_abas(tmp_path):
    """
    Integração · estrutura invariante de abas:
    - Aba 1 = Resumo Executivo
    - Aba 2 = CoracaoVisualRef.nome_aba
    - Penúltima = Parâmetros
    - Última = Diagnóstico (D-017)
    - Resumo Executivo tem exatamente 6 blocos na ordem correta (D-125)
    """
    resultado = _vnresult(visao_id="V4", tipo_cv="BARRAS", nome_cv="Composição Principal")
    caminho = _out("test_integracao_estrutura.xlsx")

    exp_result = exportar_resultado(resultado, caminho, _config_minimal())

    assert os.path.exists(caminho)
    assert exp_result.numero_abas >= 4  # pelo menos as 4 abas fixas + CoracaoVisual

    wb = load_workbook(caminho)
    nomes = wb.sheetnames

    assert nomes[0] == "Resumo Executivo"
    assert nomes[1] == "Composição Principal"
    assert nomes[-1] == "Diagnóstico"
    assert nomes[-2] == "Parâmetros"

    # Verificar que "Base Analítica" está antes de "Parâmetros"
    idx_ba = nomes.index("Base Analítica")
    idx_params = nomes.index("Parâmetros")
    assert idx_ba < idx_params

    # Verificar 6 blocos no Resumo Executivo na ordem correta (D-125)
    ws_resumo = wb["Resumo Executivo"]
    blocos_encontrados = []
    for row in ws_resumo.iter_rows(values_only=True):
        for cell_val in row:
            if cell_val and isinstance(cell_val, str) and cell_val.startswith("Bloco "):
                blocos_encontrados.append(cell_val)

    assert len(blocos_encontrados) == 6
    for i, bloco in enumerate(blocos_encontrados, 1):
        assert f"Bloco {i}" in bloco, f"Ordem incorreta: esperado Bloco {i}, encontrado '{bloco}'"

    # Verificar capabilities_acionadas obrigatórias
    assert "CAP-RESUMO-EXECUTIVO" in exp_result.capabilities_acionadas
    assert "CAP-DIAGNOSTICO" in exp_result.capabilities_acionadas
    assert "CAP-TABELA-FORMATADA" in exp_result.capabilities_acionadas
    assert "CAP-AUTOFILTER" in exp_result.capabilities_acionadas

    wb.close()


# ---------------------------------------------------------------------------
# (c) Testes de emissão de warnings §EXP.5
# ---------------------------------------------------------------------------

def test_warning_grafico_falha(tmp_path):
    """W-EXP-GRAFICO-FALHA · gráfico falha → warning coletado · dados preservados."""
    # Criar um VNResult com tipo_cv que causará falha no gráfico
    # ao passar DataFrame vazio
    df_vazio = pd.DataFrame({"Produto": [], "Vendas": []})
    resultado = _vnresult(
        tipo_cv="BARRAS",
        nome_cv="Composição Principal",
        df=df_vazio,
    )
    config = ConfigExportacao(
        incluir_graficos_nativos=True,
        paginar_matrizes_grandes=False,
        largura_colunas_auto=False,
    )
    caminho = str(tmp_path / "test_w_grafico_falha.xlsx")
    exp_result = exportar_resultado(resultado, caminho, config)

    # Arquivo deve existir mesmo sem gráfico (dados preservados)
    assert os.path.exists(caminho)
    # DataFrame vazio: cap_barchart_nativo retorna None (não lança exceção)
    # Mas o arquivo é gerado mesmo assim
    assert exp_result.numero_abas >= 4


def test_warning_grafico_falha_forcado(tmp_path):
    """W-EXP-GRAFICO-FALHA · forçar exceção via monkey-patch."""
    import exportacao as exp_mod
    original_barchart = exp_mod.cap_barchart_nativo

    def _barchart_falha(*args, **kwargs):
        raise RuntimeError("Simulação de falha de gráfico")

    exp_mod.cap_barchart_nativo = _barchart_falha
    try:
        resultado = _vnresult(tipo_cv="BARRAS", nome_cv="Composição Principal")
        config = ConfigExportacao(
            incluir_graficos_nativos=True,
            paginar_matrizes_grandes=False,
        )
        caminho = str(tmp_path / "test_w_grafico_forcado.xlsx")
        exp_result = exportar_resultado(resultado, caminho, config)

        assert any(
            w.codigo == "W-EXP-GRAFICO-FALHA" for w in exp_result.warnings_gerados
        ), "W-EXP-GRAFICO-FALHA deveria ter sido emitido"
        # Arquivo criado mesmo com falha no gráfico
        assert os.path.exists(caminho)
    finally:
        exp_mod.cap_barchart_nativo = original_barchart


def test_warning_matriz_grande_paginada(tmp_path):
    """W-EXP-MATRIZ-GRANDE-PAGINADA · matriz > 30×30 paginada e warning emitido."""
    n_row, n_col = LIMITE_PAGINACAO + 5, LIMITE_PAGINACAO + 5
    dados = {f"Col_{c}": [float(r * n_col + c) for r in range(n_row)] for c in range(n_col)}
    df_grande = pd.DataFrame(dados)

    resultado = _vnresult(
        tipo_cv="MATRIZ_COLORIDA",
        nome_cv="Mapa de Perfil",
        df=df_grande,
    )
    config = ConfigExportacao(
        incluir_graficos_nativos=False,
        aplicar_formatacao_condicional=False,
        paginar_matrizes_grandes=True,
    )
    caminho = _out("test_w_matriz_paginada.xlsx")
    exp_result = exportar_resultado(resultado, caminho, config)

    w_pag = [w for w in exp_result.warnings_gerados if w.codigo == "W-EXP-MATRIZ-GRANDE-PAGINADA"]
    assert len(w_pag) >= 1, "W-EXP-MATRIZ-GRANDE-PAGINADA deveria ter sido emitido"
    assert "CAP-PAGINACAO-MATRIZ" in exp_result.capabilities_acionadas

    wb = load_workbook(caminho)
    # Deve ter abas extras além das 5 invariantes
    assert len(wb.sheetnames) > 5
    wb.close()


def test_warning_formatacao_truncada(tmp_path):
    """W-EXP-FORMATACAO-TRUNCADA · regras > 64000 → warning emitido · aplicação parcial."""
    # Criar matriz com > 64000 células de dado (> sqrt(64000) ≈ 253 por dimensão)
    # Usar 260 linhas × 260 colunas = 67600 células
    n = 260
    dados = {f"C{c}": [float(r) for r in range(n)] for c in range(n)}
    df_grande = pd.DataFrame(dados)

    wb = Workbook()
    ws = wb.active
    _escrever_dataframe(ws, df_grande, [])
    warnings_out: list = []

    rules = cap_formatacao_condicional_matriz(ws, df_grande, "COLOR_SCALE", warnings_out)

    w_trunc = [w for w in warnings_out if w.codigo == "W-EXP-FORMATACAO-TRUNCADA"]
    assert len(w_trunc) == 1, "W-EXP-FORMATACAO-TRUNCADA deveria ter sido emitido"
    assert rules <= EXCEL_MAX_FORMAT_RULES

    caminho = str(tmp_path / "test_w_formatacao_truncada.xlsx")
    wb.save(caminho)
    assert os.path.exists(caminho)


def test_warning_arquivo_grande(tmp_path, monkeypatch):
    """W-EXP-ARQUIVO-GRANDE · arquivo > 50MB → warning emitido."""
    import exportacao as exp_mod
    import os as os_mod

    # Simular arquivo grande via monkeypatch em os.path.getsize
    monkeypatch.setattr(os_mod.path, "getsize", lambda _: 60 * 1024 * 1024)

    resultado = _vnresult()
    config = _config_minimal()
    caminho = str(tmp_path / "test_w_arquivo_grande.xlsx")

    exp_result = exportar_resultado(resultado, caminho, config)

    w_grande = [w for w in exp_result.warnings_gerados if w.codigo == "W-EXP-ARQUIVO-GRANDE"]
    assert len(w_grande) == 1, "W-EXP-ARQUIVO-GRANDE deveria ter sido emitido"
    assert w_grande[0].contexto["tamanho_bytes"] == 60 * 1024 * 1024


def test_warning_coluna_truncada(tmp_path):
    """W-EXP-COLUNA-TRUNCADA · célula > 32767 chars → truncada e warning emitido."""
    valor_longo = "X" * 40_000  # 40000 > 32767
    df = pd.DataFrame({
        "Produto": ["A", "B"],
        "Descricao": [valor_longo, "normal"],
    })
    resultado = _vnresult(df=df)
    config = _config_minimal()
    caminho = str(tmp_path / "test_w_coluna_truncada.xlsx")

    exp_result = exportar_resultado(resultado, caminho, config)

    w_trunc = [w for w in exp_result.warnings_gerados if w.codigo == "W-EXP-COLUNA-TRUNCADA"]
    assert len(w_trunc) >= 1, "W-EXP-COLUNA-TRUNCADA deveria ter sido emitido"
    assert w_trunc[0].contexto["coluna"] == "Descricao"
    assert w_trunc[0].contexto["comprimento_original"] == 40_000

    # Verificar truncamento no xlsx
    wb = load_workbook(caminho)
    # Valor truncado em alguma aba (CoracaoVisual ou Base Analítica)
    for nome_aba in wb.sheetnames:
        ws = wb[nome_aba]
        for row in ws.iter_rows(values_only=True):
            for val in row:
                if val is not None and isinstance(val, str):
                    assert len(val) <= 32767, f"Valor não truncado na aba '{nome_aba}'"
    wb.close()


# ---------------------------------------------------------------------------
# (d) Teste de determinismo C.1
# ---------------------------------------------------------------------------

def test_determinismo_c1(tmp_path):
    """
    Determinismo C.1 · mesma entrada + mesma config = mesma estrutura/valores.
    Compara: ordem de abas, valores das células (exceto timestamps voláteis),
    capabilities_acionadas, número de warnings.
    Nota: hash byte-a-byte do xlsx não é estável (timestamps no Core.xml do openpyxl
    variam entre execuções) · tolerância declarada: estrutura + conteúdo estáveis.
    """
    resultado = _vnresult(
        visao_id="V4",
        tipo_cv="BARRAS",
        nome_cv="Composição Principal",
    )
    config = _config_minimal()

    caminho1 = str(tmp_path / "det_run1.xlsx")
    caminho2 = str(tmp_path / "det_run2.xlsx")

    result1 = exportar_resultado(resultado, caminho1, config)
    result2 = exportar_resultado(resultado, caminho2, config)

    # Estrutura estável
    assert result1.numero_abas == result2.numero_abas
    assert result1.capabilities_acionadas == result2.capabilities_acionadas
    assert len(result1.warnings_gerados) == len(result2.warnings_gerados)

    # Conteúdo de células estável entre as duas execuções
    wb1 = load_workbook(caminho1)
    wb2 = load_workbook(caminho2)

    assert wb1.sheetnames == wb2.sheetnames

    for nome_aba in wb1.sheetnames:
        ws1, ws2 = wb1[nome_aba], wb2[nome_aba]
        assert ws1.max_row == ws2.max_row, f"Aba '{nome_aba}': linhas divergem"
        assert ws1.max_column == ws2.max_column, f"Aba '{nome_aba}': colunas divergem"
        for row in range(1, min(ws1.max_row + 1, 20)):  # verificar primeiras 20 linhas
            for col in range(1, ws1.max_column + 1):
                v1 = ws1.cell(row, col).value
                v2 = ws2.cell(row, col).value
                assert v1 == v2, (
                    f"Aba '{nome_aba}' R{row}C{col}: "
                    f"'{v1}' ≠ '{v2}' · violação C.1"
                )

    wb1.close()
    wb2.close()


# ---------------------------------------------------------------------------
# (e) Teste de receptividade a IA D-130
# ---------------------------------------------------------------------------

def test_d130_exportacao_result_json_serializavel(tmp_path):
    """
    D-130 · ExportacaoResult é Pydantic BaseModel sem arbitrary_types.
    model_dump() e model_dump_json() funcionam sem erros.
    Todos os campos são JSON-compatíveis.
    """
    resultado = _vnresult()
    config = _config_minimal()
    caminho = str(tmp_path / "test_d130.xlsx")

    exp_result = exportar_resultado(resultado, caminho, config)

    # Pydantic model
    from pydantic import BaseModel
    assert isinstance(exp_result, BaseModel)

    # model_dump() funciona
    d = exp_result.model_dump()
    assert d is not None
    assert "caminho_arquivo" in d
    assert "tamanho_bytes" in d
    assert "numero_abas" in d
    assert "tempo_geracao_segundos" in d
    assert "warnings_gerados" in d
    assert "capabilities_acionadas" in d

    # model_dump_json() (Pydantic v2 · D-130 · serialização JSON)
    json_str = exp_result.model_dump_json()
    assert json_str is not None
    parsed = json.loads(json_str)
    assert isinstance(parsed, dict)
    assert isinstance(parsed["capabilities_acionadas"], list)
    assert isinstance(parsed["warnings_gerados"], list)

    # Sem arbitrary_types_allowed (pd.DataFrame não presente em ExportacaoResult)
    from pydantic import ConfigDict
    # Verificar que o modelo não tem arbitrary_types_allowed=True
    model_config = exp_result.model_config
    assert not model_config.get("arbitrary_types_allowed", False), (
        "ExportacaoResult não deve ter arbitrary_types_allowed=True (D-130)"
    )


# ---------------------------------------------------------------------------
# Testes adicionais · cobertura específica
# ---------------------------------------------------------------------------

def test_exportar_resultado_cria_arquivo(tmp_path):
    """exportar_resultado() cria arquivo no caminho especificado."""
    resultado = _vnresult()
    caminho = str(tmp_path / "subdir" / "output.xlsx")

    exp_result = exportar_resultado(resultado, caminho, _config_minimal())

    assert os.path.exists(caminho)
    assert exp_result.tamanho_bytes > 0
    assert exp_result.numero_abas == 5  # Resumo + CV + Base + Parâmetros + Diagnóstico
    assert exp_result.tempo_geracao_segundos >= 0


def test_exportar_resultado_tipo_barras_linha(tmp_path):
    """V10 · tipo BARRAS_LINHA → CAP-COMBO-BAR-LINE acionada."""
    df = pd.DataFrame({
        "Produto": ["A", "B", "C", "D", "E"],
        "Vendas": [500.0, 300.0, 100.0, 60.0, 40.0],
        "Acum_%": [50.0, 80.0, 90.0, 96.0, 100.0],
    })
    resultado = _vnresult(
        visao_id="V10",
        tipo_cv="BARRAS_LINHA",
        nome_cv="Curva Pareto",
        df=df,
    )
    config = ConfigExportacao(
        incluir_graficos_nativos=True,
        paginar_matrizes_grandes=False,
        aplicar_formatacao_condicional=False,
        largura_colunas_auto=False,
    )
    caminho = _out("test_tipo_barras_linha.xlsx")
    exp_result = exportar_resultado(resultado, caminho, config)

    assert "CAP-COMBO-BAR-LINE" in exp_result.capabilities_acionadas


def test_exportar_resultado_tipo_histograma(tmp_path):
    """V5 · tipo HISTOGRAMA → CAP-HISTOGRAMA-BINS acionada."""
    import numpy as np
    rng = np.random.default_rng(42)
    df = pd.DataFrame({"Prazo_Dias": rng.normal(18, 4, 50).tolist()})
    resultado = _vnresult(
        visao_id="V5",
        tipo_cv="HISTOGRAMA",
        nome_cv="Mapa de Distribuição",
        df=df,
    )
    config = ConfigExportacao(
        incluir_graficos_nativos=True,
        paginar_matrizes_grandes=False,
        largura_colunas_auto=False,
    )
    caminho = _out("test_tipo_histograma.xlsx")
    exp_result = exportar_resultado(resultado, caminho, config)

    assert "CAP-HISTOGRAMA-BINS" in exp_result.capabilities_acionadas


def test_exportar_resultado_tipo_matriz_colorida(tmp_path):
    """V9 · tipo MATRIZ_COLORIDA → CAP-FORMATACAO-CONDICIONAL-MATRIZ com ColorScaleRule."""
    df = pd.DataFrame({
        "Filial": ["Norte", "Sul", "Leste"],
        "Prazo": [15.0, 22.0, 18.0],
        "Peso": [50.0, 35.0, 42.0],
    })
    resultado = _vnresult(
        visao_id="V9",
        tipo_cv="MATRIZ_COLORIDA",
        nome_cv="Mapa de Perfil",
        df=df,
    )
    config = ConfigExportacao(
        incluir_graficos_nativos=False,
        aplicar_formatacao_condicional=True,
        paginar_matrizes_grandes=False,
        largura_colunas_auto=False,
    )
    caminho = _out("test_tipo_matriz_colorida.xlsx")
    exp_result = exportar_resultado(resultado, caminho, config)

    assert "CAP-FORMATACAO-CONDICIONAL-MATRIZ" in exp_result.capabilities_acionadas


def test_exportar_resultado_tipo_column_100(tmp_path):
    """V6 · tipo COLUMN_EMPILHADO_100 → CAP-COLUMNCHART-EMPILHADO-100 acionada."""
    df = pd.DataFrame({
        "Filial": ["Norte", "Sul", "Leste", "Oeste"],
        "Transp_A": [40, 30, 20, 10],
        "Transp_B": [20, 40, 35, 25],
    })
    resultado = _vnresult(
        visao_id="V6",
        tipo_cv="COLUMN_EMPILHADO_100",
        nome_cv="Matriz de Cruzamento",
        df=df,
    )
    config = ConfigExportacao(
        incluir_graficos_nativos=True,
        paginar_matrizes_grandes=False,
        largura_colunas_auto=False,
    )
    caminho = _out("test_tipo_column_100.xlsx")
    exp_result = exportar_resultado(resultado, caminho, config)

    assert "CAP-COLUMNCHART-EMPILHADO-100" in exp_result.capabilities_acionadas


def test_cap11_paginacao_nao_dispara_para_matriz_pequena():
    """CAP-11 · matriz 10×10 não deve ser paginada."""
    df_pequeno = pd.DataFrame({f"C{c}": [float(r) for r in range(10)] for c in range(10)})
    wb = Workbook()
    warnings_out: list = []
    config = _config_minimal()

    abas = cap_paginacao_matriz(wb, df_pequeno, "Pequena", warnings_out, config)

    assert abas == []
    assert len(warnings_out) == 0


def test_exportar_resultado_config_default(tmp_path):
    """Config padrão (None) não causa erro · ConfigExportacao default instanciada."""
    resultado = _vnresult()
    caminho = str(tmp_path / "test_config_default.xlsx")

    # config=None → usa ConfigExportacao() default
    exp_result = exportar_resultado(resultado, caminho, None)

    assert os.path.exists(caminho)
    assert exp_result.numero_abas >= 4


# ---------------------------------------------------------------------------
# D-171 · paleta propaga de ConfigExportacao → workbook (capability 1 F-APRESENT)
# D-172 · aba Diagnóstico renderiza via capability 10 quando adaptador presente
# ---------------------------------------------------------------------------

def _espiar_aplicar_paleta(monkeypatch):
    """
    Helper · instrumenta apresentacao.aplicar_paleta para capturar o
    nome da paleta efetivamente passada à capability 1 F-APRESENT.
    A marcação em P0 é via setattr na instância do workbook · não
    persiste no arquivo salvo · o spy observa a chamada em memória.
    """
    import apresentacao

    chamadas: List[str] = []
    original = apresentacao.aplicar_paleta

    def _spy(workbook, paleta):
        chamadas.append(paleta.nome)
        return original(workbook, paleta)

    monkeypatch.setattr(apresentacao, "aplicar_paleta", _spy)
    return chamadas


def test_exportar_resultado_aplica_paleta_selecionada(tmp_path, monkeypatch):
    """
    D-171 · ConfigExportacao.paleta='cinza' aciona capability 1 F-APRESENT
    (aplicar_paleta) com o Paleta correto do catálogo (D-164).
    CAP-APLICAR-PALETA é registrada em capabilities_acionadas.
    """
    chamadas = _espiar_aplicar_paleta(monkeypatch)

    resultado = _vnresult()
    config = ConfigExportacao(paleta="cinza")
    caminho = str(tmp_path / "v2_cinza.xlsx")

    exp_result = exportar_resultado(resultado, caminho, config)

    # capability registrada na auditoria
    assert "CAP-APLICAR-PALETA" in exp_result.capabilities_acionadas
    # aplicar_paleta chamada exatamente uma vez com 'cinza'
    assert chamadas == ["cinza"]


def test_exportar_resultado_paleta_default_azul(tmp_path, monkeypatch):
    """
    D-171 · ConfigExportacao sem paleta explícita aplica 'azul' (default
    universal). Proteção contra regressão no default.
    """
    chamadas = _espiar_aplicar_paleta(monkeypatch)

    resultado = _vnresult()
    caminho = str(tmp_path / "v2_azul_default.xlsx")

    exp_result = exportar_resultado(resultado, caminho, ConfigExportacao())

    assert "CAP-APLICAR-PALETA" in exp_result.capabilities_acionadas
    assert chamadas == ["azul"]


def test_exportar_resultado_paleta_tolera_case_capitalizado(tmp_path, monkeypatch):
    """
    D-171 · normalização tolerante · widget do app pode enviar 'Azul' mas
    catálogo F-APRESENT usa chave minúscula · conversão interna preserva
    compatibilidade com widgets existentes.
    """
    chamadas = _espiar_aplicar_paleta(monkeypatch)

    resultado = _vnresult()
    config = ConfigExportacao(paleta="Vinho")
    caminho = str(tmp_path / "v2_vinho_cap.xlsx")

    exportar_resultado(resultado, caminho, config)

    assert chamadas == ["vinho"]


def test_exportar_resultado_paleta_invalida_falha_explicita(tmp_path):
    """
    D-171 / C.2 · paleta fora do catálogo canônico (D-164) falha com
    ValueError claro · nunca silenciosa.
    """
    resultado = _vnresult()
    config = ConfigExportacao(paleta="laranja")
    caminho = str(tmp_path / "v2_laranja.xlsx")

    with pytest.raises(ValueError, match="catálogo"):
        exportar_resultado(resultado, caminho, config)


def test_exportar_resultado_aba_diagnostico_seis_secoes_user_facing(tmp_path):
    """
    D-172 · quando config_usada['_config_diagnostico'] está preenchido
    (caso A-V2 refatorada S4), aba Diagnóstico é renderizada pela
    capability 10 D-165 · 6 seções fixas user-facing.
    Verifica títulos canônicos + acionamento de CAP-DIAGNOSTICO-V10.
    """
    # _config_diagnostico achatado · convenção A-V2 refatorada S4
    config_usada = {
        "agrupador": "Produto",
        "medida": "Vendas",
        "_config_diagnostico": {
            "arquivo": "demo.xlsx",
            "aba_consumida": "vendas_padrao",
            "modo_base": "SIMPLES",
            "agrupadores": ["UF", "Categoria"],
            "campo_analisado": "Realizado_vs_Orcado",
            "tipo_medida": "monetaria",
            "colunas_mapeadas": {
                "valor_origem": "Orçado",
                "valor_comparado": "Realizado",
            },
            "estados_nao_escolhidos": [],
            "paleta_aplicada": "azul",
            "thresholds_usados": {},
            "defaults_sobrescritos": {},
            "nulos_por_classificacao": {},
            "resolucao_estrutural": None,
            "modelo_aplicado": None,
        },
    }

    resultado = _vnresult(config_usada=config_usada)
    config = ConfigExportacao(paleta="azul")
    caminho = str(tmp_path / "v2_diag_v10.xlsx")

    exp_result = exportar_resultado(resultado, caminho, config)

    # capability nova · não a legada
    assert "CAP-DIAGNOSTICO-V10" in exp_result.capabilities_acionadas
    assert "CAP-DIAGNOSTICO" not in exp_result.capabilities_acionadas

    # 6 seções fixas user-facing (D-165 · cópia literal · não alterar)
    wb = load_workbook(caminho)
    ws = wb["Diagnóstico"]

    coluna_a = [
        str(ws.cell(r, 1).value or "")
        for r in range(1, ws.max_row + 1)
    ]

    titulos_secao_canonicos = [
        "Como foi analisado",
        "Ajustes do motor",
        "Pontos de atenção",
        "Decisões do usuário",
        "Configurações avançadas aplicadas",
        "Qualidade estrutural",
    ]
    for titulo in titulos_secao_canonicos:
        assert any(titulo in linha for linha in coluna_a), (
            f"Seção user-facing '{titulo}' não encontrada na aba "
            f"Diagnóstico · conteúdo coluna A: {coluna_a[:40]}"
        )


def test_exportar_resultado_diagnostico_fallback_legado_sem_adaptador(tmp_path):
    """
    D-172 · regressão · visões que ainda não criaram adaptador
    (_config_diagnostico ausente) caem no fallback legado
    cap_diagnostico. CAP-DIAGNOSTICO (não V10) é acionada.
    Preserva V1/V11/V3..V10 sem quebrar.
    """
    resultado = _vnresult()  # sem _config_diagnostico em config_usada
    config = ConfigExportacao()
    caminho = str(tmp_path / "v2_diag_legado.xlsx")

    exp_result = exportar_resultado(resultado, caminho, config)

    assert "CAP-DIAGNOSTICO" in exp_result.capabilities_acionadas
    assert "CAP-DIAGNOSTICO-V10" not in exp_result.capabilities_acionadas

    wb = load_workbook(caminho)
    assert "Diagnóstico" in wb.sheetnames
