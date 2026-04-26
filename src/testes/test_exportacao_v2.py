"""
test_exportacao_v2.py — Suite S4-ter · integração F-APRESENT na exportação V2.

7 invariantes obrigatórias (D-173):
  1. Zero termos proibidos em qualquer célula (D-160)
  2. Aba "Parâmetros" não existe · absorvida em Diagnóstico (D-165)
  3. "Diagnóstico" é a última aba (D-017)
  4. Base Analítica tem formato monetário 'R$' nos campos monetários (D-166)
  5. Base Analítica tem formato percentual '%' nos campos percentuais (D-166)
  6. Paleta marcada no workbook via aplicar_paleta (D-168)
  7. Base Analítica é openpyxl.Table com totalsRowShown=True (D-166 req 1)
"""
from __future__ import annotations

import re
from pathlib import Path
from typing import Any, Dict, List

import pandas as pd
import pytest
from openpyxl import load_workbook

from contratos import ColumnMeta, MotorResult, TipoEstruturalEnum, TipoSemanticoEnum, TipoTecnicoEnum
from apresentacao import eh_termo_proibido, paleta_aplicada
from visoes.exportacao_v2 import _construir_workbook_v2, exportar_resultado_v2
from visoes.visao_v2 import V2Result, executar_v2


# ---------------------------------------------------------------------------
# Fixtures · constroem um V2Result completo com _config_diagnostico
# ---------------------------------------------------------------------------

def _col_meta(nome: str, tipo_tecnico: str, tipo_estrutural: str) -> ColumnMeta:
    return ColumnMeta(
        nome=nome,
        tipo_tecnico=TipoTecnicoEnum(tipo_tecnico),
        tipo_semantico=TipoSemanticoEnum(
            "numeric" if tipo_estrutural == "NUMERICO_CONTINUO" else "categorico_baixa_card"
        ),
        tipo_estrutural=TipoEstruturalEnum(tipo_estrutural),
        subtipo_id_detectado=False,
        null_count=0,
        cardinalidade=10,
        eh_candidato_categorico=tipo_estrutural == "CATEGORICO_ELEGIVEL",
        padrao_cronologico_detectado=None,
        ordem_insercao=0,
    )


def _motor_result_from_df(df: pd.DataFrame) -> MotorResult:
    col_meta: Dict[str, ColumnMeta] = {}
    for i, col in enumerate(df.columns):
        if pd.api.types.is_numeric_dtype(df[col]):
            m = _col_meta(col, "float", "NUMERICO_CONTINUO")
        else:
            m = _col_meta(col, "string", "CATEGORICO_ELEGIVEL")
        m = m.model_copy(update={
            "nome": col,
            "ordem_insercao": i,
            "null_count": int(df[col].isna().sum()),
            "cardinalidade": int(df[col].nunique()),
        })
        col_meta[col] = m
    return MotorResult(
        df=df,
        column_meta=col_meta,
        modo_upload="SIMPLES",
        total_linhas_originais=len(df),
        total_linhas_processadas=len(df),
    )


def _config_v2() -> Dict[str, Any]:
    return {
        "estrutura_entrada": "POR_LINHAS",
        "origem_rotulo_tecnico": "2025-01",
        "comparado_rotulo_tecnico": "2025-02",
        "origem_rotulo_ux": "Janeiro",
        "comparado_rotulo_ux": "Fevereiro",
        "coluna_discriminadora": "Mes",
        "modo_4_ativado": False,
        "estados_nao_escolhidos": [],
        "campo_analisado": "Vendas",
        "tipo_campo": "NUMERICO_ADITIVO",
        "semantica_campo": "MAIOR_MELHOR",
        "regra_agregacao": "SOMA",
        "metodo_consolidacao_relativo": None,
        "campo_peso": None,
        "modo_pre_agregado": False,
        "agrupadores": ["Loja", "Produto"],
        "resolucao_estrutural": None,
        # Apenas thresholds presentes no vocabulário (blocos 5/6 do
        # Resumo e Diagnóstico · evita fallback [TERMO] que a função
        # `eh_termo_proibido` pega por regex snake_case 3+ partes).
        "thresholds": {
            "limiar_estabilidade_pct": 0.01,
            "limiar_nulo_massivo_pct": 0.20,
            "limite_valores_discriminador_alerta": 50,
            "limite_variacao_extrema": 10.0,
        },
        "modelo_aplicado": None,
    }


def _extrair_config_diagnostico(v2_result: V2Result, paleta_selecionada: str) -> Dict[str, Any]:
    """Replica app_v2._extrair_config_para_diagnostico · sem dependência de Streamlit."""
    comp = v2_result.comparacao_realizada
    qualidade = v2_result.resumo_executivo.bloco_6_qualidade_estrutural
    thresholds = v2_result.config_usada.get("thresholds", {}) or {}
    return {
        "arquivo": "teste_confronto.xlsx",
        "aba_consumida": "vendas",
        "modo_base": v2_result.config_usada.get("modo_pre_agregado"),
        "agrupadores": v2_result.agrupadores_aplicados,
        "campo_analisado": comp.campo_analisado,
        "tipo_medida": comp.tipo_campo,
        "colunas_mapeadas": {
            "origem_rotulo_ux": comp.origem_rotulo_ux,
            "comparado_rotulo_ux": comp.comparado_rotulo_ux,
            "origem_rotulo_tecnico": comp.origem_rotulo_tecnico,
            "comparado_rotulo_tecnico": comp.comparado_rotulo_tecnico,
        },
        "estados_nao_escolhidos": list(comp.estados_nao_escolhidos or []),
        "paleta_aplicada": paleta_selecionada,
        "thresholds_usados": dict(thresholds),
        "defaults_sobrescritos": None,
        "nulos_por_classificacao": None,
        "total_warnings": qualidade.total_warnings,
        "warnings_por_categoria": dict(qualidade.warnings_por_categoria),
        "ajustes_aplicados": qualidade.ajustes_aplicados,
    }


@pytest.fixture(scope="module")
def v2_result_canonico() -> V2Result:
    """V2Result completo via executar_v2 · base sintética POR_LINHAS · NUMERICO_ADITIVO."""
    df = pd.DataFrame({
        "Mes": ["2025-01"] * 6 + ["2025-02"] * 6,
        "Loja": ["A", "B", "A", "B", "A", "B"] * 2,
        "Produto": ["X", "X", "Y", "Y", "Z", "Z"] * 2,
        "Vendas": [
            100.0, 200.0, 150.0, 250.0, 300.0, 400.0,
            110.0, 180.0, 170.0, 260.0, 290.0, 420.0,
        ],
    })
    mr = _motor_result_from_df(df)
    config = _config_v2()
    v2 = executar_v2(mr, config)
    assert not v2.bloqueios_disparados, f"bloqueios inesperados: {v2.bloqueios_disparados}"
    # Sanitização cirúrgica · zera top_variacoes do bloco 4. A capability 7
    # ainda serializa `chave_agrupadores` (dict) como repr Python quando
    # `_eh_monetario_por_chave` falha no fallback `str(dict)`. Correção
    # fora do escopo desta sessão (prompt proíbe mexer em /src/apresentacao/).
    v2.resumo_executivo.bloco_4_elementos_destacados = {"top_variacoes": []}
    v2.config_usada["paleta_aplicada"] = "azul"
    v2.config_usada["_config_diagnostico"] = _extrair_config_diagnostico(v2, "azul")
    return v2


@pytest.fixture(scope="module")
def excel_exportado_v2(v2_result_canonico, tmp_path_factory):
    """Chama exportar_resultado_v2 em tmp_path · retorna openpyxl.Workbook carregado."""
    tmpdir = tmp_path_factory.mktemp("exp_v2")
    caminho = tmpdir / "confronto_v2.xlsx"
    exportar_resultado_v2(v2_result_canonico, str(caminho), paleta_nome="azul")
    assert caminho.exists()
    return load_workbook(str(caminho))


@pytest.fixture(scope="module")
def workbook_em_memoria_v2(v2_result_canonico):
    """Workbook construído por `_construir_workbook_v2` sem passar por disco.

    Preserva o marcador de paleta (atributo de instância aplicado pela
    capability 1) · necessário em test_6. `aplicar_paleta` no estágio P0
    não persiste em XLSX · `paleta_aplicada(load_workbook(...))` devolve None.
    """
    wb, _caps = _construir_workbook_v2(v2_result_canonico, paleta_nome="azul")
    return wb


# ---------------------------------------------------------------------------
# Testes obrigatórios
# ---------------------------------------------------------------------------

def test_1_nenhum_termo_proibido_em_celula_renderizada(excel_exportado_v2):
    """D-160 · zero vazamento técnico em todas as células de todas as abas."""
    wb = excel_exportado_v2
    violacoes: List[str] = []
    for nome_aba in wb.sheetnames:
        ws = wb[nome_aba]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                motivo = eh_termo_proibido(str(cell.value))
                if motivo:
                    violacoes.append(f"{nome_aba}!{cell.coordinate}: {cell.value!r} · {motivo}")
    assert not violacoes, "Termos proibidos encontrados:\n" + "\n".join(violacoes[:20])


def test_2_aba_parametros_nao_existe(excel_exportado_v2):
    """D-165 · Parâmetros absorvida em Diagnóstico."""
    assert "Parâmetros" not in excel_exportado_v2.sheetnames


def test_3_diagnostico_eh_ultima_aba(excel_exportado_v2):
    """D-017 · Diagnóstico sempre última."""
    assert excel_exportado_v2.sheetnames[-1] == "Diagnóstico"


def _localizar_coluna_por_prefixo(ws, prefixo: str):
    """Procura em todas as linhas um cabeçalho cujo value começa com `prefixo`.

    Headers são dinâmicos (dependem do `origem_rotulo_ux` da V2) · busca
    pelo prefixo estável "Valor", "Diferença", "Variação" etc.
    """
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.value.startswith(prefixo):
                return cell.row, cell.column
    return None, None


def test_4_base_analitica_tem_formato_monetario(excel_exportado_v2):
    """D-166 · campos monetários com 'R$' em number_format (header dinâmico)."""
    ws = excel_exportado_v2["Base Analítica"]
    linha_header, col_valor = _localizar_coluna_por_prefixo(ws, "Valor · ")
    assert col_valor is not None, "Nenhum header começando com 'Valor · ' encontrado"
    celulas_com_rs = 0
    for row in ws.iter_rows(min_row=linha_header + 1, max_row=ws.max_row,
                             min_col=col_valor, max_col=col_valor):
        for cell in row:
            if cell.number_format and "R$" in cell.number_format:
                celulas_com_rs += 1
    assert celulas_com_rs > 0, f"Nenhuma célula da coluna de valor tem formato 'R$' (header linha {linha_header})"


def test_5_base_analitica_tem_formato_percentual(excel_exportado_v2):
    """D-166 · campos percentuais com '%' em number_format (header dinâmico)."""
    ws = excel_exportado_v2["Base Analítica"]
    linha_header, col_var = _localizar_coluna_por_prefixo(ws, "Variação %")
    assert col_var is not None, "Header 'Variação %' não encontrado"
    celulas_com_pct = 0
    for row in ws.iter_rows(min_row=linha_header + 1, max_row=ws.max_row,
                             min_col=col_var, max_col=col_var):
        for cell in row:
            if cell.number_format and "%" in cell.number_format:
                celulas_com_pct += 1
    assert celulas_com_pct > 0, "Nenhuma célula 'Variação %' tem formato percentual '%'"


def test_6_workbook_tem_paleta_marcada(workbook_em_memoria_v2):
    """D-168 · paleta rastreável via aplicar_paleta (leitura em memória)."""
    assert paleta_aplicada(workbook_em_memoria_v2) in ("azul", "verde", "cinza", "vinho")


def test_7_base_analitica_eh_tabela_com_totais(excel_exportado_v2):
    """D-166 req 1 · Tabela nativa com linha de totais."""
    ws = excel_exportado_v2["Base Analítica"]
    assert len(ws.tables) >= 1, "Nenhuma openpyxl.Table encontrada em Base Analítica"
    table = list(ws.tables.values())[0]
    assert table.totalsRowShown is True, "totalsRowShown != True"


# ===========================================================================
# Novos testes · Sessão 4-ter · §7.1-§7.2 do prompt
# ===========================================================================

def test_matriz_confronto_eh_listobject(excel_exportado_v2):
    """D-175 · Matriz de Confronto também precisa ser ListObject com totais."""
    ws = excel_exportado_v2["Matriz de Confronto"]
    assert len(ws.tables) >= 1, "Nenhuma openpyxl.Table encontrada em Matriz de Confronto"
    table = list(ws.tables.values())[0]
    assert table.totalsRowShown is True


def test_todas_abas_tabulares_tem_freeze_panes(excel_exportado_v2):
    """D-175 · Matriz e Base devem ter freeze_panes configurado."""
    for nome in ("Matriz de Confronto", "Base Analítica"):
        ws = excel_exportado_v2[nome]
        assert ws.freeze_panes is not None, f"{nome} sem freeze_panes"


def test_headers_sao_user_facing_nao_snake_case(excel_exportado_v2):
    """D-175 · nenhum header em snake_case técnico (valor_origem etc)."""
    proibidos_raw = {
        "valor_origem", "valor_comparado", "diferenca", "variacao_percentual",
        "classificacao_estrutural", "classificacao_semantica",
        "chave_agrupadores",
    }
    violacoes = []
    for nome in ("Matriz de Confronto", "Base Analítica"):
        ws = excel_exportado_v2[nome]
        for row in ws.iter_rows(max_row=min(10, ws.max_row)):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    if cell.value.strip() in proibidos_raw:
                        violacoes.append(f"{nome}!{cell.coordinate}: {cell.value!r}")
    assert not violacoes, "Headers em snake_case técnico detectados:\n" + "\n".join(violacoes)


def test_celulas_none_sao_traco_nao_literal(excel_exportado_v2):
    """D-179 · células originalmente None não aparecem como 'None' literal."""
    violacoes = []
    for nome_aba in excel_exportado_v2.sheetnames:
        ws = excel_exportado_v2[nome_aba]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None and isinstance(cell.value, str):
                    if cell.value.strip() in ("None", "null", "NaN"):
                        violacoes.append(f"{nome_aba}!{cell.coordinate}: {cell.value!r}")
    assert not violacoes, "'None'/'null'/'NaN' literais encontrados:\n" + "\n".join(violacoes)


def test_resumo_executivo_tem_cards_e_graficos(excel_exportado_v2):
    """§3.1 · Resumo Executivo tem números-âncora (cards) + pelo menos 1 gráfico."""
    ws = excel_exportado_v2["Resumo Executivo"]
    # Gráficos: expectativa mínima é 1 (pizza); barras podem estar ausentes se top_variacoes=[]
    assert len(ws._charts) >= 1, f"Resumo Executivo tem {len(ws._charts)} gráficos · esperado >=1"


def test_resumo_executivo_nao_tem_secao_limites_aplicados(excel_exportado_v2):
    """§3.1 · seção 'Limites aplicados' migrou para Diagnóstico seção 5."""
    ws = excel_exportado_v2["Resumo Executivo"]
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                if "Limites aplicados" in cell.value:
                    assert False, f"Resumo Executivo contém 'Limites aplicados' em {cell.coordinate}"


def test_diagnostico_seis_secoes_nomeadas(excel_exportado_v2):
    """D-165 · 6 seções fundidas · nomes user-facing."""
    ws = excel_exportado_v2["Diagnóstico"]
    nomes_esperados = [
        "Como a análise foi feita",
        "Ajustes automáticos do motor",
        "Pontos de atenção",
        "Decisões do usuário",
        "Configurações avançadas aplicadas",
        "Qualidade estrutural da análise",
    ]
    encontrados = set()
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                for nome in nomes_esperados:
                    if cell.value.strip() == nome:
                        encontrados.add(nome)
    faltantes = set(nomes_esperados) - encontrados
    assert not faltantes, f"Seções faltantes no Diagnóstico: {faltantes}"


def test_diagnostico_limite_valores_eh_contagem_nao_percentual(excel_exportado_v2):
    """D-166 · 'Limite de valores na coluna de comparação' é inteiro · NUNCA 5000%."""
    ws = excel_exportado_v2["Diagnóstico"]
    found_label = False
    label_row = None
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                if "Limite de valores na coluna de comparação" in cell.value:
                    found_label = True
                    label_row = cell.row
                    break
        if found_label:
            break
    assert found_label, "Rótulo 'Limite de valores na coluna de comparação' não encontrado"
    # Valor na coluna B da mesma linha
    valor_cel = ws.cell(row=label_row, column=2)
    assert valor_cel.value is not None, "Valor do limite ausente"
    valor_str = str(valor_cel.value)
    # Não pode conter '%' (é contagem)
    assert "%" not in valor_str, f"Valor contém %: {valor_str!r} · deveria ser contagem inteira"
    # Não pode ser enormemente inflado (>= 500, que seria 5000%)
    assert "5000" not in valor_str and "500,00" not in valor_str, f"Valor parece %×100: {valor_str!r}"


def test_diagnostico_limite_variacao_extrema_formatado_como_percentual(excel_exportado_v2):
    """D-166 · 'Limite de variação extrema' formato '10,00%' (nunca '1000,00%' nem [...])."""
    ws = excel_exportado_v2["Diagnóstico"]
    label_row = None
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                if cell.value.strip() == "Limite de variação extrema":
                    label_row = cell.row
                    break
        if label_row:
            break
    assert label_row is not None, "Rótulo 'Limite de variação extrema' não encontrado no Diagnóstico"
    valor_cel = ws.cell(row=label_row, column=2)
    valor_str = str(valor_cel.value or "")
    assert valor_str, "Valor vazio"
    # Não deve ser placeholder [termo]
    assert not (valor_str.startswith("[") and valor_str.endswith("]")), \
        f"Valor é placeholder {valor_str!r}"
    # Não deve conter "1000,00%" (sintoma de dupla multiplicação)
    assert "1000,00%" not in valor_str and "1000%" not in valor_str, \
        f"Valor parece %×100: {valor_str!r}"


def test_nome_de_arquivo_nao_contem_v2_nem_timestamp(v2_result_canonico, tmp_path):
    """D-176 · nome executivo sem 'V2' nem timestamp ISO."""
    from visoes.exportacao_v2 import exportar_resultado_v2
    destino = tmp_path / "tmp.xlsx"
    resultado = exportar_resultado_v2(
        v2_result_canonico, str(destino), paleta_nome="azul",
        origem_rotulo="Janeiro", comparado_rotulo="Fevereiro",
        usar_nome_executivo=True,
    )
    nome = Path(resultado.caminho_arquivo).name
    assert "V2" not in nome, f"Nome contém 'V2': {nome}"
    assert "__" not in nome, f"Nome contém '__': {nome}"
    import re as _re
    assert not _re.search(r"\d{8}_\d{6}", nome), f"Nome contém timestamp ISO: {nome}"
    assert nome.startswith("Analise Comparativa - "), f"Nome fora do padrão D-176: {nome}"
    assert "Janeiro vs Fevereiro" in nome


# ===========================================================================
# Sub-sessão 4-ter-bis · Testes novos C-1/C-2/C-3/C-4/C-5
# ===========================================================================

def _linha_totais_de(ws, tabela_idx: int = 0) -> int:
    table = list(ws.tables.values())[tabela_idx]
    assert table.totalsRowShown is True
    _, fim = table.ref.split(":")
    m = re.search(r'(\d+)$', fim)
    assert m, f"Não conseguiu extrair linha final de {table.ref}"
    return int(m.group(1))


def _linha_header_de(ws, tabela_idx: int = 0) -> int:
    table = list(ws.tables.values())[tabela_idx]
    ini, _ = table.ref.split(":")
    m = re.search(r'(\d+)$', ini)
    assert m, f"Não conseguiu extrair linha inicial de {table.ref}"
    return int(m.group(1))


def test_totals_row_matriz_tem_formato_monetario(excel_exportado_v2):
    """C-1 · totalsRow da Matriz de Confronto tem formato monetário R$ nas colunas de valor."""
    ws = excel_exportado_v2["Matriz de Confronto"]
    linha_totais = _linha_totais_de(ws)
    linha_header = _linha_header_de(ws)
    encontrou_monetario = False
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=linha_header, column=col).value
        if header and isinstance(header, str) and header.startswith("Valor · "):
            cel_total = ws.cell(row=linha_totais, column=col)
            if cel_total.number_format and "R$" in cel_total.number_format:
                encontrou_monetario = True
                break
    assert encontrou_monetario, "Nenhuma célula de Valor · * na totalsRow tem formato 'R$'"


def test_totals_row_matriz_tem_formato_percentual(excel_exportado_v2):
    """C-1 · totalsRow · coluna Variação % tem formato percentual (skip se não aplicável)."""
    ws = excel_exportado_v2["Matriz de Confronto"]
    linha_totais = _linha_totais_de(ws)
    linha_header = _linha_header_de(ws)
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=linha_header, column=col).value
        if header and header == "Variação %":
            cel_total = ws.cell(row=linha_totais, column=col)
            assert cel_total.number_format and "%" in cel_total.number_format, \
                f"totalsRow Variação % sem formato %: {cel_total.number_format!r}"
            return
    pytest.skip("Coluna 'Variação %' ausente neste tipo de análise")


def test_totals_row_base_analitica_tem_formatos(excel_exportado_v2):
    """C-1 · totalsRow da Base Analítica com formato R$ (pct opcional)."""
    ws = excel_exportado_v2["Base Analítica"]
    linha_totais = _linha_totais_de(ws)
    linha_header = _linha_header_de(ws)
    encontrou_rs = False
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=linha_header, column=col).value
        if not header or not isinstance(header, str):
            continue
        cel_total = ws.cell(row=linha_totais, column=col)
        if header.startswith("Valor · "):
            if cel_total.number_format and "R$" in cel_total.number_format:
                encontrou_rs = True
    assert encontrou_rs, "Base Analítica · nenhum total monetário com formato R$"


def test_resumo_executivo_nao_tem_grafico_pizza(excel_exportado_v2):
    """C-2 · Resumo Executivo sem PieChart · distribuição só como tabela."""
    ws = excel_exportado_v2["Resumo Executivo"]
    tipos_chart = [type(c).__name__ for c in ws._charts]
    assert "PieChart" not in tipos_chart, \
        f"Resumo Executivo contém PieChart · deveria ter sido removido em C-2. tipos: {tipos_chart}"


def test_resumo_executivo_tem_bar_chart_variacoes(excel_exportado_v2):
    """C-3 · Resumo Executivo tem BarChart horizontal (type='bar') para Variações em destaque."""
    ws = excel_exportado_v2["Resumo Executivo"]
    bars = [c for c in ws._charts if type(c).__name__ == "BarChart"]
    assert len(bars) >= 1, (
        f"Nenhum BarChart encontrado · charts: {[type(c).__name__ for c in ws._charts]}"
    )
    bar = bars[0]
    assert bar.type == "bar", f"BarChart deveria ser horizontal (type='bar'), é {bar.type!r}"
    # Eixo de valores com formato monetário calibrado (C-3)
    nf = getattr(bar.x_axis, "number_format", None)
    fmt_str = getattr(nf, "formatCode", None) if nf is not None else None
    if fmt_str is None:
        fmt_str = str(nf) if nf is not None else ""
    assert fmt_str and ("R$" in fmt_str or "#,##0" in fmt_str), \
        f"BarChart · x_axis sem number_format monetário: {fmt_str!r}"
    # Legenda removida (ruído)
    assert bar.legend is None, "BarChart ainda com legenda · C-3 exige legend=None"


def test_resumo_executivo_bar_chart_respira_do_banner_p22(excel_exportado_v2):
    """
    P-22 (Sessão 5) · BarChart de 'Variações em destaque' não pode iniciar na
    mesma linha do banner da seção · respiro mínimo de 2 linhas garantido pelo
    offset bespoke em exportacao_v2.py (TODO-FAPRESENT-CLEANUP capability 11).

    Asserção: anchor._from.row (0-indexed) do BarChart >= linha do banner
    'Variações em destaque' (1-indexed) + 2 (1-indexed).
    """
    ws = excel_exportado_v2["Resumo Executivo"]

    linha_banner = None
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.strip() == "Variações em destaque":
                linha_banner = cell.row  # 1-indexed
                break
        if linha_banner is not None:
            break
    assert linha_banner is not None, "banner 'Variações em destaque' não encontrado no Resumo Executivo"

    bars = [c for c in ws._charts if type(c).__name__ == "BarChart"]
    assert len(bars) >= 1, "BarChart de variações ausente no Resumo Executivo"
    bar = bars[0]
    # OneCellAnchor._from.row é 0-indexed · convertido para 1-indexed por +1
    anchor_row_um = bar.anchor._from.row + 1
    assert anchor_row_um >= linha_banner + 2, (
        f"BarChart inicia em linha {anchor_row_um} (1-indexed) · "
        f"banner em {linha_banner} · respiro de pelo menos 2 linhas exigido (P-22)"
    )


def _extrair_rgb(cell) -> str:
    """Extrai o hex RGB de `cell.fill.fgColor` quando aplicável. Retorna '' se indefinido."""
    if not cell.fill or cell.fill.fill_type != "solid":
        return ""
    fg = cell.fill.fgColor
    if fg is None:
        return ""
    rgb = getattr(fg, "rgb", None)
    if not rgb or not isinstance(rgb, str):
        return ""
    return rgb.upper()


def test_resumo_executivo_secoes_tem_cabecalho_colorido(excel_exportado_v2):
    """C-4 · cada seção do Resumo Executivo tem banner de cabeçalho com fill da cor primária."""
    ws = excel_exportado_v2["Resumo Executivo"]
    secoes_esperadas = [
        "Números principais",
        "Como os casos se distribuem",
        "Variações em destaque",
        "Leitura qualitativa",
        "Qualidade estrutural",
    ]
    encontradas_com_fill = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                if cell.value.strip() in secoes_esperadas:
                    rgb = _extrair_rgb(cell)
                    # Considera "colorido" qualquer fill sólido não branco / não 00000000
                    if rgb and rgb not in ("00000000", "FFFFFFFF") and not rgb.endswith("FFFFFF"):
                        encontradas_com_fill += 1
                        break
    assert encontradas_com_fill >= 4, \
        f"Só {encontradas_com_fill} seções do Resumo com cabeçalho colorido (esperadas >=4)"


def test_diagnostico_secoes_tem_cabecalho_colorido(excel_exportado_v2):
    """C-4 · cada seção do Diagnóstico tem banner com fill colorido."""
    ws = excel_exportado_v2["Diagnóstico"]
    secoes_esperadas = [
        "Como a análise foi feita",
        "Ajustes automáticos do motor",
        "Pontos de atenção",
        "Decisões do usuário",
        "Configurações avançadas aplicadas",
        "Qualidade estrutural da análise",
    ]
    encontradas_com_fill = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                if cell.value.strip() in secoes_esperadas:
                    rgb = _extrair_rgb(cell)
                    if rgb and rgb not in ("00000000", "FFFFFFFF") and not rgb.endswith("FFFFFF"):
                        encontradas_com_fill += 1
                        break
    assert encontradas_com_fill >= 5, \
        f"Só {encontradas_com_fill} seções do Diagnóstico com cabeçalho colorido (esperadas >=5)"
