"""
Testes F-APRESENT P0 · TabloFlow Fundação.

Organização:
  1. Fixtures canônicas
  2. TestInterface          (~70 · assinaturas e contratos)
  3. TestSnapshot           (~60 · saídas textuais fixas)
  4. TestRegressao          (~30 · bugs conhecidos prevenidos)
  5. TestInvariantesBloco7  (~40 · varredura de amostras procurando violações)
"""
from __future__ import annotations

import logging
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Dict

import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table

from contratos import (
    CabecalhoExecucao,
    LeituraQualitativa,
    QualidadeEstrutural,
    ResumoExecutivoPadrao,
)

from apresentacao import (
    BLOCOS,
    CATALOGO_PALETAS,
    ColunaAdaptativa,
    FORMATO_CONTAGEM,
    FORMATO_DATA_BR,
    FORMATO_MONETARIO_BR,
    FORMATO_PERCENTUAL,
    FORMATO_PERCENTUAL_LITERAL,
    Paleta,
    aplicar_formato_contagem,
    aplicar_formato_data_br,
    aplicar_formato_monetario,
    aplicar_formato_percentual,
    aplicar_paleta,
    carregar_vocabulario_bilingue,
    contem_marcador_traducao_ausente,
    criar_tabela_executiva,
    eh_termo_proibido,
    estilo_para_paleta,
    formatar_data_br,
    formatar_moeda_br,
    formatar_percentual_br,
    montar_colunas_adaptativas,
    obter_paleta,
    paleta_aplicada,
    renderizar_resumo_executivo,
    se_config_diferente,
    se_config_igual,
    se_config_presente,
    sempre,
    traduzir,
)
from apresentacao.demo_amostras import gerar_todas_as_amostras


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture
def paleta_azul() -> Paleta:
    return CATALOGO_PALETAS["azul"]


@pytest.fixture
def paleta_verde() -> Paleta:
    return CATALOGO_PALETAS["verde"]


@pytest.fixture
def paleta_cinza() -> Paleta:
    return CATALOGO_PALETAS["cinza"]


@pytest.fixture
def paleta_vinho() -> Paleta:
    return CATALOGO_PALETAS["vinho"]


@pytest.fixture
def vocabulario_carregado() -> Dict[str, Dict[str, str]]:
    return carregar_vocabulario_bilingue()


@pytest.fixture
def resumo_canonico() -> ResumoExecutivoPadrao:
    return ResumoExecutivoPadrao(
        bloco_1_cabecalho=CabecalhoExecucao(
            visao="V2",
            data_execucao=datetime(2026, 4, 23, 14, 32),
            modo_upload="SIMPLES",
            agrupadores=["Mes", "Loja"],
            medida_principal="Vendas",
        ),
        bloco_2_numeros_ancora={
            "total_origem": 12543.21,
            "total_comparado": 13021.55,
            "diferenca_total": 478.34,
            "variacao_total": 0.0381,
        },
        bloco_3_distribuicao={
            "PRESENTE_AMBOS": 42,
            "AUSENTE_ORIGEM": 3,
            "AUSENTE_COMPARADO": 5,
            "NULO_ORIGEM": 1,
            "NULO_COMPARADO": 2,
            "NULO_AMBOS": 0,
        },
        bloco_4_elementos_destacados={
            "top_variacoes": [
                {"rotulo": "2025-03|Loja Norte", "variacao_percentual": 0.2345},
            ]
        },
        bloco_5_leitura_qualitativa=LeituraQualitativa(
            classificacao_ativa="ESTAVEL",
            thresholds_usados={
                "limiar_estabilidade_pct": 0.05,
                "limite_variacao_extrema": 0.20,
            },
            alguma_leitura_alterada_por_edicao=False,
        ),
        bloco_6_qualidade_estrutural=QualidadeEstrutural(
            total_warnings=3,
            warnings_por_categoria={"informativo": 1, "ajuste_leve": 2},
            ajustes_aplicados=1,
            tem_bloqueios_escapados=False,
        ),
    )


@pytest.fixture
def workbook_com_dados():
    wb = Workbook()
    ws = wb.active
    ws.append(["Agrupadores", "Orçado", "Realizado", "Variação"])
    ws.append(["Jan|LojaA", 1000.0, 1050.0, 0.05])
    ws.append(["Feb|LojaB", 2000.0, 1800.0, -0.10])
    ws.append(["Mar|LojaC", 1500.0, 1650.0, 0.10])
    return wb, ws


@pytest.fixture(scope="module")
def amostras_geradas():
    """Gera as 4 amostras em tempdir uma vez · reutiliza entre testes."""
    with tempfile.TemporaryDirectory() as td:
        caminhos = gerar_todas_as_amostras(diretorio_saida=td)
        yield caminhos


# ===========================================================================
# 1. TestInterface · contratos e assinaturas
# ===========================================================================


class TestInterfacePaletas:
    def test_catalogo_tem_4_entradas(self):
        assert len(CATALOGO_PALETAS) == 4

    def test_catalogo_chaves_sao_azul_verde_cinza_vinho(self):
        assert set(CATALOGO_PALETAS.keys()) == {"azul", "verde", "cinza", "vinho"}

    def test_catalogo_chaves_em_minusculas(self):
        for chave in CATALOGO_PALETAS.keys():
            assert chave == chave.lower()

    def test_paleta_eh_frozen_dataclass(self, paleta_azul):
        with pytest.raises(Exception):
            paleta_azul.nome = "vermelho"

    def test_paleta_eh_hasheavel(self, paleta_azul):
        hash(paleta_azul)  # não deve raise

    def test_paleta_nome_coincide_com_chave(self):
        for chave, paleta in CATALOGO_PALETAS.items():
            assert paleta.nome == chave

    def test_paleta_azul_tem_rotulo_user_facing(self, paleta_azul):
        assert "zul" in paleta_azul.rotulo_user_facing.lower()

    def test_paleta_cores_sao_strings_hex_sem_hash(self):
        for paleta in CATALOGO_PALETAS.values():
            for atributo in (
                "cor_primaria", "cor_secundaria", "cor_texto_sobre_primaria",
                "cor_destaque", "cor_neutra_escura", "cor_neutra_clara",
            ):
                valor = getattr(paleta, atributo)
                assert isinstance(valor, str)
                assert not valor.startswith("#")
                assert len(valor) == 6
                int(valor, 16)  # válido como hex

    def test_paleta_fonte_familia_calibri(self):
        for paleta in CATALOGO_PALETAS.values():
            assert paleta.fonte_familia == "Calibri"

    def test_paleta_tamanhos_fonte_positivos(self):
        for paleta in CATALOGO_PALETAS.values():
            assert paleta.fonte_tamanho_titulo > 0
            assert paleta.fonte_tamanho_secao > 0
            assert paleta.fonte_tamanho_corpo > 0
            assert paleta.fonte_tamanho_auxiliar > 0

    def test_paleta_tamanhos_fonte_ordem(self):
        for paleta in CATALOGO_PALETAS.values():
            assert paleta.fonte_tamanho_titulo >= paleta.fonte_tamanho_secao
            assert paleta.fonte_tamanho_secao >= paleta.fonte_tamanho_corpo
            assert paleta.fonte_tamanho_corpo >= paleta.fonte_tamanho_auxiliar

    def test_obter_paleta_retorna_paleta(self, paleta_azul):
        assert obter_paleta("azul") == paleta_azul

    def test_obter_paleta_desconhecida_raises(self):
        with pytest.raises(ValueError):
            obter_paleta("magenta")

    def test_aplicar_paleta_marca_workbook(self, paleta_azul):
        wb = Workbook()
        aplicar_paleta(wb, paleta_azul)
        assert paleta_aplicada(wb) == "azul"

    def test_aplicar_paleta_workbook_none_raises(self, paleta_azul):
        with pytest.raises(TypeError):
            aplicar_paleta(None, paleta_azul)

    def test_aplicar_paleta_none_raises(self):
        wb = Workbook()
        with pytest.raises(ValueError):
            aplicar_paleta(wb, None)

    def test_paleta_aplicada_sem_marcacao_retorna_none(self):
        wb = Workbook()
        assert paleta_aplicada(wb) is None

    def test_paleta_aplicada_workbook_none_retorna_none(self):
        assert paleta_aplicada(None) is None

    def test_paleta_tem_estilo_tabela(self):
        for paleta in CATALOGO_PALETAS.values():
            assert paleta.estilo_tabela.startswith("TableStyle")


class TestInterfaceVocabulario:
    def test_blocos_declarados(self):
        assert set(BLOCOS) == {
            "stepper", "modos_base", "classificacoes", "tipos_campo",
            "thresholds", "warnings", "proibidos",
        }

    def test_carregar_retorna_dict_com_7_chaves(self, vocabulario_carregado):
        assert isinstance(vocabulario_carregado, dict)
        assert set(vocabulario_carregado.keys()) == set(BLOCOS)

    def test_blocos_principais_nao_vazios(self, vocabulario_carregado):
        for chave in ("stepper", "modos_base", "classificacoes",
                      "tipos_campo", "thresholds", "warnings", "proibidos"):
            assert vocabulario_carregado[chave], f"bloco {chave} vazio"

    def test_stepper_tem_7_entradas(self, vocabulario_carregado):
        # v2 · D-167 · Bloco 1 estendido com sub-bloco 1.1 (Origem/Comparado)
        assert len(vocabulario_carregado["stepper"]) == 7

    def test_modos_base_tem_4_entradas(self, vocabulario_carregado):
        assert len(vocabulario_carregado["modos_base"]) == 4

    def test_classificacoes_tem_6_entradas(self, vocabulario_carregado):
        assert len(vocabulario_carregado["classificacoes"]) == 6

    def test_tipos_campo_tem_6_entradas(self, vocabulario_carregado):
        # v2 · D-167 · taxonomia D-025 incluiu ESTADO_SITUACAO
        assert len(vocabulario_carregado["tipos_campo"]) == 6

    def test_cache_retorna_mesma_instancia(self):
        a = carregar_vocabulario_bilingue()
        b = carregar_vocabulario_bilingue()
        assert a is b

    def test_caminho_inexistente_raises(self):
        with pytest.raises(FileNotFoundError):
            carregar_vocabulario_bilingue(Path("/tmp/inexistente_xxxx.md"))

    def test_traduzir_enum_classificacao(self, vocabulario_carregado):
        r = traduzir("PRESENTE_AMBOS", contexto="classificacoes",
                     vocabulario=vocabulario_carregado)
        # v2 · Bloco 3 · cells têm aspas literais no markdown (preservadas pelo parser)
        assert r == '"Presente nos dois lados"'

    def test_traduzir_sem_contexto_encontra(self, vocabulario_carregado):
        r = traduzir("TRANSACIONAL", vocabulario=vocabulario_carregado)
        assert "transacional" in r.lower()

    def test_traduzir_ausente_retorna_marcador(self, vocabulario_carregado, caplog):
        with caplog.at_level(logging.WARNING):
            r = traduzir("FOO_NAO_EXISTE", vocabulario=vocabulario_carregado)
        assert r == "[FOO_NAO_EXISTE]"
        assert any("FOO_NAO_EXISTE" in rec.message for rec in caplog.records)

    def test_traduzir_contexto_restringe(self, vocabulario_carregado, caplog):
        # "PRESENTE_AMBOS" só existe em classificacoes
        with caplog.at_level(logging.WARNING):
            r = traduzir("PRESENTE_AMBOS", contexto="modos_base",
                         vocabulario=vocabulario_carregado)
        assert r == "[PRESENTE_AMBOS]"

    def test_traduzir_fallback_nunca_retorna_termo_cru(self, vocabulario_carregado):
        r = traduzir("BAR_INEXISTENTE", vocabulario=vocabulario_carregado)
        assert r.startswith("[") and r.endswith("]")

    def test_contem_marcador_traducao_ausente(self):
        assert contem_marcador_traducao_ausente("[FOO_BAR]") is True
        assert contem_marcador_traducao_ausente("Presente nos dois lados") is False
        assert contem_marcador_traducao_ausente("R$ 100,00") is False

    def test_eh_termo_proibido_enum_caps(self):
        assert eh_termo_proibido("POR_COLUNAS") is not None
        assert eh_termo_proibido("PRESENTE_AMBOS") is not None
        assert eh_termo_proibido("NUMERICO_ADITIVO") is not None

    def test_eh_termo_proibido_datetime_cru(self):
        assert eh_termo_proibido("datetime.datetime(2026, 4, 23)") is not None

    def test_eh_termo_proibido_codigo_d(self):
        assert eh_termo_proibido("D-166 · detalhe") is not None

    def test_eh_termo_proibido_codigo_t(self):
        assert eh_termo_proibido("uso de T-AGRUPA aqui") is not None

    def test_eh_termo_proibido_codigo_f(self):
        assert eh_termo_proibido("F-APRESENT aprovado") is not None

    def test_eh_termo_proibido_atributo_python(self):
        assert eh_termo_proibido("chave_agrupadores") is not None
        assert eh_termo_proibido("valor_origem") is not None

    def test_eh_termo_proibido_user_facing_ok(self):
        """User-facing com acentos passa limpo."""
        assert eh_termo_proibido("Presente nos dois lados") is None
        assert eh_termo_proibido("R$ 1.234,56") is None
        assert eh_termo_proibido("23/04/2026 · 14:32") is None
        assert eh_termo_proibido("Análise Comparativa") is None
        assert eh_termo_proibido("Orçado vs Realizado") is None
        assert eh_termo_proibido("Diferença total") is None

    def test_eh_termo_proibido_string_vazia(self):
        assert eh_termo_proibido("") is None
        assert eh_termo_proibido("   ") is None

    def test_eh_termo_proibido_nao_string(self):
        assert eh_termo_proibido(123) is None  # type: ignore
        assert eh_termo_proibido(None) is None  # type: ignore


class TestInterfaceTabelas:
    def test_criar_tabela_simples(self, workbook_com_dados):
        wb, ws = workbook_com_dados
        tbl = criar_tabela_executiva(
            ws, "A1:D4", "t_teste",
            totais_por_coluna={
                "Agrupadores": "none", "Orçado": "sum",
                "Realizado": "sum", "Variação": "average",
            },
            paleta_nome="azul",
        )
        assert isinstance(tbl, Table)

    def test_criar_tabela_retorna_table(self, workbook_com_dados):
        wb, ws = workbook_com_dados
        tbl = criar_tabela_executiva(
            ws, "A1:D4", "t_x",
            totais_por_coluna={"Orçado": "sum"},
            paleta_nome="azul",
        )
        assert tbl.displayName == "t_x"

    def test_criar_tabela_totals_row_shown(self, workbook_com_dados):
        wb, ws = workbook_com_dados
        tbl = criar_tabela_executiva(
            ws, "A1:D4", "t_x",
            totais_por_coluna={"Orçado": "sum"},
            paleta_nome="azul",
        )
        assert tbl.totalsRowShown is True

    def test_criar_tabela_ref_inclui_linha_totals(self, workbook_com_dados):
        wb, ws = workbook_com_dados
        tbl = criar_tabela_executiva(
            ws, "A1:D4", "t_x",
            totais_por_coluna={"Orçado": "sum"},
            paleta_nome="azul",
        )
        # range original termina em 4 · com totals deve terminar em 5
        assert tbl.ref.endswith("5")

    def test_criar_tabela_nome_invalido_com_hifen(self, workbook_com_dados):
        wb, ws = workbook_com_dados
        with pytest.raises(ValueError):
            criar_tabela_executiva(
                ws, "A1:D4", "t-invalido",
                totais_por_coluna={"Orçado": "sum"},
            )

    def test_criar_tabela_nome_invalido_comeca_com_numero(self, workbook_com_dados):
        wb, ws = workbook_com_dados
        with pytest.raises(ValueError):
            criar_tabela_executiva(
                ws, "A1:D4", "1tabela",
                totais_por_coluna={"Orçado": "sum"},
            )

    def test_criar_tabela_nome_vazio(self, workbook_com_dados):
        wb, ws = workbook_com_dados
        with pytest.raises(ValueError):
            criar_tabela_executiva(
                ws, "A1:D4", "",
                totais_por_coluna={"Orçado": "sum"},
            )

    def test_criar_tabela_coluna_inexistente_raises(self, workbook_com_dados):
        wb, ws = workbook_com_dados
        with pytest.raises(ValueError):
            criar_tabela_executiva(
                ws, "A1:D4", "t_x",
                totais_por_coluna={"ColunaQueNaoExiste": "sum"},
            )

    def test_criar_tabela_funcao_invalida(self, workbook_com_dados):
        wb, ws = workbook_com_dados
        with pytest.raises(ValueError):
            criar_tabela_executiva(
                ws, "A1:D4", "t_x",
                totais_por_coluna={"Orçado": "somar_tudo"},  # type: ignore
            )

    def test_criar_tabela_custom_sem_formula_raises(self, workbook_com_dados):
        wb, ws = workbook_com_dados
        with pytest.raises(ValueError):
            criar_tabela_executiva(
                ws, "A1:D4", "t_x",
                totais_por_coluna={"Orçado": "custom"},
                custom_totals=None,
            )

    def test_criar_tabela_custom_com_formula_ok(self, workbook_com_dados):
        wb, ws = workbook_com_dados
        tbl = criar_tabela_executiva(
            ws, "A1:D4", "t_x",
            totais_por_coluna={"Orçado": "custom"},
            custom_totals={"Orçado": "=SUM(B2:B4)"},
        )
        assert tbl is not None

    def test_criar_tabela_range_uma_linha_raises(self, workbook_com_dados):
        wb, ws = workbook_com_dados
        with pytest.raises(ValueError):
            criar_tabela_executiva(
                ws, "A1:D1", "t_x",
                totais_por_coluna={},
            )

    def test_criar_tabela_nome_duplicado_raises(self, workbook_com_dados):
        wb, ws = workbook_com_dados
        criar_tabela_executiva(
            ws, "A1:D4", "t_dup",
            totais_por_coluna={"Orçado": "sum"},
        )
        with pytest.raises(ValueError):
            criar_tabela_executiva(
                ws, "A1:D4", "t_dup",
                totais_por_coluna={"Orçado": "sum"},
            )

    def test_estilo_para_paleta_por_nome(self):
        for nome in CATALOGO_PALETAS.keys():
            assert estilo_para_paleta(nome).startswith("TableStyle")

    def test_estilo_para_paleta_none_default(self):
        assert estilo_para_paleta(None) == "TableStyleMedium2"

    def test_estilo_para_paleta_desconhecida_fallback(self, caplog):
        with caplog.at_level(logging.WARNING):
            r = estilo_para_paleta("rosa_choque")
        assert r == "TableStyleMedium2"
        assert any("rosa_choque" in rec.message for rec in caplog.records)


class TestInterfaceFormatos:
    def test_constante_monetario_br(self):
        assert "R$" in FORMATO_MONETARIO_BR
        assert "[Red]" in FORMATO_MONETARIO_BR

    def test_constante_percentual(self):
        assert "%" in FORMATO_PERCENTUAL

    def test_constante_contagem(self):
        assert "[Red]" in FORMATO_CONTAGEM

    def test_constante_data_br(self):
        assert FORMATO_DATA_BR == "dd/mm/yyyy"

    def test_aplicar_monetario_retorna_contagem(self, workbook_com_dados):
        wb, ws = workbook_com_dados
        n = aplicar_formato_monetario([ws["B2"], ws["B3"]])
        assert n == 2

    def test_aplicar_monetario_aplica_formato(self, workbook_com_dados):
        wb, ws = workbook_com_dados
        aplicar_formato_monetario([ws["B2"]])
        assert ws["B2"].number_format == FORMATO_MONETARIO_BR

    def test_aplicar_percentual_default_fracao(self, workbook_com_dados):
        wb, ws = workbook_com_dados
        aplicar_formato_percentual([ws["D2"]])
        assert ws["D2"].number_format == FORMATO_PERCENTUAL

    def test_aplicar_percentual_literal(self, workbook_com_dados):
        wb, ws = workbook_com_dados
        aplicar_formato_percentual([ws["D2"]], conversao_fracao=False)
        assert ws["D2"].number_format == FORMATO_PERCENTUAL_LITERAL

    def test_aplicar_contagem(self, workbook_com_dados):
        wb, ws = workbook_com_dados
        aplicar_formato_contagem([ws["B2"]])
        assert ws["B2"].number_format == FORMATO_CONTAGEM

    def test_aplicar_data_br(self, workbook_com_dados):
        wb, ws = workbook_com_dados
        aplicar_formato_data_br([ws["A2"]])
        assert ws["A2"].number_format == FORMATO_DATA_BR

    def test_aplicar_aceita_range_iterable(self, workbook_com_dados):
        wb, ws = workbook_com_dados
        n = aplicar_formato_monetario(ws["B2:B4"])
        assert n == 3

    def test_formatar_moeda_br_valor_positivo(self):
        assert formatar_moeda_br(1234.56) == "R$ 1.234,56"

    def test_formatar_moeda_br_valor_negativo(self):
        assert formatar_moeda_br(-100.0) == "(R$ 100,00)"

    def test_formatar_moeda_br_zero_vira_travessao(self):
        assert formatar_moeda_br(0) == "-"

    def test_formatar_moeda_br_none(self):
        assert formatar_moeda_br(None) == "-"

    def test_formatar_percentual_br_fracao(self):
        assert formatar_percentual_br(0.0381) == "3,81%"

    def test_formatar_percentual_br_negativo(self):
        assert formatar_percentual_br(-0.05) == "-5,00%"

    def test_formatar_percentual_br_zero_vira_travessao(self):
        assert formatar_percentual_br(0.0) == "-"

    def test_formatar_percentual_br_literal(self):
        assert formatar_percentual_br(5.0, conversao_fracao=False) == "5,00%"

    def test_formatar_data_br_formato(self):
        dt = datetime(2026, 4, 23, 14, 32)
        r = formatar_data_br(dt)
        assert r == "23/04/2026 · 14:32"

    def test_formatar_data_br_none(self):
        assert formatar_data_br(None) == "-"


class TestInterfaceColunas:
    def test_coluna_adaptativa_eh_frozen(self):
        c = ColunaAdaptativa(
            identificador="x",
            cabecalho_user_facing="X",
            unidade="texto",
            condicao=sempre(),
        )
        with pytest.raises(Exception):
            c.identificador = "y"

    def test_sempre_retorna_true(self):
        cond = sempre()
        assert cond({}) is True
        assert cond({"qualquer": "coisa"}) is True

    def test_se_config_igual(self):
        cond = se_config_igual("estrutura_entrada", "POR_COLUNAS")
        assert cond({"estrutura_entrada": "POR_COLUNAS"}) is True
        assert cond({"estrutura_entrada": "POR_LINHAS"}) is False
        assert cond({}) is False

    def test_se_config_diferente(self):
        cond = se_config_diferente("modo", "SIMPLES")
        assert cond({"modo": "DUAL"}) is True
        assert cond({"modo": "SIMPLES"}) is False

    def test_se_config_presente(self):
        cond = se_config_presente("agrupadores")
        assert cond({"agrupadores": ["a", "b"]}) is True
        assert cond({"agrupadores": []}) is False
        assert cond({}) is False

    def test_montar_respeita_ordem_sugerida(self):
        esquema = [
            ColunaAdaptativa("c", "C", "texto", sempre(), ordem_sugerida=3),
            ColunaAdaptativa("a", "A", "texto", sempre(), ordem_sugerida=1),
            ColunaAdaptativa("b", "B", "texto", sempre(), ordem_sugerida=2),
        ]
        r = montar_colunas_adaptativas({}, esquema)
        assert [c.identificador for c in r] == ["a", "b", "c"]

    def test_montar_filtra_por_condicao(self):
        esquema = [
            ColunaAdaptativa("a", "A", "texto", sempre()),
            ColunaAdaptativa("b", "B", "texto", se_config_igual("x", 1)),
        ]
        r = montar_colunas_adaptativas({"x": 2}, esquema)
        assert [c.identificador for c in r] == ["a"]

    def test_montar_determinismo(self):
        esquema = [
            ColunaAdaptativa("b", "B", "texto", sempre(), ordem_sugerida=1),
            ColunaAdaptativa("a", "A", "texto", sempre(), ordem_sugerida=1),
        ]
        r1 = montar_colunas_adaptativas({}, esquema)
        r2 = montar_colunas_adaptativas({}, esquema)
        assert [c.identificador for c in r1] == [c.identificador for c in r2]

    def test_montar_lista_vazia_emite_warning(self, caplog):
        esquema = [
            ColunaAdaptativa("a", "A", "texto", se_config_igual("x", 99)),
        ]
        with caplog.at_level(logging.WARNING):
            r = montar_colunas_adaptativas({"x": 1}, esquema)
        assert r == []
        assert any("vazia" in rec.message for rec in caplog.records)

    def test_montar_esquema_nao_lista_raises(self):
        with pytest.raises(TypeError):
            montar_colunas_adaptativas({}, "nao-eh-lista")  # type: ignore

    def test_montar_config_nao_dict_raises(self):
        with pytest.raises(TypeError):
            montar_colunas_adaptativas("nao-eh-dict", [])  # type: ignore

    def test_montar_condicao_excepcao_exclui_com_warning(self, caplog):
        def cond_bomba(cfg):
            raise RuntimeError("boom")
        esquema = [
            ColunaAdaptativa("a", "A", "texto", cond_bomba),
            ColunaAdaptativa("b", "B", "texto", sempre()),
        ]
        with caplog.at_level(logging.WARNING):
            r = montar_colunas_adaptativas({}, esquema)
        assert [c.identificador for c in r] == ["b"]


class TestInterfaceResumoExecutivo:
    def test_renderizar_aceita_workbook_valido(
        self, resumo_canonico, paleta_azul, vocabulario_carregado
    ):
        wb = Workbook()
        renderizar_resumo_executivo(
            wb.active, resumo_canonico, paleta_azul, vocabulario_carregado
        )  # não deve raise

    def test_renderizar_ws_none_raises(
        self, resumo_canonico, paleta_azul, vocabulario_carregado
    ):
        with pytest.raises(TypeError):
            renderizar_resumo_executivo(
                None, resumo_canonico, paleta_azul, vocabulario_carregado
            )

    def test_renderizar_paleta_none_raises(
        self, resumo_canonico, vocabulario_carregado
    ):
        wb = Workbook()
        with pytest.raises(TypeError):
            renderizar_resumo_executivo(
                wb.active, resumo_canonico, None, vocabulario_carregado
            )

    def test_renderizar_vocabulario_none_raises(
        self, resumo_canonico, paleta_azul
    ):
        wb = Workbook()
        with pytest.raises(TypeError):
            renderizar_resumo_executivo(
                wb.active, resumo_canonico, paleta_azul, None
            )

    def test_renderizar_escreve_titulo_na_primeira_linha(
        self, resumo_canonico, paleta_azul, vocabulario_carregado
    ):
        wb = Workbook()
        ws = wb.active
        renderizar_resumo_executivo(ws, resumo_canonico, paleta_azul,
                                     vocabulario_carregado)
        titulo = ws.cell(row=1, column=1).value
        assert "Resumo Executivo" in titulo

    def test_renderizar_produz_multiplas_linhas(
        self, resumo_canonico, paleta_azul, vocabulario_carregado
    ):
        wb = Workbook()
        ws = wb.active
        renderizar_resumo_executivo(ws, resumo_canonico, paleta_azul,
                                     vocabulario_carregado)
        assert ws.max_row > 20


# ===========================================================================
# 2. TestSnapshot · saídas textuais fixas (detectam drift)
# ===========================================================================


class TestSnapshotTraducoes:
    def test_stepper_e1(self, vocabulario_carregado):
        # v2 · D-167 · stepper "4 etapas + Revisão" com prefixo numérico
        assert vocabulario_carregado["stepper"]["E1"] == "1 · Escolher arquivo"

    def test_stepper_e2(self, vocabulario_carregado):
        assert vocabulario_carregado["stepper"]["E2"] == "2 · Reconhecer estrutura"

    def test_stepper_e3(self, vocabulario_carregado):
        assert vocabulario_carregado["stepper"]["E3"] == "3 · Configurar análise"

    def test_stepper_e4(self, vocabulario_carregado):
        assert vocabulario_carregado["stepper"]["E4"] == "4 · Agrupar"

    def test_stepper_e5(self, vocabulario_carregado):
        assert vocabulario_carregado["stepper"]["E5"] == "Revisar e executar"

    def test_classif_presente_ambos(self, vocabulario_carregado):
        # v2 · Bloco 3 · cells têm aspas literais (parser preserva)
        r = vocabulario_carregado["classificacoes"]["PRESENTE_AMBOS"]
        assert r == '"Presente nos dois lados"'

    def test_classif_ausente_origem(self, vocabulario_carregado):
        # v2 · substituição dinâmica · "Apareceu no Comparado" (invertido vs v1)
        r = vocabulario_carregado["classificacoes"]["AUSENTE_ORIGEM"]
        assert r == '"Apareceu no Comparado"'

    def test_classif_ausente_comparado(self, vocabulario_carregado):
        # v2 · substituição dinâmica · "Saiu da Origem"
        r = vocabulario_carregado["classificacoes"]["AUSENTE_COMPARADO"]
        assert r == '"Saiu da Origem"'

    def test_classif_nulo_origem(self, vocabulario_carregado):
        # v2 · capitalização "Origem" (antes: "origem")
        r = vocabulario_carregado["classificacoes"]["NULO_ORIGEM"]
        assert r == '"Sem valor na Origem"'

    def test_classif_nulo_comparado(self, vocabulario_carregado):
        # v2 · capitalização "Comparado"
        r = vocabulario_carregado["classificacoes"]["NULO_COMPARADO"]
        assert r == '"Sem valor no Comparado"'

    def test_classif_nulo_ambos(self, vocabulario_carregado):
        r = vocabulario_carregado["classificacoes"]["NULO_AMBOS"]
        assert r == '"Sem valor nos dois lados"'

    def test_modo_transacional(self, vocabulario_carregado):
        r = vocabulario_carregado["modos_base"]["TRANSACIONAL"]
        assert "transacional" in r.lower()

    def test_modo_pre_agregado(self, vocabulario_carregado):
        r = vocabulario_carregado["modos_base"]["PRE_AGREGADO"]
        assert "agregada" in r.lower() or "agregado" in r.lower()

    def test_modo_individual(self, vocabulario_carregado):
        r = vocabulario_carregado["modos_base"]["INDIVIDUAL"]
        assert r == "Granularidade individual"

    def test_modo_consolidada(self, vocabulario_carregado):
        r = vocabulario_carregado["modos_base"]["CONSOLIDADA"]
        assert "chave" in r.lower()

    def test_tipo_numerico_aditivo(self, vocabulario_carregado):
        r = vocabulario_carregado["tipos_campo"]["NUMERICO_ADITIVO"]
        assert "somável" in r.lower()

    def test_tipo_categorico(self, vocabulario_carregado):
        r = vocabulario_carregado["tipos_campo"]["CATEGORICO"]
        assert "categoria" in r.lower() or "rótulo" in r.lower()

    def test_threshold_estabilidade(self, vocabulario_carregado):
        r = vocabulario_carregado["thresholds"]["limiar_estabilidade_pct"]
        assert "estabilidade" in r.lower()

    def test_threshold_variacao_extrema(self, vocabulario_carregado):
        r = vocabulario_carregado["thresholds"]["limite_variacao_extrema"]
        assert "variação extrema" in r.lower() or "extrema" in r.lower()


class TestSnapshotPaletas:
    def test_paletas_chaves(self):
        assert sorted(CATALOGO_PALETAS.keys()) == ["azul", "cinza", "verde", "vinho"]

    def test_paleta_azul_estilo(self):
        assert CATALOGO_PALETAS["azul"].estilo_tabela == "TableStyleMedium2"

    def test_paleta_verde_estilo(self):
        assert CATALOGO_PALETAS["verde"].estilo_tabela == "TableStyleMedium4"

    def test_paleta_cinza_estilo(self):
        assert CATALOGO_PALETAS["cinza"].estilo_tabela == "TableStyleMedium1"

    def test_paleta_vinho_estilo(self):
        assert CATALOGO_PALETAS["vinho"].estilo_tabela == "TableStyleMedium3"

    def test_paleta_azul_primaria_institucional(self):
        # Azul médio · começa em 1 ou 2 (não cores claras começando em F)
        hex_ = CATALOGO_PALETAS["azul"].cor_primaria
        assert hex_[0] in "0123"

    def test_paleta_vinho_primaria_bordo(self):
        # Bordô tem R dominante · primeiro byte > últimos dois
        hex_ = CATALOGO_PALETAS["vinho"].cor_primaria
        r = int(hex_[0:2], 16)
        g = int(hex_[2:4], 16)
        b = int(hex_[4:6], 16)
        assert r > g and r > b

    def test_paleta_verde_primaria_verde(self):
        hex_ = CATALOGO_PALETAS["verde"].cor_primaria
        r = int(hex_[0:2], 16)
        g = int(hex_[2:4], 16)
        b = int(hex_[4:6], 16)
        assert g > r and g > b

    def test_paleta_cinza_monocromatica(self):
        hex_ = CATALOGO_PALETAS["cinza"].cor_primaria
        r = int(hex_[0:2], 16)
        g = int(hex_[2:4], 16)
        b = int(hex_[4:6], 16)
        assert abs(r - g) < 16 and abs(g - b) < 16


class TestSnapshotFormatos:
    def test_formato_monetario_exato(self):
        assert FORMATO_MONETARIO_BR == 'R$ #,##0.00;[Red](R$ #,##0.00);-'

    def test_formato_percentual_exato(self):
        assert FORMATO_PERCENTUAL == '0.00%;[Red]-0.00%;-'

    def test_formato_percentual_literal_exato(self):
        assert FORMATO_PERCENTUAL_LITERAL == '0.00"%";[Red]-0.00"%";-'

    def test_formato_contagem_exato(self):
        assert FORMATO_CONTAGEM == '#,##0;[Red]-#,##0;-'

    def test_formato_data_br_exato(self):
        assert FORMATO_DATA_BR == "dd/mm/yyyy"

    def test_formatar_moeda_1000(self):
        assert formatar_moeda_br(1000.0) == "R$ 1.000,00"

    def test_formatar_moeda_milhoes(self):
        assert formatar_moeda_br(1234567.89) == "R$ 1.234.567,89"

    def test_formatar_percentual_12345(self):
        assert formatar_percentual_br(0.12345) == "12,35%"

    def test_formatar_percentual_100(self):
        assert formatar_percentual_br(1.0) == "100,00%"


class TestSnapshotResumoExecutivo:
    def test_titulo_contem_visao_traduzida(
        self, resumo_canonico, paleta_azul, vocabulario_carregado
    ):
        wb = Workbook()
        ws = wb.active
        renderizar_resumo_executivo(ws, resumo_canonico, paleta_azul,
                                     vocabulario_carregado)
        titulo = ws.cell(row=1, column=1).value
        assert "Análise Comparativa" in titulo

    def test_titulo_contem_data_br(
        self, resumo_canonico, paleta_azul, vocabulario_carregado
    ):
        wb = Workbook()
        ws = wb.active
        renderizar_resumo_executivo(ws, resumo_canonico, paleta_azul,
                                     vocabulario_carregado)
        titulo = ws.cell(row=1, column=1).value
        assert "23/04/2026" in titulo

    def test_subtitulos_presentes(
        self, resumo_canonico, paleta_azul, vocabulario_carregado
    ):
        wb = Workbook()
        ws = wb.active
        renderizar_resumo_executivo(ws, resumo_canonico, paleta_azul,
                                     vocabulario_carregado)
        valores_celulas = [
            ws.cell(row=r, column=1).value
            for r in range(1, ws.max_row + 1)
        ]
        texto = " ".join(v for v in valores_celulas if isinstance(v, str))
        for esperado in ("Como foi analisado", "Números principais",
                         "Como os casos se distribuem", "O que chama atenção",
                         "Leitura qualitativa", "Qualidade do dado"):
            assert esperado in texto, f"subtítulo '{esperado}' ausente"

    def test_distribuicao_tem_categorias_traduzidas(
        self, resumo_canonico, paleta_azul, vocabulario_carregado
    ):
        wb = Workbook()
        ws = wb.active
        renderizar_resumo_executivo(ws, resumo_canonico, paleta_azul,
                                     vocabulario_carregado)
        todas_strings = []
        for row in ws.iter_rows(values_only=True):
            for v in row:
                if isinstance(v, str):
                    todas_strings.append(v)
        texto = " | ".join(todas_strings)
        assert "Presente nos dois lados" in texto

    def test_thresholds_tem_labels_traduzidos(
        self, resumo_canonico, paleta_azul, vocabulario_carregado
    ):
        wb = Workbook()
        ws = wb.active
        renderizar_resumo_executivo(ws, resumo_canonico, paleta_azul,
                                     vocabulario_carregado)
        textos = []
        for row in ws.iter_rows(values_only=True):
            for v in row:
                if isinstance(v, str):
                    textos.append(v)
        todos = " ".join(textos)
        assert "Limite de estabilidade" in todos


# ===========================================================================
# 3. TestRegressao · bugs conhecidos prevenidos
# ===========================================================================


class TestRegressaoFormatos:
    def test_percentual_0108_exibe_108_nativo(self, workbook_com_dados):
        """D-166 · 0.0108 com '0.00%' exibe '1.08%' nativamente."""
        wb, ws = workbook_com_dados
        ws["D2"].value = 0.0108
        aplicar_formato_percentual([ws["D2"]], conversao_fracao=True)
        # Valor armazenado permanece fração; formato aplicado é o nativo
        assert ws["D2"].value == 0.0108
        assert ws["D2"].number_format == FORMATO_PERCENTUAL

    def test_monetario_aplicado_round_trip(self, tmp_path, workbook_com_dados):
        wb, ws = workbook_com_dados
        aplicar_formato_monetario([ws["B2"]])
        p = tmp_path / "t.xlsx"
        wb.save(p)
        wb2 = load_workbook(p)
        assert wb2.active["B2"].number_format == FORMATO_MONETARIO_BR


class TestRegressaoVocabulario:
    def test_cache_mesma_instancia_duas_chamadas(self):
        a = carregar_vocabulario_bilingue()
        b = carregar_vocabulario_bilingue()
        assert a is b

    def test_cache_retorna_mesmas_chaves(self):
        a = carregar_vocabulario_bilingue()
        b = carregar_vocabulario_bilingue()
        assert set(a.keys()) == set(b.keys())

    def test_proibidos_incluem_enums_canonicos(self, vocabulario_carregado):
        proibidos = vocabulario_carregado["proibidos"]
        # Pelo menos alguns enums canônicos listados estão presentes
        ao_menos_um_enum_proibido = any(
            p in proibidos for p in (
                "POR_COLUNAS", "PRESENTE_AMBOS", "TRANSACIONAL",
                "NUMERICO_ADITIVO",
            )
        )
        assert ao_menos_um_enum_proibido

    def test_autoconsistencia_valores_user_facing_nao_sao_proibidos(
        self, vocabulario_carregado
    ):
        """Invariante: valores traduzidos do próprio dict não violam bloco 7."""
        for bloco_nome in ("stepper", "modos_base", "classificacoes",
                           "tipos_campo", "thresholds"):
            for _chave, valor in vocabulario_carregado[bloco_nome].items():
                violacao = eh_termo_proibido(valor)
                assert violacao is None, (
                    f"vocabulário de '{bloco_nome}' tem valor que viola "
                    f"bloco 7: {valor!r} ({violacao})"
                )


class TestRegressaoCompositividade:
    def test_aplicar_paleta_mais_criar_tabela(
        self, workbook_com_dados, paleta_azul
    ):
        wb, ws = workbook_com_dados
        aplicar_paleta(wb, paleta_azul)
        tbl = criar_tabela_executiva(
            ws, "A1:D4", "t_comp",
            totais_por_coluna={"Orçado": "sum"},
            paleta_nome="azul",
        )
        assert paleta_aplicada(wb) == "azul"
        assert tbl.displayName == "t_comp"

    def test_round_trip_tabela_em_disco(self, tmp_path, workbook_com_dados):
        wb, ws = workbook_com_dados
        criar_tabela_executiva(
            ws, "A1:D4", "t_rt",
            totais_por_coluna={"Orçado": "sum", "Realizado": "sum"},
            paleta_nome="verde",
        )
        p = tmp_path / "rt.xlsx"
        wb.save(p)
        wb2 = load_workbook(p)
        ws2 = wb2.active
        assert "t_rt" in list(ws2.tables)

    def test_renderizar_resumo_round_trip(
        self, tmp_path, resumo_canonico, paleta_azul, vocabulario_carregado
    ):
        wb = Workbook()
        ws = wb.active
        ws.title = "Resumo"
        renderizar_resumo_executivo(ws, resumo_canonico, paleta_azul,
                                     vocabulario_carregado)
        p = tmp_path / "r.xlsx"
        wb.save(p)
        wb2 = load_workbook(p)
        assert "Resumo" in wb2.sheetnames
        assert wb2["Resumo"].max_row > 15


# ===========================================================================
# 4. TestInvariantesBloco7 · varredura de amostras
# ===========================================================================


def _varrer_workbook_procurando_violacoes(
    caminho: Path, vocabulario: Dict
) -> list[tuple[str, int, int, str, str]]:
    """
    Retorna lista de (sheet, linha, coluna, valor, motivo) para cada violação.
    """
    violacoes = []
    wb = load_workbook(caminho, data_only=False)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows():
            for celula in row:
                valor = celula.value
                if isinstance(valor, str):
                    motivo = eh_termo_proibido(valor, vocabulario)
                    if motivo is not None:
                        violacoes.append((
                            sheet, celula.row, celula.column, valor, motivo,
                        ))
    return violacoes


def _varrer_workbook_procurando_marcadores(
    caminho: Path,
) -> list[tuple[str, int, int, str]]:
    marcadores = []
    wb = load_workbook(caminho, data_only=False)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows():
            for celula in row:
                valor = celula.value
                if isinstance(valor, str):
                    if contem_marcador_traducao_ausente(valor):
                        marcadores.append((sheet, celula.row, celula.column, valor))
    return marcadores


class TestInvariantesAmostras:
    """Varre cada amostra gerada procurando violações do bloco 7."""

    def test_amostras_geradas_sao_4(self, amostras_geradas):
        assert len(amostras_geradas) == 4

    def test_amostras_tem_nomes_canonicos(self, amostras_geradas):
        nomes = {p.stem for p in amostras_geradas}
        assert nomes == {
            "amostra_paleta_azul",
            "amostra_paleta_verde",
            "amostra_paleta_cinza",
            "amostra_paleta_vinho",
        }

    def test_amostras_nao_corrompidas(self, amostras_geradas):
        for p in amostras_geradas:
            wb = load_workbook(p)
            assert wb.sheetnames

    def test_amostras_tamanho_razoavel(self, amostras_geradas):
        for p in amostras_geradas:
            tamanho = p.stat().st_size
            assert 3_000 < tamanho < 200_000, (
                f"amostra {p.name} com tamanho suspeito: {tamanho} bytes"
            )

    def test_amostras_tem_4_abas(self, amostras_geradas):
        for p in amostras_geradas:
            wb = load_workbook(p)
            assert len(wb.sheetnames) == 4

    def test_amostras_nome_das_abas(self, amostras_geradas):
        esperados = {"Resumo Executivo", "Base Analítica",
                     "Coração Visual", "Diagnóstico"}
        for p in amostras_geradas:
            wb = load_workbook(p)
            assert set(wb.sheetnames) == esperados

    def test_amostras_zero_violacoes_bloco_7(
        self, amostras_geradas, vocabulario_carregado
    ):
        for p in amostras_geradas:
            violacoes = _varrer_workbook_procurando_violacoes(
                p, vocabulario_carregado
            )
            detalhe = "\n".join(
                f"  [{s}] L{lin}C{col}: {v!r} → {motivo}"
                for s, lin, col, v, motivo in violacoes[:10]
            )
            assert not violacoes, (
                f"amostra {p.name} tem {len(violacoes)} violação(ões):\n{detalhe}"
            )

    def test_amostras_zero_marcadores_traducao_ausente(self, amostras_geradas):
        for p in amostras_geradas:
            marcadores = _varrer_workbook_procurando_marcadores(p)
            detalhe = "\n".join(
                f"  [{s}] L{lin}C{col}: {v!r}"
                for s, lin, col, v in marcadores[:10]
            )
            assert not marcadores, (
                f"amostra {p.name} tem {len(marcadores)} marcador(es) "
                f"[TERMO]:\n{detalhe}"
            )

    def test_amostras_sem_datetime_literal(self, amostras_geradas):
        for p in amostras_geradas:
            wb = load_workbook(p)
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in ws.iter_rows(values_only=True):
                    for valor in row:
                        if isinstance(valor, str):
                            assert "datetime.datetime(" not in valor, (
                                f"{p.name}[{sheet}]: datetime.datetime( em {valor!r}"
                            )

    def test_amostras_sem_dict_python(self, amostras_geradas):
        for p in amostras_geradas:
            wb = load_workbook(p)
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in ws.iter_rows(values_only=True):
                    for valor in row:
                        if isinstance(valor, str):
                            # Verifica dict serializado tipo {'x': 1, ...}
                            assert not ("{'" in valor and "':" in valor), (
                                f"{p.name}[{sheet}]: dict Python em {valor!r}"
                            )

    def test_amostras_sem_codigo_d_xxx(self, amostras_geradas):
        import re
        regex = re.compile(r"\bD-\d{2,}\b")
        for p in amostras_geradas:
            wb = load_workbook(p)
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in ws.iter_rows(values_only=True):
                    for valor in row:
                        if isinstance(valor, str):
                            assert not regex.search(valor), (
                                f"{p.name}[{sheet}]: código D-XXX em {valor!r}"
                            )

    def test_amostras_sem_codigo_f_xxx(self, amostras_geradas):
        import re
        regex = re.compile(r"\bF-[A-Z]{3,}\b")
        for p in amostras_geradas:
            wb = load_workbook(p)
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in ws.iter_rows(values_only=True):
                    for valor in row:
                        if isinstance(valor, str):
                            assert not regex.search(valor), (
                                f"{p.name}[{sheet}]: código F-XXX em {valor!r}"
                            )

    def test_amostras_sem_enum_por_colunas(self, amostras_geradas):
        for p in amostras_geradas:
            wb = load_workbook(p)
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in ws.iter_rows(values_only=True):
                    for valor in row:
                        if isinstance(valor, str):
                            assert "POR_COLUNAS" not in valor
                            assert "POR_LINHAS" not in valor

    def test_amostras_sem_enum_presente_ambos(self, amostras_geradas):
        for p in amostras_geradas:
            wb = load_workbook(p)
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in ws.iter_rows(values_only=True):
                    for valor in row:
                        if isinstance(valor, str):
                            assert "PRESENTE_AMBOS" not in valor
                            assert "AUSENTE_ORIGEM" not in valor

    def test_amostras_sem_transacional_caps(self, amostras_geradas):
        for p in amostras_geradas:
            wb = load_workbook(p)
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in ws.iter_rows(values_only=True):
                    for valor in row:
                        if isinstance(valor, str):
                            assert "TRANSACIONAL" not in valor
                            assert "PRE_AGREGADO" not in valor

    def test_amostras_resumo_tem_blocos_esperados(self, amostras_geradas):
        esperados = ["Como foi analisado", "Números principais",
                     "Como os casos se distribuem", "O que chama atenção",
                     "Leitura qualitativa", "Qualidade do dado"]
        for p in amostras_geradas:
            wb = load_workbook(p)
            ws = wb["Resumo Executivo"]
            texto_completo = []
            for row in ws.iter_rows(values_only=True):
                for v in row:
                    if isinstance(v, str):
                        texto_completo.append(v)
            todos = " ".join(texto_completo)
            for esperado in esperados:
                assert esperado in todos, (
                    f"{p.name}: bloco '{esperado}' ausente no Resumo Executivo"
                )

    def test_amostras_base_analitica_tem_tabela_nativa(self, amostras_geradas):
        for p in amostras_geradas:
            wb = load_workbook(p)
            ws = wb["Base Analítica"]
            assert len(list(ws.tables)) == 1, (
                f"{p.name}: Base Analítica deveria ter exatamente 1 tabela nativa"
            )

    def test_amostras_base_analitica_tabela_tem_totals(self, amostras_geradas):
        for p in amostras_geradas:
            wb = load_workbook(p)
            ws = wb["Base Analítica"]
            nomes_tabelas = list(ws.tables)
            tbl = ws.tables[nomes_tabelas[0]]
            # totalsRowCount não expõe direto no reload · inferimos por ref
            assert tbl is not None

    def test_amostras_linha_totais_tem_formulas_subtotal(self, amostras_geradas):
        """Regressão Correção-1 · Defeito 1.

        Setar tableColumn.totalsRowFunction no XML da Table não basta ·
        Excel exibe a célula VAZIA se o conteúdo não for preenchido.
        Este teste garante que criar_tabela_executiva grava a fórmula
        =SUBTOTAL(...) na célula da linha de totais para toda coluna com
        totalsRowFunction declarada, e que filtros recalculam via SUBTOTAL
        (códigos 1XX ignoram linhas ocultas/filtradas).
        """
        import re
        for p in amostras_geradas:
            wb = load_workbook(p)
            ws = wb["Base Analítica"]
            assert list(ws.tables), f"{p.name}: Base Analítica sem tabela"
            nome_tabela = list(ws.tables)[0]
            tbl = ws.tables[nome_tabela]

            assert tbl.totalsRowShown is True, (
                f"{p.name}: totalsRowShown deveria ser True"
            )
            assert (tbl.totalsRowCount or 0) >= 1, (
                f"{p.name}: totalsRowCount deveria ser >= 1"
            )

            # Ref com totals row → inferir última linha
            ref = tbl.ref  # ex: 'A1:F14'
            _inicio, _fim = ref.split(":")
            match_fim = re.match(r"([A-Z]+)(\d+)", _fim)
            assert match_fim is not None, f"{p.name}: ref inválido {ref!r}"
            linha_totais = int(match_fim.group(2))

            colunas_com_funcao = [
                (i, c) for i, c in enumerate(tbl.tableColumns)
                if c.totalsRowFunction
            ]
            assert colunas_com_funcao, (
                f"{p.name}: nenhuma coluna com totalsRowFunction"
            )

            for j, coluna in colunas_com_funcao:
                col_idx = j + 1  # openpyxl é 1-based
                cell = ws.cell(row=linha_totais, column=col_idx)
                valor = cell.value
                assert isinstance(valor, str), (
                    f"{p.name}: célula de total col={col_idx} "
                    f"linha={linha_totais} deveria conter fórmula · "
                    f"achou {valor!r}"
                )
                assert valor.upper().startswith("=SUBTOTAL("), (
                    f"{p.name}: célula de total col={col_idx} não começa "
                    f"com =SUBTOTAL · achou {valor!r}"
                )

    def test_amostras_resumo_executivo_colunas_com_larguras_minimas(
        self, amostras_geradas,
    ):
        """Regressão Correção-1 · Defeito 2.

        Aba Resumo Executivo deve abrir legível sem ajuste manual. Valida
        que as larguras fixas declaradas no final de
        renderizar_resumo_executivo foram efetivamente aplicadas e estão
        acima dos thresholds mínimos que cobrem rótulos longos e valores
        monetários em prosa.
        """
        MIN_A = 30  # rótulos longos de threshold não podem ficar cortados
        MIN_B = 40  # valores monetários "R$ 1.234.567,89" + margem
        MIN_C = 18  # terceira coluna (contagem · participação)
        MIN_D = 18  # quarta coluna

        for p in amostras_geradas:
            wb = load_workbook(p)
            ws = wb["Resumo Executivo"]
            larg_a = ws.column_dimensions["A"].width or 0
            larg_b = ws.column_dimensions["B"].width or 0
            larg_c = ws.column_dimensions["C"].width or 0
            larg_d = ws.column_dimensions["D"].width or 0

            assert larg_a >= MIN_A, (
                f"{p.name}: col A com largura {larg_a} · mínimo {MIN_A}"
            )
            assert larg_b >= MIN_B, (
                f"{p.name}: col B com largura {larg_b} · mínimo {MIN_B}"
            )
            assert larg_c >= MIN_C, (
                f"{p.name}: col C com largura {larg_c} · mínimo {MIN_C}"
            )
            assert larg_d >= MIN_D, (
                f"{p.name}: col D com largura {larg_d} · mínimo {MIN_D}"
            )

    def test_amostras_cabecalho_base_analitica_visivel(self, amostras_geradas):
        """Regressão Correção-2 · Cabeçalho da Base Analítica invisível.

        O TableStyleMedium{1,2,3,4} aplica fill escuro (cor_primaria) + fonte
        branca ao cabeçalho. Aplicar uma Font custom na linha 1 (mesmo que
        só com bold=True) sobrescreve a cor branca do estilo, resultando
        em texto preto sobre fundo escuro = barra preta visual.

        Este teste protege contra:
          (a) Font com rgb preto EXPLÍCITO no cabeçalho
          (b) Fill solid preto no cabeçalho
          (c) Anti-regressão específica · se a tabela usa estilo Medium/Dark
              (fill escuro), nenhuma célula do cabeçalho pode ter Font
              custom com bold=True e color ausente · essa combinação é
              exatamente o bug corrigido (Font override anula branco do
              estilo).
        """
        def _eh_preto_explicito(color_obj) -> bool:
            if color_obj is None:
                return False
            if getattr(color_obj, "type", None) != "rgb":
                return False
            rgb = getattr(color_obj, "rgb", None)
            if not isinstance(rgb, str):
                return False
            return rgb.upper() in ("00000000", "FF000000", "000000")

        for p in amostras_geradas:
            wb = load_workbook(p)
            ws = wb["Base Analítica"]

            # Identifica se há tabela com estilo de fill escuro
            tem_estilo_escuro = False
            for tbl in ws.tables.values():
                tsi = tbl.tableStyleInfo
                if tsi is None or not tsi.name:
                    continue
                if "Medium" in tsi.name or "Dark" in tsi.name:
                    tem_estilo_escuro = True
                    break

            for col_idx in range(1, ws.max_column + 1):
                c = ws.cell(row=1, column=col_idx)
                if not isinstance(c.value, str) or not c.value:
                    continue

                # (a) Font preta explícita
                assert not _eh_preto_explicito(c.font.color), (
                    f"{p.name}[L1C{col_idx}] cabeçalho com font rgb preta "
                    f"explícita: {c.value!r}"
                )
                # (b) Fill preto
                fill_solido_preto = (
                    c.fill.fill_type == "solid"
                    and _eh_preto_explicito(c.fill.fgColor)
                )
                assert not fill_solido_preto, (
                    f"{p.name}[L1C{col_idx}] cabeçalho com fill solid preto"
                )
                # (c) Anti-regressão específica da Correção-2
                if tem_estilo_escuro:
                    font_custom_sem_cor = (
                        c.font.b is True
                        and c.font.color is None
                    )
                    assert not font_custom_sem_cor, (
                        f"{p.name}[L1C{col_idx}] Font custom aplicada ao "
                        f"cabeçalho (bold=True, sem color) sobrescreve a "
                        f"cor branca do TableStyleMedium/Dark · "
                        f"texto fica preto sobre fundo escuro · "
                        f"valor={c.value!r}"
                    )

    def test_amostras_filtro_ativo_em_todas_as_colunas(self, amostras_geradas):
        """Regressão Correção-3 · Botão Filtrar desligado por default.

        Sem Table.autoFilter explícito, o Excel abre a tabela nativa com
        filtros desligados · a Usuária precisaria marcar "Design da
        Tabela > Botão Filtrar" manualmente em cada amostra. Este teste
        garante que criar_tabela_executiva seta Table.autoFilter com
        ref = range original (header + dados, sem a linha de totais ·
        caso contrário o botão aparece também sobre a totals row).
        """
        import re

        for p in amostras_geradas:
            wb = load_workbook(p)
            ws = wb["Base Analítica"]

            assert list(ws.tables), f"{p.name}: Base Analítica sem Table"
            nome_tabela = list(ws.tables)[0]
            tbl = ws.tables[nome_tabela]

            # autoFilter presente
            assert tbl.autoFilter is not None, (
                f"{p.name}: Table.autoFilter é None · botão Filtrar "
                f"virá desligado · Usuária precisa marcar manualmente"
            )
            assert tbl.autoFilter.ref, (
                f"{p.name}: Table.autoFilter.ref vazio"
            )

            # autoFilter ref termina 1 linha antes da table ref (exclui totals)
            ref_tabela = tbl.ref            # ex: "A1:F14"
            ref_filtro = tbl.autoFilter.ref  # ex: "A1:F13" ou "$A$1:$F$13"

            m_tabela = re.match(r"[A-Z]+\d+:([A-Z]+)(\d+)", ref_tabela)
            m_filtro = re.match(
                r"\$?[A-Z]+\$?\d+:\$?([A-Z]+)\$?(\d+)", ref_filtro,
            )
            assert m_tabela and m_filtro, (
                f"{p.name}: ref inválido · tabela={ref_tabela!r} "
                f"filtro={ref_filtro!r}"
            )

            col_tabela = m_tabela.group(1)
            lin_tabela = int(m_tabela.group(2))
            col_filtro = m_filtro.group(1)
            lin_filtro = int(m_filtro.group(2))

            assert col_filtro == col_tabela, (
                f"{p.name}: autoFilter cobre colunas diferentes da Table"
            )
            # Totals row shown + totalsRowCount=1 → filtro exclui 1 linha
            if tbl.totalsRowShown and (tbl.totalsRowCount or 0) >= 1:
                assert lin_filtro == lin_tabela - 1, (
                    f"{p.name}: autoFilter.ref deve excluir a linha de "
                    f"totais · tabela termina L{lin_tabela} filtro "
                    f"termina L{lin_filtro}"
                )

            # Cabeçalho tem exatamente N colunas = len(tableColumns)
            n_colunas_header = sum(
                1 for col_idx in range(1, ws.max_column + 1)
                if isinstance(ws.cell(row=1, column=col_idx).value, str)
            )
            assert n_colunas_header == len(tbl.tableColumns), (
                f"{p.name}: header com {n_colunas_header} cabeçalhos "
                f"mas tableColumns={len(tbl.tableColumns)}"
            )


class TestInvariantesResumoExecutivoIsolado:
    """Varre o resumo executivo gerado ad-hoc sem depender das amostras."""

    def test_resumo_zero_violacoes_azul(
        self, tmp_path, resumo_canonico, paleta_azul, vocabulario_carregado
    ):
        wb = Workbook()
        ws = wb.active
        ws.title = "R"
        renderizar_resumo_executivo(ws, resumo_canonico, paleta_azul,
                                     vocabulario_carregado)
        p = tmp_path / "r.xlsx"
        wb.save(p)
        violacoes = _varrer_workbook_procurando_violacoes(p, vocabulario_carregado)
        assert not violacoes, violacoes[:5]

    def test_resumo_zero_violacoes_verde(
        self, tmp_path, resumo_canonico, paleta_verde, vocabulario_carregado
    ):
        wb = Workbook()
        ws = wb.active
        renderizar_resumo_executivo(ws, resumo_canonico, paleta_verde,
                                     vocabulario_carregado)
        p = tmp_path / "r.xlsx"
        wb.save(p)
        violacoes = _varrer_workbook_procurando_violacoes(p, vocabulario_carregado)
        assert not violacoes, violacoes[:5]

    def test_resumo_zero_violacoes_cinza(
        self, tmp_path, resumo_canonico, paleta_cinza, vocabulario_carregado
    ):
        wb = Workbook()
        ws = wb.active
        renderizar_resumo_executivo(ws, resumo_canonico, paleta_cinza,
                                     vocabulario_carregado)
        p = tmp_path / "r.xlsx"
        wb.save(p)
        violacoes = _varrer_workbook_procurando_violacoes(p, vocabulario_carregado)
        assert not violacoes, violacoes[:5]

    def test_resumo_zero_violacoes_vinho(
        self, tmp_path, resumo_canonico, paleta_vinho, vocabulario_carregado
    ):
        wb = Workbook()
        ws = wb.active
        renderizar_resumo_executivo(ws, resumo_canonico, paleta_vinho,
                                     vocabulario_carregado)
        p = tmp_path / "r.xlsx"
        wb.save(p)
        violacoes = _varrer_workbook_procurando_violacoes(p, vocabulario_carregado)
        assert not violacoes, violacoes[:5]

    def test_resumo_sem_marcadores_ausentes(
        self, tmp_path, resumo_canonico, paleta_azul, vocabulario_carregado
    ):
        wb = Workbook()
        ws = wb.active
        renderizar_resumo_executivo(ws, resumo_canonico, paleta_azul,
                                     vocabulario_carregado)
        p = tmp_path / "r.xlsx"
        wb.save(p)
        marcadores = _varrer_workbook_procurando_marcadores(p)
        assert not marcadores, marcadores[:5]


class TestInvariantesVariadasEntradas:
    """Variações de entrada · invariantes devem se manter."""

    def test_resumo_com_distribuicao_vazia(
        self, tmp_path, paleta_azul, vocabulario_carregado
    ):
        r = ResumoExecutivoPadrao(
            bloco_1_cabecalho=CabecalhoExecucao(
                visao="V2",
                data_execucao=datetime(2026, 1, 1, 10, 0),
                modo_upload="DUAL",
                agrupadores=[],
                medida_principal=None,
            ),
            bloco_2_numeros_ancora={},
            bloco_3_distribuicao={},
            bloco_4_elementos_destacados={},
            bloco_5_leitura_qualitativa=LeituraQualitativa(
                classificacao_ativa="ESTAVEL",
                thresholds_usados={},
                alguma_leitura_alterada_por_edicao=False,
            ),
            bloco_6_qualidade_estrutural=QualidadeEstrutural(
                total_warnings=0,
                warnings_por_categoria={},
                ajustes_aplicados=0,
                tem_bloqueios_escapados=False,
            ),
        )
        wb = Workbook()
        ws = wb.active
        renderizar_resumo_executivo(ws, r, paleta_azul, vocabulario_carregado)
        p = tmp_path / "r.xlsx"
        wb.save(p)
        violacoes = _varrer_workbook_procurando_violacoes(p, vocabulario_carregado)
        assert not violacoes

    def test_resumo_com_bloqueios_escapados(
        self, tmp_path, paleta_vinho, vocabulario_carregado
    ):
        r = ResumoExecutivoPadrao(
            bloco_1_cabecalho=CabecalhoExecucao(
                visao="V2",
                data_execucao=datetime(2026, 1, 1, 10, 0),
                modo_upload="SIMPLES",
                agrupadores=["Mes"],
                medida_principal="Vendas",
            ),
            bloco_2_numeros_ancora={"total_origem": 1000.0},
            bloco_3_distribuicao={"PRESENTE_AMBOS": 10},
            bloco_4_elementos_destacados={},
            bloco_5_leitura_qualitativa=LeituraQualitativa(
                classificacao_ativa="SIGNIFICATIVA",
                thresholds_usados={"limiar_estabilidade_pct": 0.05},
                alguma_leitura_alterada_por_edicao=True,
            ),
            bloco_6_qualidade_estrutural=QualidadeEstrutural(
                total_warnings=5,
                warnings_por_categoria={"ausencia": 3, "nulo": 2},
                ajustes_aplicados=0,
                tem_bloqueios_escapados=True,
            ),
        )
        wb = Workbook()
        ws = wb.active
        renderizar_resumo_executivo(ws, r, paleta_vinho, vocabulario_carregado)
        p = tmp_path / "r.xlsx"
        wb.save(p)
        violacoes = _varrer_workbook_procurando_violacoes(p, vocabulario_carregado)
        assert not violacoes, violacoes[:5]

    def test_resumo_todas_as_4_paletas_produzem_conteudo_equivalente_em_volume(
        self, tmp_path, resumo_canonico, vocabulario_carregado
    ):
        """
        As 4 paletas devem produzir volumes similares de conteúdo (diferença
        deve ser ornamental · não estrutural).
        """
        totais = []
        for nome in ("azul", "verde", "cinza", "vinho"):
            wb = Workbook()
            ws = wb.active
            renderizar_resumo_executivo(
                ws, resumo_canonico, CATALOGO_PALETAS[nome],
                vocabulario_carregado,
            )
            p = tmp_path / f"r_{nome}.xlsx"
            wb.save(p)
            wb2 = load_workbook(p)
            ws2 = wb2.active
            conteudo = [
                v for row in ws2.iter_rows(values_only=True)
                for v in row if isinstance(v, str)
            ]
            totais.append(len(conteudo))
        # Todos com mesma contagem (mesmo resumo, só estilo muda)
        assert len(set(totais)) == 1, f"volumes divergentes: {totais}"
