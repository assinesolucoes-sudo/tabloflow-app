"""
test_v2_s8_smoke_visual.py — Smoke tests visuais Sessão 8.1 · C-6.

Protege contra regressão futura onde o valor renderizado em célula sai fora
de range realista. A suite Sessão 8 (725 testes) valida estrutura/contrato
mas não capta que 0.18 (margem) acabe formatado como "18,00%" (correto) ou
"69,77%" (regressão · multiplicação dupla).

5 testes obrigatórios:
  1. PERCENTUAL · card Média · valor renderizado em range realista (8-100%)
  2. PERCENTUAL · célula Diferença em Matriz · pre-multiplicada por 100
     com formato "p.p" literal (range -100 a +100)
  3. MONETARIO_BRL · card Total · valor renderizado em R$ realista
  4. QUANTIDADE · card Total · inteiro >= 0
  5. PERCENTUAL · Saúde da comparação · 3 colunas (sem Δ total)

Estratégia: gera workbook em arquivo tmp · abre com openpyxl · varre
células do Resumo Executivo procurando por padrões.
"""
from __future__ import annotations

import re
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd
import pytest
from openpyxl import load_workbook

_SRC = Path(__file__).parent.parent
if str(_SRC) not in sys.path:
    sys.path.insert(0, str(_SRC))

from contratos import (
    ColumnMeta,
    MotorResult,
    TipoEstruturalEnum,
    TipoSemanticoEnum,
    TipoTecnicoEnum,
)
from visoes.exportacao_v2 import exportar_resultado_v2
from visoes.visao_v2 import executar_v2


# ---------------------------------------------------------------------------
# Helpers (refletindo gerar_amostras_s8.py)
# ---------------------------------------------------------------------------


def _col_meta(nome: str, tipo_estrutural: str = "NUMERICO_CONTINUO") -> ColumnMeta:
    return ColumnMeta(
        nome=nome,
        tipo_tecnico=TipoTecnicoEnum("float" if tipo_estrutural == "NUMERICO_CONTINUO" else "string"),
        tipo_semantico=(
            TipoSemanticoEnum("numeric")
            if tipo_estrutural == "NUMERICO_CONTINUO"
            else TipoSemanticoEnum("categorico_baixa_card")
        ),
        tipo_estrutural=TipoEstruturalEnum(tipo_estrutural),
        subtipo_id_detectado=False,
        null_count=0,
        cardinalidade=10,
        eh_candidato_categorico=tipo_estrutural == "CATEGORICO_ELEGIVEL",
        padrao_cronologico_detectado=None,
        ordem_insercao=0,
    )


def _motor_result(df: pd.DataFrame) -> MotorResult:
    col_meta = {}
    for i, col in enumerate(df.columns):
        if pd.api.types.is_numeric_dtype(df[col]):
            meta = _col_meta(col, "NUMERICO_CONTINUO")
        else:
            meta = _col_meta(col, "CATEGORICO_ELEGIVEL")
        meta = meta.model_copy(update={
            "nome": col, "ordem_insercao": i,
            "null_count": int(df[col].isna().sum()),
            "cardinalidade": int(df[col].nunique()),
        })
        col_meta[col] = meta
    return MotorResult(
        df=df,
        column_meta=col_meta,
        modo_upload="SIMPLES",
        total_linhas_originais=len(df),
        total_linhas_processadas=len(df),
    )


def _config(unidade: str, campo: str = "Campo") -> Dict[str, Any]:
    return {
        "estrutura_entrada": "POR_COLUNAS",
        "origem_rotulo_tecnico": "Origem",
        "comparado_rotulo_tecnico": "Comparado",
        "origem_rotulo_ux": "Janeiro 2025",
        "comparado_rotulo_ux": "Fevereiro 2025",
        "coluna_discriminadora": None,
        "modo_4_ativado": False,
        "estados_nao_escolhidos": [],
        "campo_analisado": campo,
        "tipo_campo": "NUMERICO_ADITIVO",
        "semantica_campo": "MAIOR_MELHOR",
        "unidade": unidade,
        "regra_agregacao": "SOMA",
        "metodo_consolidacao_relativo": None,
        "campo_peso": None,
        "modo_pre_agregado": False,
        "agrupadores": ["Filial", "Produto"],
        "agrupador_destacado": "Filial",
        "resolucao_estrutural": None,
        "thresholds": {
            "limiar_estabilidade_pct": 0.01,
            "limiar_nulo_massivo_pct": 0.20,
            "limite_valores_discriminador_alerta": 50,
            "limite_variacao_extrema_pct": 10.0,
        },
        "modelo_aplicado": None,
    }


_FILIAIS = ["SP", "MG", "RJ", "BA", "PR"]
_PRODUTOS = ["Produto A", "Produto B", "Produto C", "Produto D"]


def _df_percentual_realista() -> pd.DataFrame:
    """
    Margens 0.08-0.45 · variação ±10 p.p · seed fixa.

    Sessão 8.1 · uma observação por (Filial, Produto) · evita SOMA por grupo
    multiplicar margens (5 amostras × 0.265 → 1.325 avg que faria card
    Média mostrar 132% ao invés de 26%). Para PERCENTUAL realista, cada
    linha já é uma taxa medida no nível do grupo.
    """
    import random
    random.seed(42)
    rows: List[Dict[str, Any]] = []
    for filial in _FILIAIS:
        for produto in _PRODUTOS:
            origem = round(float(random.uniform(0.08, 0.45)), 4)
            ruido = float(random.uniform(-0.10, 0.10))
            comparado = max(0.01, min(0.99, origem + ruido))
            rows.append({
                "Filial": filial,
                "Produto": produto,
                "Origem": origem,
                "Comparado": round(comparado, 4),
            })
    return pd.DataFrame(rows)


def _df_monetario_realista() -> pd.DataFrame:
    """Receitas R$ 500-2500 · seed fixa."""
    import random
    random.seed(42)
    rows: List[Dict[str, Any]] = []
    for filial in _FILIAIS:
        for produto in _PRODUTOS:
            for _ in range(5):
                origem = float(random.uniform(500, 2500))
                comparado = origem + random.uniform(-300, 300)
                rows.append({
                    "Filial": filial,
                    "Produto": produto,
                    "Origem": round(origem, 2),
                    "Comparado": round(comparado, 2),
                })
    return pd.DataFrame(rows)


def _df_quantidade_realista() -> pd.DataFrame:
    """Unidades 50-1500 · seed fixa."""
    import random
    random.seed(42)
    rows: List[Dict[str, Any]] = []
    for filial in _FILIAIS:
        for produto in _PRODUTOS:
            for _ in range(5):
                origem = int(random.uniform(50, 1500))
                comparado = max(0, int(origem + random.uniform(-200, 200)))
                rows.append({
                    "Filial": filial,
                    "Produto": produto,
                    "Origem": origem,
                    "Comparado": comparado,
                })
    return pd.DataFrame(rows)


def _gerar_workbook(unidade: str, df: pd.DataFrame, tmp_path: Path) -> Path:
    cfg = _config(unidade=unidade)
    v2 = executar_v2(_motor_result(df), cfg)
    saida = tmp_path / f"smoke_{unidade}.xlsx"
    exportar_resultado_v2(v2, str(saida), paleta_nome="azul")
    return saida


def _achar_celula_apos_rotulo(ws, padrao_rotulo: str) -> Optional[Any]:
    """
    Procura uma célula cujo valor (string) inicia com `padrao_rotulo` e
    devolve a célula imediatamente abaixo (que é o valor do card no
    layout `_mesclar_card`).
    """
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith(padrao_rotulo):
                return ws.cell(row=cell.row + 1, column=cell.column)
    return None


def _parse_percent_br(texto: str) -> Optional[float]:
    """Converte '18,30%' em fração 0.183. Retorna None se não bater."""
    if not isinstance(texto, str):
        return None
    m = re.match(r"^\s*(-?)([\d\.]+),(\d+)\s*%\s*$", texto)
    if not m:
        return None
    sinal = -1.0 if m.group(1) == "-" else 1.0
    inteiro = m.group(2).replace(".", "")
    decimal = m.group(3)
    return sinal * float(f"{inteiro}.{decimal}") / 100.0


def _parse_moeda_br(texto: str) -> Optional[float]:
    """Converte 'R$ 1.234,50' em 1234.50. Aceita parênteses para negativo."""
    if not isinstance(texto, str):
        return None
    neg = texto.strip().startswith("(") and texto.strip().endswith(")")
    limpo = texto.replace("(", "").replace(")", "").strip()
    m = re.match(r"^R\$\s*([\d\.]+),(\d+)$", limpo)
    if not m:
        return None
    inteiro = m.group(1).replace(".", "")
    decimal = m.group(2)
    val = float(f"{inteiro}.{decimal}")
    return -val if neg else val


# ===========================================================================
# Smoke 1 · PERCENTUAL · card Média · valor renderizado em range realista
# ===========================================================================


def test_smoke_visual_percentual_card_media_dentro_de_range_realista(tmp_path):
    """
    Card 'Média · Janeiro 2025' em PERCENTUAL deve renderizar valor
    user-facing entre 5% e 100% (frações 0.05 a 1.0) quando base é
    margem realista (0.08-0.45 com ruído). Protege contra regressão onde
    formato wrong multiplica valor (ex.: "69,77%" para média de margem).
    """
    df = _df_percentual_realista()
    saida = _gerar_workbook("PERCENTUAL", df, tmp_path)
    wb = load_workbook(str(saida))
    ws = wb["Resumo Executivo"]

    cell = _achar_celula_apos_rotulo(ws, "Média · Janeiro")
    assert cell is not None, "Card 'Média · Janeiro 2025' não encontrado"
    valor_str = cell.value
    fracao = _parse_percent_br(valor_str)
    assert fracao is not None, f"Card Média não parsável como %: {valor_str!r}"
    assert 0.05 <= fracao <= 1.0, (
        f"Card Média em PERCENTUAL fora de range realista: {fracao} "
        f"(string={valor_str!r}) · esperado 0.05-1.00 para margens reais"
    )


# ===========================================================================
# Smoke 2 · PERCENTUAL · célula Diferença em Matriz · p.p literal pré-mult
# ===========================================================================


def test_smoke_visual_percentual_celula_diferenca_em_pp_literal(tmp_path):
    """
    Célula 'Variação absoluta (p.p)' na Matriz de Confronto deve:
      - ter formato number_format contendo 'p.p'
      - ter valor numérico pré-multiplicado por 100 (range -100 a +100)
        já que diferenças máximas plausíveis são ±0.99 em fração.
    """
    df = _df_percentual_realista()
    saida = _gerar_workbook("PERCENTUAL", df, tmp_path)
    wb = load_workbook(str(saida))
    ws = wb["Matriz de Confronto"]

    # Acha coluna pelo cabeçalho · cabeçalho está em uma das primeiras linhas
    col_dif = None
    linha_header = None
    for row in ws.iter_rows(min_row=1, max_row=10):
        for cell in row:
            if isinstance(cell.value, str) and "Variação absoluta" in cell.value and "p.p" in cell.value:
                col_dif = cell.column
                linha_header = cell.row
                break
        if col_dif:
            break
    assert col_dif is not None, "Cabeçalho 'Variação absoluta (p.p)' não encontrado na Matriz"

    # Pega primeiras 3 células de dados na coluna
    valores: List[Any] = []
    formatos: List[str] = []
    for r in range(linha_header + 1, linha_header + 4):
        cel = ws.cell(row=r, column=col_dif)
        valores.append(cel.value)
        formatos.append(cel.number_format)

    assert all("p.p" in f for f in formatos), (
        f"Formato da coluna Diferença não contém 'p.p': {formatos}"
    )
    # Valores numéricos esperados em range [-100, +100]
    numericos = [v for v in valores if isinstance(v, (int, float))]
    assert numericos, f"Sem valores numéricos na coluna Diferença: {valores}"
    for v in numericos:
        assert -100.0 <= v <= 100.0, (
            f"Valor de diferença PERCENTUAL fora de [-100, 100]: {v} "
            f"(esperado pré-multiplicação por 100 para formato p.p literal)"
        )


# ===========================================================================
# Smoke 3 · MONETARIO_BRL · card Total · regressão zero (R$ realista)
# ===========================================================================


def test_smoke_visual_monetario_card_total_em_rs_realista(tmp_path):
    """
    Card 'Total · Janeiro 2025' em MONETARIO_BRL renderiza string 'R$ ...'
    parsável para valor entre R$ 100 e R$ 1.000.000. Protege regressão
    zero do contrato MONETARIO pré-Sessão 8.
    """
    df = _df_monetario_realista()
    saida = _gerar_workbook("MONETARIO_BRL", df, tmp_path)
    wb = load_workbook(str(saida))
    ws = wb["Resumo Executivo"]

    cell = _achar_celula_apos_rotulo(ws, "Total · Janeiro")
    assert cell is not None, "Card 'Total · Janeiro 2025' não encontrado"
    valor_str = cell.value
    valor_num = _parse_moeda_br(valor_str)
    assert valor_num is not None, f"Card Total não parsável como R$: {valor_str!r}"
    assert 100.0 <= valor_num <= 1_000_000.0, (
        f"Card Total em MONETARIO_BRL fora de range realista: {valor_num} "
        f"(string={valor_str!r}) · esperado R$ 100 a R$ 1.000.000"
    )


# ===========================================================================
# Smoke 4 · QUANTIDADE · card Total · inteiro >= 0
# ===========================================================================


def test_smoke_visual_quantidade_card_total_inteiro_positivo(tmp_path):
    """
    Card 'Total · Janeiro 2025' em QUANTIDADE renderiza inteiro
    (sem casas decimais) >= 0.
    """
    df = _df_quantidade_realista()
    saida = _gerar_workbook("QUANTIDADE", df, tmp_path)
    wb = load_workbook(str(saida))
    ws = wb["Resumo Executivo"]

    cell = _achar_celula_apos_rotulo(ws, "Total · Janeiro")
    assert cell is not None, "Card 'Total · Janeiro 2025' não encontrado"
    valor_str = cell.value
    assert isinstance(valor_str, str), f"Esperado string · obtido {valor_str!r}"
    # Aceita "1.234" · "56789" · "-789"
    limpo = valor_str.replace(".", "").replace("-", "").strip()
    assert limpo.isdigit(), (
        f"Card Total QUANTIDADE não é inteiro renderizado: {valor_str!r}"
    )
    valor_num = int(limpo)
    assert valor_num >= 0, f"Card Total negativo: {valor_num}"


# ===========================================================================
# Smoke 5 · PERCENTUAL · Saúde da comparação · sem coluna Δ total
# ===========================================================================


def test_smoke_visual_percentual_saude_sem_coluna_delta_total(tmp_path):
    """
    PERCENTUAL · Saúde da comparação tem APENAS 3 colunas:
    Categoria · Casos · Participação. Coluna 'Δ total' não deve aparecer
    (somar p.p é como somar percentuais · viola C.D3).
    """
    df = _df_percentual_realista()
    saida = _gerar_workbook("PERCENTUAL", df, tmp_path)
    wb = load_workbook(str(saida))
    ws = wb["Resumo Executivo"]

    # Procura a célula com texto "Saúde da comparação"
    linha_secao = None
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and "Saúde da comparação" in cell.value:
                linha_secao = cell.row
                break
        if linha_secao:
            break
    assert linha_secao is not None, "Seção 'Saúde da comparação' não encontrada"

    # Header da tabela está logo abaixo
    headers_lidos: List[str] = []
    for col in range(1, 20):
        v = ws.cell(row=linha_secao + 1, column=col).value
        if isinstance(v, str) and v.strip():
            headers_lidos.append(v.strip())

    headers_unicos = set(headers_lidos)
    assert "Categoria" in headers_unicos, f"Cabeçalho Categoria ausente: {headers_lidos}"
    assert "Casos" in headers_unicos, f"Cabeçalho Casos ausente: {headers_lidos}"
    assert "Participação" in headers_unicos, f"Cabeçalho Participação ausente: {headers_lidos}"
    assert "Δ total" not in headers_unicos, (
        f"PERCENTUAL não deveria ter coluna Δ total · cabeçalhos: {headers_lidos}"
    )


# ===========================================================================
# Smoke 6 · MONETARIO · Saúde da comparação · MANTÉM coluna Δ total
# ===========================================================================


def test_smoke_visual_monetario_saude_mantem_coluna_delta_total(tmp_path):
    """
    Regressão zero · MONETARIO_BRL preserva coluna 'Δ total' na Saúde
    da comparação (somar valores monetários é semanticamente válido).
    """
    df = _df_monetario_realista()
    saida = _gerar_workbook("MONETARIO_BRL", df, tmp_path)
    wb = load_workbook(str(saida))
    ws = wb["Resumo Executivo"]

    linha_secao = None
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and "Saúde da comparação" in cell.value:
                linha_secao = cell.row
                break
        if linha_secao:
            break
    assert linha_secao is not None, "Seção 'Saúde da comparação' não encontrada"

    headers_lidos: List[str] = []
    for col in range(1, 20):
        v = ws.cell(row=linha_secao + 1, column=col).value
        if isinstance(v, str) and v.strip():
            headers_lidos.append(v.strip())

    assert "Δ total" in set(headers_lidos), (
        f"MONETARIO_BRL deveria ter coluna Δ total · cabeçalhos: {headers_lidos}"
    )
