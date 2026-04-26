"""
Testes F-APRESENT P1 · capability 8 · Badges semânticos.

4 tipos:
  - Interface (assinatura / raises)
  - Snapshot (fills e fontes corretos)
  - Regressão (compatibilidade com Table nativa)
  - Invariantes (catálogo completo · semânticas distintas)
"""
from __future__ import annotations

import logging

import pytest
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

from apresentacao import (
    BadgeStyle,
    CATALOGO_BADGES,
    CATALOGO_PALETAS,
    MAPEAMENTO_V2,
    Paleta,
    SEMANTICAS_CANONICAS,
    aplicar_badge,
    criar_tabela_executiva,
)


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture
def paleta_azul() -> Paleta:
    return CATALOGO_PALETAS["azul"]


@pytest.fixture
def paleta_verde() -> Paleta:
    return CATALOGO_PALETAS["verde"]


@pytest.fixture
def paleta_cinza() -> Paleta:
    return CATALOGO_PALETAS["cinza"]


@pytest.fixture
def paleta_vinho() -> Paleta:
    return CATALOGO_PALETAS["vinho"]


@pytest.fixture
def ws():
    wb = Workbook()
    return wb.active


# ===========================================================================
# Interface
# ===========================================================================

class TestInterfaceBadges:
    def test_semanticas_canonicas_tem_4(self):
        assert len(SEMANTICAS_CANONICAS) == 4

    def test_semanticas_canonicas_nomes(self):
        assert set(SEMANTICAS_CANONICAS) == {"positivo", "negativo", "neutro", "atencao"}

    def test_catalogo_cobre_4_paletas(self):
        assert set(CATALOGO_BADGES.keys()) == {"azul", "verde", "cinza", "vinho"}

    def test_catalogo_cada_paleta_tem_4_semanticas(self):
        for nome, por_sem in CATALOGO_BADGES.items():
            assert set(por_sem.keys()) == set(SEMANTICAS_CANONICAS), (
                f"paleta {nome} não tem catálogo completo"
            )

    def test_badgestyle_frozen(self):
        b = BadgeStyle(fill_hex="AABBCC", font_hex="112233")
        with pytest.raises(Exception):
            b.fill_hex = "DEADBE"  # type: ignore[misc]

    def test_badgestyle_hasheavel(self):
        b = BadgeStyle(fill_hex="AABBCC", font_hex="112233")
        hash(b)  # não deve raise

    def test_badgestyle_validacao_hex(self):
        with pytest.raises(ValueError):
            BadgeStyle(fill_hex="#AABBCC", font_hex="112233")  # leading '#'
        with pytest.raises(ValueError):
            BadgeStyle(fill_hex="ZZZZZZ", font_hex="112233")  # não hex
        with pytest.raises(ValueError):
            BadgeStyle(fill_hex="AABB", font_hex="112233")  # tamanho errado

    def test_aplicar_badge_celula_none_raises(self, paleta_azul):
        with pytest.raises(TypeError):
            aplicar_badge(None, "PRESENTE_AMBOS", paleta_azul)

    def test_aplicar_badge_paleta_none_raises(self, ws):
        c = ws.cell(row=1, column=1)
        with pytest.raises(TypeError):
            aplicar_badge(c, "PRESENTE_AMBOS", None)

    def test_mapeamento_v2_cobre_6_classificacoes_estruturais(self):
        canonicas = {
            "PRESENTE_AMBOS", "AUSENTE_ORIGEM", "AUSENTE_COMPARADO",
            "NULO_ORIGEM", "NULO_COMPARADO", "NULO_AMBOS",
        }
        assert canonicas.issubset(MAPEAMENTO_V2.keys())

    def test_mapeamento_v2_cobre_4_classificacoes_semanticas(self):
        canonicas = {"POSITIVO", "NEGATIVO", "NEUTRO", "NAO_APLICAVEL"}
        assert canonicas.issubset(MAPEAMENTO_V2.keys())

    def test_aplicar_badge_aceita_mapeamento_custom(self, ws, paleta_azul):
        mapa = {"X": "positivo"}
        c = ws.cell(row=1, column=1, value="X")
        aplicar_badge(c, "X", paleta_azul, mapeamento=mapa)
        esperado = CATALOGO_BADGES["azul"]["positivo"]
        assert c.fill.start_color.rgb.endswith(esperado.fill_hex)


# ===========================================================================
# Snapshot
# ===========================================================================

class TestSnapshotBadges:
    def test_positivo_azul_fill_correto(self, ws, paleta_azul):
        c = ws.cell(row=1, column=1, value="POSITIVO")
        aplicar_badge(c, "POSITIVO", paleta_azul)
        esp = CATALOGO_BADGES["azul"]["positivo"]
        assert c.fill.start_color.rgb.endswith(esp.fill_hex)

    def test_positivo_azul_font_correto(self, ws, paleta_azul):
        c = ws.cell(row=1, column=1, value="POSITIVO")
        aplicar_badge(c, "POSITIVO", paleta_azul)
        esp = CATALOGO_BADGES["azul"]["positivo"]
        assert c.font.color.rgb.endswith(esp.font_hex)

    def test_negativo_verde(self, ws, paleta_verde):
        c = ws.cell(row=1, column=1, value="NEGATIVO")
        aplicar_badge(c, "NEGATIVO", paleta_verde)
        esp = CATALOGO_BADGES["verde"]["negativo"]
        assert c.fill.start_color.rgb.endswith(esp.fill_hex)
        assert c.font.color.rgb.endswith(esp.font_hex)

    def test_neutro_cinza(self, ws, paleta_cinza):
        c = ws.cell(row=1, column=1, value="PRESENTE_AMBOS")
        aplicar_badge(c, "PRESENTE_AMBOS", paleta_cinza)
        esp = CATALOGO_BADGES["cinza"]["neutro"]
        assert c.fill.start_color.rgb.endswith(esp.fill_hex)

    def test_atencao_vinho_ausente_origem(self, ws, paleta_vinho):
        c = ws.cell(row=1, column=1, value="AUSENTE_ORIGEM")
        aplicar_badge(c, "AUSENTE_ORIGEM", paleta_vinho)
        esp = CATALOGO_BADGES["vinho"]["atencao"]
        assert c.fill.start_color.rgb.endswith(esp.fill_hex)

    def test_ausente_comparado_mapa_atencao(self, ws, paleta_azul):
        c = ws.cell(row=1, column=1, value="AUSENTE_COMPARADO")
        aplicar_badge(c, "AUSENTE_COMPARADO", paleta_azul)
        esp = CATALOGO_BADGES["azul"]["atencao"]
        assert c.fill.start_color.rgb.endswith(esp.fill_hex)

    def test_nulo_ambos_atencao(self, ws, paleta_verde):
        c = ws.cell(row=1, column=1, value="NULO_AMBOS")
        aplicar_badge(c, "NULO_AMBOS", paleta_verde)
        esp = CATALOGO_BADGES["verde"]["atencao"]
        assert c.fill.start_color.rgb.endswith(esp.fill_hex)

    def test_nao_aplicavel_cai_em_neutro(self, ws, paleta_cinza):
        c = ws.cell(row=1, column=1, value="NAO_APLICAVEL")
        aplicar_badge(c, "NAO_APLICAVEL", paleta_cinza)
        esp = CATALOGO_BADGES["cinza"]["neutro"]
        assert c.fill.start_color.rgb.endswith(esp.fill_hex)

    def test_aumento_vira_positivo(self, ws, paleta_azul):
        c = ws.cell(row=1, column=1, value="AUMENTO")
        aplicar_badge(c, "AUMENTO", paleta_azul)
        esp = CATALOGO_BADGES["azul"]["positivo"]
        assert c.fill.start_color.rgb.endswith(esp.fill_hex)

    def test_reducao_vira_negativo(self, ws, paleta_azul):
        c = ws.cell(row=1, column=1, value="REDUCAO")
        aplicar_badge(c, "REDUCAO", paleta_azul)
        esp = CATALOGO_BADGES["azul"]["negativo"]
        assert c.fill.start_color.rgb.endswith(esp.fill_hex)

    def test_classificacao_case_insensitive(self, ws, paleta_azul):
        c = ws.cell(row=1, column=1, value="positivo")
        aplicar_badge(c, "positivo", paleta_azul)
        esp = CATALOGO_BADGES["azul"]["positivo"]
        assert c.fill.start_color.rgb.endswith(esp.fill_hex)

    def test_fonte_respeita_familia_da_paleta(self, ws, paleta_azul):
        c = ws.cell(row=1, column=1, value="POSITIVO")
        aplicar_badge(c, "POSITIVO", paleta_azul)
        assert c.font.name == paleta_azul.fonte_familia


# ===========================================================================
# Regressão
# ===========================================================================

class TestRegressaoBadges:
    def test_classificacao_desconhecida_fallback_neutro(
        self, ws, paleta_azul, caplog,
    ):
        c = ws.cell(row=1, column=1, value="PANTONE_INVENTADO")
        with caplog.at_level(logging.WARNING):
            aplicar_badge(c, "PANTONE_INVENTADO", paleta_azul)
        esp = CATALOGO_BADGES["azul"]["neutro"]
        assert c.fill.start_color.rgb.endswith(esp.fill_hex)
        assert any("PANTONE_INVENTADO" in r.message for r in caplog.records)

    def test_classificacao_none_fallback_neutro(self, ws, paleta_azul, caplog):
        c = ws.cell(row=1, column=1)
        with caplog.at_level(logging.WARNING):
            aplicar_badge(c, None, paleta_azul)  # type: ignore[arg-type]
        esp = CATALOGO_BADGES["azul"]["neutro"]
        assert c.fill.start_color.rgb.endswith(esp.fill_hex)

    def test_classificacao_vazia_fallback_neutro(self, ws, paleta_azul):
        c = ws.cell(row=1, column=1, value="")
        aplicar_badge(c, "", paleta_azul)
        esp = CATALOGO_BADGES["azul"]["neutro"]
        assert c.fill.start_color.rgb.endswith(esp.fill_hex)

    def test_badge_dentro_de_table_nao_quebra_autofilter(self, paleta_azul):
        wb = Workbook()
        ws = wb.active
        ws.append(["Agrup", "Classificação", "Total"])
        ws.append(["Jan|Norte", "POSITIVO", 1000.0])
        ws.append(["Fev|Sul",   "NEGATIVO", 800.0])

        # Aplica badges ANTES de criar Table · cenário típico
        c2 = ws.cell(row=2, column=2)
        c3 = ws.cell(row=3, column=2)
        aplicar_badge(c2, "POSITIVO", paleta_azul)
        aplicar_badge(c3, "NEGATIVO", paleta_azul)

        tabela = criar_tabela_executiva(
            ws=ws,
            range_ref="A1:C3",
            nome="TabelaTeste",
            totais_por_coluna={"Total": "sum"},
            paleta_nome="azul",
        )
        assert tabela.autoFilter is not None
        # Linha de totais + 1 (linha 4) · autoFilter cobre 1:3 (inclui cabeçalho)
        assert "A1:C3" in tabela.autoFilter.ref

    def test_aplicar_badge_nao_altera_valor_celula(self, ws, paleta_azul):
        c = ws.cell(row=1, column=1, value="POSITIVO")
        aplicar_badge(c, "POSITIVO", paleta_azul)
        assert c.value == "POSITIVO"

    def test_reaplicacao_mesmo_badge_idempotente(self, ws, paleta_azul):
        c = ws.cell(row=1, column=1, value="POSITIVO")
        aplicar_badge(c, "POSITIVO", paleta_azul)
        fill1 = c.fill.start_color.rgb
        aplicar_badge(c, "POSITIVO", paleta_azul)
        fill2 = c.fill.start_color.rgb
        assert fill1 == fill2


# ===========================================================================
# Invariantes
# ===========================================================================

class TestInvariantesBadges:
    def test_cada_paleta_tem_positivo_distinto_de_negativo(self):
        for nome, mapa in CATALOGO_BADGES.items():
            assert mapa["positivo"].fill_hex != mapa["negativo"].fill_hex, (
                f"paleta {nome} tem positivo e negativo com mesmo fill"
            )
            assert mapa["positivo"].font_hex != mapa["negativo"].font_hex, (
                f"paleta {nome} tem positivo e negativo com mesma font"
            )

    def test_cada_paleta_tem_atencao_distinto_de_neutro(self):
        for nome, mapa in CATALOGO_BADGES.items():
            assert mapa["atencao"].fill_hex != mapa["neutro"].fill_hex, (
                f"paleta {nome}: atenção indistinguível de neutro"
            )

    def test_todos_badges_sao_hex_validos(self):
        for mapa in CATALOGO_BADGES.values():
            for style in mapa.values():
                int(style.fill_hex, 16)  # não deve raise
                int(style.font_hex, 16)
                assert len(style.fill_hex) == 6
                assert len(style.font_hex) == 6

    def test_semanticas_sao_strings_minusculas(self):
        for mapa in CATALOGO_BADGES.values():
            for chave in mapa.keys():
                assert chave == chave.lower()

    def test_todas_paletas_aplicam_todas_semanticas_sem_exception(self, ws):
        classificacoes = ["POSITIVO", "NEGATIVO", "PRESENTE_AMBOS", "AUSENTE_ORIGEM"]
        for paleta in CATALOGO_PALETAS.values():
            for i, cls in enumerate(classificacoes, start=1):
                c = ws.cell(row=i, column=1, value=cls)
                aplicar_badge(c, cls, paleta)

    def test_mapeamento_v2_so_entrega_semanticas_canonicas(self):
        for chave, sem in MAPEAMENTO_V2.items():
            assert sem in SEMANTICAS_CANONICAS, (
                f"entrada {chave!r} do MAPEAMENTO_V2 entrega semântica fora do catálogo"
            )
