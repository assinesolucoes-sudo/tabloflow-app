"""
test_graficos.py — F-APRESENT capability 11 · gráficos nativos + nomenclatura.

Cobertura mínima (D-176):
  - criar_grafico_distribuicao insere 1 PieChart na worksheet
  - criar_grafico_top_variacoes insere 1 BarChart horizontal
  - gerar_nome_arquivo com e sem contexto produz nome executivo correto
  - caracteres inválidos de filesystem são higienizados
"""
from __future__ import annotations

from datetime import date
from pathlib import Path

import pytest
from openpyxl import Workbook

from apresentacao import (
    criar_grafico_distribuicao,
    criar_grafico_top_variacoes,
    gerar_nome_arquivo,
)


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture
def ws_com_distribuicao():
    wb = Workbook()
    ws = wb.active
    ws["B5"] = "Categoria"
    ws["C5"] = "Casos"
    ws["B6"], ws["C6"] = "Presente nos dois lados", 18
    ws["B7"], ws["C7"] = "Ausente na origem", 4
    ws["B8"], ws["C8"] = "Ausente no comparado", 2
    return ws


@pytest.fixture
def ws_com_variacoes():
    wb = Workbook()
    ws = wb.active
    ws["B10"] = "Rótulo"
    ws["C10"] = "Diferença"
    ws["B11"], ws["C11"] = "2025-02 · Filial RJ · X", 1500.0
    ws["B12"], ws["C12"] = "2025-02 · Filial SP · Y", -800.0
    ws["B13"], ws["C13"] = "2025-02 · Filial RJ · Z", 0.0
    return ws


# ---------------------------------------------------------------------------
# Testes · pizza de distribuição
# ---------------------------------------------------------------------------

def test_grafico_distribuicao_inserido(ws_com_distribuicao):
    """Depois de chamar criar_grafico_distribuicao, ws._charts tem 1 item."""
    n_antes = len(ws_com_distribuicao._charts)
    chart = criar_grafico_distribuicao(
        ws_com_distribuicao,
        anchor_cell="E5",
        dados_range="B5:C8",
        paleta_nome="azul",
    )
    assert chart is not None
    assert len(ws_com_distribuicao._charts) == n_antes + 1


def test_grafico_distribuicao_aceita_paletas_do_catalogo(ws_com_distribuicao):
    """Cada paleta do catálogo D-164 é aceita sem raise."""
    for nome in ("azul", "verde", "cinza", "vinho"):
        criar_grafico_distribuicao(
            ws_com_distribuicao,
            anchor_cell="E5",
            dados_range="B5:C8",
            paleta_nome=nome,
        )


def test_grafico_distribuicao_falha_em_range_invalido(ws_com_distribuicao):
    with pytest.raises(ValueError):
        criar_grafico_distribuicao(
            ws_com_distribuicao,
            anchor_cell="E5",
            dados_range="B5",  # sem ":"
            paleta_nome="azul",
        )


# ---------------------------------------------------------------------------
# Testes · barras de top variações
# ---------------------------------------------------------------------------

def test_grafico_top_variacoes_inserido(ws_com_variacoes):
    """Depois de chamar criar_grafico_top_variacoes, ws._charts tem 1 item."""
    n_antes = len(ws_com_variacoes._charts)
    chart = criar_grafico_top_variacoes(
        ws_com_variacoes,
        anchor_cell="E10",
        dados_range="B10:C13",
        paleta_nome="azul",
    )
    assert chart is not None
    assert len(ws_com_variacoes._charts) == n_antes + 1


# ---------------------------------------------------------------------------
# Testes · nomenclatura (D-176 · P-12)
# ---------------------------------------------------------------------------

def test_gerar_nome_arquivo_com_contexto(tmp_path: Path):
    caminho = gerar_nome_arquivo(
        nome_visao_user_facing="Analise Comparativa",
        contexto="Orcado vs Realizado",
        data=date(2026, 4, 24),
        diretorio_saida=tmp_path,
    )
    assert caminho.name == "Analise Comparativa - Orcado vs Realizado - 24-04-2026.xlsx"
    assert caminho.parent == tmp_path


def test_gerar_nome_arquivo_sem_contexto(tmp_path: Path):
    caminho = gerar_nome_arquivo(
        nome_visao_user_facing="Analise Comparativa",
        contexto=None,
        data=date(2026, 4, 24),
        diretorio_saida=tmp_path,
    )
    assert caminho.name == "Analise Comparativa - 24-04-2026.xlsx"


def test_gerar_nome_arquivo_higieniza_caracteres_invalidos(tmp_path: Path):
    caminho = gerar_nome_arquivo(
        nome_visao_user_facing="Analise Comparativa",
        contexto="Plano A/B: 2025 * teste",
        data=date(2026, 4, 24),
        diretorio_saida=tmp_path,
    )
    nome = caminho.name
    # Caracteres proibidos substituídos por hífen
    for proibido in ("\\", "/", ":", "*", "?", '"', "<", ">", "|"):
        assert proibido not in nome
    assert nome.endswith(".xlsx")
    assert "24-04-2026" in nome


def test_gerar_nome_arquivo_nao_contem_v2_nem_timestamp(tmp_path: Path):
    caminho = gerar_nome_arquivo(
        nome_visao_user_facing="Analise Comparativa",
        contexto="Janeiro vs Fevereiro",
        data=date(2026, 4, 24),
        diretorio_saida=tmp_path,
    )
    nome = caminho.name
    # D-176 · sem código V2 nem timestamp ISO
    assert "V2" not in nome
    assert "v2" not in nome.lower() or nome.lower().count("v2") == 1 and "v2" in "vs"
    # Sem formato ISO_BR (DD-MM-AAAA é o padrão, mas não HH:MM)
    import re as _re
    assert not _re.search(r"\d{8}_\d{6}", nome)


def test_gerar_nome_arquivo_exige_nome_visao(tmp_path: Path):
    with pytest.raises(ValueError):
        gerar_nome_arquivo(
            nome_visao_user_facing="",
            contexto="x",
            data=date(2026, 4, 24),
            diretorio_saida=tmp_path,
        )
