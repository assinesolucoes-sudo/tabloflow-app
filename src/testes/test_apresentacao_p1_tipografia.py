"""
Testes F-APRESENT P1 · capability 9 · Hierarquia tipográfica.

4 tipos:
  - Interface (assinatura / raises)
  - Snapshot (estilo aplicado é coerente com paleta)
  - Regressão (bugs potenciais prevenidos)
  - Invariante (propriedades que devem valer sempre)
"""
from __future__ import annotations

from typing import Dict

import pytest
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from apresentacao import (
    CATALOGO_PALETAS,
    NIVEIS_CANONICOS,
    Paleta,
    aplicar_campo,
    aplicar_hierarquia_tipografica,
    aplicar_nivel,
    aplicar_secao,
    aplicar_titulo_aba,
    aplicar_valor,
    escrever_campo_valor,
    escrever_secao,
    escrever_titulo_aba,
)


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture
def paleta_azul() -> Paleta:
    return CATALOGO_PALETAS["azul"]


@pytest.fixture
def paleta_verde() -> Paleta:
    return CATALOGO_PALETAS["verde"]


@pytest.fixture
def paleta_cinza() -> Paleta:
    return CATALOGO_PALETAS["cinza"]


@pytest.fixture
def paleta_vinho() -> Paleta:
    return CATALOGO_PALETAS["vinho"]


@pytest.fixture
def ws():
    wb = Workbook()
    return wb.active


# ===========================================================================
# Interface
# ===========================================================================

class TestInterfaceTipografia:
    def test_niveis_canonicos_tem_4(self):
        assert len(NIVEIS_CANONICOS) == 4

    def test_niveis_canonicos_nomes(self):
        assert set(NIVEIS_CANONICOS) == {"titulo_aba", "secao", "campo", "valor"}

    def test_aplicar_titulo_aba_aceita_cell_e_paleta(self, ws, paleta_azul):
        c = ws.cell(row=1, column=1, value="X")
        aplicar_titulo_aba(c, paleta_azul)
        assert c.font.name == "Calibri"

    def test_aplicar_secao_aceita_cell_e_paleta(self, ws, paleta_azul):
        c = ws.cell(row=2, column=1, value="Seção")
        aplicar_secao(c, paleta_azul)
        assert c.font.bold is True

    def test_aplicar_campo_aceita_cell_e_paleta(self, ws, paleta_azul):
        c = ws.cell(row=3, column=1, value="Rótulo")
        aplicar_campo(c, paleta_azul)
        assert c.font.bold is True

    def test_aplicar_valor_aceita_cell_e_paleta(self, ws, paleta_azul):
        c = ws.cell(row=4, column=1, value="Valor")
        aplicar_valor(c, paleta_azul)
        assert c.font.bold is False

    def test_aplicar_titulo_aba_cell_none_raises(self, paleta_azul):
        with pytest.raises(TypeError):
            aplicar_titulo_aba(None, paleta_azul)

    def test_aplicar_secao_paleta_none_raises(self, ws):
        c = ws.cell(row=1, column=1)
        with pytest.raises(TypeError):
            aplicar_secao(c, None)

    def test_aplicar_campo_paleta_none_raises(self, ws):
        c = ws.cell(row=1, column=1)
        with pytest.raises(TypeError):
            aplicar_campo(c, None)

    def test_aplicar_valor_paleta_none_raises(self, ws):
        c = ws.cell(row=1, column=1)
        with pytest.raises(TypeError):
            aplicar_valor(c, None)

    def test_aplicar_hierarquia_tipografica_ws_none_raises(self, paleta_azul):
        with pytest.raises(TypeError):
            aplicar_hierarquia_tipografica(None, paleta_azul)

    def test_aplicar_hierarquia_tipografica_paleta_none_raises(self, ws):
        with pytest.raises(TypeError):
            aplicar_hierarquia_tipografica(ws, None)

    def test_aplicar_hierarquia_tipografica_noop_nao_quebra(self, ws, paleta_azul):
        # Capacity é wrapper de conveniência; no P1 é no-op seguro
        aplicar_hierarquia_tipografica(ws, paleta_azul)

    def test_aplicar_nivel_dispatcher_valido(self, ws, paleta_azul):
        c = ws.cell(row=5, column=1, value="T")
        aplicar_nivel("titulo_aba", c, paleta_azul)
        assert c.font.bold is True

    def test_aplicar_nivel_invalido_raises(self, ws, paleta_azul):
        c = ws.cell(row=5, column=1)
        with pytest.raises(ValueError):
            aplicar_nivel("XPTO", c, paleta_azul)  # type: ignore[arg-type]

    def test_escrever_titulo_aba_retorna_proxima_linha(self, ws, paleta_azul):
        prox = escrever_titulo_aba(ws, 1, 1, 4, "Diagnóstico", paleta_azul)
        assert prox == 2

    def test_escrever_titulo_aba_sem_merge_quando_col_unica(self, ws, paleta_azul):
        prox = escrever_titulo_aba(ws, 1, 1, 1, "Único", paleta_azul)
        assert prox == 2

    def test_escrever_titulo_aba_col_fim_menor_raises(self, ws, paleta_azul):
        with pytest.raises(ValueError):
            escrever_titulo_aba(ws, 1, 3, 1, "X", paleta_azul)

    def test_escrever_secao_retorna_proxima_linha(self, ws, paleta_azul):
        prox = escrever_secao(ws, 5, 1, "Seção", paleta_azul)
        assert prox == 6

    def test_escrever_campo_valor_retorna_proxima_linha(self, ws, paleta_azul):
        prox = escrever_campo_valor(ws, 7, 1, 2, "Rótulo", "Valor", paleta_azul)
        assert prox == 8


# ===========================================================================
# Snapshot
# ===========================================================================

class TestSnapshotTipografia:
    def test_titulo_aba_usa_cor_primaria_da_paleta(self, ws, paleta_azul):
        c = ws.cell(row=1, column=1, value="T")
        aplicar_titulo_aba(c, paleta_azul)
        assert c.font.color.rgb.endswith(paleta_azul.cor_primaria)

    def test_titulo_aba_tamanho_conforme_paleta(self, ws, paleta_azul):
        c = ws.cell(row=1, column=1, value="T")
        aplicar_titulo_aba(c, paleta_azul)
        assert c.font.sz == paleta_azul.fonte_tamanho_titulo

    def test_secao_cor_de_destaque(self, ws, paleta_vinho):
        c = ws.cell(row=2, column=1, value="S")
        aplicar_secao(c, paleta_vinho)
        assert c.font.color.rgb.endswith(paleta_vinho.cor_destaque)

    def test_secao_tamanho_conforme_paleta(self, ws, paleta_verde):
        c = ws.cell(row=2, column=1, value="S")
        aplicar_secao(c, paleta_verde)
        assert c.font.sz == paleta_verde.fonte_tamanho_secao

    def test_campo_bold(self, ws, paleta_cinza):
        c = ws.cell(row=3, column=1, value="R")
        aplicar_campo(c, paleta_cinza)
        assert c.font.bold is True

    def test_campo_cor_neutra_escura(self, ws, paleta_cinza):
        c = ws.cell(row=3, column=1, value="R")
        aplicar_campo(c, paleta_cinza)
        assert c.font.color.rgb.endswith(paleta_cinza.cor_neutra_escura)

    def test_valor_regular_nao_bold(self, ws, paleta_azul):
        c = ws.cell(row=4, column=1, value="V")
        aplicar_valor(c, paleta_azul)
        assert c.font.bold is False

    def test_valor_cor_neutra_escura(self, ws, paleta_vinho):
        c = ws.cell(row=4, column=1, value="V")
        aplicar_valor(c, paleta_vinho)
        assert c.font.color.rgb.endswith(paleta_vinho.cor_neutra_escura)

    def test_todos_niveis_usam_fonte_familia_da_paleta(self, ws):
        for paleta in CATALOGO_PALETAS.values():
            for i, fn in enumerate(
                [aplicar_titulo_aba, aplicar_secao, aplicar_campo, aplicar_valor],
                start=1,
            ):
                c = ws.cell(row=100 + i, column=1, value="X")
                fn(c, paleta)
                assert c.font.name == paleta.fonte_familia, (
                    f"paleta {paleta.nome} / nível {fn.__name__} não respeitou fonte_familia"
                )

    def test_titulo_aba_merge_aplica_quando_col_fim_maior(self, ws, paleta_azul):
        escrever_titulo_aba(ws, 1, 1, 4, "Diagnóstico", paleta_azul)
        merges = {str(mr) for mr in ws.merged_cells.ranges}
        assert any("A1" in m and "D1" in m for m in merges)

    def test_escrever_campo_valor_estilos_aplicados(self, ws, paleta_azul):
        escrever_campo_valor(ws, 7, 1, 2, "Arquivo", "vendas.xlsx", paleta_azul)
        c_campo = ws.cell(row=7, column=1)
        c_valor = ws.cell(row=7, column=2)
        assert c_campo.font.bold is True
        assert c_valor.font.bold is False
        assert c_campo.value == "Arquivo"
        assert c_valor.value == "vendas.xlsx"

    def test_titulo_aba_altura_linha_definida(self, ws, paleta_azul):
        escrever_titulo_aba(ws, 1, 1, 4, "T", paleta_azul)
        assert ws.row_dimensions[1].height == 28


# ===========================================================================
# Regressão
# ===========================================================================

class TestRegressaoTipografia:
    def test_encadeamento_titulo_secao_campo_valor_sem_conflito(
        self, ws, paleta_azul,
    ):
        linha = 1
        linha = escrever_titulo_aba(ws, linha, 1, 4, "Aba X", paleta_azul)
        linha += 1
        linha = escrever_secao(ws, linha, 1, "Seção", paleta_azul)
        linha = escrever_campo_valor(ws, linha, 1, 2, "Campo", "Valor", paleta_azul)
        # Não deve levantar · estilos independentes
        assert linha > 1

    def test_helpers_nao_alteram_valor_da_celula(self, ws, paleta_azul):
        c = ws.cell(row=1, column=1, value="Texto Original")
        aplicar_valor(c, paleta_azul)
        assert c.value == "Texto Original"

    def test_reaplicacao_idempotente(self, ws, paleta_azul):
        c = ws.cell(row=1, column=1, value="X")
        aplicar_titulo_aba(c, paleta_azul)
        sz1 = c.font.sz
        aplicar_titulo_aba(c, paleta_azul)
        sz2 = c.font.sz
        assert sz1 == sz2

    def test_alignment_wrap_text_em_valor(self, ws, paleta_azul):
        c = ws.cell(row=1, column=1, value="X")
        aplicar_valor(c, paleta_azul)
        assert c.alignment.wrap_text is True

    def test_alignment_wrap_text_em_campo(self, ws, paleta_azul):
        c = ws.cell(row=1, column=1, value="X")
        aplicar_campo(c, paleta_azul)
        assert c.alignment.wrap_text is True


# ===========================================================================
# Invariantes
# ===========================================================================

class TestInvariantesTipografia:
    def test_ordem_tamanhos_titulo_secao_campo_valor(self, ws):
        for paleta in CATALOGO_PALETAS.values():
            cel_t = ws.cell(row=1, column=1, value="t"); aplicar_titulo_aba(cel_t, paleta)
            cel_s = ws.cell(row=2, column=1, value="s"); aplicar_secao(cel_s, paleta)
            cel_c = ws.cell(row=3, column=1, value="c"); aplicar_campo(cel_c, paleta)
            cel_v = ws.cell(row=4, column=1, value="v"); aplicar_valor(cel_v, paleta)
            assert cel_t.font.sz >= cel_s.font.sz >= cel_c.font.sz == cel_v.font.sz

    def test_todas_paletas_aceitam_4_niveis(self, ws):
        for paleta in CATALOGO_PALETAS.values():
            for fn in (aplicar_titulo_aba, aplicar_secao, aplicar_campo, aplicar_valor):
                c = ws.cell(row=5, column=1, value="x")
                fn(c, paleta)

    def test_secao_cor_nao_igual_a_valor_cor(self, ws):
        # Seção deve ter cor de destaque · valor tem neutra escura ·
        # garantir contraste hierárquico
        for paleta in CATALOGO_PALETAS.values():
            cel_s = ws.cell(row=1, column=1, value="s"); aplicar_secao(cel_s, paleta)
            cel_v = ws.cell(row=2, column=1, value="v"); aplicar_valor(cel_v, paleta)
            assert cel_s.font.color.rgb != cel_v.font.color.rgb, (
                f"paleta {paleta.nome}: seção e valor com mesma cor · hierarquia indistinguível"
            )
