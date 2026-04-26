"""
Testes F-APRESENT P1 · capability 10 · Diagnóstico narrativo em 6 seções.

4 tipos:
  - Interface (assinatura / raises)
  - Snapshot (6 seções presentes · nomes literais · ordem)
  - Regressão (vocabulário traduzido · None → —)
  - Invariantes (zero código técnico · zero enum caps · zero None literal)
"""
from __future__ import annotations

import re
from datetime import datetime
from typing import Any, Dict, List

import pytest
from openpyxl import Workbook

from contratos import (
    AjusteMotor,
    BloqueioOperacional,
    CategoriaWarning,
    DecisaoUsuario,
    DiagnosticoVN,
    IntegridadeEstrutural,
    WarningEstrutural,
)
from apresentacao import (
    CATALOGO_PALETAS,
    NOMES_SECOES,
    Paleta,
    carregar_vocabulario_bilingue,
    eh_termo_proibido,
    formatar_valor_ou_traco,
    renderizar_diagnostico,
)


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture
def paleta_azul() -> Paleta:
    return CATALOGO_PALETAS["azul"]


@pytest.fixture
def paleta_cinza() -> Paleta:
    return CATALOGO_PALETAS["cinza"]


@pytest.fixture
def paleta_verde() -> Paleta:
    return CATALOGO_PALETAS["verde"]


@pytest.fixture
def paleta_vinho() -> Paleta:
    return CATALOGO_PALETAS["vinho"]


@pytest.fixture
def vocabulario() -> Dict[str, Dict[str, str]]:
    return carregar_vocabulario_bilingue()


@pytest.fixture
def ws():
    wb = Workbook()
    return wb.active


@pytest.fixture
def config_usada_canonica() -> Dict[str, Any]:
    """config_usada canônica de uma execução V2 completa."""
    return {
        "arquivo": "vendas_2026.xlsx",
        "aba_consumida": "Plan1",
        "modo_base": "TRANSACIONAL",
        "agrupadores": ["Mes", "Loja"],
        "campo_analisado": "Vendas",
        "tipo_medida": "NUMERICO_ADITIVO",
        "colunas_mapeadas": {
            "origem_rotulo_tecnico": "Orçado",
            "comparado_rotulo_tecnico": "Realizado",
        },
        "paleta_aplicada": "azul",
        "thresholds_usados": {
            "limiar_estabilidade_pct": 0.05,
            "limite_variacao_extrema": 0.20,
        },
        "nulos_por_classificacao": {
            "NULO_ORIGEM": 3,
            "NULO_COMPARADO": 1,
        },
        "estados_nao_escolhidos": [],
    }


@pytest.fixture
def t_diag_canonico() -> DiagnosticoVN:
    return DiagnosticoVN(
        ajustes_aplicados=[
            AjusteMotor(
                tipo_ajuste="LINHA_EXCLUIDA_NULO",
                linhas_afetadas=4,
                descricao="4 linhas com chave ausente foram excluídas antes da análise.",
            ),
        ],
        warnings_por_categoria={},
        decisoes_usuario=[
            DecisaoUsuario(
                contexto="threshold_dominante_editado",
                valor_default=0.05,
                valor_escolhido=0.03,
            ),
        ],
        bloqueios_informativos=[],
        integridade=IntegridadeEstrutural(
            total_linhas_base_original=120,
            total_linhas_processadas=116,
            total_linhas_excluidas=4,
            motivo_exclusao_por_categoria={"NULO_CHAVE": 4},
        ),
    )


@pytest.fixture
def warnings_mistos() -> List[WarningEstrutural]:
    return [
        WarningEstrutural(
            codigo="W-B-INFO-01",
            categoria=CategoriaWarning.INFORMATIVO,
            microcopy="Arquivo carregado com sucesso.",
            contexto={"total": 116},
        ),
        WarningEstrutural(
            codigo="W-V2-AUSENCIA-01",
            categoria=CategoriaWarning.ALERTA_ESTRUTURAL,
            microcopy="Chaves presentes apenas em um dos lados.",
            contexto={"count": 8},
        ),
        WarningEstrutural(
            codigo="W-V2-NULO-01",
            categoria=CategoriaWarning.ALERTA_ESTRUTURAL_LEVE,
            microcopy="Valores ausentes no campo analisado.",
            contexto={"count": 3},
        ),
    ]


def _colher_textos_da_worksheet(ws) -> List[str]:
    """Percorre todas as células e coleta os valores string renderizados."""
    textos: List[str] = []
    for row in ws.iter_rows(values_only=True):
        for v in row:
            if isinstance(v, str):
                textos.append(v)
    return textos


# ===========================================================================
# Interface
# ===========================================================================

class TestInterfaceDiagnostico:
    def test_renderizar_assinatura_completa(
        self, ws, config_usada_canonica, t_diag_canonico, warnings_mistos,
        paleta_azul, vocabulario,
    ):
        renderizar_diagnostico(
            ws=ws,
            config_usada=config_usada_canonica,
            resolucao_estrutural=None,
            modelo_aplicado=None,
            t_diag=t_diag_canonico,
            warnings=warnings_mistos,
            paleta=paleta_azul,
            vocabulario=vocabulario,
        )

    def test_ws_none_raises(self, paleta_azul, vocabulario):
        with pytest.raises(TypeError):
            renderizar_diagnostico(
                None, {}, None, None, None, [], paleta_azul, vocabulario,
            )

    def test_paleta_none_raises(self, ws, vocabulario):
        with pytest.raises(TypeError):
            renderizar_diagnostico(
                ws, {}, None, None, None, [], None, vocabulario,
            )

    def test_vocabulario_none_raises(self, ws, paleta_azul):
        with pytest.raises(TypeError):
            renderizar_diagnostico(
                ws, {}, None, None, None, [], paleta_azul, None,
            )

    def test_warnings_none_nao_quebra(
        self, ws, config_usada_canonica, t_diag_canonico, paleta_azul, vocabulario,
    ):
        renderizar_diagnostico(
            ws, config_usada_canonica, None, None, t_diag_canonico,
            None, paleta_azul, vocabulario,
        )

    def test_t_diag_none_nao_quebra(
        self, ws, config_usada_canonica, paleta_azul, vocabulario,
    ):
        renderizar_diagnostico(
            ws, config_usada_canonica, None, None, None, [], paleta_azul, vocabulario,
        )

    def test_config_vazia_nao_quebra(
        self, ws, paleta_azul, vocabulario,
    ):
        renderizar_diagnostico(
            ws, {}, None, None, None, [], paleta_azul, vocabulario,
        )

    def test_config_como_objeto_aceito(
        self, ws, t_diag_canonico, paleta_azul, vocabulario,
    ):
        class ConfigObj:
            arquivo = "teste.xlsx"
            aba_consumida = "Plan1"
            modo_base = "TRANSACIONAL"
            agrupadores = ["X"]
            campo_analisado = "Valor"
            tipo_medida = "NUMERICO_ADITIVO"
            colunas_mapeadas = {}
            paleta_aplicada = "azul"
            thresholds_usados = {}
            nulos_por_classificacao = {}
            estados_nao_escolhidos = []

        renderizar_diagnostico(
            ws, ConfigObj(), None, None, t_diag_canonico, [],
            paleta_azul, vocabulario,
        )

    def test_nomes_secoes_canonico_6(self):
        assert len(NOMES_SECOES) == 6

    def test_formatar_valor_ou_traco_none(self):
        assert formatar_valor_ou_traco(None) == "—"

    def test_formatar_valor_ou_traco_origens_contextuais(self):
        assert formatar_valor_ou_traco(None, "ausencia") == "— (não consta)"
        assert formatar_valor_ou_traco(None, "nulo") == "— (sem valor)"
        assert formatar_valor_ou_traco(None, "impossivel") == "— (não calculável)"

    def test_formatar_valor_ou_traco_valor_presente(self):
        assert formatar_valor_ou_traco("Valor X") == "Valor X"
        assert formatar_valor_ou_traco(42) == "42"

    def test_formatar_valor_ou_traco_nan(self):
        assert formatar_valor_ou_traco(float("nan")) == "—"

    def test_formatar_valor_ou_traco_string_vazia(self):
        assert formatar_valor_ou_traco("") == "—"
        assert formatar_valor_ou_traco("   ") == "—"


# ===========================================================================
# Snapshot · nomes e ordem das 6 seções
# ===========================================================================

class TestSnapshotSecoes:
    def test_6_secoes_renderizadas_em_ordem(
        self, ws, config_usada_canonica, t_diag_canonico, warnings_mistos,
        paleta_azul, vocabulario,
    ):
        renderizar_diagnostico(
            ws, config_usada_canonica, None, None, t_diag_canonico,
            warnings_mistos, paleta_azul, vocabulario,
        )
        textos = _colher_textos_da_worksheet(ws)
        blob = "\n".join(textos)
        posicoes = [blob.find(n) for n in NOMES_SECOES]
        # Todas presentes
        assert all(p >= 0 for p in posicoes), f"seções faltando: {posicoes}"
        # Ordem preservada
        assert posicoes == sorted(posicoes)

    def test_titulo_aba_diagnostico(
        self, ws, config_usada_canonica, t_diag_canonico,
        paleta_azul, vocabulario,
    ):
        renderizar_diagnostico(
            ws, config_usada_canonica, None, None, t_diag_canonico,
            [], paleta_azul, vocabulario,
        )
        assert ws.cell(row=1, column=1).value == "Diagnóstico"

    def test_secao_1_como_foi_analisado_literal(
        self, ws, config_usada_canonica, t_diag_canonico,
        paleta_azul, vocabulario,
    ):
        renderizar_diagnostico(
            ws, config_usada_canonica, None, None, t_diag_canonico,
            [], paleta_azul, vocabulario,
        )
        assert "Como foi analisado" in _colher_textos_da_worksheet(ws)

    def test_secao_2_ajustes_do_motor_literal(
        self, ws, config_usada_canonica, t_diag_canonico,
        paleta_azul, vocabulario,
    ):
        renderizar_diagnostico(
            ws, config_usada_canonica, None, None, t_diag_canonico,
            [], paleta_azul, vocabulario,
        )
        assert "Ajustes do motor" in _colher_textos_da_worksheet(ws)

    def test_secao_3_pontos_de_atencao_literal(
        self, ws, config_usada_canonica, t_diag_canonico,
        paleta_azul, vocabulario,
    ):
        renderizar_diagnostico(
            ws, config_usada_canonica, None, None, t_diag_canonico,
            [], paleta_azul, vocabulario,
        )
        assert "Pontos de atenção" in _colher_textos_da_worksheet(ws)

    def test_secao_4_decisoes_do_usuario_literal(
        self, ws, config_usada_canonica, t_diag_canonico,
        paleta_azul, vocabulario,
    ):
        renderizar_diagnostico(
            ws, config_usada_canonica, None, None, t_diag_canonico,
            [], paleta_azul, vocabulario,
        )
        assert "Decisões do usuário" in _colher_textos_da_worksheet(ws)

    def test_secao_5_configuracoes_avancadas_literal(
        self, ws, config_usada_canonica, t_diag_canonico,
        paleta_azul, vocabulario,
    ):
        renderizar_diagnostico(
            ws, config_usada_canonica, None, None, t_diag_canonico,
            [], paleta_azul, vocabulario,
        )
        assert "Configurações avançadas aplicadas" in _colher_textos_da_worksheet(ws)

    def test_secao_6_qualidade_estrutural_literal(
        self, ws, config_usada_canonica, t_diag_canonico,
        paleta_azul, vocabulario,
    ):
        renderizar_diagnostico(
            ws, config_usada_canonica, None, None, t_diag_canonico,
            [], paleta_azul, vocabulario,
        )
        assert "Qualidade estrutural" in _colher_textos_da_worksheet(ws)


# ===========================================================================
# Snapshot · seções sem dados · frase explicativa C.2
# ===========================================================================

class TestSnapshotFrasesVazias:
    def test_secao_2_sem_ajustes_frase_explicativa(
        self, ws, paleta_azul, vocabulario,
    ):
        t_diag_vazio = DiagnosticoVN()
        renderizar_diagnostico(
            ws, {}, None, None, t_diag_vazio, [], paleta_azul, vocabulario,
        )
        textos = _colher_textos_da_worksheet(ws)
        assert any("Nenhum ajuste estrutural" in t for t in textos)

    def test_secao_3_sem_atencao_frase_explicativa(
        self, ws, paleta_azul, vocabulario,
    ):
        # warnings só informativos → nada para seção 3
        infos = [
            WarningEstrutural(
                codigo="W-INFO-01",
                categoria=CategoriaWarning.INFORMATIVO,
                microcopy="Info X",
            )
        ]
        t_diag_vazio = DiagnosticoVN()
        renderizar_diagnostico(
            ws, {}, None, None, t_diag_vazio, infos, paleta_azul, vocabulario,
        )
        textos = _colher_textos_da_worksheet(ws)
        assert any("Nenhum ponto de atenção" in t for t in textos)

    def test_secao_4_sem_decisoes_frase_explicativa(
        self, ws, paleta_azul, vocabulario,
    ):
        t_diag_vazio = DiagnosticoVN()
        renderizar_diagnostico(
            ws, {}, None, None, t_diag_vazio, [], paleta_azul, vocabulario,
        )
        textos = _colher_textos_da_worksheet(ws)
        assert any("configurações padrão" in t.lower() for t in textos)


# ===========================================================================
# Snapshot · traduções aplicadas
# ===========================================================================

class TestSnapshotTraducoes:
    def test_modo_base_traduzido(
        self, ws, config_usada_canonica, t_diag_canonico, paleta_azul, vocabulario,
    ):
        renderizar_diagnostico(
            ws, config_usada_canonica, None, None, t_diag_canonico,
            [], paleta_azul, vocabulario,
        )
        textos = _colher_textos_da_worksheet(ws)
        blob = "\n".join(textos)
        # Modo TRANSACIONAL traduzido
        assert "transacional" in blob.lower()

    def test_tipo_medida_traduzido(
        self, ws, config_usada_canonica, t_diag_canonico, paleta_azul, vocabulario,
    ):
        renderizar_diagnostico(
            ws, config_usada_canonica, None, None, t_diag_canonico,
            [], paleta_azul, vocabulario,
        )
        textos = _colher_textos_da_worksheet(ws)
        blob = "\n".join(textos)
        # NUMERICO_ADITIVO → "Valor somável..."
        assert "somável" in blob.lower() or "somavel" in blob.lower()

    def test_paleta_selecionada_exibida_user_facing(
        self, ws, config_usada_canonica, t_diag_canonico, paleta_azul, vocabulario,
    ):
        renderizar_diagnostico(
            ws, config_usada_canonica, None, None, t_diag_canonico,
            [], paleta_azul, vocabulario,
        )
        textos = _colher_textos_da_worksheet(ws)
        blob = "\n".join(textos)
        assert "Azul executivo" in blob

    def test_thresholds_como_percentual(
        self, ws, config_usada_canonica, t_diag_canonico, paleta_azul, vocabulario,
    ):
        renderizar_diagnostico(
            ws, config_usada_canonica, None, None, t_diag_canonico,
            [], paleta_azul, vocabulario,
        )
        textos = _colher_textos_da_worksheet(ws)
        blob = "\n".join(textos)
        # 0.05 vira "5,00%"
        assert re.search(r"5,00\s*%", blob)

    def test_categoria_warning_traduzida(
        self, ws, config_usada_canonica, t_diag_canonico, warnings_mistos,
        paleta_azul, vocabulario,
    ):
        renderizar_diagnostico(
            ws, config_usada_canonica, None, None, t_diag_canonico,
            warnings_mistos, paleta_azul, vocabulario,
        )
        textos = _colher_textos_da_worksheet(ws)
        blob = "\n".join(textos)
        assert "Informativo" in blob or "Ajuste automático" in blob


# ===========================================================================
# Snapshot · resolução estrutural + modelo aplicado
# ===========================================================================

class TestSnapshotResolucaoEModelo:
    def test_resolucao_estrutural_renderizada(
        self, ws, config_usada_canonica, t_diag_canonico, paleta_azul, vocabulario,
    ):
        from visoes.visao_v2 import ResolucaoEstruturalV2
        resolucao = ResolucaoEstruturalV2(
            tipo_caso="NIVEL_AGRUPAMENTO_DIFERENTE",
            opcoes_oferecidas=["opção A", "opção B"],
            escolha_usuario="opção A",
            contexto_caso={},
        )
        renderizar_diagnostico(
            ws, config_usada_canonica, resolucao, None, t_diag_canonico,
            [], paleta_azul, vocabulario,
        )
        textos = _colher_textos_da_worksheet(ws)
        blob = "\n".join(textos)
        assert "Resolução estrutural" in blob
        assert "opção A" in blob

    def test_modelo_aplicado_renderizado(
        self, ws, config_usada_canonica, t_diag_canonico, paleta_azul, vocabulario,
    ):
        from visoes.visao_v2 import ModeloAplicadoV2
        modelo = ModeloAplicadoV2(
            nome_modelo="Modelo Vendas Q2",
            data_criacao_modelo=datetime(2026, 1, 10),
            campos_casados=5,
            campos_nao_casados=[],
            tipo_aplicacao="COMPLETA",
        )
        renderizar_diagnostico(
            ws, config_usada_canonica, None, modelo, t_diag_canonico,
            [], paleta_azul, vocabulario,
        )
        textos = _colher_textos_da_worksheet(ws)
        blob = "\n".join(textos)
        assert "Modelo Vendas Q2" in blob
        assert "Totalmente compatível" in blob


# ===========================================================================
# Regressão
# ===========================================================================

class TestRegressaoDiagnostico:
    def test_nao_gera_excecao_com_v2result_canonico(
        self, ws, config_usada_canonica, t_diag_canonico, warnings_mistos,
        paleta_azul, vocabulario,
    ):
        # regressão · chamada com argumentos completos não pode quebrar
        renderizar_diagnostico(
            ws, config_usada_canonica, None, None, t_diag_canonico,
            warnings_mistos, paleta_azul, vocabulario,
        )

    def test_funciona_em_todas_as_4_paletas(
        self, config_usada_canonica, t_diag_canonico, warnings_mistos, vocabulario,
    ):
        for paleta in CATALOGO_PALETAS.values():
            wb = Workbook()
            ws = wb.active
            renderizar_diagnostico(
                ws, config_usada_canonica, None, None, t_diag_canonico,
                warnings_mistos, paleta, vocabulario,
            )

    def test_categoria_warning_como_enum_ou_string_equivalente(
        self, ws, paleta_azul, vocabulario,
    ):
        # Warning com categoria Enum
        w_enum = WarningEstrutural(
            codigo="W-X-01",
            categoria=CategoriaWarning.ALERTA_ESTRUTURAL,
            microcopy="Teste enum",
        )
        renderizar_diagnostico(
            ws, {}, None, None, None, [w_enum], paleta_azul, vocabulario,
        )
        textos = _colher_textos_da_worksheet(ws)
        assert any("Teste enum" in t for t in textos)

    def test_warning_sem_microcopy_nao_quebra(
        self, ws, paleta_azul, vocabulario,
    ):
        w = WarningEstrutural(
            codigo="W-X-01",
            categoria=CategoriaWarning.ALERTA_ESTRUTURAL,
            microcopy="",
        )
        renderizar_diagnostico(
            ws, {}, None, None, None, [w], paleta_azul, vocabulario,
        )

    def test_multiplos_warnings_mesma_categoria_agregam(
        self, ws, paleta_azul, vocabulario,
    ):
        ws_lista = [
            WarningEstrutural(
                codigo=f"W-X-{i:02d}",
                categoria=CategoriaWarning.ALERTA_ESTRUTURAL,
                microcopy=f"Ocorrência {i}",
            )
            for i in range(3)
        ]
        renderizar_diagnostico(
            ws, {}, None, None, None, ws_lista, paleta_azul, vocabulario,
        )
        textos = _colher_textos_da_worksheet(ws)
        blob = "\n".join(textos)
        # Agrega com "3 ocorrências · exemplo:"
        assert "3 ocorrências" in blob


# ===========================================================================
# Invariantes · bloco 7 e bloco 8 do vocabulario_bilingue.md v2
# ===========================================================================

def _coletar_strings_diagnostico(
    config_usada, resolucao, modelo, t_diag, warnings, paleta, vocabulario,
) -> List[str]:
    wb = Workbook()
    ws = wb.active
    renderizar_diagnostico(
        ws, config_usada, resolucao, modelo, t_diag, warnings,
        paleta, vocabulario,
    )
    return _colher_textos_da_worksheet(ws)


class TestInvariantesBloco7:
    def test_zero_codigos_d_xxx(
        self, config_usada_canonica, t_diag_canonico, warnings_mistos,
        paleta_azul, vocabulario,
    ):
        textos = _coletar_strings_diagnostico(
            config_usada_canonica, None, None, t_diag_canonico,
            warnings_mistos, paleta_azul, vocabulario,
        )
        padrao = re.compile(r"\bD-\d{2,}\b")
        for t in textos:
            assert not padrao.search(t), f"código D-XXX vazou: {t!r}"

    def test_zero_codigos_b_v2(
        self, config_usada_canonica, t_diag_canonico, warnings_mistos,
        paleta_azul, vocabulario,
    ):
        textos = _coletar_strings_diagnostico(
            config_usada_canonica, None, None, t_diag_canonico,
            warnings_mistos, paleta_azul, vocabulario,
        )
        padrao = re.compile(r"\bB-V\d+-[A-Z]+")
        for t in textos:
            assert not padrao.search(t), f"código B-V*-* vazou: {t!r}"

    def test_zero_codigos_w_v2(
        self, config_usada_canonica, t_diag_canonico, warnings_mistos,
        paleta_azul, vocabulario,
    ):
        textos = _coletar_strings_diagnostico(
            config_usada_canonica, None, None, t_diag_canonico,
            warnings_mistos, paleta_azul, vocabulario,
        )
        padrao = re.compile(r"\bW-V\d+-[A-Z]+")
        for t in textos:
            assert not padrao.search(t), f"código W-V*-* vazou: {t!r}"

    def test_zero_codigos_t_xxx(
        self, config_usada_canonica, t_diag_canonico, warnings_mistos,
        paleta_azul, vocabulario,
    ):
        textos = _coletar_strings_diagnostico(
            config_usada_canonica, None, None, t_diag_canonico,
            warnings_mistos, paleta_azul, vocabulario,
        )
        padrao = re.compile(r"\bT-[A-Z]{3,}\b")
        for t in textos:
            assert not padrao.search(t), f"código T-* vazou: {t!r}"

    def test_zero_codigos_f_xxx(
        self, config_usada_canonica, t_diag_canonico, warnings_mistos,
        paleta_azul, vocabulario,
    ):
        textos = _coletar_strings_diagnostico(
            config_usada_canonica, None, None, t_diag_canonico,
            warnings_mistos, paleta_azul, vocabulario,
        )
        padrao = re.compile(r"\bF-[A-Z]{3,}\b")
        for t in textos:
            assert not padrao.search(t), f"código F-* vazou: {t!r}"

    def test_zero_enums_caps_conhecidos(
        self, config_usada_canonica, t_diag_canonico, warnings_mistos,
        paleta_azul, vocabulario,
    ):
        textos = _coletar_strings_diagnostico(
            config_usada_canonica, None, None, t_diag_canonico,
            warnings_mistos, paleta_azul, vocabulario,
        )
        proibidos = [
            "POR_COLUNAS", "POR_LINHAS",
            "PRESENTE_AMBOS", "AUSENTE_ORIGEM", "AUSENTE_COMPARADO",
            "NULO_ORIGEM", "NULO_COMPARADO", "NULO_AMBOS",
            "TRANSACIONAL", "PRE_AGREGADO",
            "NUMERICO_ADITIVO", "NUMERICO_RELATIVO",
            "ESTADO_SITUACAO",
            "NIVEL_AGRUPAMENTO_DIFERENTE", "COLUNA_PRESENTE_EM_UM_LADO",
            "LINHA_EXCLUIDA_NULO", "INTERVALO_AJUSTADO_INICIO",
        ]
        blob = "\n".join(textos)
        for p in proibidos:
            assert p not in blob, f"enum técnico '{p}' vazou para superfície cliente"

    def test_zero_none_literal(
        self, config_usada_canonica, t_diag_canonico, warnings_mistos,
        paleta_azul, vocabulario,
    ):
        textos = _coletar_strings_diagnostico(
            config_usada_canonica, None, None, t_diag_canonico,
            warnings_mistos, paleta_azul, vocabulario,
        )
        for t in textos:
            # Aceita "None" só se for substring de palavra maior (ex.: "Nonentidade"
            # não existe em vocabulário · nada aceita). Regex ensures word boundary.
            assert not re.search(r"\bNone\b", t), f"'None' literal vazou: {t!r}"
            assert not re.search(r"\bnull\b", t), f"'null' literal vazou: {t!r}"
            assert not re.search(r"\bNaN\b", t), f"'NaN' literal vazou: {t!r}"

    def test_zero_dict_python_serializado(
        self, config_usada_canonica, t_diag_canonico, warnings_mistos,
        paleta_azul, vocabulario,
    ):
        textos = _coletar_strings_diagnostico(
            config_usada_canonica, None, None, t_diag_canonico,
            warnings_mistos, paleta_azul, vocabulario,
        )
        padrao = re.compile(r"\{\s*['\"]\w+['\"]\s*:")
        for t in textos:
            assert not padrao.search(t), f"dict Python vazou: {t!r}"

    def test_zero_datetime_cru(
        self, config_usada_canonica, t_diag_canonico, warnings_mistos,
        paleta_azul, vocabulario,
    ):
        textos = _coletar_strings_diagnostico(
            config_usada_canonica, None, None, t_diag_canonico,
            warnings_mistos, paleta_azul, vocabulario,
        )
        for t in textos:
            assert "datetime.datetime(" not in t, f"datetime cru vazou: {t!r}"

    def test_zero_termos_proibidos_bloco7(
        self, config_usada_canonica, t_diag_canonico, warnings_mistos,
        paleta_azul, vocabulario,
    ):
        textos = _coletar_strings_diagnostico(
            config_usada_canonica, None, None, t_diag_canonico,
            warnings_mistos, paleta_azul, vocabulario,
        )
        for t in textos:
            viol = eh_termo_proibido(t)
            assert viol is None, f"termo proibido detectado em {t!r}: {viol}"


class TestInvariantesBloco8:
    def test_none_origem_ausencia_microcopy(self):
        assert formatar_valor_ou_traco(None, "ausencia") == "— (não consta)"

    def test_none_origem_nulo_microcopy(self):
        assert formatar_valor_ou_traco(None, "nulo") == "— (sem valor)"

    def test_none_origem_impossivel_microcopy(self):
        assert formatar_valor_ou_traco(None, "impossivel") == "— (não calculável)"

    def test_config_sem_arquivo_renderiza_traco(
        self, ws, paleta_azul, vocabulario,
    ):
        # config sem arquivo · seção 1 deve exibir "— (não consta)"
        renderizar_diagnostico(
            ws, {}, None, None, None, [], paleta_azul, vocabulario,
        )
        textos = _colher_textos_da_worksheet(ws)
        blob = "\n".join(textos)
        assert "— (não consta)" in blob or "—" in blob
