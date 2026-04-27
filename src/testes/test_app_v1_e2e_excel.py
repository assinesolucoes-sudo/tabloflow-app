"""
test_app_v1_e2e_excel.py — E2E completo · A-V1 Fase 8 · gate duplo Camada 1.

Cobre:
  - Pipeline real upload → motor_base → executar_v1 → exportar_resultado_v1
  - 2 ramos lógicos (ABAS_DISTINTAS · MESMA_ABA cobertura placeholder)
  - 4 paletas regenerando Excel
  - Estrutura de 6 abas validada com openpyxl
  - Validação cruzada com casos_esperados.yaml entrada V1 (V1-A01..V1-A12)
"""
from __future__ import annotations

import tempfile
from decimal import Decimal
from pathlib import Path

import pytest
import yaml
from openpyxl import load_workbook

from motor_base import processar_base
from motor_upload import ArquivoEntrada, processar_upload
from visoes.exportacao_v1 import exportar_resultado_v1
from visoes.visao_v1 import (
    AgrupadorMatchV1,
    CampoComparadoV1,
    CasoLogicoV1,
    ClassificacaoRegistroV1,
    ModoMatchV1,
    StatusPonteV1,
    TipoCampoV1,
    UnidadeCanonica,
    executar_v1,
)


BASE_PATH = str(Path(__file__).resolve().parents[2] / "bases" / "base_fundacao.xlsx")
YAML_PATH = str(Path(__file__).resolve().parents[2] / "bases" / "casos_esperados.yaml")


def _config_canonica_v1():
    """Config canônica V1 baseada em casos_esperados.yaml entrada V1."""
    return {
        "agrupadores_match": [
            AgrupadorMatchV1(
                nome_origem="Conta", nome_comparado="Conta",
                rotulo_analitico="Conta", modo_match=ModoMatchV1.EXATO,
            ),
            AgrupadorMatchV1(
                nome_origem="Centro_Custo", nome_comparado="Centro_Custo",
                rotulo_analitico="Centro de Custo", modo_match=ModoMatchV1.EXATO,
            ),
        ],
        "campos_comparados": [
            CampoComparadoV1(
                nome_origem="Valor", nome_comparado="Valor",
                nome_analitico="Valor",
                tipo_logico=TipoCampoV1.VALOR_MONETARIO,
                unidade=UnidadeCanonica.MONETARIO_BRL,
                tolerancia=Decimal("0.01"),
            ),
        ],
        "agrupadores_executivos": ["Centro_Custo"],
        "epsilon_por_unidade": {UnidadeCanonica.MONETARIO_BRL: Decimal("0.01")},
        "thresholds": {},
        "origem_ux": "Razão", "comparado_ux": "Balancete",
        "arquivo_origem": "base_fundacao.xlsx", "arquivo_comparado": "base_fundacao.xlsx",
        "aba_origem": "dual_origem_crm", "aba_comparado": "dual_comparado_erp",
        "n_arquivos": 1, "paleta_aplicada": "Azul executivo",
    }


@pytest.fixture(scope="module")
def pipeline_completo_canonico():
    """Roda pipeline canônico end-to-end (motor_upload + motor_base + executar_v1)."""
    upload_result = processar_upload(
        [
            ArquivoEntrada(
                caminho_fisico=BASE_PATH,
                caminho_logico="origem",
                aba_solicitada="dual_origem_crm",
            ),
            ArquivoEntrada(
                caminho_fisico=BASE_PATH,
                caminho_logico="comparado",
                aba_solicitada="dual_comparado_erp",
            ),
        ],
        modo="DUAL",
    )
    motor_result = processar_base(upload_result)
    v1_result = executar_v1(motor_result, _config_canonica_v1())
    return upload_result, motor_result, v1_result


# ---------------------------------------------------------------------------
# TestE2ECompleto · 6 testes críticos
# ---------------------------------------------------------------------------


class TestE2ECompleto:
    def test_e2e_abas_distintas_pipeline_completo(self, pipeline_completo_canonico):
        _, motor_result, v1_result = pipeline_completo_canonico
        assert v1_result is not None
        assert v1_result.conciliacao_realizada.caso_logico_inferido == CasoLogicoV1.ABAS_DISTINTAS
        assert v1_result.conciliacao_realizada.n_processados > 0

    def test_e2e_excel_gerado_6_abas(self, pipeline_completo_canonico, tmp_path):
        _, _, v1_result = pipeline_completo_canonico
        caminho = str(tmp_path / "e2e_6abas.xlsx")
        res = exportar_resultado_v1(
            v1_result=v1_result, caminho_saida=caminho, paleta_nome="azul",
        )
        wb = load_workbook(res.caminho_arquivo)
        assert len(wb.sheetnames) == 6
        # Aba 6 sempre Diagnóstico (D-017)
        assert wb.sheetnames[-1] == "Diagnóstico"

    def test_e2e_paleta_troca_4_paletas(self, pipeline_completo_canonico, tmp_path):
        _, _, v1_result = pipeline_completo_canonico
        for paleta in ["azul", "verde", "cinza", "vinho"]:
            caminho = str(tmp_path / f"e2e_{paleta}.xlsx")
            res = exportar_resultado_v1(
                v1_result=v1_result, caminho_saida=caminho, paleta_nome=paleta,
            )
            assert Path(res.caminho_arquivo).exists()
            assert res.tamanho_bytes > 5000

    def test_e2e_assertions_yaml_v1(self, pipeline_completo_canonico):
        """Valida que os totals batem com casos_esperados.yaml entrada V1."""
        _, _, v1_result = pipeline_completo_canonico
        with open(YAML_PATH, "r", encoding="utf-8") as f:
            casos = yaml.safe_load(f)
        v1_assertions = casos["visoes"]["V1"]["assertions"]
        contagem = v1_result.contagem_por_classificacao

        # V1-A02 · SO_ORIGEM faixa empírica (34-38)
        a02 = next((a for a in v1_assertions if a["id"] == "V1-A02"), None)
        assert a02 is not None
        n_so_o = contagem.get(ClassificacaoRegistroV1.SO_ORIGEM, 0)
        assert a02["esperado"]["min"] <= n_so_o <= a02["esperado"]["max"], (
            f"V1-A02 · SO_ORIGEM={n_so_o} fora da faixa "
            f"[{a02['esperado']['min']}, {a02['esperado']['max']}]"
        )

        # V1-A06 · DIVERGENCIA_AMBIGUIDADE em modo EXATO == 0
        a06 = next((a for a in v1_assertions if a["id"] == "V1-A06"), None)
        assert a06 is not None
        n_amb = contagem.get(ClassificacaoRegistroV1.DIVERGENCIA_AMBIGUIDADE, 0)
        assert n_amb == a06["esperado"]["valor"], (
            f"V1-A06 · DIVERGENCIA_AMBIGUIDADE={n_amb} esperado {a06['esperado']['valor']}"
        )

        # V1-A11 · Status Ponte FECHA
        a11 = next((a for a in v1_assertions if a["id"] == "V1-A11"), None)
        assert a11 is not None
        assert v1_result.status_ponte_geral == StatusPonteV1.FECHA, (
            f"V1-A11 · Status Ponte={v1_result.status_ponte_geral.value} "
            f"esperado FECHA"
        )

    def test_e2e_diagnostico_aba_existe_e_tem_conteudo(self, pipeline_completo_canonico, tmp_path):
        _, _, v1_result = pipeline_completo_canonico
        caminho = str(tmp_path / "e2e_diag.xlsx")
        exportar_resultado_v1(
            v1_result=v1_result, caminho_saida=caminho, paleta_nome="azul",
        )
        wb = load_workbook(caminho)
        ws_diag = wb["Diagnóstico"]
        # Diagnóstico tem pelo menos 6 seções · cada uma com pelo menos 2 linhas
        # estimativa conservadora: max_row >= 20
        assert ws_diag.max_row >= 20

    def test_e2e_button_confirmar_e1_ok_processa_motor(self, tmp_path):
        """E2E exercitando o botão Confirmar do E1_OK · que invoca
        _processar_bases_pos_e1_ok (cobre paths de upload físico)."""
        from streamlit.testing.v1 import AppTest
        APP_PATH = str(Path(__file__).resolve().parents[1] / "app_v1.py")
        at = AppTest.from_file(APP_PATH, default_timeout=30)
        at.run()
        with open(BASE_PATH, "rb") as f:
            bytes_base = f.read()
        # Simula pós-upload caso1 · 2 abas (Caso 2 D-213)
        at.session_state["n_arquivos"] = 1
        at.session_state["upload_unico_bytes"] = bytes_base
        at.session_state["upload_unico_nome"] = "base_fundacao.xlsx"
        at.session_state["abas_origem_disponiveis"] = [
            "dual_origem_crm", "dual_comparado_erp",
        ]
        at.session_state["abas_comparado_disponiveis"] = [
            "dual_origem_crm", "dual_comparado_erp",
        ]
        at.session_state["abas_escolhidas_caso1_2abas"] = [
            "dual_origem_crm", "dual_comparado_erp",
        ]
        at.session_state["aba_escolhida_unica_caso1"] = None
        at.session_state["etapa"] = "E1_OK"
        at.run()
        # Clica Confirmar · roda motor_upload + motor_base internamente
        at.button(key="btn_confirmar_e1_ok").click().run()
        assert at.session_state["etapa"] == "E2"
        assert at.session_state["motor_result"] is not None
        assert at.session_state["motor_result"].modo_upload == "DUAL"

    def test_e2e_resumo_executivo_tem_taxa_conciliacao(
        self, pipeline_completo_canonico, tmp_path
    ):
        _, _, v1_result = pipeline_completo_canonico
        caminho = str(tmp_path / "e2e_taxa.xlsx")
        exportar_resultado_v1(
            v1_result=v1_result, caminho_saida=caminho, paleta_nome="azul",
        )
        wb = load_workbook(caminho)
        ws_re = wb["Resumo Executivo"]
        # Procurar a célula com "Taxa de Conciliação Geral" (rótulo do card)
        encontrou = False
        for row in ws_re.iter_rows(min_row=1, max_row=15, values_only=True):
            for val in row:
                if val and "Taxa de Conciliação" in str(val):
                    encontrou = True
                    break
            if encontrou:
                break
        assert encontrou, "Card 'Taxa de Conciliação' não encontrado no Resumo Executivo"
