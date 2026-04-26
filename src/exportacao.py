"""
exportacao.py — Exportação Excel · Fundação TabloFlow · F-EXP · D-136
11 capabilities em 4 camadas topológicas · API pública exportar_resultado()

Estrutura de abas invariante cross-visão (§EXP.1):
  1ª  · Resumo Executivo  (D-125 · 6 blocos fixos)
  2ª  · Coração Visual    (D-126 · nome específico por visão)
  N   · Abas analíticas específicas (Fase 2 · visão_vN.py adiciona)
  N+1 · Base Analítica    (C.D3 · D-124)
  N+2 · Parâmetros        (penúltima)
  Última · Diagnóstico    (D-017 · invariante transversal)

Decisões técnicas (§EXP.6):
  - openpyxl 3.1+ exclusivo · xlsxwriter descartado
  - ws.append() para DataFrames grandes (> LIMITE_WRITE_EFFICIENT linhas)
    em vez de write_only mode (incompatível com gráficos/formatação no mesmo wb)
  - Gráficos: try/except → W-EXP-GRAFICO-FALHA · dados sempre preservados
  - Formatação condicional: limite 64000 regras observado · W-EXP-FORMATACAO-TRUNCADA
"""
from __future__ import annotations

import math
import os
import time
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from contratos import (
    CategoriaWarning,
    ConfigExportacao,
    DiagnosticoVN,
    ExportacaoResult,
    ResumoExecutivoPadrao,
    VNResultBase,
    WarningEstrutural,
)

# ---------------------------------------------------------------------------
# Constantes
# ---------------------------------------------------------------------------

EXCEL_MAX_CELL_LEN = 32_767
EXCEL_MAX_FORMAT_RULES = 64_000
LIMITE_PAGINACAO = 30               # linhas E colunas acima deste → CAP-PAGINACAO-MATRIZ
LIMITE_WRITE_EFFICIENT = 10_000     # usar ws.append() eficiente acima deste limite
LIMITE_BYTES_GRANDE = 50 * 1_024 * 1_024  # 50 MB → W-EXP-ARQUIVO-GRANDE
LIMITE_SEGUNDOS_GRANDE = 30.0

# Tipos de CoracaoVisualRef que representam matrizes paginável
_TIPOS_PAGINAVEL = {"MATRIZ_COLORIDA", "TABELA_HEATMAP", "COLUMN_EMPILHADO_100"}

# Estilos compartilhados
_H_FONT = Font(bold=True, color="FFFFFFFF", size=11)
_H_FILL = PatternFill(fill_type="solid", fgColor="2F4F8F")
_H_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_TITULO_FONT = Font(bold=True, color="2F4F8F", size=12)
_SUBTITULO_FONT = Font(bold=True, size=11)
_BORDER_THIN = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
_ALT_FILL = PatternFill(fill_type="solid", fgColor="EEF2F8")


# ---------------------------------------------------------------------------
# Utilitários internos
# ---------------------------------------------------------------------------

def _emit_warning(
    codigo: str,
    categoria: CategoriaWarning,
    microcopy: str,
    **contexto: Any,
) -> WarningEstrutural:
    return WarningEstrutural(
        codigo=codigo,
        categoria=categoria,
        microcopy=microcopy,
        contexto=dict(contexto),
    )


def _safe_val(
    val: Any,
    col_name: str,
    row_num: int,
    warnings_out: List[WarningEstrutural],
) -> Any:
    """Trunca strings > 32767 chars e emite W-EXP-COLUNA-TRUNCADA."""
    if isinstance(val, str) and len(val) > EXCEL_MAX_CELL_LEN:
        warnings_out.append(_emit_warning(
            "W-EXP-COLUNA-TRUNCADA",
            CategoriaWarning.ALERTA_ESTRUTURAL,
            (
                f"Coluna '{col_name}' linha {row_num}: valor truncado de "
                f"{len(val)} para {EXCEL_MAX_CELL_LEN} caracteres."
            ),
            coluna=col_name,
            linha=row_num,
            comprimento_original=len(val),
        ))
        return val[:EXCEL_MAX_CELL_LEN]
    return val


def _escrever_dataframe(
    ws: Worksheet,
    df: pd.DataFrame,
    warnings_out: List[WarningEstrutural],
) -> Tuple[int, int]:
    """
    Escreve DataFrame no ws usando ws.append() (eficiente para bases grandes).
    O header é anexado à próxima linha disponível após conteúdo existente.
    Tipos numéricos e bool são preservados; NaN → None.
    Retorna (n_linhas_dado, n_colunas).
    """
    n_rows, n_cols = len(df), len(df.columns)

    # Header
    ws.append([str(c) for c in df.columns])

    # Linhas de dado
    for row_num, row in enumerate(df.itertuples(index=False, name=None), start=1):
        processed: List[Any] = []
        for col_name, val in zip(df.columns, row):
            safe = _safe_val(val, str(col_name), row_num, warnings_out)
            if isinstance(safe, float) and math.isnan(safe):
                processed.append(None)
            else:
                processed.append(safe)
        ws.append(processed)

    return n_rows, n_cols


def _aplicar_estilos_header(ws: Worksheet, row: int, num_cols: int) -> None:
    for col in range(1, num_cols + 1):
        cell = ws.cell(row, col)
        cell.font = _H_FONT
        cell.fill = _H_FILL
        cell.alignment = _H_ALIGN
        cell.border = _BORDER_THIN


def _aplicar_estilos_dados(
    ws: Worksheet,
    row_inicio: int,
    row_fim: int,
    num_cols: int,
) -> None:
    for row in range(row_inicio, row_fim + 1):
        fill = _ALT_FILL if row % 2 == 0 else None
        for col in range(1, num_cols + 1):
            cell = ws.cell(row, col)
            cell.border = _BORDER_THIN
            cell.alignment = Alignment(vertical="center")
            if fill:
                cell.fill = fill


def _ajustar_larguras(ws: Worksheet, num_cols: int, max_width: int = 50) -> None:
    for col in range(1, num_cols + 1):
        col_letter = get_column_letter(col)
        max_len = max(
            (len(str(cell.value)) for cell in ws[col_letter] if cell.value is not None),
            default=8,
        )
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 8), max_width)


# ---------------------------------------------------------------------------
# CAMADA 1 · Universais
# ---------------------------------------------------------------------------

def cap_tabela_formatada(
    ws: Worksheet,
    num_colunas: int,
    config: ConfigExportacao,
    header_row: int = 1,
    total_rows: int = 0,
) -> None:
    """
    CAP-01 · Formata tabela com estilos profissionais:
    cabeçalho bold+cor · bordas · fill alternado · largura auto · freeze_panes.
    Consumida por toda aba tabular.
    """
    _aplicar_estilos_header(ws, header_row, num_colunas)

    if total_rows > 0:
        _aplicar_estilos_dados(ws, header_row + 1, header_row + total_rows, num_colunas)

    if config.largura_colunas_auto:
        _ajustar_larguras(ws, num_colunas)

    if config.congelar_cabecalho:
        ws.freeze_panes = ws.cell(header_row + 1, 1)


def cap_autofilter(
    ws: Worksheet,
    num_colunas: int,
    header_row: int = 1,
) -> None:
    """
    CAP-02 · Ativa AutoFilter em todas as abas tabulares.
    Default True via ConfigExportacao · invariante cross-visão (§EXP.1).
    """
    if num_colunas > 0:
        col_fim = get_column_letter(num_colunas)
        ws.auto_filter.ref = f"A{header_row}:{col_fim}{header_row}"


def cap_diagnostico(
    ws: Worksheet,
    diagnostico: DiagnosticoVN,
    config: ConfigExportacao,
) -> None:
    """
    CAP-03 · Aba Diagnóstico · sempre última (D-017).
    Renderiza DiagnosticoVN em 5 seções: Ajustes · Warnings · Decisões ·
    Bloqueios Escapados · Integridade Estrutural.
    """
    row = 1

    def _secao(titulo: str) -> None:
        nonlocal row
        cell = ws.cell(row, 1, titulo)
        cell.font = _SUBTITULO_FONT
        cell.fill = PatternFill(fill_type="solid", fgColor="D9E2F0")
        row += 1

    def _kv(chave: str, valor: Any) -> None:
        nonlocal row
        ws.cell(row, 1, chave).font = Font(bold=True)
        ws.cell(row, 2, str(valor) if not isinstance(valor, (int, float, bool)) else valor)
        row += 1

    def _br() -> None:
        nonlocal row
        row += 1

    # Seção 1 · Ajustes
    _secao("Seção 1 · Ajustes Aplicados pelo Motor")
    if diagnostico.ajustes_aplicados:
        for col, label in enumerate(["Tipo Ajuste", "Linhas Afetadas", "Descrição"], 1):
            ws.cell(row, col, label)
        _aplicar_estilos_header(ws, row, 3)
        row += 1
        for aj in diagnostico.ajustes_aplicados:
            ws.cell(row, 1, aj.tipo_ajuste)
            ws.cell(row, 2, aj.linhas_afetadas)
            ws.cell(row, 3, aj.descricao)
            row += 1
    else:
        _kv("(sem ajustes)", "")
    _br()

    # Seção 2 · Warnings por categoria
    _secao("Seção 2 · Warnings por Categoria")
    total_w = sum(len(v) for v in diagnostico.warnings_por_categoria.values())
    if total_w > 0:
        for col, label in enumerate(["Código", "Categoria", "Mensagem", "Contexto"], 1):
            ws.cell(row, col, label)
        _aplicar_estilos_header(ws, row, 4)
        row += 1
        for categoria, lista in diagnostico.warnings_por_categoria.items():
            for w in lista:
                ws.cell(row, 1, w.codigo)
                ws.cell(row, 2, categoria)
                ws.cell(row, 3, w.microcopy)
                ws.cell(row, 4, str(w.contexto) if w.contexto else "")
                row += 1
    else:
        _kv("(sem warnings)", "")
    _br()

    # Seção 3 · Decisões do usuário
    _secao("Seção 3 · Decisões do Usuário")
    if diagnostico.decisoes_usuario:
        for col, label in enumerate(["Contexto", "Default", "Escolhido", "Justificativa"], 1):
            ws.cell(row, col, label)
        _aplicar_estilos_header(ws, row, 4)
        row += 1
        for d in diagnostico.decisoes_usuario:
            ws.cell(row, 1, d.contexto)
            ws.cell(row, 2, str(d.valor_default))
            ws.cell(row, 3, str(d.valor_escolhido))
            ws.cell(row, 4, d.justificativa_opcional or "")
            row += 1
    else:
        _kv("(sem decisões do usuário)", "")
    _br()

    # Seção 4 · Bloqueios escapados
    _secao("Seção 4 · Bloqueios com Escape Acionado")
    if diagnostico.bloqueios_informativos:
        for col, label in enumerate(["Código", "Condição", "Escape", "Warning Pós-Escape"], 1):
            ws.cell(row, col, label)
        _aplicar_estilos_header(ws, row, 4)
        row += 1
        for b in diagnostico.bloqueios_informativos:
            ws.cell(row, 1, b.codigo)
            ws.cell(row, 2, b.condicao_disparo)
            ws.cell(row, 3, str(b.escape_acionado))
            ws.cell(row, 4, b.warning_pos_escape or "")
            row += 1
    else:
        _kv("(sem bloqueios escapados)", "")
    _br()

    # Seção 5 · Integridade
    _secao("Seção 5 · Integridade Estrutural")
    if diagnostico.integridade:
        integ = diagnostico.integridade
        _kv("Total linhas originais", integ.total_linhas_base_original)
        _kv("Total linhas processadas", integ.total_linhas_processadas)
        _kv("Total linhas excluídas", integ.total_linhas_excluidas)
        for motivo, qtd in integ.motivo_exclusao_por_categoria.items():
            _kv(f"  → {motivo}", qtd)
    else:
        _kv("(sem dados de integridade)", "")

    if config.largura_colunas_auto:
        _ajustar_larguras(ws, 4)


def cap_resumo_executivo(
    ws: Worksheet,
    resumo: ResumoExecutivoPadrao,
    visao_id: str,
    config: ConfigExportacao,
) -> None:
    """
    CAP-04 · Aba Resumo Executivo · sempre primeira (D-125).
    6 blocos fixos na ordem canônica.
    Supressão de bloco inteiro proibida (D-125).
    """
    row = 1
    MAX_COLS = 3

    def _titulo(label: str) -> None:
        nonlocal row
        cell = ws.cell(row, 1, label)
        cell.font = _TITULO_FONT
        cell.fill = PatternFill(fill_type="solid", fgColor="D9E2F0")
        ws.merge_cells(
            start_row=row, start_column=1, end_row=row, end_column=MAX_COLS
        )
        row += 1

    def _kv(chave: str, valor: Any) -> None:
        nonlocal row
        ws.cell(row, 1, chave).font = Font(bold=True)
        ws.cell(row, 2, str(valor) if not isinstance(valor, (int, float, bool)) else valor)
        row += 1

    def _br() -> None:
        nonlocal row
        row += 1

    # ── Bloco 1 · Cabeçalho ──────────────────────────────────────────────
    _titulo("Bloco 1 · Cabeçalho da Execução")
    cab = resumo.bloco_1_cabecalho
    _kv("Visão", cab.visao)
    _kv("Data de Execução", cab.data_execucao.strftime("%d/%m/%Y %H:%M:%S"))
    _kv("Modo Upload", cab.modo_upload)
    _kv("Agrupadores", ", ".join(cab.agrupadores) if cab.agrupadores else "(nenhum)")
    _kv("Medida Principal", cab.medida_principal or "(não definida)")
    _br()

    # ── Bloco 2 · Números-âncora ─────────────────────────────────────────
    _titulo("Bloco 2 · Números-Âncora")
    for chave, valor in resumo.bloco_2_numeros_ancora.items():
        _kv(chave, valor)
    _br()

    # ── Bloco 3 · Distribuição ───────────────────────────────────────────
    _titulo("Bloco 3 · Distribuição")
    for chave, valor in resumo.bloco_3_distribuicao.items():
        if isinstance(valor, dict):
            _kv(chave, "(detalhe)")
            for sk, sv in valor.items():
                _kv(f"  · {sk}", sv)
        elif isinstance(valor, list):
            _kv(chave, f"({len(valor)} itens)")
        else:
            _kv(chave, valor)
    _br()

    # ── Bloco 4 · Elementos destacados ───────────────────────────────────
    _titulo("Bloco 4 · Elementos Destacados")
    for chave, valor in resumo.bloco_4_elementos_destacados.items():
        if isinstance(valor, list):
            _kv(chave, f"({len(valor)} itens)")
            for item in valor[:10]:
                _kv("  →", str(item))
        elif isinstance(valor, dict):
            _kv(chave, "(detalhe)")
            for sk, sv in valor.items():
                _kv(f"  · {sk}", sv)
        else:
            _kv(chave, valor)
    _br()

    # ── Bloco 5 · Leitura qualitativa ────────────────────────────────────
    _titulo("Bloco 5 · Leitura Qualitativa")
    lq = resumo.bloco_5_leitura_qualitativa
    _kv("Classificação Ativa", lq.classificacao_ativa)
    _kv("Leitura Alterada por Edição", "Sim" if lq.alguma_leitura_alterada_por_edicao else "Não")
    for th_nome, th_val in lq.thresholds_usados.items():
        _kv(f"Threshold · {th_nome}", th_val)
    _br()

    # ── Bloco 6 · Qualidade estrutural ───────────────────────────────────
    _titulo("Bloco 6 · Qualidade Estrutural (BAD · C.D3)")
    qe = resumo.bloco_6_qualidade_estrutural
    _kv("Total de Warnings", qe.total_warnings)
    _kv("Ajustes Aplicados pelo Motor", qe.ajustes_aplicados)
    _kv("Tem Bloqueios Escapados", "Sim" if qe.tem_bloqueios_escapados else "Não")
    for cat, qtd in qe.warnings_por_categoria.items():
        _kv(f"  Warnings ({cat})", qtd)

    if config.largura_colunas_auto:
        _ajustar_larguras(ws, MAX_COLS)


# ---------------------------------------------------------------------------
# CAMADA 2 · Gráficos nativos simples
# ---------------------------------------------------------------------------

def cap_barchart_nativo(
    ws: Worksheet,
    n_linhas_dado: int,
    n_colunas: int,
    titulo: str = "Composição",
) -> Optional[BarChart]:
    """
    CAP-05 · BarChart nativo openpyxl.
    Assume dados em ws: linha 1 = cabeçalhos (ou header_row calculado),
    col 1 = rótulos (categorias), cols 2..n = valores de séries.
    Âncora: coluna n_colunas+2.
    Consumida por V4 · V10 (auxiliar) · V5 (auxiliar) · V7 (auxiliar).
    """
    if n_linhas_dado == 0 or n_colunas < 2:
        return None

    # Detectar a linha real do header (última linha do header + 1 para dados)
    # Convenção: header em max_row - n_linhas_dado
    header_row_real = (ws.max_row or (n_linhas_dado + 1)) - n_linhas_dado
    data_end_row = header_row_real + n_linhas_dado

    chart = BarChart()
    chart.type = "col"
    chart.title = titulo
    chart.y_axis.title = "Valor"
    chart.x_axis.title = "Categoria"
    chart.shape = 4
    chart.height = 14
    chart.width = 20

    cats = Reference(ws, min_col=1, min_row=header_row_real + 1, max_row=data_end_row)
    chart.set_categories(cats)

    for col_idx in range(2, n_colunas + 1):
        data = Reference(
            ws, min_col=col_idx,
            min_row=header_row_real, max_row=data_end_row,
        )
        chart.add_data(data, titles_from_data=True)

    anchor_col = get_column_letter(n_colunas + 2)
    ws.add_chart(chart, f"{anchor_col}1")
    return chart


def cap_linechart_nativo(
    ws: Worksheet,
    n_linhas_dado: int,
    n_colunas: int,
    titulo: str = "Evolução",
) -> Optional[LineChart]:
    """
    CAP-06 · LineChart nativo openpyxl.
    Mesma convenção de dados de cap_barchart_nativo.
    Consumida por V10 (linha acumulada) · V8 (presença ao longo do tempo) · V3.
    """
    if n_linhas_dado == 0 or n_colunas < 2:
        return None

    header_row_real = (ws.max_row or (n_linhas_dado + 1)) - n_linhas_dado
    data_end_row = header_row_real + n_linhas_dado

    chart = LineChart()
    chart.title = titulo
    chart.y_axis.title = "Valor"
    chart.x_axis.title = "Período"
    chart.height = 14
    chart.width = 20

    cats = Reference(ws, min_col=1, min_row=header_row_real + 1, max_row=data_end_row)
    chart.set_categories(cats)

    for col_idx in range(2, n_colunas + 1):
        data = Reference(
            ws, min_col=col_idx,
            min_row=header_row_real, max_row=data_end_row,
        )
        chart.add_data(data, titles_from_data=True)

    anchor_col = get_column_letter(n_colunas + 2)
    ws.add_chart(chart, f"{anchor_col}1")
    return chart


def cap_histograma_bins(
    ws: Worksheet,
    df: pd.DataFrame,
    coluna_numerica: Optional[str],
    titulo: str = "Distribuição de Frequência",
) -> Optional[BarChart]:
    """
    CAP-07 · Histograma com bins Sturges (k = ⌈1 + log₂(n)⌉, mínimo 3).
    Escreve bins em área separada do ws (após dados existentes).
    Consumida exclusivamente por V5 (Mapa de Distribuição · D-108).
    """
    if df.empty:
        return None

    # Identificar coluna numérica
    if coluna_numerica and coluna_numerica in df.columns:
        serie = df[coluna_numerica].dropna()
    else:
        serie = next(
            (df[c].dropna() for c in df.columns if pd.api.types.is_numeric_dtype(df[c])),
            None,
        )
    if serie is None or len(serie) < 3:
        return None

    # Sturges
    n = len(serie)
    k = max(3, math.ceil(1 + math.log2(n)))
    mn, mx = float(serie.min()), float(serie.max())
    if mn == mx:
        return None
    bin_width = (mx - mn) / k

    bins: List[Tuple[str, int]] = []
    for i in range(k):
        lower = mn + i * bin_width
        upper = mn + (i + 1) * bin_width
        mask = (serie >= lower) & (serie <= upper if i == k - 1 else serie < upper)
        bins.append((f"{lower:.2f}–{upper:.2f}", int(mask.sum())))

    # Área de bins: após os dados existentes no ws
    offset_col = (ws.max_column or 0) + 2

    ws.cell(1, offset_col, "Intervalo")
    ws.cell(1, offset_col + 1, "Frequência")
    for i, (label, count) in enumerate(bins, 2):
        ws.cell(i, offset_col, label)
        ws.cell(i, offset_col + 1, count)

    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = titulo
    chart.y_axis.title = "Frequência"
    chart.x_axis.title = "Intervalo"
    chart.height = 14
    chart.width = 20

    cats = Reference(ws, min_col=offset_col, min_row=2, max_row=k + 1)
    chart.set_categories(cats)
    data = Reference(ws, min_col=offset_col + 1, min_row=1, max_row=k + 1)
    chart.add_data(data, titles_from_data=True)

    anchor_col = get_column_letter(offset_col + 3)
    ws.add_chart(chart, f"{anchor_col}1")
    return chart


# ---------------------------------------------------------------------------
# CAMADA 3 · Gráficos compostos e formatação avançada
# ---------------------------------------------------------------------------

def cap_combo_bar_line(
    ws: Worksheet,
    n_linhas_dado: int,
    n_colunas: int,
    titulo: str = "Curva Pareto",
) -> None:
    """
    CAP-08 · Combo BarChart + LineChart com eixo Y secundário.
    Col 1 = rótulos, col 2 = valores (barras, eixo primário),
    col 3 = acumulado % (linha, eixo secundário).
    Linhas de referência a 80% e 95% adicionadas como séries constantes.
    Consumida exclusivamente por V10 (Curva Pareto · Coração Visual).
    """
    if n_linhas_dado == 0 or n_colunas < 2:
        return

    # Colunas de referência (80% e 95%) após dados existentes
    ref_col_80 = n_colunas + 2
    ref_col_95 = n_colunas + 3
    ws.cell(1, ref_col_80, "Ref 80%")
    ws.cell(1, ref_col_95, "Ref 95%")
    for r in range(2, n_linhas_dado + 2):
        ws.cell(r, ref_col_80, 80.0)
        ws.cell(r, ref_col_95, 95.0)

    header_row_real = (ws.max_row - n_linhas_dado) or 1
    data_end_row = header_row_real + n_linhas_dado

    # BarChart (valores, eixo primário)
    bar = BarChart()
    bar.type = "col"
    bar.title = titulo
    bar.y_axis.title = "Valor"
    bar.x_axis.title = "Elemento"
    bar.height = 15
    bar.width = 25

    cats = Reference(ws, min_col=1, min_row=header_row_real + 1, max_row=data_end_row)
    bar.set_categories(cats)
    bar_data = Reference(ws, min_col=2, min_row=header_row_real, max_row=data_end_row)
    bar.add_data(bar_data, titles_from_data=True)

    # LineChart (acumulado % + referências, eixo secundário)
    line = LineChart()
    line.y_axis.axId = 200
    line.y_axis.title = "Acumulado %"
    line.y_axis.crosses = "max"

    if n_colunas >= 3:
        acum = Reference(ws, min_col=3, min_row=header_row_real, max_row=data_end_row)
        line.add_data(acum, titles_from_data=True)

    ref80 = Reference(ws, min_col=ref_col_80, min_row=1, max_row=n_linhas_dado + 1)
    ref95 = Reference(ws, min_col=ref_col_95, min_row=1, max_row=n_linhas_dado + 1)
    line.add_data(ref80, titles_from_data=True)
    line.add_data(ref95, titles_from_data=True)

    bar += line

    anchor_col = get_column_letter(n_colunas + 5)
    ws.add_chart(bar, f"{anchor_col}1")


def cap_formatacao_condicional_matriz(
    ws: Worksheet,
    df: pd.DataFrame,
    tipo_formatacao: str,
    warnings_out: List[WarningEstrutural],
) -> int:
    """
    CAP-09 · Formatação condicional de matriz.
    tipo_formatacao:
      "COLOR_SCALE"      → ColorScaleRule vermelho-amarelo-verde (V9 Mapa de Perfil)
      "FORMULA_DISCRETA" → FormulaRule presença/ausência (V6 densidade · V8 presença)
    Limite Excel 64000 regras observado · emite W-EXP-FORMATACAO-TRUNCADA.
    Retorna número de regras aplicadas.
    """
    if df.empty:
        return 0

    n_rows, n_cols = df.shape

    # Coluna 1 = rótulos → formatação começa na col 2
    data_start_col = 2
    data_end_col = n_cols  # total cols incluindo col-rótulo

    # Linhas de dado começam na linha 2 (linha 1 = header)
    data_start_row = 2
    data_end_row = n_rows + 1

    total_cells = (data_end_row - data_start_row + 1) * (data_end_col - data_start_col + 1)

    if total_cells > EXCEL_MAX_FORMAT_RULES:
        warnings_out.append(_emit_warning(
            "W-EXP-FORMATACAO-TRUNCADA",
            CategoriaWarning.ALERTA_ESTRUTURAL,
            (
                f"Formatação condicional truncada: {total_cells} células excedem o "
                f"limite Excel de {EXCEL_MAX_FORMAT_RULES} regras. "
                "Aplicadas parcialmente."
            ),
            total_celulas=total_cells,
            limite=EXCEL_MAX_FORMAT_RULES,
        ))
        # Truncar: limitar colunas
        max_cols_usavel = EXCEL_MAX_FORMAT_RULES // (data_end_row - data_start_row + 1)
        data_end_col = data_start_col + max_cols_usavel - 1

    start_ref = f"{get_column_letter(data_start_col)}{data_start_row}"
    end_ref = f"{get_column_letter(data_end_col)}{data_end_row}"
    cell_range = f"{start_ref}:{end_ref}"

    if tipo_formatacao == "COLOR_SCALE":
        rule = ColorScaleRule(
            start_type="min", start_color="F8696B",
            mid_type="percentile", mid_value=50, mid_color="FFEB84",
            end_type="max", end_color="63BE7B",
        )
        ws.conditional_formatting.add(cell_range, rule)
    else:
        # FORMULA_DISCRETA: célula preenchida → verde
        rule = FormulaRule(
            formula=[f'{get_column_letter(data_start_col)}{data_start_row}<>""'],
            fill=PatternFill(fill_type="solid", fgColor="C6EFCE"),
        )
        ws.conditional_formatting.add(cell_range, rule)

    return min(total_cells, EXCEL_MAX_FORMAT_RULES)


def cap_columnchart_empilhado_100(
    ws: Worksheet,
    n_linhas_dado: int,
    n_colunas: int,
    titulo: str = "Matriz de Cruzamento",
) -> Optional[BarChart]:
    """
    CAP-10 · ColumnChart empilhado 100%.
    Alternativa declarada ao heatmap real não suportado em openpyxl (D-118).
    Consumida exclusivamente por V6 (Coração Visual · Matriz de Cruzamento).
    """
    if n_linhas_dado == 0 or n_colunas < 2:
        return None

    header_row_real = (ws.max_row or (n_linhas_dado + 1)) - n_linhas_dado
    data_end_row = header_row_real + n_linhas_dado

    chart = BarChart()
    chart.type = "col"
    chart.grouping = "percentStacked"
    chart.overlap = 100
    chart.title = titulo
    chart.y_axis.title = "% da Distribuição"
    chart.x_axis.title = "Categoria"
    chart.height = 14
    chart.width = 24

    cats = Reference(ws, min_col=1, min_row=header_row_real + 1, max_row=data_end_row)
    chart.set_categories(cats)

    for col_idx in range(2, n_colunas + 1):
        data = Reference(
            ws, min_col=col_idx,
            min_row=header_row_real, max_row=data_end_row,
        )
        chart.add_data(data, titles_from_data=True)

    anchor_col = get_column_letter(n_colunas + 2)
    ws.add_chart(chart, f"{anchor_col}1")
    return chart


# ---------------------------------------------------------------------------
# CAMADA 4 · Orquestração cross-capability
# ---------------------------------------------------------------------------

def cap_paginacao_matriz(
    wb: Workbook,
    df: pd.DataFrame,
    nome_base: str,
    warnings_out: List[WarningEstrutural],
    config: ConfigExportacao,
) -> List[str]:
    """
    CAP-11 · Paginação de matriz grande (> 30 linhas × 30 colunas).
    Gera abas secundárias com notas de navegação.
    Emite W-EXP-MATRIZ-GRANDE-PAGINADA.
    Consumida por V6 (Matriz de Cruzamento > 30×30) e V8 (Presença > 30×30).
    Retorna lista de nomes de abas criadas.
    """
    n_rows, n_cols = df.shape
    if n_rows <= LIMITE_PAGINACAO and n_cols <= LIMITE_PAGINACAO:
        return []

    paginas_row = math.ceil(n_rows / LIMITE_PAGINACAO)
    paginas_col = math.ceil(n_cols / LIMITE_PAGINACAO)
    total_paginas = paginas_row * paginas_col

    if total_paginas <= 1:
        return []

    abas_criadas: List[str] = []
    pagina_num = 1

    for pr in range(paginas_row):
        for pc in range(paginas_col):
            r0 = pr * LIMITE_PAGINACAO
            r1 = min((pr + 1) * LIMITE_PAGINACAO, n_rows)
            c0 = pc * LIMITE_PAGINACAO
            c1 = min((pc + 1) * LIMITE_PAGINACAO, n_cols)

            df_page = df.iloc[r0:r1, c0:c1]
            nome_aba = f"{nome_base[:25]} ({pagina_num})"
            ws_page = wb.create_sheet(nome_aba)
            abas_criadas.append(nome_aba)

            # Linha 1: nota de navegação
            ws_page.cell(1, 1, (
                f"[Parte {pagina_num}/{total_paginas}] "
                f"Linhas {r0 + 1}–{r1} · Colunas {c0 + 1}–{c1}"
            )).font = Font(italic=True, color="888888")

            # ws.max_row == 1 → _escrever_dataframe vai para linha 2
            n_rows_p, n_cols_p = _escrever_dataframe(ws_page, df_page, warnings_out)
            header_row_p = 2  # append colocou header na linha 2
            cap_tabela_formatada(ws_page, n_cols_p, config, header_row=header_row_p, total_rows=n_rows_p)
            if config.aplicar_filtros_automaticos:
                cap_autofilter(ws_page, n_cols_p, header_row=header_row_p)

            pagina_num += 1

    warnings_out.append(_emit_warning(
        "W-EXP-MATRIZ-GRANDE-PAGINADA",
        CategoriaWarning.INFORMATIVO,
        (
            f"Matriz '{nome_base}' ({n_rows}×{n_cols}) excede 30×30 e foi "
            f"paginada em {total_paginas} abas secundárias."
        ),
        nome_base=nome_base,
        dimensoes=f"{n_rows}x{n_cols}",
        total_paginas=total_paginas,
        abas_geradas=abas_criadas,
    ))

    return abas_criadas


# ---------------------------------------------------------------------------
# Helpers internos de exportação
# ---------------------------------------------------------------------------

def _escrever_parametros(
    ws: Worksheet,
    config_usada: Dict[str, Any],
    config: ConfigExportacao,
) -> None:
    ws.cell(1, 1, "Parâmetro")
    ws.cell(1, 2, "Valor")
    _aplicar_estilos_header(ws, 1, 2)
    for row_idx, (chave, valor) in enumerate(config_usada.items(), 2):
        ws.cell(row_idx, 1, str(chave))
        val = str(valor) if not isinstance(valor, (int, float, bool)) else valor
        ws.cell(row_idx, 2, val)
    if config.largura_colunas_auto:
        _ajustar_larguras(ws, 2)
    if config.congelar_cabecalho:
        ws.freeze_panes = "A2"


def _render_coracao_visual(
    wb: Workbook,
    ws_cv: Worksheet,
    resultado: VNResultBase,
    config: ConfigExportacao,
    capabilities_out: List[str],
    warnings_out: List[WarningEstrutural],
) -> None:
    """
    Renderiza o Coração Visual em 4 etapas independentes:
    1. Tabela + AutoFilter (sempre)
    2. Formatação condicional (controlado por aplicar_formatacao_condicional)
    3. Gráficos nativos (controlado por incluir_graficos_nativos)
    4. Paginação de matriz grande (controlado por paginar_matrizes_grandes + tipo)
    """
    df = resultado.base_analitica
    tipo = resultado.coracao_visual.tipo
    nome = resultado.coracao_visual.nome_aba
    caps_req = resultado.coracao_visual.capabilities_requeridas

    # ── Etapa 1: Tabela + AutoFilter (sempre) ─────────────────────────────
    n_rows, n_cols = _escrever_dataframe(ws_cv, df, warnings_out)
    cap_tabela_formatada(ws_cv, n_cols, config, total_rows=n_rows)
    if "CAP-TABELA-FORMATADA" not in capabilities_out:
        capabilities_out.append("CAP-TABELA-FORMATADA")

    if config.aplicar_filtros_automaticos:
        cap_autofilter(ws_cv, n_cols)
        if "CAP-AUTOFILTER" not in capabilities_out:
            capabilities_out.append("CAP-AUTOFILTER")

    # ── Etapa 2: Formatação condicional (independente de gráficos) ────────
    if config.aplicar_formatacao_condicional:
        if tipo == "MATRIZ_COLORIDA":
            cap_formatacao_condicional_matriz(ws_cv, df, "COLOR_SCALE", warnings_out)
            if "CAP-FORMATACAO-CONDICIONAL-MATRIZ" not in capabilities_out:
                capabilities_out.append("CAP-FORMATACAO-CONDICIONAL-MATRIZ")
        elif tipo == "TABELA_HEATMAP":
            cap_formatacao_condicional_matriz(ws_cv, df, "FORMULA_DISCRETA", warnings_out)
            if "CAP-FORMATACAO-CONDICIONAL-MATRIZ" not in capabilities_out:
                capabilities_out.append("CAP-FORMATACAO-CONDICIONAL-MATRIZ")
        elif tipo == "COLUMN_EMPILHADO_100" and "formatacao_condicional" in caps_req:
            cap_formatacao_condicional_matriz(ws_cv, df, "FORMULA_DISCRETA", warnings_out)
            if "CAP-FORMATACAO-CONDICIONAL-MATRIZ" not in capabilities_out:
                capabilities_out.append("CAP-FORMATACAO-CONDICIONAL-MATRIZ")

    # ── Etapa 3: Gráficos nativos ─────────────────────────────────────────
    if config.incluir_graficos_nativos:
        try:
            if tipo == "BARRAS":
                chart = cap_barchart_nativo(ws_cv, n_rows, n_cols, titulo=nome)
                if chart is not None:
                    capabilities_out.append("CAP-BARCHART-NATIVO")

            elif tipo == "BARRAS_LINHA":
                cap_combo_bar_line(ws_cv, n_rows, n_cols, titulo=nome)
                capabilities_out.append("CAP-COMBO-BAR-LINE")

            elif tipo == "COLUMN_EMPILHADO_100":
                chart = cap_columnchart_empilhado_100(ws_cv, n_rows, n_cols, titulo=nome)
                if chart is not None:
                    capabilities_out.append("CAP-COLUMNCHART-EMPILHADO-100")

            elif tipo == "HISTOGRAMA":
                col_num = next(
                    (c for c in df.columns if pd.api.types.is_numeric_dtype(df[c])),
                    None,
                )
                chart = cap_histograma_bins(ws_cv, df, col_num, titulo=nome)
                if chart is not None:
                    capabilities_out.append("CAP-HISTOGRAMA-BINS")

        except Exception as exc:
            warnings_out.append(_emit_warning(
                "W-EXP-GRAFICO-FALHA",
                CategoriaWarning.ALERTA_ESTRUTURAL,
                (
                    f"Capability de gráfico falhou para '{nome}' ({tipo}): {exc!s}. "
                    "Aba gerada sem gráfico · dados preservados em tabela."
                ),
                nome_aba=nome,
                tipo_visual=tipo,
                erro=str(exc),
            ))

    # ── Etapa 4: Paginação de matriz grande (independente de gráficos) ────
    if config.paginar_matrizes_grandes and tipo in _TIPOS_PAGINAVEL:
        if n_rows > LIMITE_PAGINACAO and n_cols > LIMITE_PAGINACAO:
            abas = cap_paginacao_matriz(wb, df, nome, warnings_out, config)
            if abas and "CAP-PAGINACAO-MATRIZ" not in capabilities_out:
                capabilities_out.append("CAP-PAGINACAO-MATRIZ")


# ---------------------------------------------------------------------------
# API pública
# ---------------------------------------------------------------------------

def exportar_resultado(
    resultado_visao: VNResultBase,
    caminho_saida: str,
    configuracao_exportacao: Optional[ConfigExportacao] = None,
) -> ExportacaoResult:
    """
    Materializa VNResultBase em arquivo .xlsx.
    Invocado pelo app_vN.py quando o usuário clica em "Exportar Excel".

    Estrutura invariante de abas (§EXP.1):
      1ª  Resumo Executivo  (D-125 · 6 blocos fixos)
      2ª  Coração Visual    (D-126 · nome de CoracaoVisualRef.nome_aba)
      N   Base Analítica    (C.D3 · D-124)
      N+1 Parâmetros        (penúltima)
      Última Diagnóstico    (D-017 · invariante transversal)

    Retorna ExportacaoResult com auditoria de capabilities e warnings (C.2).
    """
    t0 = time.monotonic()
    config = configuracao_exportacao or ConfigExportacao()
    capabilities_acionadas: List[str] = []
    warnings_gerados: List[WarningEstrutural] = []

    # Garantir diretório de saída
    os.makedirs(os.path.dirname(os.path.abspath(caminho_saida)), exist_ok=True)

    wb = Workbook()
    # Remove Sheet padrão gerado pelo Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # D-171/D-168 · capability 1 de F-APRESENT · marca a paleta selecionada
    # no workbook como marcação rastreável (P0 · não reformata abas já escritas).
    # Fonte da verdade da paleta: ConfigExportacao.paleta (default 'azul').
    # Normalização tolerante · aceita 'Azul' ou 'azul' vindos do widget do app.
    from apresentacao import CATALOGO_PALETAS, aplicar_paleta as _aplicar_paleta
    _nome_paleta_normalizado = (config.paleta or "azul").strip().lower()
    if _nome_paleta_normalizado not in CATALOGO_PALETAS:
        raise ValueError(
            f"ConfigExportacao.paleta='{config.paleta}' não está no catálogo "
            f"canônico F-APRESENT (D-164) · opções: "
            f"{sorted(CATALOGO_PALETAS.keys())}"
        )
    _paleta_obj = CATALOGO_PALETAS[_nome_paleta_normalizado]
    _aplicar_paleta(wb, _paleta_obj)
    capabilities_acionadas.append("CAP-APLICAR-PALETA")

    # ── Aba 1: Resumo Executivo (SEMPRE PRIMEIRA · D-125) ─────────────────
    ws_resumo = wb.create_sheet("Resumo Executivo")
    cap_resumo_executivo(
        ws_resumo, resultado_visao.resumo_executivo, resultado_visao.visao_id, config
    )
    capabilities_acionadas.append("CAP-RESUMO-EXECUTIVO")

    # ── Aba 2: Coração Visual (SEMPRE SEGUNDA · D-126) ──────────────────
    nome_cv = resultado_visao.coracao_visual.nome_aba
    ws_cv = wb.create_sheet(nome_cv)
    _render_coracao_visual(wb, ws_cv, resultado_visao, config, capabilities_acionadas, warnings_gerados)

    # ── Base Analítica (C.D3 · D-124 · antes de Parâmetros) ─────────────
    ws_ba = wb.create_sheet("Base Analítica")
    n_rows_ba, n_cols_ba = _escrever_dataframe(ws_ba, resultado_visao.base_analitica, warnings_gerados)
    cap_tabela_formatada(ws_ba, n_cols_ba, config, total_rows=n_rows_ba)
    if "CAP-TABELA-FORMATADA" not in capabilities_acionadas:
        capabilities_acionadas.append("CAP-TABELA-FORMATADA")
    if config.aplicar_filtros_automaticos:
        cap_autofilter(ws_ba, n_cols_ba)
        if "CAP-AUTOFILTER" not in capabilities_acionadas:
            capabilities_acionadas.append("CAP-AUTOFILTER")

    # ── Parâmetros (SEMPRE PENÚLTIMA) ─────────────────────────────────────
    ws_params = wb.create_sheet("Parâmetros")
    _escrever_parametros(ws_params, resultado_visao.config_usada, config)

    # ── Diagnóstico (SEMPRE ÚLTIMA · D-017) ──────────────────────────────
    ws_diag = wb.create_sheet("Diagnóstico")

    # D-172 · router · cap_diagnostico legado vs capability 10 F-APRESENT P1
    # (D-165 · 6 seções narrativas user-facing). A visão produz o adaptador
    # via A-VN e achata os campos em config_usada["_config_diagnostico"]
    # (convenção A-V2 refatorada · S4). Visões que ainda não criaram
    # adaptador caem no fallback legado · continuam funcionando.
    _config_diag_achatado = resultado_visao.config_usada.get(
        "_config_diagnostico"
    )

    if _config_diag_achatado is not None:
        # Caminho novo · capability 10 D-165 · 6 seções user-facing
        from apresentacao import (
            carregar_vocabulario_bilingue as _carregar_vocabulario,
        )
        from apresentacao.diagnostico_narrativo import (
            renderizar_diagnostico as _renderizar_diagnostico,
        )

        _vocabulario = _carregar_vocabulario()
        _resolucao = _config_diag_achatado.get("resolucao_estrutural")
        _modelo = _config_diag_achatado.get("modelo_aplicado")
        _renderizar_diagnostico(
            ws_diag,
            _config_diag_achatado,
            _resolucao,
            _modelo,
            resultado_visao.diagnostico,
            resultado_visao.warnings,
            _paleta_obj,
            _vocabulario,
        )
        capabilities_acionadas.append("CAP-DIAGNOSTICO-V10")
    else:
        # Fallback · cap_diagnostico legado · preserva visões antigas
        # (V1, V11, V3..V10 ainda não criaram seu adaptador)
        cap_diagnostico(ws_diag, resultado_visao.diagnostico, config)
        capabilities_acionadas.append("CAP-DIAGNOSTICO")

    # Capturar número de abas antes de salvar
    numero_abas = len(wb.sheetnames)

    wb.save(caminho_saida)

    t1 = time.monotonic()
    tempo = round(t1 - t0, 4)
    tamanho = os.path.getsize(caminho_saida)

    # ── W-EXP-ARQUIVO-GRANDE ─────────────────────────────────────────────
    if tamanho > LIMITE_BYTES_GRANDE or tempo > LIMITE_SEGUNDOS_GRANDE:
        warnings_gerados.append(_emit_warning(
            "W-EXP-ARQUIVO-GRANDE",
            CategoriaWarning.INFORMATIVO,
            (
                f"Arquivo gerado é grande: {tamanho / (1024 * 1024):.1f} MB · "
                f"{tempo:.1f}s de geração. Considere usar filtros para reduzir o volume."
            ),
            tamanho_bytes=tamanho,
            tempo_segundos=tempo,
        ))

    return ExportacaoResult(
        caminho_arquivo=caminho_saida,
        tamanho_bytes=tamanho,
        numero_abas=numero_abas,
        tempo_geracao_segundos=tempo,
        warnings_gerados=warnings_gerados,
        capabilities_acionadas=capabilities_acionadas,
    )
