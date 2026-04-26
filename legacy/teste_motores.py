"""
Bloco V-0 — Validação dos Motores
Testa motor_upload.py e motor_base.py sem arquivos reais de disco.
Cada caso retorna True (passou) ou False (falhou).
"""
from __future__ import annotations

import io
import sys
import os
import traceback
from datetime import date
from typing import Any, Optional

import inspect
import tempfile
import sys as _sys_enc
if hasattr(_sys_enc.stdout, "reconfigure"):
    _sys_enc.stdout.reconfigure(encoding="utf-8", errors="replace")
if hasattr(_sys_enc.stderr, "reconfigure"):
    _sys_enc.stderr.reconfigure(encoding="utf-8", errors="replace")
import openpyxl
from openpyxl import Workbook
import pandas as pd

# ── Importação dos motores ──────────────────────────────────────────────────

_UPLOAD_MOD = None
_BASE_MOD = None

# Tenta importar de dentro de src/ (quando rodado do projeto) e direto (quando rodado de src/)
for _attempt in ("motor_upload", "src.motor_upload"):
    try:
        import importlib as _il
        _UPLOAD_MOD = _il.import_module(_attempt)
        break
    except Exception:
        pass

if _UPLOAD_MOD is None:
    print("[IMPORT] motor_upload indisponível — casos de upload retornarão FALHOU")

for _attempt in ("motor_base", "src.motor_base"):
    try:
        import importlib as _il
        _BASE_MOD = _il.import_module(_attempt)
        break
    except Exception:
        pass

if _BASE_MOD is None:
    print("[IMPORT] motor_base indisponível — casos de base retornarão FALHOU")

# ── Helpers ─────────────────────────────────────────────────────────────────


def _excel_bytes(*sheets: tuple[str, list[dict]]) -> bytes:
    """Cria arquivo .xlsx em memória com openpyxl."""
    wb = Workbook()
    for i, (sheet_name, rows) in enumerate(sheets):
        ws = wb.active if i == 0 else wb.create_sheet()
        ws.title = sheet_name
        if rows:
            headers = list(rows[0].keys())
            ws.append(headers)
            for row in rows:
                ws.append([row.get(h) for h in headers])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _getattr_multi(obj: Any, *names: str) -> tuple[Any, Optional[str]]:
    """Tenta obj.name para cada name. Retorna (valor, nome_usado) ou (None, None)."""
    for n in names:
        if hasattr(obj, n):
            return getattr(obj, n), n
    return None, None


def _call_upload(
    file_bytes: bytes,
    file_name: str,
    aba: Optional[str] = None,
    divs: Optional[list[str]] = None,
) -> Any:
    """Chama motor_upload tentando assinaturas comuns. Documenta divergências em divs."""
    if divs is None:
        divs = []
    if _UPLOAD_MOD is None:
        divs.append("motor_upload não importado")
        return None

    fn = None
    for candidate in ("processar", "processar_upload", "process_file", "process_files", "upload", "ler_arquivo", "executar", "run"):
        fn = getattr(_UPLOAD_MOD, candidate, None)
        if fn is not None:
            break

    if fn is None:
        publics = [x for x in dir(_UPLOAD_MOD) if not x.startswith("_") and callable(getattr(_UPLOAD_MOD, x, None))]
        divs.append(f"Nenhuma função de upload encontrada. Funções públicas: {publics}")
        return None

    # Inspeciona parâmetros para adaptar a chamada
    try:
        sig = inspect.signature(fn)
        param_names = list(sig.parameters.keys())
    except (ValueError, TypeError):
        param_names = []

    # Estratégia 1: (bytes, nome)
    # Estratégia 2: (BytesIO, nome)
    # Estratégia 3: (Path, nome) — escreve em arquivo temporário
    # Estratégia 4: (nome, bytes)  — parâmetros invertidos
    # Estratégia 5: somente Path ou somente BytesIO como posicional

    from pathlib import Path as _Path
    import tempfile as _tmpfile

    def _try(payload, extra_args=()):
        try:
            args = (payload,) + extra_args
            if aba is not None:
                args = args + (aba,)
            result = fn(*args)
            return result, None
        except TypeError as e:
            return None, ("TypeError", str(e))
        except Exception as e:
            return None, (type(e).__name__, str(e))

    def _success(r: Any) -> bool:
        """Retorna True se o resultado indica sucesso."""
        if r is None:
            return False
        s = getattr(r, "success", None)
        return s is True

    # Tentativas com arquivo em memória (sem file_name como argumento)
    for payload in (file_bytes, io.BytesIO(file_bytes)):
        r, err = _try(payload)
        if _success(r):
            return r
        r2, err2 = _try(payload, (file_name,))
        if _success(r2):
            return r2

    # Tentativa com Path (arquivo temporário)
    with _tmpfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(file_bytes)
        tmp_path = _Path(tmp.name)
    try:
        # Primeiro tenta só com o path (sem file_name extra)
        r, err = _try(tmp_path)
        if _success(r):
            return r
        # Tenta renomear para o file_name esperado (motor pode usar o nome do arquivo)
        named_path = tmp_path.parent / file_name
        try:
            tmp_path.rename(named_path)
            r, err = _try(named_path)
            if _success(r):
                return r
            r, err = _try(named_path, (aba,)) if aba else (None, None)
            if _success(r):
                return r
        except Exception as e:
            err = (type(e).__name__, str(e))
        # Última tentativa: path + file_name
        for p in (tmp_path, named_path):
            r, err = _try(p, (file_name,))
            if r is not None:
                return r  # retorna mesmo que success=False para documentar
    except Exception as e:
        err = (type(e).__name__, str(e))
    finally:
        for p in [tmp_path, tmp_path.parent / file_name]:
            try:
                p.unlink(missing_ok=True)
            except Exception:
                pass

    if err:
        divs.append(f"Último erro em {fn.__name__}: {err[0]}: {err[1]}")
    divs.append(f"Nenhuma assinatura compatível para {fn.__name__}. Params detectados: {param_names}")
    return None


def _call_base(
    upload_result: Any,
    aba: Optional[str] = None,
    divs: Optional[list[str]] = None,
) -> Any:
    """Chama motor_base tentando assinaturas comuns. Documenta divergências em divs."""
    if divs is None:
        divs = []
    if _BASE_MOD is None:
        divs.append("motor_base não importado")
        return None

    fn = None
    for candidate in ("processar", "processar_base", "processar_motor_base", "process_base", "process", "analisar", "executar", "run", "base"):
        fn = getattr(_BASE_MOD, candidate, None)
        if fn is not None:
            break

    if fn is None:
        publics = [x for x in dir(_BASE_MOD) if not x.startswith("_") and callable(getattr(_BASE_MOD, x, None))]
        divs.append(f"Nenhuma função base encontrada. Funções públicas: {publics}")
        return None

    try:
        result = fn(upload_result) if aba is None else fn(upload_result, aba)
        return result
    except Exception as e:
        divs.append(f"Erro em {fn.__name__}: {type(e).__name__}: {e}")
        return None


def _check(condition: bool, msg: str, divs: list[str]) -> bool:
    if not condition:
        divs.append(msg)
    return condition


def _report(case_name: str, divs: list[str], ok: bool) -> bool:
    if divs:
        tag = "[INFO]"
        # Só imprime se houver divergências reais (não INFO)
        real_divs = [d for d in divs if not d.startswith("[INFO]")]
        info_divs = [d for d in divs if d.startswith("[INFO]")]
        if real_divs:
            print(f"    Divergências em {case_name}:")
            for d in real_divs:
                print(f"      • {d}")
        if info_divs:
            for d in info_divs:
                print(f"      {d}")
    return ok


def _upload_simple(
    rows: list[dict], sheet: str = "Sheet1", fname: str = "test.xlsx"
) -> tuple[Any, list[str]]:
    """Cria Excel em memória, faz upload, retorna (UploadResult, divs)."""
    divs: list[str] = []
    fb = _excel_bytes((sheet, rows))
    r = _call_upload(fb, fname, divs=divs)
    return r, divs


def _find_col(motor_result: Any, col_name: str) -> Optional[Any]:
    """Localiza ColumnMeta pelo nome no MotorResult."""
    cols, _ = _getattr_multi(motor_result, "colunas", "columns", "column_metas")
    if cols is None:
        return None
    for c in cols:
        name = getattr(c, "nome", getattr(c, "name", ""))
        if name == col_name:
            return c
    return None


def _col_tipo(col: Any) -> Optional[str]:
    """Lê tipo da ColumnMeta tentando campos alternativos."""
    return getattr(col, "tipo", getattr(col, "type", getattr(col, "inferred_type", None)))


# ── GRUPO A — Motor de Upload ───────────────────────────────────────────────


def caso_A01_excel_simples_1_aba() -> bool:
    """Excel simples, 1 aba, colunas mistas (int, str, float, date, bool)."""
    divs: list[str] = []
    try:
        rows = [
            {
                "id": i,
                "nome": f"Nome{i}",
                "valor": float(i) * 1.5,
                "data": date(2024, 1, 1),
                "ativo": i % 2 == 0,
            }
            for i in range(1, 21)
        ]
        fb = _excel_bytes(("Dados", rows))
        r = _call_upload(fb, "A01.xlsx", divs=divs)
        if r is None:
            return _report("A01", divs, False)

        ok = True
        ok &= _check(getattr(r, "success", False) is True, "success não é True", divs)

        abas, abas_f = _getattr_multi(r, "abas", "available_sheets")
        n_abas, n_abas_f = _getattr_multi(r, "n_abas")
        if n_abas is None and abas is not None:
            n_abas = len(abas)
            divs.append(f"[INFO] 'n_abas' ausente; derivado de '{abas_f}' -> {n_abas}")
        ok &= _check(n_abas == 1, f"n_abas esperado=1, encontrado={n_abas}", divs)
        ok &= _check(
            not getattr(r, "errors", []),
            f"errors não vazio: {getattr(r, 'errors', [])}",
            divs,
        )
        return _report("A01", divs, ok)
    except Exception:
        divs.append(traceback.format_exc())
        return _report("A01", divs, False)


def caso_A02_excel_multiplas_abas() -> bool:
    """Excel com 2 abas: 'Vendas' e 'Metas'."""
    divs: list[str] = []
    try:
        rows = [{"produto": f"P{i}", "qtd": i * 10} for i in range(1, 11)]
        fb = _excel_bytes(("Vendas", rows), ("Metas", rows))
        r = _call_upload(fb, "A02.xlsx", divs=divs)
        if r is None:
            return _report("A02", divs, False)

        ok = True
        ok &= _check(getattr(r, "success", False) is True, "success não é True", divs)

        abas, abas_f = _getattr_multi(r, "abas", "available_sheets")
        n_abas, _ = _getattr_multi(r, "n_abas")
        if n_abas is None and abas is not None:
            n_abas = len(abas)
            divs.append(f"[INFO] 'n_abas' ausente; derivado de '{abas_f}'")

        ok &= _check(n_abas == 2, f"n_abas esperado=2, encontrado={n_abas}", divs)
        if abas is not None:
            ok &= _check(
                set(abas) == {"Vendas", "Metas"},
                f"abas esperadas={{'Vendas','Metas'}}, encontradas={set(abas)}",
                divs,
            )
        return _report("A02", divs, ok)
    except Exception:
        divs.append(traceback.format_exc())
        return _report("A02", divs, False)


def caso_A03_coluna_sem_nome() -> bool:
    """Header com 1 coluna sem nome ->warning ou renomeação automática."""
    divs: list[str] = []
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append([None, "valor"])
        for i in range(5):
            ws.append([i, f"v{i}"])
        buf = io.BytesIO()
        wb.save(buf)
        fb = buf.getvalue()

        r = _call_upload(fb, "A03.xlsx", divs=divs)
        if r is None:
            return _report("A03", divs, False)

        warnings = getattr(r, "warnings", [])
        cols, _ = _getattr_multi(r, "columns", "colunas")
        has_warning = any(
            "coluna" in w.lower() or "sem nome" in w.lower() or "renomea" in w.lower()
            for w in warnings
        )
        renamed = cols is not None and any(
            "Coluna_" in str(getattr(c, "name", getattr(c, "nome", ""))) for c in cols
        )
        ok = has_warning or renamed
        if not ok:
            divs.append(
                f"Sem warning de coluna sem nome e sem renomeação. warnings={warnings}"
            )
        return _report("A03", divs, ok)
    except Exception:
        divs.append(traceback.format_exc())
        return _report("A03", divs, False)


def caso_A04_coluna_com_acento() -> bool:
    """Coluna 'Descrição' — encoding preservado corretamente."""
    divs: list[str] = []
    try:
        rows = [{"Descrição": f"item {i}", "valor": i} for i in range(5)]
        fb = _excel_bytes(("Dados", rows))
        r = _call_upload(fb, "A04.xlsx", divs=divs)
        if r is None:
            return _report("A04", divs, False)

        ok = True
        ok &= _check(getattr(r, "success", False) is True, "success não é True", divs)
        cols, _ = _getattr_multi(r, "columns", "colunas")
        if cols is not None:
            names = [getattr(c, "name", getattr(c, "nome", "")) for c in cols]
            ok &= _check(
                "Descrição" in names,
                f"Coluna 'Descrição' não preservada. Colunas: {names}",
                divs,
            )
        return _report("A04", divs, ok)
    except Exception:
        divs.append(traceback.format_exc())
        return _report("A04", divs, False)


def caso_A05_arquivo_vazio() -> bool:
    """Apenas header, sem linhas de dados ->success=False ou warning crítico."""
    divs: list[str] = []
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["id", "nome"])
        buf = io.BytesIO()
        wb.save(buf)
        fb = buf.getvalue()

        r = _call_upload(fb, "A05.xlsx", divs=divs)
        if r is None:
            return _report("A05", divs, False)

        success = getattr(r, "success", None)
        warnings = getattr(r, "warnings", [])
        errors = getattr(r, "errors", [])
        row_count, _ = _getattr_multi(r, "row_count", "n_linhas")

        ok = (not success) or bool(warnings) or bool(errors) or (row_count == 0)
        if not ok:
            divs.append(
                f"Arquivo vazio sem sinalização: success={success}, "
                f"warnings={warnings}, errors={errors}, row_count={row_count}"
            )
        return _report("A05", divs, ok)
    except Exception:
        divs.append(traceback.format_exc())
        return _report("A05", divs, False)


def caso_A06_arquivo_grande() -> bool:
    """100.001 linhas ->warning de volume alto ou error de limite."""
    divs: list[str] = []
    try:
        df = pd.DataFrame({"id": range(100_001), "valor": range(100_001)})
        buf = io.BytesIO()
        df.to_excel(buf, index=False, engine="openpyxl")
        fb = buf.getvalue()

        r = _call_upload(fb, "A06.xlsx", divs=divs)
        if r is None:
            return _report("A06", divs, False)

        warnings = getattr(r, "warnings", [])
        errors = getattr(r, "errors", [])
        ok = bool(warnings) or bool(errors)
        if not ok:
            divs.append(
                f"100.001 linhas sem warning/error de volume. "
                f"warnings={warnings}, errors={errors}"
            )
        return _report("A06", divs, ok)
    except Exception:
        divs.append(traceback.format_exc())
        return _report("A06", divs, False)


# ── GRUPO B — Motor Base ────────────────────────────────────────────────────


def caso_B01_tipo_numeric() -> bool:
    """Coluna numérica com 1 nulo ->tipo='numeric', null_count=1, null_pct=0.2."""
    divs: list[str] = []
    try:
        rows = [{"valor": v} for v in [1.0, 2.5, 3.0, None, 4.0]]
        ur, udivs = _upload_simple(rows, fname="B01.xlsx")
        divs.extend(udivs)
        if ur is None:
            return _report("B01", divs, False)

        mr = _call_base(ur, divs=divs)
        if mr is None:
            return _report("B01", divs, False)

        col = _find_col(mr, "valor")
        if col is None:
            divs.append("Coluna 'valor' não encontrada no MotorResult")
            return _report("B01", divs, False)

        ok = True
        tipo = _col_tipo(col)
        ok &= _check(tipo == "numeric", f"tipo esperado='numeric', encontrado='{tipo}'", divs)
        nc = getattr(col, "null_count", None)
        ok &= _check(nc == 1, f"null_count esperado=1, encontrado={nc}", divs)
        np_ = getattr(col, "null_pct", None)
        ok &= _check(np_ == 0.2, f"null_pct esperado=0.2, encontrado={np_}", divs)
        return _report("B01", divs, ok)
    except Exception:
        divs.append(traceback.format_exc())
        return _report("B01", divs, False)


def caso_B02_tipo_text_categorico() -> bool:
    """Coluna texto, baixa cardinalidade ->tipo='text', is_candidate_categorical=True."""
    divs: list[str] = []
    try:
        rows = [{"estado": v} for v in ["SP", "RJ", "MG", "SP", None]]
        ur, udivs = _upload_simple(rows, fname="B02.xlsx")
        divs.extend(udivs)
        if ur is None:
            return _report("B02", divs, False)

        mr = _call_base(ur, divs=divs)
        if mr is None:
            return _report("B02", divs, False)

        col = _find_col(mr, "estado")
        if col is None:
            divs.append("Coluna 'estado' não encontrada")
            return _report("B02", divs, False)

        ok = True
        tipo = _col_tipo(col)
        ok &= _check(tipo == "text", f"tipo esperado='text', encontrado='{tipo}'", divs)
        is_cat = getattr(col, "is_candidate_categorical", None)
        ok &= _check(is_cat is True, f"is_candidate_categorical esperado=True, encontrado={is_cat}", divs)
        return _report("B02", divs, ok)
    except Exception:
        divs.append(traceback.format_exc())
        return _report("B02", divs, False)


def caso_B03_tipo_date() -> bool:
    """Coluna com objetos date ->tipo='date'."""
    divs: list[str] = []
    try:
        rows = [{"data": v} for v in [date(2024, 1, 1), date(2024, 2, 1), date(2024, 3, 1)]]
        ur, udivs = _upload_simple(rows, fname="B03.xlsx")
        divs.extend(udivs)
        if ur is None:
            return _report("B03", divs, False)

        mr = _call_base(ur, divs=divs)
        if mr is None:
            return _report("B03", divs, False)

        col = _find_col(mr, "data")
        if col is None:
            divs.append("Coluna 'data' não encontrada")
            return _report("B03", divs, False)

        tipo = _col_tipo(col)
        ok = _check(tipo == "date", f"tipo esperado='date', encontrado='{tipo}'", divs)
        return _report("B03", divs, ok)
    except Exception:
        divs.append(traceback.format_exc())
        return _report("B03", divs, False)


def caso_B04_tipo_boolean() -> bool:
    """Coluna com booleans ->tipo='boolean'."""
    divs: list[str] = []
    try:
        rows = [{"ativo": v} for v in [True, False, True, True, False]]
        ur, udivs = _upload_simple(rows, fname="B04.xlsx")
        divs.extend(udivs)
        if ur is None:
            return _report("B04", divs, False)

        mr = _call_base(ur, divs=divs)
        if mr is None:
            return _report("B04", divs, False)

        col = _find_col(mr, "ativo")
        if col is None:
            divs.append("Coluna 'ativo' não encontrada")
            return _report("B04", divs, False)

        tipo = _col_tipo(col)
        ok = _check(tipo == "boolean", f"tipo esperado='boolean', encontrado='{tipo}'", divs)
        return _report("B04", divs, ok)
    except Exception:
        divs.append(traceback.format_exc())
        return _report("B04", divs, False)


def caso_B05_tipo_mixed() -> bool:
    """Coluna com int + str + float + bool ->tipo='mixed', warning disparado."""
    divs: list[str] = []
    try:
        rows = [{"campo": v} for v in [1, "texto", 2.5, None, True]]
        ur, udivs = _upload_simple(rows, fname="B05.xlsx")
        divs.extend(udivs)
        if ur is None:
            return _report("B05", divs, False)

        mr = _call_base(ur, divs=divs)
        if mr is None:
            return _report("B05", divs, False)

        col = _find_col(mr, "campo")
        if col is None:
            divs.append("Coluna 'campo' não encontrada")
            return _report("B05", divs, False)

        ok = True
        tipo = _col_tipo(col)
        ok &= _check(tipo == "mixed", f"tipo esperado='mixed', encontrado='{tipo}'", divs)
        warnings_mr = getattr(mr, "warnings", [])
        if not any("mist" in w.lower() or "campo" in w.lower() for w in warnings_mr):
            divs.append(f"[INFO] Sem warning de tipo misto. warnings={warnings_mr}")
        return _report("B05", divs, ok)
    except Exception:
        divs.append(traceback.format_exc())
        return _report("B05", divs, False)


def caso_B06_null_strings() -> bool:
    """Null-strings ('N/A', '-', '', 'NULL', 'none', ' ') ->null_count >= 6."""
    divs: list[str] = []
    try:
        rows = [{"texto": v} for v in ["N/A", "n/a", "-", "", " ", "NULL", "none", "valor"]]
        ur, udivs = _upload_simple(rows, fname="B06.xlsx")
        divs.extend(udivs)
        if ur is None:
            return _report("B06", divs, False)

        mr = _call_base(ur, divs=divs)
        if mr is None:
            return _report("B06", divs, False)

        col = _find_col(mr, "texto")
        if col is None:
            divs.append("Coluna 'texto' não encontrada")
            return _report("B06", divs, False)

        nc = getattr(col, "null_count", None)
        ok = _check(
            nc is not None and nc >= 6,
            f"null_count esperado>=6, encontrado={nc}",
            divs,
        )
        return _report("B06", divs, ok)
    except Exception:
        divs.append(traceback.format_exc())
        return _report("B06", divs, False)


def caso_B07_categorico_cardinalidade_baixa() -> bool:
    """50 linhas, 3 valores únicos ->is_candidate_categorical=True."""
    divs: list[str] = []
    try:
        vals = ["A", "B", "C"] * 16 + ["A", "B"]  # 50 total, 3 únicos
        rows = [{"cat": v} for v in vals]
        ur, udivs = _upload_simple(rows, fname="B07.xlsx")
        divs.extend(udivs)
        if ur is None:
            return _report("B07", divs, False)

        mr = _call_base(ur, divs=divs)
        if mr is None:
            return _report("B07", divs, False)

        col = _find_col(mr, "cat")
        if col is None:
            divs.append("Coluna 'cat' não encontrada")
            return _report("B07", divs, False)

        is_cat = getattr(col, "is_candidate_categorical", None)
        ok = _check(
            is_cat is True,
            f"is_candidate_categorical esperado=True (3 únicos/50 linhas), encontrado={is_cat}",
            divs,
        )
        return _report("B07", divs, ok)
    except Exception:
        divs.append(traceback.format_exc())
        return _report("B07", divs, False)


def caso_B08_nao_categorico_cardinalidade_alta() -> bool:
    """60 valores únicos em 60 linhas ->is_candidate_categorical=False (threshold=50)."""
    divs: list[str] = []
    try:
        rows = [{"id_str": f"ID_{i:04d}"} for i in range(60)]
        ur, udivs = _upload_simple(rows, fname="B08.xlsx")
        divs.extend(udivs)
        if ur is None:
            return _report("B08", divs, False)

        mr = _call_base(ur, divs=divs)
        if mr is None:
            return _report("B08", divs, False)

        col = _find_col(mr, "id_str")
        if col is None:
            divs.append("Coluna 'id_str' não encontrada")
            return _report("B08", divs, False)

        is_cat = getattr(col, "is_candidate_categorical", None)
        ok = _check(
            is_cat is False,
            f"is_candidate_categorical esperado=False (threshold=50), encontrado={is_cat}",
            divs,
        )
        return _report("B08", divs, ok)
    except Exception:
        divs.append(traceback.format_exc())
        return _report("B08", divs, False)


def caso_B09_selecao_aba_unica() -> bool:
    """Arquivo com 1 aba ->selecionada automaticamente, sem error."""
    divs: list[str] = []
    try:
        rows = [{"x": i} for i in range(5)]
        fb = _excel_bytes(("Única", rows))
        r = _call_upload(fb, "B09.xlsx", divs=divs)
        if r is None:
            return _report("B09", divs, False)

        mr = _call_base(r, divs=divs)
        if mr is None:
            return _report("B09", divs, False)

        ok = True
        ok &= _check(getattr(mr, "success", False) is True, "success não é True", divs)
        aba_proc, _ = _getattr_multi(mr, "aba_processada", "sheet_name", "aba")
        ok &= _check(aba_proc is not None, "aba_processada ausente no MotorResult", divs)
        ok &= _check(
            not getattr(mr, "errors", []),
            f"errors inesperados: {getattr(mr, 'errors', [])}",
            divs,
        )
        return _report("B09", divs, ok)
    except Exception:
        divs.append(traceback.format_exc())
        return _report("B09", divs, False)


def caso_B10_selecao_aba_multiplas_sem_especificacao() -> bool:
    """2 abas sem aba_selecionada ->warning de ambiguidade OU primeira aba selecionada."""
    divs: list[str] = []
    try:
        rows = [{"x": i} for i in range(5)]
        fb = _excel_bytes(("Aba1", rows), ("Aba2", rows))
        r = _call_upload(fb, "B10.xlsx", divs=divs)
        if r is None:
            return _report("B10", divs, False)

        mr = _call_base(r, divs=divs)
        if mr is None:
            return _report("B10", divs, False)

        warnings = getattr(mr, "warnings", [])
        aba_proc, _ = _getattr_multi(mr, "aba_processada", "sheet_name", "aba")
        has_warning = bool(warnings)
        selected_first = aba_proc == "Aba1"

        ok = has_warning or selected_first
        if not ok:
            divs.append(
                f"Sem warning de ambiguidade e aba ≠ primeira. "
                f"aba_processada={aba_proc}, warnings={warnings}"
            )
        if has_warning:
            divs.append(f"[INFO] Warning de ambiguidade: {warnings}")
        if selected_first:
            divs.append(f"[INFO] Primeira aba selecionada automaticamente: {aba_proc}")
        return _report("B10", divs, ok)
    except Exception:
        divs.append(traceback.format_exc())
        return _report("B10", divs, False)


def caso_B11_coluna_completamente_vazia() -> bool:
    """Coluna com todos os valores None ->tipo='empty', null_pct=1.0."""
    divs: list[str] = []
    try:
        rows = [{"vazia": None, "normal": i} for i in range(5)]
        ur, udivs = _upload_simple(rows, fname="B11.xlsx")
        divs.extend(udivs)
        if ur is None:
            return _report("B11", divs, False)

        mr = _call_base(ur, divs=divs)
        if mr is None:
            return _report("B11", divs, False)

        col = _find_col(mr, "vazia")
        if col is None:
            divs.append("Coluna 'vazia' não encontrada")
            return _report("B11", divs, False)

        ok = True
        tipo = _col_tipo(col)
        ok &= _check(tipo == "empty", f"tipo esperado='empty', encontrado='{tipo}'", divs)
        np_ = getattr(col, "null_pct", None)
        ok &= _check(np_ == 1.0, f"null_pct esperado=1.0, encontrado={np_}", divs)
        return _report("B11", divs, ok)
    except Exception:
        divs.append(traceback.format_exc())
        return _report("B11", divs, False)


def caso_B13_nao_regressao_volume() -> bool:
    """
    Bloco V-0c — Teste de não-regressão para o bug de truncamento.
    Constrói base com 60 linhas (acima do preview de 5) e verifica que
    motor_result.df contém TODAS as 60 linhas, não apenas as primeiras 5.
    """
    divs: list[str] = []
    try:
        rows = [
            {"id": i, "categoria": f"cat_{i % 3}", "valor": float(i) * 1.5}
            for i in range(1, 61)
        ]
        fb = _excel_bytes(("Dados", rows))
        ur = _call_upload(fb, "B13.xlsx", divs=divs)
        if ur is None:
            return _report("B13", divs, False)

        mr = _call_base(ur, divs=divs)
        if mr is None:
            return _report("B13", divs, False)

        ok = True
        ok &= _check(
            getattr(mr, "success", False) is True,
            f"success esperado=True, encontrado={getattr(mr, 'success', None)}",
            divs,
        )

        n_linhas = getattr(mr, "n_linhas", None)
        ok &= _check(
            n_linhas == 60,
            f"n_linhas esperado=60, encontrado={n_linhas}",
            divs,
        )

        df = getattr(mr, "df", None)
        if df is None:
            divs.append("motor_result.df ausente")
            return _report("B13", divs, False)

        # ASSERÇÃO CRÍTICA: df não pode estar truncado no preview (5 linhas)
        ok &= _check(
            len(df) == 60,
            f"len(df) esperado=60 (bug de truncamento), encontrado={len(df)}",
            divs,
        )

        if "id" in df.columns:
            id_min = df["id"].min()
            id_max = df["id"].max()
            ok &= _check(
                id_min == 1,
                f"id.min() esperado=1, encontrado={id_min}",
                divs,
            )
            ok &= _check(
                id_max == 60,
                f"id.max() esperado=60, encontrado={id_max}",
                divs,
            )
        else:
            divs.append("Coluna 'id' não encontrada em df.columns")
            ok = False

        return _report("B13", divs, ok)
    except Exception:
        divs.append(traceback.format_exc())
        return _report("B13", divs, False)


def caso_B12_motoresult_campos_obrigatorios() -> bool:
    """MotorResult deve conter todos os campos obrigatórios do contrato."""
    divs: list[str] = []
    try:
        rows = [{"id": i, "nome": f"N{i}"} for i in range(5)]
        ur, udivs = _upload_simple(rows, fname="B12.xlsx")
        divs.extend(udivs)
        if ur is None:
            return _report("B12", divs, False)

        mr = _call_base(ur, divs=divs)
        if mr is None:
            return _report("B12", divs, False)

        ok = True
        _SENTINEL = object()
        required_fields = {
            "success": bool,
            "errors": list,
            "warnings": list,
            "n_linhas": int,
            "n_colunas": int,
            "aba_processada": str,
            "colunas": list,
        }
        for field, expected_type in required_fields.items():
            val = getattr(mr, field, _SENTINEL)
            if val is _SENTINEL:
                divs.append(f"Campo obrigatório ausente: '{field}'")
                ok = False
            elif not isinstance(val, expected_type):
                divs.append(
                    f"Campo '{field}': tipo esperado={expected_type.__name__}, "
                    f"encontrado={type(val).__name__} (valor={val!r})"
                )
                ok = False

        # Verifica campos de ColumnMeta
        colunas, _ = _getattr_multi(mr, "colunas", "columns")
        if colunas and len(colunas) > 0:
            col = colunas[0]
            for cm_field in ["nome", "tipo", "null_count", "null_pct", "is_candidate_categorical", "n_unique"]:
                if not hasattr(col, cm_field):
                    alt = {"nome": "name", "tipo": "type"}.get(cm_field)
                    if alt and hasattr(col, alt):
                        divs.append(f"[INFO] ColumnMeta usa '{alt}' em vez de '{cm_field}'")
                    else:
                        divs.append(f"ColumnMeta: campo '{cm_field}' ausente")
                        ok = False
        return _report("B12", divs, ok)
    except Exception:
        divs.append(traceback.format_exc())
        return _report("B12", divs, False)


# ── GRUPO C — Integração Upload ->Base ─────────────────────────────────────


def caso_C01_pipeline_completo() -> bool:
    """Excel em memória ->UploadResult ->MotorResult com n_linhas e n_colunas corretos."""
    divs: list[str] = []
    try:
        N_LINHAS, N_COLUNAS = 15, 3
        rows = [{"id": i, "nome": f"N{i}", "valor": float(i)} for i in range(N_LINHAS)]
        fb = _excel_bytes(("Sheet1", rows))

        ur = _call_upload(fb, "C01.xlsx", divs=divs)
        if ur is None:
            return _report("C01", divs, False)

        mr = _call_base(ur, divs=divs)
        if mr is None:
            return _report("C01", divs, False)

        ok = True
        n_linhas = getattr(mr, "n_linhas", None)
        n_colunas = getattr(mr, "n_colunas", None)
        ok &= _check(
            n_linhas == N_LINHAS,
            f"n_linhas esperado={N_LINHAS}, encontrado={n_linhas}",
            divs,
        )
        ok &= _check(
            n_colunas == N_COLUNAS,
            f"n_colunas esperado={N_COLUNAS}, encontrado={n_colunas}",
            divs,
        )
        return _report("C01", divs, ok)
    except Exception:
        divs.append(traceback.format_exc())
        return _report("C01", divs, False)


def caso_C02_data_como_string_iso() -> bool:
    """Coluna 'data' com strings ISO ->tipo='date' ou 'text' (documentar comportamento)."""
    divs: list[str] = []
    try:
        rows = [
            {"data": v, "val": i}
            for i, v in enumerate(["2024-01-01", "2024-02-01", "2024-03-01"])
        ]
        ur, udivs = _upload_simple(rows, fname="C02.xlsx")
        divs.extend(udivs)
        if ur is None:
            return _report("C02", divs, False)

        mr = _call_base(ur, divs=divs)
        if mr is None:
            return _report("C02", divs, False)

        col = _find_col(mr, "data")
        if col is None:
            divs.append("Coluna 'data' não encontrada")
            return _report("C02", divs, False)

        tipo = _col_tipo(col)
        ok = tipo in ("date", "text")
        if not ok:
            divs.append(f"tipo inesperado para string ISO '2024-01-01': '{tipo}'")
        else:
            divs.append(f"[INFO] String ISO '2024-01-01' inferida como tipo='{tipo}'")
        return _report("C02", divs, ok)
    except Exception:
        divs.append(traceback.format_exc())
        return _report("C02", divs, False)


def caso_C03_pipeline_misto() -> bool:
    """5 colunas de tipos distintos, 30 linhas, ~10% nulos — tipos e null_count corretos."""
    divs: list[str] = []
    try:
        import random
        random.seed(42)

        def maybe_none(v: Any, pct: float = 0.1) -> Any:
            return None if random.random() < pct else v

        rows = [
            {
                "id": maybe_none(i),
                "nome": maybe_none(f"Nome{i}"),
                "valor": maybe_none(float(i) * 1.1),
                "data": maybe_none(date(2024, 1, 1)),
                "ativo": maybe_none(i % 2 == 0),
            }
            for i in range(30)
        ]
        fb = _excel_bytes(("Sheet1", rows))
        ur = _call_upload(fb, "C03.xlsx", divs=divs)
        if ur is None:
            return _report("C03", divs, False)

        mr = _call_base(ur, divs=divs)
        if mr is None:
            return _report("C03", divs, False)

        ok = True
        expected_tipos = {
            "id": "numeric",
            "nome": "text",
            "valor": "numeric",
            "data": "date",
            "ativo": "boolean",
        }
        for col_name, expected_tipo in expected_tipos.items():
            col = _find_col(mr, col_name)
            if col is None:
                divs.append(f"Coluna '{col_name}' não encontrada")
                ok = False
                continue
            tipo = _col_tipo(col)
            if tipo != expected_tipo:
                divs.append(
                    f"Coluna '{col_name}': tipo esperado='{expected_tipo}', encontrado='{tipo}'"
                )
                ok = False
            nc = getattr(col, "null_count", None)
            if nc is None:
                divs.append(f"Coluna '{col_name}': null_count ausente")
                ok = False
        return _report("C03", divs, ok)
    except Exception:
        divs.append(traceback.format_exc())
        return _report("C03", divs, False)


# ── Main ────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    casos = [
        caso_A01_excel_simples_1_aba,
        caso_A02_excel_multiplas_abas,
        caso_A03_coluna_sem_nome,
        caso_A04_coluna_com_acento,
        caso_A05_arquivo_vazio,
        caso_A06_arquivo_grande,
        caso_B01_tipo_numeric,
        caso_B02_tipo_text_categorico,
        caso_B03_tipo_date,
        caso_B04_tipo_boolean,
        caso_B05_tipo_mixed,
        caso_B06_null_strings,
        caso_B07_categorico_cardinalidade_baixa,
        caso_B08_nao_categorico_cardinalidade_alta,
        caso_B09_selecao_aba_unica,
        caso_B10_selecao_aba_multiplas_sem_especificacao,
        caso_B11_coluna_completamente_vazia,
        caso_B12_motoresult_campos_obrigatorios,
        caso_B13_nao_regressao_volume,
        caso_C01_pipeline_completo,
        caso_C02_data_como_string_iso,
        caso_C03_pipeline_misto,
    ]

    resultados: list[tuple[str, bool]] = []
    for caso in casos:
        ok = caso()
        resultados.append((caso.__name__, ok))

    print()
    print("=" * 60)
    print("RESUMO")
    print("=" * 60)
    for nome, ok in resultados:
        print(f"{'[OK]  ' if ok else '[FAIL]'} {nome}")

    total = sum(1 for _, ok in resultados if ok)
    print(f"\n{total}/{len(resultados)} casos passaram")
